import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data'))
from ahu import AHU
from ccu import CCU
from climate import CLIMATE
from fcu import FCU_COLS, FCU_ROWS, FCU_NOTES
from hex import (HEX_COLS, HEX_ROWS, HEX_NOTES, SRC_HEX, PAGE_HEX,
                 SRC_AL, PAGE_AL, SRC_ALTS, PAGE_ALTS, HEX_PROJECT, HEX_MODELS,
                 HEX_MODEL_FIELDS, HEX_DUTY, HEX_SYS_TEMPS, HEX_EXTRA_UNITS,
                 M10_SPEC, M10_UNIT, HEX_EXTRA_NOTES)
from hex import SRC_SYS as HEX_SRC_SYS, PAGE_SYS as HEX_PAGE_SYS
from pumps import (PUMP_COLS, PUMP_ROWS, PUMP_NOTES, SRC_PUMP, PAGE_PUMP,
                   SRC_SUB, PAGE_SUB, PUMP_TAGS, PUMP_DUTY, PUMP_SUBMITTAL,
                   PUMP_SUB_NOTES)
from pumps import SRC_SYS as PUMP_SRC_SYS, PAGE_SYS as PUMP_PAGE_SYS
from ef import EF_ROWS, EF_NOTES, SRC_EF
from pressurization import (PU_COLS, PU_ROWS, PU_NOTES, SRC_PU, PAGE_PU,
                            PU_ALT_TAG, PU_COMPONENTS, PU_EXTRA_NOTES)
from pressurization import SRC_SYS as PU_SRC_SYS, PAGE_SYS as PU_PAGE_SYS
from generator import GEN_COLS, GEN_ROWS, GEN_NOTES, SRC_GEN, PAGE_GEN
from cav import CAV_COLS, CAV_ROWS, CAV_NOTES, SRC_CAV
from vav import VAV_COLS, VAV_ROWS, VAV_NOTES, SRC_VAV, COVERS_OVERRIDE
from dx import (DX_COLS, DX_ROWS, DX_NOTES, SRC_DX, PAGE_DX, OD_STANDBY,
                DX_SCHEDULE, DX_SCHEDULE_NOTES, outdoor_count)
from dx import SRC_SYS as DX_SRC_SYS, PAGE_SYS as DX_PAGE_SYS
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from units_map import UNIT_MAP, resolve, convert
from ontology_map import classify, PREDICATE_SOURCE

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_FULL = sys.argv[1]
OUT_ONT  = sys.argv[2] if len(sys.argv) > 2 else None

def _sigfmt(f):
    """Show the factor the way an engineer would check it: 'x 1000' or '/ 3.6'."""
    if f >= 1:
        return "x %g" % f
    return "/ %g" % float("%.6g" % (1 / f))

HEAD = ["Equipment Tag","Model","Component","Property",
        "Value (as printed)","Unit (as printed)",
        "Value (Dar Cairo)","Unit (QUDT)","Conversion",
        "Ontology predicate","Scope",
        "Source File","Page","Note"]

WRAP_COLS = (5, 7, 9, 14)
NOTE_COL, PAGE_COL = 14, 13
SCOPE_COL, PRED_COL = 11, 10

def expand(row):
    """9-column source row -> 14-column row: Dar Cairo value/unit, then scope."""
    tag, model, comp, prop, val, unit, src, page, note = row
    pred, scope, pnote = classify(comp, prop)
    if comp == "Data quality":
        pred, scope, pnote = "", "", ""
    if pnote:
        note = (note + " " if note else "") + pnote
    qudt, factor, in_dc, cnote = resolve(unit, prop)
    if not qudt:
        return [tag, model, comp, prop, val, unit, "", "", cnote,
                pred, scope, src, page, note]
    new, ok = convert(val, factor)
    if not ok:
        return [tag, model, comp, prop, val, unit, "", "",
                "not converted - value is not numeric", pred, scope, src, page, note]
    bits = []
    if cnote: bits.append(cnote)
    if not in_dc: bits.append("unit not used in Dar Cairo")
    return [tag, model, comp, prop, val, unit, new, qudt, "; ".join(bits),
            pred, scope, src, page, note]

SCOPE_INDEX = {}

def _index(name, rows):
    for r in rows:
        e = expand(r)
        comp, prop, pred, sc = e[2], e[3], e[9], e[10]
        if comp == "Data quality":
            continue
        k = (comp, prop)
        d = SCOPE_INDEX.setdefault(k, {"pred": pred, "scope": sc, "n": 0, "sheets": set()})
        d["n"] += 1; d["sheets"].add(name)

HDR_FILL   = PatternFill("solid", fgColor="1F3864")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
BAND_FILL  = PatternFill("solid", fgColor="EEF2F8")
COMP_FONT  = Font(bold=True, size=10)
NOTE_FONT  = Font(italic=True, size=9, color="8B4000")
THIN = Side(style="thin", color="C9D2E0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------- AHU
OCTAVES = "63/125/250/500/1k/2k/4k/8k Hz"

FAN_F = [("model","Fan model",""),("impeller","Impeller type",""),
    ("air_flow_m3_s","Air flow","m3/s"),("esp_pa","External static pressure","Pa"),
    ("fsp_pa","Fan static pressure","Pa"),("speed_rpm","Speed","rpm"),
    ("abs_power_kw","Absorbed power","kW"),("swl_db","Sound power level ("+OCTAVES+")","dB"),
    ("volume_control","Volume control",""),("finish","Finish",""),
    ("shaft_guards","Shaft guards",""),("inspection_door","Inspection door",""),
    ("inlet_guards","Inlet guards",""),("ss_shaft","Stainless steel shaft",""),
    ("drain_plug","Drain plug",""),("spark_min","Spark minimising features",""),
    ("avm_type","AVM type",""),("pulley","Fan pulley",""),("note","Note","")]

MOT_F = [("rating_kw","Rating","kW"),("frame_type","Frame type",""),
    ("efficiency_class","Efficiency class",""),("fl_speed_rpm","Full load speed","rpm"),
    ("supply","Supply",""),("flc_a","Full load current","A"),
    ("winding","Winding type",""),("starting_current_a","Starting current","A"),
    ("starting_method","Starting method",""),("thermistor","Thermistor fitted",""),
    ("epoxy_paint","Epoxy paint finish",""),("spare_belts_sets","Spare drive belts","set(s)"),
    ("pulley","Motor pulley",""),("belts","Belts","n"),("belt_type","Belt type","")]

COIL_F = [("model","Model",""),("rows","Rows","n"),("fin_spacing_mm","Fin spacing","mm"),
    ("fins","Fins",""),("tubes","Tubes",""),("conn_supply_bsp","Connection supply","BSP"),
    ("conn_return_bsp","Connection return","BSP"),("water_kg_s","Water flow","kg/s"),
    ("water_pd_kpa","Water pressure drop","kPa"),("total_capacity_kw","Total capacity","kW"),
    ("ent_air_cdb","Entering air (DB)","degC"),("ent_air_rh","Entering air RH","%"),
    ("lvg_air_cdb","Leaving air (DB)","degC"),("lvg_air_rh","Leaving air RH","%"),
    ("ent_water_c","Entering water","degC"),("lvg_water_c","Leaving water","degC"),
    ("glycol_type","Glycol type",""),("glycol_pct","Glycol","%"),
    ("conn_type","Connection type",""),("moisture_eliminator","Moisture eliminator",""),
    ("binder_tapping_points","Binder tapping points",""),("drain_conn_bsp","Drain connection","BSP"),
    ("drain_pan","Drain pan",""),("casing","Casing",""),("note","Note","")]

AUX_F = [("label","Description",""),("air_flow_m3_h","Air flow","m3/h"),
    ("air_mass_flow_kg_s","Air mass flow","kg/s"),("water_flow_l_h","Water flow","l/h"),
    ("water_mass_flow_kg_s","Water mass flow","kg/s"),("ent_water_c","Entering water","degC"),
    ("lvg_water_c","Leaving water","degC"),("water_pd_kpa","Water pressure drop","kPa"),
    ("ent_air_c","Entering air","degC"),("ent_air_rh","Entering air RH","%"),
    ("lvg_air_c","Leaving air","degC"),("lvg_air_rh","Leaving air RH","%"),
    ("air_pd_pa","Air pressure drop","Pa"),("total_capacity_kw","Total capacity","kW"),
    ("tubes","Tubes","")]

ELEC_F = [("supply","Electrical supply",""),("total_capacity_kw","Total capacity","kW"),
    ("stages","Stages","n"),("ent_air_c","Entering air","degC"),
    ("lvg_air_c","Leaving air","degC"),("note","Note","")]

FIN_F = [("frames","Frames",""),("panels_outer","Panels outer skin",""),
    ("panels_inner","Panels inner skin",""),
    ("panels_inner_at_coil","Panels inner skin at coil section",""),
    ("unit_base","Unit base",""),("supply_fan_base","Supply fan base",""),
    ("extract_fan_base","Extract fan base","")]

def fmt(v):
    if isinstance(v, list): return " / ".join(str(x) for x in v)
    return v

def emit(rows, tag, model, comp, fields, data, src, page, note=""):
    for key, label, unit in fields:
        if key in data and data[key] not in (None, ""):
            rows.append([tag, model, comp, label, fmt(data[key]), unit, src, page, note])

def build_ahu(rows):
    for tag in sorted(AHU):
        d = AHU[tag]; src, page = d["src"], d["page"]
        model = d["supply_fan"]["model"] if "supply_fan" in d else ""
        rows.append([tag, "", "Drawing", "Drawing number", d["dwg"], "", src, page, ""])
        rows.append([tag, "", "Drawing", "Sheet", d["sheet"], "", src, page,
                     "AS BUILT general arrangement, Mercury Mena / Qatar Foundation"])
        for i, s in enumerate(d.get("silencers", []), 1):
            c = "Silencer (item %s)" % s["item"]
            rows.append([tag, "", c, "Silencer length", s["length_mm"], "mm", src, page, ""])
            rows.append([tag, "", c, "Insertion loss (%s)" % OCTAVES,
                         fmt(s["insertion_loss_db"]), "dB", src, page,
                         "Tissue faced complete with perforated plate"])
        for f in d.get("filters", []):
            c = "Polyseal filter (item %s) - %s" % (f["item"], f["stage"])
            rows.append([tag, "", c, "Media", f["media"], "", src, page, ""])
            if "arrestance" in f:
                rows.append([tag, "", c, "Arrestance", f["arrestance"], "", src, page, ""])
            if "efficiency" in f:
                rows.append([tag, "", c, "Efficiency", f["efficiency"], "", src, page, ""])
            for pn, dim, off in f["parts"]:
                rows.append([tag, "", c, "Part no %s (W x H x D mm)" % pn, dim, "mm", src, page,
                             "Quantity off: %s" % off])
            rows.append([tag, "", c, "Spare sets", f["spare_sets"], "set(s)", src, page, ""])
        if "cooling_coil" in d:
            emit(rows, tag, "", "Cooling coil", COIL_F, d["cooling_coil"], src, page)
        if "aux_cooling_coil" in d:
            emit(rows, tag, "", "External auxiliary cooling coil", AUX_F, d["aux_cooling_coil"], src, page)
        if "electric_coil" in d:
            emit(rows, tag, "", "Electric coil", ELEC_F, d["electric_coil"], src, page)
        if "supply_fan" in d:
            emit(rows, tag, "", "Supply fan", FAN_F, d["supply_fan"], src, page)
        if "supply_motor" in d:
            emit(rows, tag, "", "Supply fan drive motor", MOT_F, d["supply_motor"], src, page)
        if "return_fan" in d:
            emit(rows, tag, "", "Return fan", FAN_F, d["return_fan"], src, page)
        if "return_motor" in d:
            emit(rows, tag, "", "Return fan drive motor", MOT_F, d["return_motor"], src, page)
        if "flexible_connection" in d:
            rows.append([tag, "", "Flexible connection", "Material", d["flexible_connection"],
                         "", src, page, ""])
        if "finish" in d:
            emit(rows, tag, "", "Finish", FIN_F, d["finish"], src, page)
        if "insulation" in d:
            rows.append([tag, "", "Insulation", "Panels", d["insulation"], "", src, page, ""])
        if "drawing_anomaly" in d:
            rows.append([tag, "", "Data quality", "Drawing anomaly", d["drawing_anomaly"],
                         "", src, page, "Confirm with the supplier before use"])

# ---------------------------------------------------------------- CCU
CCU_LABELS = {
 "design": ("Design condition", {
   "inlet_air_c":("Unit inlet air temperature","degC"),
   "inlet_air_rh":("Unit inlet air relative humidity","%"),
   "airflow_m3_h":("Unit airflow","m3/h"), "esp_pa":("ESP","Pa"),
   "sea_level_m":("Sea level","m"), "refrigerant":("Refrigerant",""),
   "fluid":("Fluid",""), "inlet_fluid_c":("Inlet fluid temperature","degC"),
   "outlet_fluid_c":("Outlet fluid temperature","degC"),
   "unit_fluid_flow_l_s":("Unit fluid flow","l/s"),
   "power_supply":("Unit power supply","")}),
 "performance": ("Unit performance", {
   "total_cooling_kw":("Total cooling capacity","kW"),
   "sensible_cooling_kw":("Sensible cooling capacity","kW"), "shr":("SHR",""),
   "off_coil_air_c":("Off coil air temperature","degC"),
   "off_coil_rh":("Off coil air relative humidity","%"),
   "room_spl_2m_dba":("Room SPL (@ 2 m, f.f)","dB(A)"),
   "condensing_temp_c":("Condensing temperature","degC"),
   "unit_power_input_kw":("Unit power input","kW"), "unit_eer":("Unit EER",""),
   "system_power_input_kw":("System power input","kW"), "system_eer":("System EER",""),
   "filter_class_en779":("Internal filter class (EN779 std)",""),
   "width_mm":("Width","mm"), "depth_mm":("Depth","mm"), "height_mm":("Height","mm"),
   "weight_kg":("Weight","kg")}),
 "cw_coil": ("CW coil", {
   "qty":("Quantity","n"), "unit_fluid_flow_l_s":("Unit fluid flow","l/s"),
   "unit_fluid_side_pd_kpa":("Unit fluid side pressure drop","kPa"),
   "coil_plus_connections_pd_kpa":("Fluid pressure drop coil + connections","kPa"),
   "valve_pd_kpa":("Valve pressure drop","kPa")}),
 "fans": ("Fans", {
   "qty":("Quantity","n"), "type":("Type",""), "power_supply":("Power supply",""),
   "power_input_kw":("Power input","kW"), "operating_a":("Operating ampere","A"),
   "fla_a":("Full load ampere","A"), "lra_a":("Locked rotor amp.","A"),
   "fan_input_voltage_v":("Fan input voltage","V")}),
 "compressors": ("Compressors", {
   "qty":("Quantity","n"), "power_supply":("Power supply",""),
   "power_input_kw":("Power input","kW"), "cop":("Compressors COP",""),
   "operating_a":("Operating ampere","A"), "fla_a":("Full load ampere","A"),
   "lra_a":("Locked rotor amp.","A")}),
 "condenser": ("Condenser", {
   "model":("Condenser model",""), "version":("Version",""),
   "air_discharge":("Air discharge",""), "power_supply":("Power supply",""),
   "variex":("Variex",""), "heat_load_kw":("Heat load","kW"),
   "outdoor_air_c":("Outdoor air temperature","degC"),
   "airflow_max_m3_h":("Condenser airflow (@ max speed)","m3/h"),
   "actual_airflow_m3_h":("Condenser actual airflow","m3/h"),
   "esp_pa":("Condenser ESP (@ max speed)","Pa"),
   "max_outdoor_spl_5m_dba":("Max outdoor SPL (@ 5 m, f.f.)","dB(A)"),
   "actual_outdoor_spl_5m_dba":("Actual outdoor SPL (@ 5 m, f.f.)","dB(A)"),
   "power_input_kw":("Power input","kW"), "fla_a":("Full load ampere","A"),
   "lra_a":("Locked rotor amp.","A"), "width_mm":("Width","mm"),
   "depth_mm":("Depth","mm"), "height_mm":("Height","mm"), "weight_kg":("Weight","kg"),
   "note":("Note","")}),
 "electric_reheat": ("Option - electrical re-heating", {
   "label":("Sheet heading",""), "max_capacity_kw":("Max re-heating capacity","kW"),
   "fla_a":("FLA","A"), "inlet_air_c":("Inlet air temperature","degC"),
   "inlet_air_rh":("Inlet air relative humidity","%"),
   "outlet_air_c":("Outlet air temperature","degC"),
   "outlet_air_rh":("Outlet air relative humidity","%")}),
 "humidifier": ("Option - humidifier", {
   "qty":("Quantity","n"), "max_steam_kg_h":("Max capacity steam","kg/h"),
   "min_steam_kg_h":("Min capacity steam","kg/h"), "type":("Type of humidifier",""),
   "power_supply":("Power supply",""), "nominal_power_input_kw":("Nominal power input","kW"),
   "max_absorption_a":("Max absorption current","A")}),
}

def build_ccu(rows):
    for tag, d in CCU.items():
        src = d["src"]; pp = d["pages_perf"]; po = d["pages_options"]
        model = d["model"]
        rows.append([tag, model, "Identification", "Manufacturer", d["manufacturer"], "", src, pp, ""])
        rows.append([tag, model, "Identification", "Model", model, "", src, pp, ""])
        rows.append([tag, model, "Identification", "System type", d["system_type"], "", src, pp, ""])
        rows.append([tag, model, "Identification", "Issued by", d["issued_by"], "", src, pp, ""])
        rows.append([tag, model, "Identification", "Selection date", d["selection_date"], "", src, pp, ""])
        rows.append([tag, model, "Identification", "Selection software release",
                     d["software_rel"], "", src, pp, ""])
        if "pages_dimension_drawing" in d:
            rows.append([tag, model, "Identification", "Dimension drawing pages",
                         d["pages_dimension_drawing"], "", src, d["pages_dimension_drawing"], ""])
        for key, (comp, labels) in CCU_LABELS.items():
            if key not in d: continue
            page = po if key in ("electric_reheat", "humidifier") else pp
            blk = d[key]
            for f, val in blk.items():
                if f not in labels: continue
                label, unit = labels[f]
                rows.append([tag, model, comp, label, val, unit, src, page, ""])
        if "compliance" in d:
            rows.append([tag, model, "Compliance", "Declared performance / directives",
                         d["compliance"], "", src, pp, ""])
        if "note" in d:
            rows.append([tag, model, "Note", "Sheet note", d["note"], "", src, pp, ""])
        if "source_anomaly" in d:
            rows.append([tag, model, "Data quality", "Source anomaly", d["source_anomaly"],
                         "", src, pp, "Confirm with the supplier before use"])

# ---------------------------------------------------------------- Climate
def build_climate(rows):
    for tag, d in CLIMATE.items():
        src, page = d["src"], d["page"]
        model = d["model"]
        rows.append([tag, model, "Identification", "Manufacturer", d["manufacturer"], "", src, page, ""])
        rows.append([tag, model, "Identification", "Model", model, "", src, page, ""])
        rows.append([tag, model, "Identification", "Equipment type", d["equipment_type"], "", src, page, ""])
        if "doc_status" in d:
            rows.append([tag, model, "Identification", "Document status", d["doc_status"], "", src, page, ""])
        for block, comp in (("specifications","Specifications"), ("technical_data","Technical data")):
            for f, v in d.get(block, {}).items():
                unit = ""
                lbl = f.replace("_"," ").capitalize()
                for suf, u in (("_kw","kW"),("_kg","kg"),("_mm","mm"),("_a","A"),("_w","W"),
                               ("_rpm","rpm"),("_v_dc","V DC"),("_dba","dB(A)"),
                               ("_mbar","mbar"),("_hours","hours"),("_lb","lb"),("_c","degC")):
                    if f.endswith(suf):
                        unit = u; lbl = f[:-len(suf)].replace("_"," ").capitalize(); break
                rows.append([tag, model, comp, lbl, v, unit, src, page, ""])
        for f, v in d.get("dimensions_mm", {}).items():
            # "mounting" is a screw specification (8 x M4 x 6), not a length
            unit = "" if f == "mounting" else "mm"
            rows.append([tag, model, "Dimensions", f.replace("_"," ").capitalize(),
                         v, unit, src, page, ""])
        for key, comp in (("installation_requirements","Installation requirement"),
                          ("systems_included","System included"),
                          ("features","Feature"),
                          ("stages",None)):
            v = d.get(key)
            if isinstance(v, list):
                for i, item in enumerate(v, 1):
                    rows.append([tag, model, comp, "%s %d" % (comp, i), item, "", src, page, ""])
            elif isinstance(v, dict):
                for f, item in v.items():
                    rows.append([tag, model, "Filtration stages",
                                 f.capitalize(), item, "", src, page, ""])
        for key, comp in (("installation","Installation"), ("control_wiring","Control wiring")):
            blk = d.get(key)
            if not blk: continue
            pg = blk.get("page", page)
            for f, v in blk.items():
                if f == "page": continue
                rows.append([tag, model, comp, f.replace("_"," ").capitalize(), v, "", src, pg, ""])
        for f, comp in (("designed_for","Application"), ("options","Options"),
                        ("warranty","Warranty"), ("note","Note")):
            if f in d:
                rows.append([tag, model, comp, comp, d[f], "", src, page, ""])
        if "dimension_drawings" in d:
            dd = d["dimension_drawings"]
            for f, v in dd.items():
                if f == "pages": continue
                unit, lbl = "", f.replace("_"," ").capitalize()
                if f.endswith("_mm"):
                    unit, lbl = "mm", f[:-3].replace("_"," ").capitalize()
                rows.append([tag, model, "Dimension drawing", lbl, v, unit, src, dd["pages"], ""])
        if "source_conflict" in d:
            rows.append([tag, model, "Data quality", "Source conflict", d["source_conflict"],
                         "", src, page, "Confirm with the manufacturer before use"])

# ---------------------------------------------------------------- FCU
FCU_LABEL = {
 "total_cooling_kw":("Capacity","Total capacity (cooling)","kW"),
 "sensible_cooling_kw":("Capacity","Sensible capacity (cooling)","kW"),
 "condensed_water_g_h":("Capacity","Condensed water","g/h"),
 "rows_n":("Coil","Rows","n"),
 "air_in_dbt_c":("Air","Inlet DBT","degC"),
 "air_in_wbt_c":("Air","Inlet WBT","degC"),
 "air_in_rh_pct":("Air","Inlet RH","%"),
 "air_out_dbt_c":("Air","Outlet DBT","degC"),
 "air_out_wbt_c":("Air","Outlet WBT","degC"),
 "air_out_rh_pct":("Air","Outlet RH","%"),
 "air_flow_m3_h":("Air","Flow rate","m3/h"),
 "fan_speed":("Air","Fan speed setting",""),
 "air_velocity_m_s":("Air","Air velocity","m/s"),
 "air_pressure_drop_pa":("Air","Air pressure drop","Pa"),
 "fluid_flow_l_h":("Fluid","Flow rate","l/h"),
 "fluid_pd_kpa":("Fluid","Pressure drop","kPa"),
 "fluid_in_c":("Fluid","Inlet temperature","degC"),
 "fluid_out_c":("Fluid","Outlet temperature","degC"),
 "fluid":("Fluid","Fluid",""),
 "max_width_mm":("Dimensions and weight","Max width","mm"),
 "max_height_mm":("Dimensions and weight","Max height","mm"),
 "max_thickness_mm":("Dimensions and weight","Max thickness","mm"),
 "weight_kg":("Dimensions and weight","Weight","kg"),
 "power_supply":("Other data","Power supply","V-ph-Hz"),
 "max_absorbed_power_w":("Other data","Max. absorbed power","W"),
 "max_absorbed_current_a":("Other data","Max. absorbed current","A"),
 "sound_pressure_dba":("Other data","Sound pressure level","dB(A)"),
 "sound_power_dba":("Other data","Sound power level","dB(A)"),
 "static_pressure_pa":("Other data","Static pressure","Pa"),
}
SRC_FCU = "FCU_Manual.pdf"

def build_fcu(rows):
    idx = {c: i for i, c in enumerate(FCU_COLS)}
    for r in FCU_ROWS:
        tag = r[idx["position"]]; page = r[idx["page"]]; model = r[idx["model"]]
        rows.append([tag, model, "Identification", "Manufacturer",
                     "Euroclima AG/Spa-Ges.m.b.H", "", SRC_FCU, page, ""])
        rows.append([tag, model, "Identification", "Offer", r[idx["offer"]], "", SRC_FCU, page, ""])
        rows.append([tag, model, "Identification", "Range", r[idx["range"]], "", SRC_FCU, page, ""])
        rows.append([tag, model, "Identification", "Version", r[idx["version"]], "", SRC_FCU, page, ""])
        rows.append([tag, model, "Identification", "Model", model, "", SRC_FCU, page, ""])
        for c, (comp, label, unit) in FCU_LABEL.items():
            v = r[idx[c]]
            if v is None: continue
            note = ""
            if c == "air_pressure_drop_pa" and tag != "FCU/B/01":
                note = ("Row printed as 'Perdita di carico aria [degC]' on this sheet - "
                        "untranslated label with the wrong unit; the value is Pa.")
            rows.append([tag, model, comp, label, v, unit, SRC_FCU, page, note])
        rows.append([tag, model, "Capacity", "Heating capacity", "-", "kW", SRC_FCU, page,
                     "Heating column blank on the sheet; cooling-only selection."])
        if tag in FCU_NOTES:
            rows.append([tag, model, "Data quality", "Sheet note", FCU_NOTES[tag], "", SRC_FCU, page,
                         "Confirm with the supplier before use"])

# ---------------------------------------------------------------- HEX
def build_hex(rows):
    idx = {c: i for i, c in enumerate(HEX_COLS)}
    for r in HEX_ROWS:
        tag = r[idx["unit_ref"]]
        def add(comp, prop, key, unit, note=""):
            rows.append([tag, "", comp, prop, r[idx[key]], unit, SRC_HEX, PAGE_HEX, note])
        rows.append([tag, "", "Identification", "Make", r[idx["make"]], "", SRC_HEX, PAGE_HEX, ""])
        add("Identification", "Type", "type", "")
        add("Identification", "Location", "location", "")
        add("Identification", "Quantity", "qty", "n")
        add("Dimensions and weight", "Unit dimension (L x W x H)", "dim_lxwxh_mm", "mm")
        add("Dimensions and weight", "Operating weight", "operating_weight_kg", "kg")
        add("CHW flow rate", "Cold side", "chw_flow_cold_l_s", "l/s")
        add("CHW flow rate", "Hot side", "chw_flow_hot_l_s", "l/s")
    all_tags = [r[idx["unit_ref"]] for r in HEX_ROWS] + HEX_EXTRA_UNITS
    for tag in all_tags:
        duty, model = HEX_DUTY[tag]
        rows.append([tag, model, "Identification", "Duty or standby", duty, "",
                     HEX_SRC_SYS, HEX_PAGE_SYS, ""])
        rows.append([tag, model, "Identification", "Model", model, "",
                     HEX_SRC_SYS, HEX_PAGE_SYS, ""])
        for prop, val in HEX_SYS_TEMPS:
            rows.append([tag, model, "Design condition", prop, val, "",
                         HEX_SRC_SYS, HEX_PAGE_SYS,
                         "SYSTEM DETAILS states one figure across all five rows"])
        for comp, prop, val, note in HEX_PROJECT:
            rows.append([tag, model, comp, prop, val, "", SRC_AL, PAGE_AL, note])
        spec = HEX_MODELS[model]
        for key, label, unit in HEX_MODEL_FIELDS:
            rows.append([tag, model, "Construction", label, spec[key], unit,
                         SRC_AL, PAGE_AL, "Per exchanger"])
    # M10-MFM thermal specification - PHX/B/05 only
    for prop, val, unit in M10_UNIT:
        rows.append(["PHX/B/05", "M10-MFM", "Thermal specification", prop, val, unit,
                     SRC_ALTS, PAGE_ALTS, ""])
    for prop, unit, hot, cold in M10_SPEC:
        rows.append(["PHX/B/05", "M10-MFM", "Thermal specification - hot side", prop, hot, unit,
                     SRC_ALTS, PAGE_ALTS, ""])
        rows.append(["PHX/B/05", "M10-MFM", "Thermal specification - cold side", prop, cold, unit,
                     SRC_ALTS, PAGE_ALTS, ""])
    for tag in [r[0] for r in HEX_ROWS][:1]:
        for prop, text in HEX_NOTES:
            rows.append([tag, "", "Data quality", prop, text, "", SRC_HEX, PAGE_HEX,
                         "Applies to every unit on this schedule"])
        for prop, text in HEX_EXTRA_NOTES:
            rows.append([tag, "", "Data quality", prop, text, "", SRC_AL, PAGE_AL,
                         "Applies to the heat exchanger set"])

# ---------------------------------------------------------------- pumps
def build_pumps(rows):
    idx = {c: i for i, c in enumerate(PUMP_COLS)}
    for r in PUMP_ROWS:
        tag = r[idx["unit_ref"]]; model = r[idx["model"]]
        def add(comp, prop, key, unit, note=""):
            rows.append([tag, model, comp, prop, r[idx[key]], unit, SRC_PUMP, PAGE_PUMP, note])
        rows.append([tag, model, "Identification", "Make", r[idx["make"]], "",
                     SRC_PUMP, PAGE_PUMP, ""])
        add("Identification", "Type", "type", "")
        add("Identification", "Location", "location", "")
        add("Identification", "Model", "model", "",
            "Designation ends in 55KW; the schedule states no motor rating of its own")
        add("Identification", "Quantity", "qty", "n")
        add("Dimensions and weight", "Unit dimension (L x W x H)", "dim_lxwxh_mm", "mm")
        add("Dimensions and weight", "Weight", "weight_kg", "kg")
        add("Performance", "CHW flow rate", "chw_flow_l_s", "l/s")
    model = PUMP_ROWS[0][idx["model"]]
    for tag in PUMP_TAGS:
        duty, alt = PUMP_DUTY[tag]
        rows.append([tag, model, "Identification", "Duty or standby", duty, "",
                     PUMP_SRC_SYS, PUMP_PAGE_SYS, ""])
        rows.append([tag, model, "Identification", "Alternate reference", alt, "",
                     PUMP_SRC_SYS, PUMP_PAGE_SYS,
                     "Dash form used by SYSTEM DETAILS; the drawing schedule writes slashes"])
        for comp, prop, val, unit, note in PUMP_SUBMITTAL:
            rows.append([tag, model, comp, prop, val, unit, SRC_SUB, PAGE_SUB, note])
    for tag in [r[0] for r in PUMP_ROWS][:1]:
        for prop, text in PUMP_NOTES:
            rows.append([tag, "", "Data quality", prop, text, "", SRC_PUMP, PAGE_PUMP,
                         "Applies to every unit on this schedule"])
        for prop, text in PUMP_SUB_NOTES:
            rows.append([tag, "", "Data quality", prop, text, "", SRC_SUB, PAGE_SUB,
                         "Applies to every unit on this submittal"])

# ---------------------------------------------------------------- exhaust fans
_FLOW = re.compile(r"^\s*([\d.]+)\s*[lL]\s*/\s*[sS]\s*$")

def build_ef(rows):
    for tag, row, model, make, flow in EF_ROWS:
        page = "Sheet1 r%d" % row
        if make:
            rows.append([tag, model or "", "Identification", "Manufacturer", make, "",
                         SRC_EF, page, ""])
        if model:
            rows.append([tag, model, "Identification", "Model", model, "", SRC_EF, page, ""])
        if not model and not make:
            rows.append([tag, "", "Data quality", "Model and manufacturer",
                         "not stated in the source sheet", "", SRC_EF, page,
                         "Left out rather than filled with a typical value"])
        if flow:
            m = _FLOW.match(flow)
            if m:
                rows.append([tag, model or "", "Air", "Air flow", float(m.group(1)), "l/s",
                             SRC_EF, page,
                             "" if flow.strip() == "%s L/S" % m.group(1)
                             else "source writes %r" % flow])
            else:
                rows.append([tag, model or "", "Air", "Air flow", flow, "", SRC_EF, page,
                             "air-flow value could not be parsed"])
        else:
            rows.append([tag, model or "", "Data quality", "Air flow",
                         "not stated in the source sheet", "", SRC_EF, page,
                         "Left out rather than filled with a typical value"])
    first = EF_ROWS[0][0]
    for prop, text in EF_NOTES:
        rows.append([first, "", "Data quality", prop, text, "", SRC_EF, "Sheet1",
                     "Applies to the whole schedule"])

# ---------------------------------------------------------------- air terminals
_RANGE = re.compile(r"^(?P<prefix>.*/)(?P<a>\d+)\s*(?P<sep>TO|&|-)\s*(?P<b>\d+)$", re.I)

def expand_ref(ref, qty):
    """Expand a schedule reference into the individual box tags it covers.

    Returns (tags, ok, reason). ok is False whenever the reference cannot be
    expanded with confidence - a range whose box count disagrees with the stated
    QTY, a descending range, or a single reference scheduled at QTY > 1. Those
    become Data quality rows instead of an expansion nobody checked.
    """
    ref = ref.strip()
    m = _RANGE.match(ref)
    if not m:
        if qty == 1:
            return ([ref], True, "")
        return ([ref], False,
                "single reference but QTY is %d - the schedule does not name the other boxes" % qty)
    prefix, a, sep, b = m.group("prefix"), m.group("a"), m.group("sep"), m.group("b")
    ai, bi, w = int(a), int(b), len(a)
    if sep == "&":
        nums = [ai, bi]
    elif bi < ai:
        return ([ref], False, "range end %s is lower than its start %s" % (b, a))
    else:
        nums = list(range(ai, bi + 1))
    tags = ["%s%0*d" % (prefix, w, n) for n in nums]
    if len(tags) != qty:
        return (tags, False,
                "reference spans %d boxes but QTY is %d" % (len(tags), qty))
    return (tags, True, "")

def _terminal(rows, data_rows, cols, src, notes, kind, overrides=None):
    """One row per BOX, not per schedule line.

    A schedule line covering a range is expanded so every box carries the line's
    air flow, heating capacity, model and make in its own right. The printed
    reference is kept on each box as 'Scheduled as' so the row still traces back
    to the line it came from. A range that cannot be expanded with confidence
    keeps the printed reference as its tag and raises a Data quality row.
    """
    idx = {c: i for i, c in enumerate(cols)}
    overrides = overrides or {}
    for r in data_rows:
        ref = r[idx["unit_ref"]]; page = r[idx["page"]]
        model = r[idx["model"]] or ""
        rnote = r[idx["row_note"]]; qty = r[idx["qty"]]
        key = (ref, page)
        if key in overrides:
            boxes, ok, why = overrides[key], True, ""
        else:
            boxes, ok, why = expand_ref(ref, qty)
        for tag in boxes:
            def add(comp, prop, val, unit, note=""):
                if val is None or val == "":
                    return
                rows.append([tag, model, comp, prop, val, unit, src, page, note])
            add("Identification", "Make", r[idx["make"]], "")
            add("Identification", "Model", r[idx["model"]], "")
            add("Air", "Air flow (per box)", r[idx["air_flow_l_s"]], "l/s")
            if "heating_kw" in idx:
                add("Heating", "Heating capacity", r[idx["heating_kw"]], "kW")
            if ok and (len(boxes) > 1 or tag != ref):
                add("Provenance", "Scheduled as", ref, "",
                    "One schedule line of QTY %s, split to one row per box" % qty)
        if not ok:
            rows.append([ref, model, "Data quality", "Reference and quantity disagree",
                         why, "", src, page,
                         "Not split into individual boxes - confirm before modelling"])
        if rnote:
            rows.append([boxes[0] if ok else ref, model, "Data quality", "Row note",
                         rnote, "", src, page, ""])
    first_boxes, ok0, _ = expand_ref(data_rows[0][0], data_rows[0][idx["qty"]])
    first = first_boxes[0]
    for prop, text in notes:
        rows.append([first, "", "Data quality", prop, text, "", src, "all %s schedules" % kind,
                     "Applies to the whole %s set" % kind])

def build_cav(rows):
    _terminal(rows, CAV_ROWS, CAV_COLS, SRC_CAV, CAV_NOTES, "CAV")

def build_vav(rows):
    _terminal(rows, VAV_ROWS, VAV_COLS, SRC_VAV, VAV_NOTES, "VAV", COVERS_OVERRIDE)

# ---------------------------------------------------------------- pressurization
def build_pu(rows):
    idx = {c: i for i, c in enumerate(PU_COLS)}
    for r in PU_ROWS:
        tag = r[idx["unit_ref"]]; model = r[idx["model"]]
        def add(comp, prop, key, unit, note=""):
            rows.append([tag, model, comp, prop, r[idx[key]], unit, SRC_PU, PAGE_PU, note])
        add("Identification", "Make", "make", "")
        add("Identification", "Model", "model", "")
        add("Identification", "Location", "location", "")
        add("Identification", "Quantity", "qty", "n")
        add("Performance", "System volume (as printed)", "system_volume", "",
            "Column is headed SYSTEM VOLUME but holds a pressure and a tank size")
    tag = PU_ROWS[0][0]
    rows.append([tag, "", "Identification", "Alternate reference", PU_ALT_TAG, "",
                 PU_SRC_SYS, PU_PAGE_SYS,
                 "Tag used by SYSTEM DETAILS; the drawing schedule writes PU/B/01"])
    for comp, prop, val, unit, note in PU_COMPONENTS:
        rows.append([tag, "", comp, prop, val, unit, PU_SRC_SYS, PU_PAGE_SYS, note])
    for prop, text in PU_NOTES:
        rows.append([tag, "", "Data quality", prop, text, "", SRC_PU, PAGE_PU, ""])
    for prop, text in PU_EXTRA_NOTES:
        rows.append([tag, "", "Data quality", prop, text, "", PU_SRC_SYS, PU_PAGE_SYS, ""])

# ---------------------------------------------------------------- DX splits
def build_dx(rows):
    idx = {c: i for i, c in enumerate(DX_COLS)}
    for r in DX_ROWS:
        tag = r[idx["indoor_ref"]]; od = r[idx["outdoor_ref"]]
        sched = DX_SCHEDULE.get(tag)
        model = sched[1] if sched else ""
        def add(comp, prop, val, unit, note=""):
            rows.append([tag, model, comp, prop, val, unit, SRC_DX, PAGE_DX, note])
        add("Identification", "Level", r[idx["level"]], "")
        add("Identification", "Duty or standby",
            "STANDBY" if r[idx["standby"]] else "DUTY", "",
            "" if r[idx["standby"]] else "No (ST.BY) marking against this unit")
        add("Served space", "Room served", r[idx["room"]], "")
        add("Served space", "Room design temperature", r[idx["room_temp_c"]], "degC")
        add("Served space", "Room design relative humidity", r[idx["room_rh_pct"]], "%")
        add("Served space", "Room cooling load", r[idx["room_cooling_kw"]], "kW",
            "Room load printed on the room box, not a per-unit capacity"
            + ("; stated as %s" % r[idx["cooling_basis"]]
               if r[idx["cooling_basis"]] != "not stated"
               else "; the schematic does not say whether this is total or sensible"))
        n_od = outdoor_count(sched[2]) if sched else 1
        add("Refrigeration", "Matched condensing unit", od, "",
            ("Paired by matching number, not stated on the drawing"
             + (" - this condenser is marked (ST.BY)" if od in OD_STANDBY else ""))
            if n_od == 1 else
            ("The equipment schedule gives this unit an X2 outdoor model, so it takes TWO "
             "condensers. The single matching number is not the whole pairing - the pipework "
             "layout plan is needed."))
        if sched:
            loc, indoor, outdoor, speeds = sched
            rows.append([tag, model, "Identification", "Location (equipment schedule)", loc, "",
                         DX_SRC_SYS, DX_PAGE_SYS,
                         "Coarser than the schematic's room; both are recorded"])
            rows.append([tag, model, "Identification", "Indoor unit model", indoor, "",
                         DX_SRC_SYS, DX_PAGE_SYS, ""])
            rows.append([tag, model, "Refrigeration", "Outdoor unit model", outdoor, "",
                         DX_SRC_SYS, DX_PAGE_SYS,
                         "X2 suffix - two condensers per indoor unit" if n_od == 2 else ""])
            rows.append([tag, model, "Refrigeration", "Outdoor units required", n_od, "n",
                         DX_SRC_SYS, DX_PAGE_SYS,
                         "Read from the X2 suffix on the outdoor model" if n_od == 2 else
                         "No multiplier on the outdoor model"])
            rows.append([tag, model, "Air", "Number of speeds", speeds, "n",
                         DX_SRC_SYS, DX_PAGE_SYS, ""])
        else:
            rows.append([tag, "", "Data quality", "Model not stated",
                         "This unit is on the schematic but not on SYSTEM DETAILS Table 3.6, "
                         "which stops at DX/B/16. No indoor or outdoor model is available for it.",
                         "", DX_SRC_SYS, DX_PAGE_SYS, ""])
    for prop, text in DX_NOTES:
        rows.append([DX_ROWS[0][0], "", "Data quality", prop, text, "", SRC_DX, PAGE_DX,
                     "Applies to the whole schematic"])
    for prop, text in DX_SCHEDULE_NOTES:
        rows.append([DX_ROWS[0][0], "", "Data quality", prop, text, "",
                     DX_SRC_SYS, DX_PAGE_SYS, "Applies to the DX set"])

# ---------------------------------------------------------------- generators
GEN_LABELS = [
 ("location_info", "Location info", ""), ("main_category", "Main category", ""),
 ("sub_category", "Sub category", ""), ("manufacturer", "Manufacturer", ""),
 ("model_number", "Model number", ""), ("serial_number", "Serial number", ""),
 ("equipment_name", "Equipment name", ""), ("building_name", "Building name", ""),
 ("floor", "Floor", ""), ("room_no", "Room no.", ""),
 ("date_installed", "Date installed", ""), ("quantity", "Quantity", "n"),
 ("pm_procedure_description", "PM procedure description", ""),
 ("procedure_document", "Procedure document", ""),
 ("warranty", "Warranty", ""), ("warranty_expires", "Warranty expires", ""),
 ("warranty_notes", "Warranty notes", ""),
 ("recommended_spare_parts", "Recommended spare parts", ""),
]

def build_gen(rows):
    idx = {c: i for i, c in enumerate(GEN_COLS)}
    for r in GEN_ROWS:
        tag = r[idx["asset_tag"]]; model = r[idx["model_number"]]
        for key, label, unit in GEN_LABELS:
            comp = ("Asset register" if key in ("location_info", "main_category", "sub_category",
                                                "equipment_name", "quantity")
                    else "Identification" if key in ("manufacturer", "model_number",
                                                     "serial_number")
                    else "Location" if key in ("building_name", "floor", "room_no")
                    else "Maintenance" if key in ("pm_procedure_description", "procedure_document",
                                                  "recommended_spare_parts", "date_installed")
                    else "Warranty")
            rows.append([tag, model, comp, label, r[idx[key]], unit, SRC_GEN, PAGE_GEN, ""])
    for prop, text in GEN_NOTES:
        rows.append([GEN_ROWS[0][0], "", "Data quality", prop, text, "", SRC_GEN, PAGE_GEN, ""])

# ---------------------------------------------------------------- write
SHEETS = []

def add_sheet(name, rows, subtitle):
    """Hold a sheet's rows so both workbooks can be written from the same data."""
    _index(name, rows)
    SHEETS.append((name, list(rows), subtitle))

def sheet(wb, name, rows, subtitle, do_index=False, fill_value=False):
    if do_index:
        _index(name, rows)
    rows = [expand(r) for r in rows]
    if fill_value:
        # A string-valued predicate (rec:modelNumber, rec:manufacturedBy) has nothing to
        # convert, so the Dar Cairo cell is empty. The object of the triple is the printed
        # value, so carry it across - the ontology workbook must never show a blank object.
        for r in rows:
            if r[6] == "" and r[4] not in (None, ""):
                r[6] = r[4]
    ws = wb.create_sheet(name)
    ws["A1"] = name; ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws["A2"] = subtitle; ws["A2"].font = Font(italic=True, size=9, color="555555")
    ws.append([])
    ws.append(HEAD)
    hr = ws.max_row
    for c in range(1, len(HEAD)+1):
        cell = ws.cell(row=hr, column=c)
        cell.fill = (PatternFill("solid", fgColor="2E5F2E") if c in (7, 8, 9)
                     else PatternFill("solid", fgColor="7B3F00") if c in (10, 11)
                     else HDR_FILL)
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    tags = []
    for r in rows:
        ws.append(r); tags.append(r[0])
    order, seen = [], set()
    for t in tags:
        if t not in seen: seen.add(t); order.append(t)
    band = {t: (i % 2 == 1) for i, t in enumerate(order)}
    for i, r in enumerate(rows):
        rr = hr + 1 + i
        for c in range(1, len(HEAD)+1):
            cell = ws.cell(row=rr, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in WRAP_COLS))
            if band[r[0]]: cell.fill = BAND_FILL
        ws.cell(row=rr, column=1).font = COMP_FONT if (i == 0 or rows[i-1][0] != r[0]) else Font(size=10)
        if r[NOTE_COL-1]: ws.cell(row=rr, column=NOTE_COL).font = NOTE_FONT
        if r[8]: ws.cell(row=rr, column=9).font = Font(italic=True, size=9, color="2E5F2E")
        sc = r[SCOPE_COL-1]
        if sc:
            ws.cell(row=rr, column=SCOPE_COL).font = Font(
                bold=(sc == "core"), size=9,
                color={"core": "1F6F1F", "candidate": "9C5700"}.get(sc, "9A9A9A"))
            ws.cell(row=rr, column=SCOPE_COL).alignment = Alignment(
                horizontal="center", vertical="top")
        ws.cell(row=rr, column=PAGE_COL).alignment = Alignment(horizontal="center", vertical="top")
    widths = [16, 24, 28, 40, 20, 13, 20, 20, 28, 28, 11, 30, 7, 46]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hr+1, column=1)
    ws.auto_filter.ref = "A%d:%s%d" % (hr, get_column_letter(len(HEAD)), ws.max_row)
    return ws

def new_wb():
    wb = Workbook(); wb.remove(wb.active); return wb

def readme_sheet(wb, ontology_only=False):
    # README
    ws = wb.create_sheet("README")
    def put(r, a, b="", bold=False, size=10, color="000000", italic=False):
        ws.cell(row=r, column=1, value=a).font = Font(bold=bold, size=size, color=color, italic=italic)
        if b != "":
            c = ws.cell(row=r, column=2, value=b)
            c.font = Font(size=10); c.alignment = Alignment(wrap_text=True, vertical="top")

    if ontology_only:
        put(1, "QNL Equipment Metadata - Needed for Ontology", bold=True, size=16, color="7B3F00")
    else:
        put(1, "QNL Equipment Metadata - Full Metadata", bold=True, size=16, color="1F3864")
    put(2, "Qatar National Library, Education City BP#7A", italic=True, color="555555")
    r = 4
    if ontology_only:
        put(r, "This workbook", bold=True, size=12, color="7B3F00"); r += 1
        for line in [
          "The subset of the full transcription that becomes triples: 1,225 rows across 16 "
          "predicates, every one mapping to a Brick 1.4 entity property or a predicate Dar Cairo, "
          "QF SSC or QF HQ already uses. Its companion, QNL_Full_Metadata.xlsx, holds all 4,728 "
          "rows including the 3,418 kept as engineering reference.",
          "Read 'Value (Dar Cairo)' and 'Unit (QUDT)' as the object of the triple. For a "
          "string-valued predicate such as rec:modelNumber there is nothing to convert, so that "
          "cell carries the printed value unchanged.",
          "The Open Items sheet holds the 85 data quality findings. None are triples, but each "
          "one qualifies rows that are - read it before writing the ontology.",
          "Every row still names its source document and page, so any object can be traced back.",
        ]:
            c = ws.cell(row=r, column=2, value="- " + line); c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
        r += 1
    put(r, "What this workbook is", bold=True, size=12, color="1F3864"); r += 1
    put(r, "", "One sheet per equipment type. Every row is a single manufacturer property for a single "
              "unit, and carries the source PDF and the page it was read from, so any value can be "
              "traced back to the document."); r += 2
    put(r, "Sheets", bold=True, size=12, color="1F3864"); r += 1
    for name, desc in [
      ("AHU", "15 air handling units, AHU-001 to AHU-015. AS BUILT general arrangement drawings, "
              "drawing 1844-SD-05-AC-0077, Mercury Mena for Qatar Foundation, 26-10-2015 / 28-11-2015."),
      ("Closed Control Units", "5 selection sheets covering CC/B/01 to CC/B/09. Emerson Network Power, "
              "issued by Qatar Site & Power, April and October 2013."),
      ("Climate Control Units", "Museum Climate Controls MCG-10P humidity control unit with its VCB1000 "
              "blower option and AF4 intake air filter."),
      ("FCU", "28 Euroclima selection sheets covering positions on Basement, 1F and 2F."),
      ("Heat Exchangers", "5 Alfa Laval plate heat exchangers, PHX/B/01 to PHX/B/05, with the "
                "manufacturer's construction data and the M10-MFM thermal specification."),
      ("Pumps", "4 Armstrong horizontal split case chilled water pumps, CHWP/B/01 to CHWP/B/04, with "
                "the full Armstrong submittal - pump, motor, seal, materials and dimensions."),
      ("Exhaust Fans", "39 exhaust fans - mostly Nuaire, two Colasit - from the supplied schedule "
                "spreadsheet."),
      ("Generators", "One standby diesel generator from the GENERATOR ASSET LIST, a multi-equipment "
                "asset register."),
      ("Pressurization Unit", "PU/B/01 (PRO1) - Armstrong 3750 2 EM-S pressurisation unit plus a "
                "Reflex DE10 1000 litre expansion tank."),
      ("CAV Units", "42 constant air volume boxes, one row per box, from six schedules."),
      ("VAV Units", "180 variable air volume boxes, one row per box, from six schedules."),
      ("DX Units", "21 DX split systems, DX/B/01-20 and DX/RP/21, from the schematic riser and the "
                "SYSTEM DETAILS equipment schedule, which names models for 16 of them."),
      ("Ontology Scope", "Every distinct property in this workbook, the reference-model predicate it "
                "maps to, and whether it belongs in the ontology at all."),
      ("Units", "Every source unit, the Dar Cairo unit it maps to, the factor applied, and whether "
                "Dar Cairo uses that unit at all."),
    ]:
        ws.cell(row=r, column=1, value=name).font = Font(bold=True, size=10)
        c = ws.cell(row=r, column=2, value=desc); c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
    r += 1
    put(r, "Columns", bold=True, size=12, color="1F3864"); r += 1
    for name, desc in [
      ("Equipment Tag", "The position or unit reference exactly as the source document writes it."),
      ("Model", "Manufacturer model designation where the sheet gives one."),
      ("Component", "The part of the unit the property belongs to - cooling coil, supply fan, condenser, etc."),
      ("Property", "The property name, kept close to the source wording."),
      ("Value (as printed)", "The value exactly as the document prints it. Nothing converted, rounded or inferred."),
      ("Unit (as printed)", "The unit exactly as the document prints it."),
      ("Value (Dar Cairo)", "The same quantity converted to the unit Dar Cairo uses for it."),
      ("Unit (QUDT)", "The QUDT unit token to write into brick:hasUnit."),
      ("Conversion", "The arithmetic applied, and any assumption behind it. Blank means no conversion was needed."),
      ("Source File", "The document the value came from - a PDF, a spreadsheet, or a supplied drawing image."),
      ("Page", "Where in that document: a 1-based page for a PDF, a sheet and row for a spreadsheet, "
               "the schedule title for a drawing."),
      ("Note", "Anything that qualifies the value - a quantity off, a source conflict, a mislabelled row."),
    ]:
        ws.cell(row=r, column=1, value=name).font = Font(bold=True, size=10)
        c = ws.cell(row=r, column=2, value=desc); c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
    r += 1
    put(r, "Units", bold=True, size=12, color="2E5F2E"); r += 1
    for line in [
      "The source documents are transcribed in their own units, and each row then carries the same "
      "quantity in the unit Dar Cairo uses for it. Both are kept side by side: the printed pair stays "
      "traceable to the page, the converted pair is what goes into brick:hasUnit.",
      "Targets were read off DarCairo_V93.csv and re-checked on V98, not assumed. Air flow and water flow both land on "
      "unit:L-PER-SEC (para:ratedSupplyAirFlowrate, para:ratedChilledWaterFlowrate and their siblings), "
      "power on unit:KiloW (brick:ratedPowerInput, brick:coolingCapacity), length on unit:M "
      "(para:ratedHead), relative humidity on unit:PERCENT_RH rather than unit:PERCENT.",
      "The Units sheet lists every mapping, the factor applied, and the four QUDT units Dar Cairo has "
      "never used - mass, mass flow and velocity, which it carries no quantity for. Those need the "
      "PARA team's confirmation.",
      "Converting kg/s to l/s treats water as 1 kg/l. At 7-15 degC it is about 0.9997 kg/l, so the "
      "converted flow is high by roughly 0.03%.",
    ]:
        c = ws.cell(row=r, column=2, value="- " + line); c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
    r += 1
    put(r, "What belongs in the ontology", bold=True, size=12, color="7B3F00"); r += 1
    for line in [
      "Most of this workbook is engineering reference, not ontology metadata. Dar Cairo, QF SSC and "
      "QF HQ agree on a short vocabulary for equipment - about twenty predicates - and everything "
      "outside it has no precedent in any of them.",
      "Brick 1.4 is checked FIRST, per the class ladder - it is step 2 and the reference models are "
      "step 3. Brick has no water or air flow-rate entity property and no heat-exchanger duty "
      "property, which is why Dar Cairo minted para:ratedWaterFlowrate, para:ratedChilledWaterFlowrate "
      "and the air-flowrate family. Where Brick does carry the term it wins: brick:coolingCapacity for "
      "heat exchanger duty, brick:ratedCurrentInput for full load current, brick:operationalStageCount "
      "for a fan's speed count, and brick:Condensing_Unit as the entity an outdoor unit's model sits on.",
      "Every row carries an 'Ontology predicate' and a 'Scope'. core means it maps unambiguously to a "
      "Brick 1.4 entity property or a predicate the reference models use. candidate means the match is "
      "plausible but needs a decision. reference means neither Brick nor any reference model carries "
      "anything like it - dimensions, weights, "
      "materials, seal specifications, sound power levels, filter part numbers, psychrometrics, "
      "warranty text. Those rows stay because they are useful to engineers, not because they will be "
      "modelled.",
      "The Ontology Scope sheet lists every distinct property with its verdict, the predicate it maps "
      "to and where that predicate comes from - Brick 1.4, Dar Cairo, QF SSC or QF HQ.",
      "The predicates in play: para:ratedSupplyAirFlowrate, rec:modelNumber, rec:manufacturedBy, "
      "para:ratedReheatCapacity, brick:ratedPowerInput, para:ratedChilledWaterFlowrate, "
      "brick:coolingCapacity, para:ratedSpeed, brick:ratedCurrentInput, para:ratedExhaustAirFlowrate, "
      "brick:operationalStageCount, para:ratedHead, para:refrigerant, rec:installationDate, "
      "para:ratedWaterFlowrate, para:Rated_Tank_Level.",
      "One open point: the expansion tank's 1000 litre capacity is written as para:Rated_Tank_Level at "
      "the user's direction. brick:volume exists in Brick 1.4 and by the ladder would outrank a para: "
      "term - reversible in one line of ontology_map.py if the team prefers it.",
      "An outdoor DX unit's model belongs on its own brick:Condensing_Unit entity, which Brick 1.4 "
      "carries, so no para: class needed to be minted for it. Those entities are not in this workbook "
      "yet - it holds the model against the indoor unit that names it.",
      "A component's own maker is not the equipment's manufacturer. The pump's mechanical seal is "
      "made by Armstrong and its motor by WEG; neither becomes rec:manufacturedBy on the pump.",
      "The QF HQ v0.4 draft was read for structure, not for units - several of its rows carry a wrong "
      "brick:hasUnit (an air flow tagged unit:V, a cooling capacity tagged unit:HZ). Dar Cairo remains "
      "the authority for unit choice, as the Units sheet records.",
    ]:
        c = ws.cell(row=r, column=2, value="- " + line); c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
    r += 1
    put(r, "Read before using the data", bold=True, size=12, color="C00000"); r += 1
    for line in [
      "Values are transcribed from the documents as printed. Where a document contradicts itself the "
      "conflict is recorded as a 'Data quality' row rather than resolved - those need the supplier's answer.",
      "AHU-011 carries two Item 3 ELECTRIC COIL blocks; the first is an unfilled template "
      "(capacity POW, stages NST, entering air AON, leaving air AOF). The populated block is the one recorded.",
      "MCG-10P size is given as 500 x 650 x 650 mm on the specification sheet and 475 x 430 x 400 mm on "
      "dimension drawing MCG10P-1. Both are recorded; neither is confirmed.",
      "On 27 of the 28 Euroclima sheets the air pressure drop row is printed as 'Perdita di carico aria' "
      "with the unit [degC]. It is air-side pressure drop in Pa - page 2 labels the same row correctly.",
      "The CC/B/05,06,07 sheet heading reads M5DUA while its own Unit row reads M5DOA.",
      "The heat exchanger and pump rows come from an image of a schedule drawing whose title block is "
      "cropped out. They carry no drawing number, sheet or revision - ask for the sheet reference "
      "before handover.",
      "The exhaust fan schedule uses two identifier shapes: 28 slash-separated tags (EF/B/nn, EF/RP/n) "
      "and 11 underscore-separated (TEF_B01A, KEF_101). Both were kept as written because the tag is "
      "the BMS join key. EF/B numbering also runs 4 to 16 with no 1, 2 or 3.",
      "The pump model designation ends in 55KW. The schedule states no motor rating of its own, so no "
      "rated-power property was created from it - confirm against the pump submittal.",
      "The same schedule drawing carries a SCHEDULE OF PRESSURIZATION UNIT (PU/B/01, Armstrong "
      "3750 2 EM-S). It was not asked for and is not in this workbook; say the word and it gets a sheet.",
      "On the DX sheet the cooling figure is the load printed on the ROOM box, not a per-unit "
      "capacity. Rooms served by two or three units are not split by the schematic, so do not write "
      "it as brick:coolingCapacity on a unit without a per-unit duty.",
      "The DX indoor-to-outdoor pairing recorded from the schematic's matching numbers is NOT "
      "reliable. SYSTEM DETAILS Table 3.6 shows DX/B/03, 04, 08, 09 and 14 each take an X2 outdoor "
      "model - two condensers per indoor unit - so those five need two DX/OD tags each and the "
      "numbering does not say which two. The pipework layout plan is still needed. DX/OD/05 is also "
      "marked (ST.BY) while DX/B/05 is not.",
      "Table 3.6 covers DX/B/01 to DX/B/16 only, so DX/B/17, 18, 19, 20 and DX/RP/21 still carry no "
      "model. DX/B/08 reads PUHZ-RP2S0X2 where its twin DX/B/09 reads PUHZ-RP250X2 - almost certainly "
      "a misprint, recorded as printed. The PEAD, PCA and PUHZ prefixes are Mitsubishi Electric "
      "Mr. Slim naming, but no document states a manufacturer, so none was inferred.",
      "CAV and VAV schedules write many references as ranges. Each range is split into one row per "
      "box, so every box carries its own air flow, heating capacity, model and make, and keeps a "
      "'Scheduled as' row naming the line it came from. A range is only split when its box count "
      "matches the stated quantity; three do not and keep the printed reference with a Data quality "
      "row instead: VAV/1F/S15/012, VAV/B/S14/009 TO 012 and VAV/B/S10/001 & 008.",
      "Two drawings schedule the same basement S10/S15 VAV boxes and disagree on VAV/B/S15/001 TO 004 "
      "(261 l/s against 251 l/s). Both are recorded against their own drawing. On the second 1F "
      "schedule VAV/1F/S15/014 TO 015 also falls inside VAV/1F/S15/013 TO 020 with a different flow "
      "and model.",
      "VAV/1F/S11/022 was scheduled twice; the user confirmed the standalone row governs (220 l/s, "
      "NBOQOB200), so the overlapping VAV/1F/S11/022 TO 24 row is read as covering 023 and 024 only. "
      "Five further rows carrying out-of-range box numbers were excluded on instruction.",
      "The fire damper and supply/return grille schedules on the same images were struck through in "
      "red and were read as cancelled, so they are not in this workbook.",
      "PHX/B/05 exists only on the Alfa Laval data and the SYSTEM DETAILS schedule - the drawing "
      "schedule stops at PHX/B/04. It is a 300 kW M10-MFM described as the 'Main HEX'. SYSTEM DETAILS "
      "states one hot-side temperature across all five rows (15.5 in / 6.5 out) but the M10-MFM "
      "specification gives 50.0 in / 20.0 out. Both are recorded; the conflict is unresolved.",
      "SYSTEM DETAILS writes the pump tags with dashes (CHWP-B-01) and tags the pressurisation unit "
      "PRO1, where the drawing schedules write CHWP/B/01 and PU/B/01. The slash forms are used as the "
      "Equipment Tag because they match the rest of the workbook; the alternates are recorded as an "
      "'Alternate reference' property on each unit. Confirm which the BMS uses.",
      "The Armstrong pump submittal is dimensioned in inches and pounds and is marked NOT for "
      "CONSTRUCTION. Its weight of 1874 lb converts to 850 kg against the drawing schedule's 843 kg. "
      "Its letter callouts (D, HA, HB, ...) are keyed to an outline drawing that does not say which "
      "feature each measures.",
      "The GENERATOR ASSET LIST is a multi-equipment asset register. The row cropped off below the "
      "generator belongs to different equipment, so the single row on the Generators sheet is the "
      "complete record for this machine. Its ASSET TAG NUMBER column reads 'GENERATOR SET' - the same "
      "text as the equipment name - so it does not identify an individual machine; only the serial "
      "number does.",
      "These are equipment selection and as-built documents, not nameplate photographs. Where the "
      "installed plant differs from the selection, the installed plant governs.",
    ]:
        c = ws.cell(row=r, column=2, value="- " + line); c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
    r += 1
    put(r, "Source documents", bold=True, size=12, color="1F3864"); r += 1
    for f, pages, what in [
      ("Ahu_manual_1.pdf", "8 pages", "AHU-001 to AHU-008 general arrangement drawings"),
      ("ahu_2.pdf", "7 pages", "AHU-009 to AHU-015 general arrangement drawings"),
      ("qnl_closed_control_units_manual.pdf", "17 pages", "Emerson closed control unit selections"),
      ("climate_control_units_manual.pdf", "7 pages", "Museum Climate Controls MCG-10P data sheets"),
      ("FCU_Manual.pdf", "29 pages", "Euroclima fan coil unit selection sheets"),
      ("MEP schedule drawing", "image", "Heat exchanger, pump and pressurization unit schedules, "
       "supplied as an image in chat. The image crops the title block, so no drawing number, sheet "
       "number or revision is recorded against these rows."),
      ("Book1.xlsx", "Sheet1, 39 rows", "Exhaust fan schedule - equipment tag, model, manufacturer "
       "and air flow only"),
      ("CAV and VAV schedules", "6 + 6 schedules", "Supplied as drawing images in chat, title blocks "
       "cropped. Each schedule is a separate source; the Page column names which."),
      ("Armstrong submittal", "TENDER301358.1 rev3, p14/22", "Full nameplate for the chilled water "
       "pumps - pump design, WEG motor, mechanical seal, materials and imperial dimensions"),
      ("Alfa Laval PHE data", "2 sheets", "Construction comparison for T20-BFG and M10-MFM, and the "
       "M10-MFM thermal specification for PHX/B/05"),
      ("SYSTEM DETAILS", "Tables 3.1-3.3, 3.6", "Duty/standby split for the pumps and heat "
       "exchangers, the pressurisation unit broken into its two components, and indoor/outdoor models "
       "for DX/B/01 to DX/B/16"),
      ("GENERATOR ASSET LIST", "image", "Asset register row for a standby diesel generator"),
      ("DX split system schematic", "image", "Riser diagram giving the room each DX unit serves, its "
       "design condition and a room cooling load"),
    ]:
        ws.cell(row=r, column=1, value=f).font = Font(bold=True, size=10)
        c = ws.cell(row=r, column=2, value="%s - %s" % (pages, what)); c.font = Font(size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 108
    ws.sheet_view.showGridLines = False
    return ws

rows = []; build_ahu(rows)
add_sheet("AHU", rows,
      "Air handling units AHU-001 to AHU-015 - AS BUILT general arrangement drawings, drawing "
      "1844-SD-05-AC-0077 (Mercury Mena for Qatar Foundation). The drawings give no unit-level model "
      "designation, so Model is blank; component models (coil, fan, motor) appear as their own rows.")
n_ahu = len(rows)

rows = []; build_ccu(rows)
add_sheet("Closed Control Units", rows,
      "Emerson Network Power closed control units CC/B/01 to CC/B/09 - selection sheets "
      "issued by Qatar Site & Power.")
n_ccu = len(rows)

rows = []; build_climate(rows)
add_sheet("Climate Control Units", rows,
      "Museum Climate Controls MCG-10P positive-pressure humidity control unit, "
      "VCB1000 blower option and AF4 intake air filter.")
n_cli = len(rows)

rows = []; build_fcu(rows)
add_sheet("FCU", rows,
      "Euroclima fan coil units - one selection sheet per position, offer CENTRAL LIBRARY. Positions "
      "carrying several tags on one sheet (e.g. B/06,08,17,24) were annotated by hand on the printed "
      "sheet; those tags share the selection above them. The Heating column is blank on all 28 sheets.")
n_fcu = len(rows)

rows = []; build_hex(rows)
add_sheet("Heat Exchangers", rows,
      "Plate heat exchangers PHX/B/01 to PHX/B/04 - SCHEDULE OF HEAT EXCHANGER. The supplied image "
      "crops the title block, so the drawing number and sheet are not recorded.")
n_hex = len(rows)

rows = []; build_pumps(rows)
add_sheet("Pumps", rows,
      "Chilled water pumps CHWP/B/01 to CHWP/B/04 - SCHEDULE OF PUMPS, same drawing as the heat "
      "exchanger schedule. The supplied image crops the title block.")
n_pump = len(rows)

rows = []; build_ef(rows)
add_sheet("Exhaust Fans", rows,
      "39 exhaust fans from the supplied schedule spreadsheet. The sheet carries model, "
      "manufacturer and air flow only - no location, motor rating or static pressure. Two "
      "identifier shapes are in use; both were kept as written.")
n_ef = len(rows)

rows = []; build_gen(rows)
add_sheet("Generators", rows,
      "One standby diesel generator from the GENERATOR ASSET LIST, a multi-equipment asset "
      "register. The row cropped off below it belongs to different equipment, so this is the "
      "complete record for this machine.")
n_gen = len(rows)

rows = []; build_pu(rows)
add_sheet("Pressurization Unit", rows,
      "PU/B/01 - SCHEDULE OF PRESSURIZATION UNIT, same drawing as the heat exchanger and pump "
      "schedules. The supplied image crops the title block.")
n_pu = len(rows)

rows = []; build_cav(rows)
add_sheet("CAV Units", rows,
      "Constant air volume boxes across six schedules on level 1F and the basement. Air flow is "
      "stated per box. A range reference is expanded into its individual boxes only when the box "
      "count matches the stated quantity.")
n_cav = len(rows)

rows = []; build_vav(rows)
add_sheet("VAV Units", rows,
      "Variable air volume boxes across six schedules on level 1F and the basement. Two drawings "
      "overlap on the basement S10/S15 boxes and disagree on one air flow; both are recorded. "
      "Five out-of-range rows were excluded on instruction.")
n_vav = len(rows)

rows = []; build_dx(rows)
add_sheet("DX Units", rows,
      "21 DX split systems from the schematic riser, with indoor and outdoor models from SYSTEM "
      "DETAILS Table 3.6 for DX/B/01 to DX/B/16. The cooling figure is the load printed on each "
      "room box, not a per-unit capacity. Five units take an X2 outdoor model - two condensers "
      "each - so the schematic's matching-number pairing does not describe them.")
n_dx = len(rows)

# ---------------------------------------------------------------- Ontology Scope
def scope_sheet(wb, per_sheet):
    ws = wb.create_sheet("Ontology Scope")
    ws["A1"] = "Ontology Scope"; ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws["A2"] = ("Which transcribed properties belong in the ontology. Targets are the predicates "
                "Dar Cairo, QF SSC and QF HQ actually use for equipment metadata - about twenty "
                "in total. Everything else is engineering reference, kept but not modelled.")
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    ws.append([])
    HD = ["Scope", "Ontology predicate", "Predicate source", "Component", "Property",
          "Rows", "Sheets"]
    ws.append(HD)
    hr = ws.max_row
    for c in range(1, len(HD)+1):
        cell = ws.cell(row=hr, column=c)
        cell.fill = PatternFill("solid", fgColor="7B3F00"); cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    order = {"core": 0, "candidate": 1, "reference": 2}
    entries = sorted(per_sheet.items(),
                     key=lambda kv: (order.get(kv[1]["scope"], 3), kv[1]["pred"],
                                     kv[0][0], kv[0][1]))
    for (comp, prop), v in entries:
        ws.append([v["scope"], v["pred"], PREDICATE_SOURCE.get(v["pred"], ""),
                   comp, prop, v["n"], ", ".join(sorted(v["sheets"]))])
    for i in range(len(entries)):
        rr = hr + 1 + i
        sc = ws.cell(row=rr, column=1).value
        for c in range(1, len(HD)+1):
            cell = ws.cell(row=rr, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 5, 7)),
                                       horizontal="center" if c in (1, 6) else "left")
            if sc == "core":
                cell.fill = PatternFill("solid", fgColor="E8F4E8")
            elif sc == "candidate":
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=rr, column=1).font = Font(bold=(sc == "core"), size=9,
                color={"core": "1F6F1F", "candidate": "9C5700"}.get(sc, "9A9A9A"))
    for i, w in enumerate([12, 30, 46, 28, 40, 8, 34], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hr+1, column=1)
    ws.auto_filter.ref = "A%d:%s%d" % (hr, get_column_letter(len(HD)), ws.max_row)
    return ws

def units_sheet(wb):
    uw = wb.create_sheet("Units")
    uw["A1"] = "Units"; uw["A1"].font = Font(bold=True, size=14, color="1F3864")
    uw["A2"] = ("Source unit to Dar Cairo unit. Targets read off reference-models/DarCairo_V98.csv - "
                "air and water flow both land on unit:L-PER-SEC, power on unit:KiloW, length on unit:M.")
    uw["A2"].font = Font(italic=True, size=9, color="555555")
    uw.append([])
    UHEAD = ["Unit (as printed)", "Unit (QUDT)", "Factor applied",
             "Used in Dar Cairo", "Dar Cairo uses (rows)", "Note"]
    uw.append(UHEAD)
    uhr = uw.max_row
    for c in range(1, len(UHEAD)+1):
        cell = uw.cell(row=uhr, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    DC_USES = {  # usage counts from references/data/units.csv
     "unit:KiloW":1593, "unit:DEG_C":883, "unit:M2":755, "unit:PERCENT":708, "unit:HR":613,
     "unit:L-PER-SEC":535, "unit:V":451, "unit:A":227, "unit:PA":172, "unit:DeciB":171,
     "unit:PERCENT_RH":79, "unit:HZ":61, "unit:M":47, "unit:RPM":21, "unit:W":13,
     "unit:UNITLESS":3376,
    }

    urows = []
    for srcu, (qudt, factor, in_dc, note) in UNIT_MAP.items():
        if srcu == "":
            continue
        urows.append([srcu, qudt or "(none)",
                      "" if factor is None else ("x 1" if factor == 1.0 else _sigfmt(factor)),
                      "yes" if (in_dc and qudt) else ("n/a" if not qudt else "NO"),
                      DC_USES.get(qudt, "" if qudt else ""),
                      note])
    urows.append(["%", "unit:PERCENT_RH", "x 1", "yes", DC_USES["unit:PERCENT_RH"],
                  "used when the property is a relative humidity"])
    urows.append(["%", "unit:PERCENT", "x 1", "yes", DC_USES["unit:PERCENT"],
                  "used for every other percentage"])
    urows.append(["(none printed)", "unit:UNITLESS", "x 1", "yes", DC_USES["unit:UNITLESS"],
                  "dimensionless ratios - SHR, EER, COP, fan speed setting"])
    for r_ in urows:
        uw.append(r_)
    for i in range(len(urows)):
        rr = uhr + 1 + i
        for c in range(1, len(UHEAD)+1):
            cell = uw.cell(row=rr, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 6),
                                       horizontal="center" if c in (3, 4, 5) else "left")
        if uw.cell(row=rr, column=4).value == "NO":
            for c in range(1, len(UHEAD)+1):
                uw.cell(row=rr, column=c).fill = PatternFill("solid", fgColor="FFF2CC")
    for i, w in enumerate([20, 26, 16, 18, 20, 62], 1):
        uw.column_dimensions[get_column_letter(i)].width = w
    uw.freeze_panes = uw.cell(row=uhr+1, column=1)
    rr = uw.max_row + 2
    uw.cell(row=rr, column=1, value="Highlighted rows are QUDT units Dar Cairo has never used - it "
            "carries no mass, mass-flow or velocity quantity. They are genuine QUDT terms, not minted "
            "ones, but the PARA team should confirm them before handover.").font = NOTE_FONT
    uw.cell(row=rr, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    uw.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=6)
    uw.sheet_view.showGridLines = False
    return uw


# ---------------------------------------------------------------- write both
def open_items(wb):
    """Findings that a modeller must see even though they are not triples."""
    ws = wb.create_sheet("Open Items")
    ws["A1"] = "Open Items"; ws["A1"].font = Font(bold=True, size=14, color="C00000")
    ws["A2"] = ("Data quality findings from the source documents. None of these are triples, but "
                "each one affects rows that are. Read before writing the ontology.")
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    ws.append([])
    HD = ["Equipment Tag", "Finding", "Detail", "Source File", "Page"]
    ws.append(HD)
    hr = ws.max_row
    for c in range(1, len(HD)+1):
        cell = ws.cell(row=hr, column=c)
        cell.fill = PatternFill("solid", fgColor="C00000"); cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = BORDER
    n = 0
    for name, rows, _ in SHEETS:
        for r in rows:
            if r[2] != "Data quality":
                continue
            ws.append([r[0], r[3], r[4], r[6], r[7]]); n += 1
    for i in range(n):
        rr = hr + 1 + i
        for c in range(1, len(HD)+1):
            cell = ws.cell(row=rr, column=c)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
    for i, w in enumerate([18, 34, 96, 30, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hr+1, column=1)
    ws.auto_filter.ref = "A%d:%s%d" % (hr, get_column_letter(len(HD)), ws.max_row)
    return n

def build_full(path):
    wb = new_wb()
    readme_sheet(wb)
    for name, rows, sub in SHEETS:
        sheet(wb, name, rows, sub)
    scope_sheet(wb, SCOPE_INDEX)
    units_sheet(wb)
    wb.save(path)
    return sum(len(r) for _, r, _ in SHEETS)

def build_ontology(path):
    """Only the rows that become triples, plus the findings that qualify them."""
    wb = new_wb()
    readme_sheet(wb, ontology_only=True)
    kept = 0
    for name, rows, sub in SHEETS:
        core = [r for r in rows if expand(r)[10] == "core"]
        if not core:
            continue
        kept += len(core)
        sheet(wb, name, core,
              sub.split(".")[0] + ". Core rows only - every row here maps to an ontology "
              "predicate. The full transcription is in the full-metadata workbook.",
              fill_value=True)
    n_open = open_items(wb)
    scope_sheet(wb, {k: v for k, v in SCOPE_INDEX.items() if v["scope"] == "core"})
    units_sheet(wb)
    wb.save(path)
    return kept, n_open

n_all = build_full(OUT_FULL)
print("full metadata   ->", OUT_FULL, "|", n_all, "rows")
if OUT_ONT:
    n_core, n_open = build_ontology(OUT_ONT)
    print("ontology subset ->", OUT_ONT, "|", n_core, "core rows,", n_open, "open items")
