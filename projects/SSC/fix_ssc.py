#!/usr/bin/env python3
"""Apply the agreed term fixes to the finalized SSC sheet, then highlight the rest.

    python3 projects/SSC/fix_ssc.py

Input  : sources/QF_SSC_Ontology_ver01.xlsx   (untouched)
Output : QF_SSC_Ontology_ver01_fixed.xlsx     (fixes applied, rest highlighted)

The five decisions came from the PARA team on 2026-08-20, after the validator
reported 354 row-level errors on the finalized sheet. Everything the team did
not settle is left exactly as it was and filled yellow for a manual pass -
guessing at the remainder would bury the ones that need a person.
"""
import collections
import pathlib
import subprocess
import sys

import openpyxl

HERE = pathlib.Path(__file__).resolve().parent
ROOT = HERE.parent.parent
SRC = HERE / "sources" / "QF_SSC_Ontology_ver01.xlsx"
OUT = HERE / "QF_SSC_Ontology_ver01_fixed.xlsx"
SHEET = "SSC_Ontology_Ver0.6"

# 1-3: wrong terms, replaced everywhere they appear.
SUBS = {
    # not a Brick term; Dar Cairo uses the _Sensor form 38 times
    "brick:Supply_Air_Flow": "brick:Supply_Air_Flow_Sensor",
    # a dropped leading O
    "brick:ccupied_Air_Temperature_Setpoint": "brick:Occupied_Air_Temperature_Setpoint",
    # brick:Fan_Status has no Dar Cairo precedent; brick:On_Off_Status has 435 rows
    "brick:Fan_Status": "brick:On_Off_Status",
}
# 4: brick:Apparent_Power_Sensor stands - recorded in references/data/accepted-terms.txt
# 5: entity:HVAC carried both brick:HVAC_System and brick:System. HVAC_System is
#    the house term, so only entity:HVAC's own type cells change - brick:System
#    is right for entity:Electrical_System and must not be swept up with it.


def main():
    if not SRC.exists():
        sys.exit(f"missing input: {SRC}")
    wb = openpyxl.load_workbook(SRC)
    ws = wb[SHEET]

    hits = collections.Counter()
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip() in SUBS:
                cell.value = SUBS[v.strip()]
                hits[v.strip()] += 1

    retyped = 0
    for row in ws.iter_rows(min_row=2):
        if row[0].value == "entity:HVAC" and row[1].value == "brick:System":
            row[1].value = "brick:HVAC_System"
            retyped += 1
        if row[3].value == "entity:HVAC" and row[4].value == "brick:System":
            row[4].value = "brick:HVAC_System"
            retyped += 1

    staged = HERE / ".staged.xlsx"
    wb.save(staged)
    for old, new in SUBS.items():
        print(f"  {hits[old]:>4}  {old} -> {new}")
    print(f"  {retyped:>4}  entity:HVAC retyped brick:HVAC_System")

    script = ROOT / "skills" / "building-ontology" / "scripts" / "highlight_findings.py"
    r = subprocess.run([sys.executable, str(script), str(staged), "--out", str(OUT),
                        "--label-style", "verbatim", "--severity", "ERROR"],
                       capture_output=True, text=True)
    print(r.stdout[r.stdout.find("rows highlighted") - 8:] if "rows highlighted" in r.stdout
          else r.stdout + r.stderr)
    staged.unlink(missing_ok=True)
    return r.returncode


if __name__ == "__main__":
    sys.exit(main())
