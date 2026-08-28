#!/usr/bin/env python3
"""Apply the house format to Assumption_Log.xlsx, without touching its content.

    python3 format_assumption_log.py                 # reformat every sheet
    python3 format_assumption_log.py --add RDC2      # scaffold a new building sheet

**The log is hand-maintained.** It is edited in Excel as decisions are taken, so this
script never writes, reorders or removes a row - it only applies the shared format so
every building's sheet reads the same way. The earlier build_assumption_log.py, which
regenerated the whole workbook from a hardcoded list, is gone: re-running it after a
hand edit would silently discard the edit.

The QNL sheet is the format reference. Whatever its header row and column widths are,
the other sheets are brought to match - so changing the columns is done once, in Excel,
on QNL, and this propagates it.
"""
import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LOG = Path(__file__).resolve().parent / "Assumption_Log.xlsx"
REFERENCE = "QNL"          # the sheet whose format the others follow
PLACEHOLDER = "(no entries yet)"

# Category -> fill. A category with no entry here simply gets no fill, so adding a
# new one in Excel is not a crash.
CATEGORY_FILL = {
    "Identifier": "DDEBF7", "Location": "FCE4D6", "Units": "E2EFDA",
    "Spelling": "FFF2CC", "Class": "E4DFEC", "Scope": "F2F2F2",
    "Source defect": "FBE5D6", "Structure": "DEEAF6",
}


def read_format(ws):
    """The reference sheet's own header, widths and cell styles."""
    ncol = ws.max_column
    head = ws.cell(1, 1)
    body = ws.cell(2, 1)
    edge = body.border.left.color.rgb if body.border.left.color else "FFBFBFBF"
    return {
        "ncol": ncol,
        "headers": [ws.cell(1, c).value for c in range(1, ncol + 1)],
        "widths": [ws.column_dimensions[get_column_letter(c)].width
                   for c in range(1, ncol + 1)],
        "hdr_fill": PatternFill("solid", fgColor=head.fill.fgColor.rgb),
        "hdr_font": Font(bold=True, color=head.font.color.rgb, size=head.font.sz),
        "hdr_height": ws.row_dimensions[1].height,
        "body_height": ws.row_dimensions[2].height,
        "border": Border(*(Side(style="thin", color=edge),) * 4),
    }


def apply_format(ws, fmt, is_reference=False):
    """Bring one sheet to the house format. Row content is never altered."""
    # columns the reference no longer has, dropped right to left
    for c in range(ws.max_column, fmt["ncol"], -1):
        ws.delete_cols(c)

    for c in range(1, fmt["ncol"] + 1):
        cell = ws.cell(1, c, fmt["headers"][c - 1])
        cell.fill = fmt["hdr_fill"]
        cell.font = fmt["hdr_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = fmt["widths"][c - 1]
    ws.row_dimensions[1].height = fmt["hdr_height"]

    body_align = Alignment(vertical="top", wrap_text=True)
    for r in range(2, ws.max_row + 1):
        for c in range(1, fmt["ncol"] + 1):
            cell = ws.cell(r, c)
            cell.alignment = body_align
            cell.border = fmt["border"]
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = fmt["body_height"]
        cat = ws.cell(r, 3).value
        if cat in CATEGORY_FILL:
            ws.cell(r, 3).fill = PatternFill("solid", fgColor=CATEGORY_FILL[cat])

    if str(ws.cell(2, 1).value or "") == PLACEHOLDER:
        ws.cell(2, 1).font = Font(italic=True, color="FF808080")

    # Freeze the header only. The reference sheet keeps whatever the author set,
    # which after a block row-deletion can end up well below row 2.
    if not is_reference:
        ws.freeze_panes = "A2"
    return ws.max_row - 1


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--add", metavar="NAME",
                    help="scaffold an empty sheet for a new building")
    args = ap.parse_args()

    if not LOG.exists():
        sys.exit(f"{LOG} not found")
    wb = openpyxl.load_workbook(LOG)
    if REFERENCE not in wb.sheetnames:
        sys.exit(f"{LOG.name} has no {REFERENCE!r} sheet to take the format from")

    fmt = read_format(wb[REFERENCE])
    print(f"format from {REFERENCE}: {fmt['ncol']} columns")
    print(f"  {' | '.join(str(h) for h in fmt['headers'])}\n")

    if args.add:
        if args.add in wb.sheetnames:
            sys.exit(f"sheet {args.add!r} already exists")
        ws = wb.create_sheet(args.add)
        ws.cell(2, 1, PLACEHOLDER)
        print(f"added sheet {args.add!r}")

    for ws in wb.worksheets:
        n = apply_format(ws, fmt, is_reference=(ws.title == REFERENCE))
        empty = str(ws.cell(2, 1).value or "") == PLACEHOLDER
        print(f"  {ws.title:6} {'no entries' if empty else str(n) + ' entries':>12}"
              f"   {ws.max_column} cols   freeze {ws.freeze_panes}")

    wb.save(LOG)
    print(f"\nsaved {LOG.name}")


if __name__ == "__main__":
    main()
