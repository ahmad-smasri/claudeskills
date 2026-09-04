import openpyxl, csv, collections

# --- QNL: class -> unit
wb=openpyxl.load_workbook('projects/QNL/QNL_Ontology.xlsx', data_only=True); ws=wb.active
qnl=collections.Counter()
for r in range(2,ws.max_row+1):
    if ws.cell(r,3).value!='brick:hasPoint': continue
    cls=ws.cell(r,5).value; u=None
    for c in range(6,27,2):
        if ws.cell(r,c).value=='brick:hasUnit': u=ws.cell(r,c+1).value
    qnl[(cls,u)]+=1

# --- Dar Cairo: class -> Counter(unit)
dar=collections.defaultdict(collections.Counter)
with open('reference-models/DarCairo_V93.csv', encoding='utf-8-sig', newline='') as fh:
    rd=csv.reader(fh); next(rd)
    for r in rd:
        if len(r)<6: continue
        ot=r[4].strip()
        if not ot: continue
        for i in range(5,len(r)-1):
            if r[i].strip()=='brick:hasUnit':
                dar[ot][r[i+1].strip()]+=1

# --- SSC: class -> Counter(unit)
sb=openpyxl.load_workbook('reference-models/QF_SSC_Ontology_V03.xlsx', data_only=True)
ss=sb['SSC_Ontology_Ver0.6']
ssc=collections.defaultdict(collections.Counter)
for r in range(2, ss.max_row+1):
    ot=str(ss.cell(r,5).value or '')
    if not ot: continue
    for c in range(6,28):
        if str(ss.cell(r,c).value or '')=='brick:hasUnit':
            ssc[ot][str(ss.cell(r,c+1).value or '')]+=1

def top(cnt):
    return cnt.most_common(1)[0] if cnt else None

print(f"{'class':46} {'QNL unit':16} {'n':>5}  {'Dar Cairo':16} {'SSC':16} verdict")
print('-'*126)
agree=disagree=noprec=0
issues=[]
for (cls,u),n in sorted(qnl.items()):
    d=top(dar.get(cls, collections.Counter()))
    s=top(ssc.get(cls, collections.Counter()))
    ds = f"{d[0]} ({d[1]})" if d else "-"
    ss_ = f"{s[0]} ({s[1]})" if s else "-"
    if d:
        verdict = "MATCH" if d[0]==u else "*** DIFFERS ***"
        if d[0]==u: agree+=1
        else: disagree+=1; issues.append((cls,u,d[0],n,'Dar Cairo'))
    elif s:
        verdict = "match (SSC)" if s[0]==u else "*** DIFFERS (SSC) ***"
        if s[0]==u: agree+=1
        else: disagree+=1; issues.append((cls,u,s[0],n,'SSC'))
    else:
        verdict = "no precedent"; noprec+=1
    print(f"{str(cls):46} {str(u):16} {n:>5}  {ds:16} {ss_:16} {verdict}")
print('-'*126)
print(f"agree with precedent: {agree}   differ: {disagree}   no precedent: {noprec}")
if issues:
    print("\nDISAGREEMENTS:")
    for cls,u,exp,n,src in issues:
        print(f"  {cls}: QNL {u} vs {src} {exp}  ({n} points)")
