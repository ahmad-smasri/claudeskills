#!/usr/bin/env python3
"""Retarget QNL's ontology onto the corrected room names.

Three changes, in this order:

1. **Every room subject is rewritten into the Dar Cairo shape** -
   ``entity:QNL_B_063_PLANT-ROOM-01`` becomes ``entity:QNL_B-063_Plant-Room-01``
   - and every row that names a room as subject or object follows it.

   Where the naming sheet has decided a reference, its name wins. Where the
   ontology holds more than one room at one reference the sheet is not used at
   all and each room keeps its own name, because the sheet knows a reference
   and the ontology knows rooms: ``L1.104`` is a male prayer room and a female
   one, and taking the sheet's single name there would merge two rooms into one.

2. **The equipment that moved room is retargeted** - ``rec:locatedIn`` for all
   of it, and ``rec:feeds`` only where the equipment currently feeds the room it
   sits in. Twenty-two exhaust fans, toilet fans and chilled-water pumps already
   feed somewhere other than where they sit, and that reading is not this
   sheet's to overwrite.

3. **A room the sheet needs and the ontology lacks is declared**, as a
   ``rec:Room`` with its ``rec:isPartOf`` level row and a label.

Nothing is deleted. A room left with no equipment stays declared, and a room
the ontology declares twice stays declared twice - both are reported instead.
"""
import argparse
import collections
import csv
import os
import re
import sys

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                "..", "dar-cairo-room-naming"))
import build_room_names as B                                    # noqa: E402

def room_re(code):
    return re.compile(r"^entity:%s_([A-Za-z0-9]+)_([A-Za-z0-9-]+)_(.+)$" % code)


def pick_sheet(wb):
    for ws in wb:
        head = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if head and str(head[0]).strip().lower() == "subject":
            return ws
    raise SystemExit("no ontology sheet")


def parse_room(building, ent):
    m = room_re(building.code).match(ent)
    if not m:
        # A room with no level segment - entity:SSC_ST-10_STAIR-10, a stairwell
        # that belongs to no floor. Its first segment is the reference.
        m2 = re.match(r"^entity:%s_([A-Za-z0-9-]+)_(.+)$" % building.code, ent)
        if not m2:
            return None
        return "", m2.group(1), m2.group(2).replace("-", " ")
    lvl, num, name = building.norm_level(m.group(1)), m.group(2), m.group(3)
    d = re.match(r"^(\d+)([A-Z]?)$", num)
    if d:
        num = building.norm_room(d.group(1), d.group(2))
    return lvl, num, name.replace("-", " ")


def load_sheet(path, code):
    ws = openpyxl.load_workbook(path)[code]
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[11]:
            out[str(r[0]).strip()] = {
                "level": r[8], "ref": r[9], "name": r[10],
                "subject": r[11], "label": r[12], "notes": r[14],
            }
    return out


def join_tags(tags, entities, crosswalk, code):
    """Asset tag -> ontology entity, via the crosswalk then a normalised match."""
    cw = {}
    if os.path.isfile(crosswalk):
        for line in csv.DictReader(open(crosswalk)):
            cw[line["old_identifier"]] = line["new_identifier"]
    def norm(x):
        """Separators and leading zeros are not differences.

        The register writes CCU-B-001B and KEF-1F-0103 where the ontology
        writes SSC_CCUB0001B and SSC_KEF1F0103. Collapsing punctuation alone
        left eight SSC units unjoined; collapsing the padding too joins them.
        """
        x = re.sub(r"[^A-Za-z0-9]", "", x.upper())
        return re.sub(r"0+(\d)", r"\1", x)

    def without_level(x):
        """The same tag with a level segment dropped.

        The register writes KEF-1F-0103 and TEF-1F-0101 where the ontology
        writes SSC_KEF0103 and SSC_TEF0101 - every other member of those two
        families matches exactly, and the odd ones differ only by the level
        they sit on. Used only when the direct match finds nothing and the
        result is unique.
        """
        parts = [p for p in re.split(r"[^A-Za-z0-9]+", x) if p]
        kept = [p for p in parts if not re.fullmatch(r"\d?F|B\d?", p.upper())]
        return norm("".join(kept)) if len(kept) < len(parts) else None

    by_norm = collections.defaultdict(list)
    for e in entities:
        by_norm[norm(e.replace("entity:%s" % code, ""))].append(e)
    out, loosened = {}, []
    for t in tags:
        cand = [f for f in ("entity:%s_%s" % (code, t),
                            cw.get("entity:%s_%s" % (code, t), ""))
                if f in entities] or by_norm.get(norm(t), [])
        if not cand:
            alt = without_level(t)
            if alt:
                cand = by_norm.get(alt, [])
                if len(cand) == 1:
                    loosened.append((t, cand[0]))
        if len(cand) == 1:
            out[t] = cand[0]
    if loosened:
        print("  joined by dropping a level segment from the tag: %s"
              % ", ".join("%s = %s" % (t, e) for t, e in loosened))
    return out


def target(building, rename, per_ref, by_ref, m):
    """The room entity a retargeted row should name.

    The ontology's own room, renamed, whenever it has exactly one at that
    reference; the sheet's subject only where the ontology has no room there.
    """
    ref = (m["level"], m["ref"])
    here = per_ref.get(ref, [])
    if len(here) == 1:
        return rename[here[0]]
    return m["subject"]


def level_entity(levels, building, lvl):
    """The level entity this building already uses.

    QNL calls its basement entity:QNL_B; SSC calls the same thing
    entity:SSC_Level-B1. Guessing entity:SSC_B put three new rooms under a
    level that does not exist, which the validator caught as orphans.
    """
    code = building.code
    for cand in ("entity:%s_%s" % (code, building.level_segment(lvl)),
                 "entity:%s_%s" % (code, lvl),
                 "entity:%s_Level-%s" % (code, lvl)):
        if cand in levels:
            return cand
    raise SystemExit(
        "no level entity in the ontology for %r - tried %s. The new room "
        "cannot be attached until the level is declared."
        % (lvl, ", ".join("entity:%s_%s" % (code, x)
                          for x in (building.level_segment(lvl), lvl))))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--code", default="QNL", choices=sorted(B.BUILDINGS))
    ap.add_argument("--ontology", default="projects/QNL/QNL_Ontology.xlsx")
    ap.add_argument("--names", default="projects/dar-cairo-room-naming/Room_Names.xlsx")
    ap.add_argument("--crosswalk", default="projects/QNL/QNL_naming_crosswalk.csv")
    ap.add_argument("--out", required=True)
    ap.add_argument("--force", action="store_true",
                    help="run even if the ontology already carries the new "
                         "room shape")
    ap.add_argument("--crosswalk-out")
    ap.add_argument("--log")
    args = ap.parse_args()

    code = args.code
    building = B.BUILDINGS[code]
    wb = openpyxl.load_workbook(args.ontology)
    ws = pick_sheet(wb)
    data = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]

    rooms, level_of, rows_by_level, level_type = set(), {}, set(), {}
    loc, feeds = {}, collections.defaultdict(list)
    for r in data:
        if r[1] == "rec:Room":
            rooms.add(str(r[0]))
            if r[2] == "rec:isPartOf":
                level_of[str(r[0])] = (str(r[3]), str(r[4]))
        if r[1] in ("rec:Level", "rec:BasementLevel", "rec:RoofLevel"):
            rows_by_level.add(str(r[0]))
            level_type[str(r[0])] = r[1]
        if r[4] in ("rec:Level", "rec:BasementLevel", "rec:RoofLevel"):
            rows_by_level.add(str(r[3]))
        if r[4] == "rec:Room":
            rooms.add(str(r[3]))
        if r[2] == "rec:locatedIn":
            loc[str(r[0])] = str(r[3])
        if r[2] == "rec:feeds":
            feeds[str(r[0])].append(str(r[3]))

    sheet = load_sheet(args.names, code)
    by_ref = {}
    for tag, m in sheet.items():
        by_ref[(m["level"], m["ref"])] = m

    # ---- 1. the rename map ------------------------------------------------
    parsed = {e: parse_room(building, e) for e in rooms}
    unparsed = [e for e, p in parsed.items() if p is None]
    # QNL_Ontology.xlsx now IS the converted file - this script overwrote it -
    # so a second run would find no old-shape rooms to rename and would quietly
    # produce a copy with nothing done. Stop instead of pretending to work.
    if rooms and len(unparsed) > len(rooms) / 2 and not args.force:
        raise SystemExit(
            "%d of %d rooms in %s are already in the new shape - this looks "
            "like a file this script has already converted. Re-run it against "
            "the pre-conversion ontology from git history, or pass --force."
            % (len(unparsed), len(rooms), args.ontology))
    per_ref = collections.defaultdict(list)
    for e, p in parsed.items():
        if p:
            per_ref[p[:2]].append(e)

    # The shape changes; the name does not. The ontology's room names have
    # been cleaned and the register's have not - taking the sheet's name would
    # have pushed ARABIC SUDIES over ARABIC STUDIES, TRANSH CHAMBER over TRASH
    # CHAMBER and SHIPPING CLIRK over SHIPPING CLERK into a delivered model.
    # The sheet's authority is which room a unit is in, not what it is called.
    rename = {}
    for e, p in parsed.items():
        if p is None:
            continue
        lvl, num, name = p
        rename[e] = B.subject(building, lvl, num, name)
    clash = {v: k for v, k in
             ((v, [a for a in rename if rename[a] == v]) for v in set(rename.values()))
             if len(k) > 1}
    if clash:
        raise SystemExit("rename would merge rooms: %s" % clash)

    # ---- 2. the equipment that moved --------------------------------------
    joined = join_tags(sheet, set(loc) | set(feeds), args.crosswalk, code)
    move_loc, move_feeds, kept_feeds, odd_location, placeholder = {}, {}, [], [], []
    for tag, ent in joined.items():
        m = sheet[tag]
        want = (m["level"], m["ref"])
        cur = loc.get(ent)
        if not cur:
            continue
        # A unit that already feeds somewhere other than where it sits is
        # carrying a reading this sheet does not have - the sheet gives one
        # room per unit, and on every one of these it is the location. Leave
        # the feeds row alone whether or not the location moves.
        diverges = ent in feeds and cur not in feeds[ent]
        if diverges:
            kept_feeds.append((tag, ent, cur, feeds[ent]))
        here = parse_room(building, cur)
        if here is None:
            if cur in rows_by_level:
                # Located on a level rather than in a room, which is a
                # deliberate modelling choice, not something to reinterpret.
                odd_location.append((tag, ent, cur, "left - it is a level"))
                continue
            # Neither a room nor a level: a placeholder. SSC points five
            # exhaust fans at entity:Level7_Office0367, which carries no label
            # and names no SSC room. A real room from the register beats it.
            placeholder.append((tag, ent, cur, m["subject"]))
            move_loc[ent] = (cur, m)
            if ent in feeds and not diverges:
                move_feeds[ent] = (cur, m)
            continue
        if here[:2] == want:
            continue
        move_loc[ent] = (cur, m)
        if ent in feeds and not diverges:
            move_feeds[ent] = (cur, m)

    # ---- 3. rooms the sheet needs and the ontology lacks -------------------
    have = set(per_ref)
    missing = {k: v for k, v in by_ref.items() if k not in have}

    # ---- write ------------------------------------------------------------
    log = []
    changed_cells = 0
    for row in ws.iter_rows(min_row=2):
        subj, obj = row[0], row[3]
        if isinstance(subj.value, str) and subj.value in rename \
                and rename[subj.value] != subj.value:
            subj.value = rename[subj.value]
            changed_cells += 1
        if isinstance(obj.value, str) and obj.value in rename \
                and rename[obj.value] != obj.value:
            obj.value = rename[obj.value]
            changed_cells += 1

    retargeted = 0
    for row in ws.iter_rows(min_row=2):
        s = row[0].value
        if not isinstance(s, str):
            continue
        pred = row[2].value
        if pred == "rec:locatedIn" and s in move_loc:
            was, m = move_loc[s]
            row[3].value = target(building, rename, per_ref, by_ref, m)
            row[4].value = "rec:Room"
            log.append(("rec:locatedIn", s, rename.get(was, was), row[3].value))
            retargeted += 1
        elif pred == "rec:feeds" and s in move_feeds:
            was, m = move_feeds[s]
            if row[3].value in (was, rename.get(was)):
                row[3].value = target(building, rename, per_ref, by_ref, m)
                row[4].value = "rec:Room"
                log.append(("rec:feeds", s, rename.get(was, was), row[3].value))
                retargeted += 1

    # one label per renamed room reference that the sheet restates
    relabelled = 0        # names are unchanged, so labels are too

    added = 0
    for (lvl, ref), m in sorted(missing.items()):
        lvl_ent = level_entity(rows_by_level, building, lvl)
        lvl_type = level_type.get(lvl_ent,
                                  "rec:BasementLevel" if lvl.startswith("B")
                                  else "rec:Level")
        ws.append([m["subject"], "rec:Room", "rec:isPartOf", lvl_ent, lvl_type,
                   "rdfs:label_en", m["label"], "", ""])
        log.append(("new room", m["subject"], "", m["label"]))
        added += 1

    wb.save(args.out)
    print("room subjects renamed      : %d (shape only - every name kept)"
          % len(rename))
    print("cells rewritten            : %d" % changed_cells)
    print("locatedIn / feeds moved    : %d" % retargeted)
    print("feeds deliberately left    : %d" % len(kept_feeds))
    print("room labels restated       : %d" % relabelled)
    print("rooms added                : %d" % added)
    if placeholder:
        print("locatedIn was a placeholder : %d moved to a real room %s"
              % (len(placeholder), [x[0] for x in placeholder[:6]]))
    if odd_location:
        print("locatedIn is a level, left  : %d %s"
              % (len(odd_location), [x[0] for x in odd_location[:6]]))
    if unparsed:
        print("room entities not parsed   : %d %s" % (len(unparsed), unparsed[:4]))
    print("wrote %s" % args.out)

    if args.crosswalk_out:
        with open(args.crosswalk_out, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["old_subject", "new_subject", "named_by"])
            for k in sorted(rename):
                p = parse_room(building, k)
                src = ("naming sheet"
                       if len(per_ref[p[:2]]) == 1 and p[:2] in by_ref
                       else "ontology (reference holds several rooms)")
                w.writerow([k, rename[k], src])
        print("wrote %s" % args.crosswalk_out)

    if args.log:
        with open(args.log, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["change", "subject", "was", "now"])
            w.writerows(log)
            w.writerow([])
            w.writerow(["feeds left untouched - the unit already serves "
                        "a room other than the one it sits in"])
            w.writerow(["tag", "entity", "sits in", "feeds"])
            for t, e, c, fs in kept_feeds:
                w.writerow([t, e, c, "; ".join(fs)])
            if placeholder:
                w.writerow([])
                w.writerow(["locatedIn pointed at a placeholder entity, not a "
                            "room, and was moved to the register's room"])
                w.writerow(["tag", "entity", "pointed at", "now"])
                w.writerows(placeholder)
            if odd_location:
                w.writerow([])
                w.writerow(["locatedIn left alone - it does not name a room "
                            "in this building's shape"])
                w.writerow(["tag", "entity", "points at", "sheet says"])
                w.writerows(odd_location)
        print("wrote %s" % args.log)


if __name__ == "__main__":
    main()
