#!/usr/bin/env python3
"""Add manufacturer metadata triples to the QNL ontology.

Reads the reviewed metadata workbook (QNL_Needed_For_Ontology.xlsx) and writes
its 16 predicates onto the equipment already in the QNL ontology, following the
row shapes Dar Cairo, QF SSC and QF HQ use for equipment metadata:

  * a LITERAL predicate (rec:modelNumber, rec:manufacturedBy, rec:installationDate)
    rides an existing triple as a subject_prop on the equipment, or as an
    object_prop on a component's brick:hasPart row.
  * a QUANTITY predicate (capacities, flows, power, head, speed, ...) is its own
    triple whose object is a <blanknode> carrying brick:value and brick:hasUnit.

Component-level properties (an AHU's cooling coil, supply/return fans and their
drive motors, a pump's motor, a DX outdoor unit) attach to a component
sub-entity created with brick:hasPart, exactly as Dar Cairo attaches
coolingCapacity to entity:<AHU>_CHW-Coil and ratedPowerInput to entity:<fan>_Motor.
Unit-level properties attach to the equipment itself.

The join from a datasheet tag to an ontology entity is by family (see NORMALISE).
Families whose datasheet tags do not correspond to the ontology's BMS tags -
FCU (Euroclima selection-sheet positions), the closed control units (design tags
CC/B/0n vs the BMS CCU_808n), the pressurisation unit and the museum climate
unit - are left without metadata and reported, rather than guessed.

Input  : a base QNL_Ontology.xlsx (default: the one beside this script)
Output : QNL_Ontology.xlsx (+ .csv), QNL_metadata_join_report.csv
"""
import argparse
import collections
import csv
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, "..", ".."))
sys.path.insert(0, os.path.join(REPO, "tools", "equipment-metadata"))
from ontology_map import classify, is_literal  # noqa: E402

HEADER = ["subject", "subjectType", "predicate", "object", "objectType"] + \
    ["subject_prop_name", "subject_prop_val", "object_prop_name", "object_prop_val"] * 5 + \
    ["subject_prop_name", "subject_prop_val"]
assert len(HEADER) == 27

SUBJ_SLOTS = [(5, 6), (9, 10), (13, 14), (17, 18), (21, 22), (25, 26)]
OBJ_SLOTS = [(7, 8), (11, 12), (15, 16), (19, 20), (23, 24)]


def row(subject, stype, pred, obj, otype="", props=()):
    """props: (side, name, value); side 's' or 'o'."""
    cells = [subject, stype, pred, obj, otype] + [""] * 22
    su = iter(SUBJ_SLOTS)
    ob = iter(OBJ_SLOTS)
    for side, name, val in props:
        n, v = next(su) if side == "s" else next(ob)
        cells[n], cells[v] = name, val
    return cells


def add_props(cells, side, pairs):
    """Add (name, value) prop pairs to free slots of an existing row in place."""
    slots = SUBJ_SLOTS if side == "s" else OBJ_SLOTS
    for name, val in pairs:
        for n, v in slots:
            if not cells[n]:
                cells[n], cells[v] = name, val
                break
        else:
            raise RuntimeError("no free %s slot on row for %s" % (side, cells[0]))


def label(text):
    return " ".join(str(text).replace("_", " ").replace("-", " ").split())


# --------------------------------------------------------------------------- join
def NORMALISE(family, tag):
    """A datasheet tag -> the ontology entity id, or None when the family's tags
    do not correspond to the ontology's."""
    t = tag.strip()
    if family == "AHU":
        m = re.match(r"AHU-0*(\d+)$", t)
        return "entity:QNL_AHU_B_%03d" % int(m.group(1)) if m else None
    if family in ("CAV", "VAV"):
        segs = t.split("/")
        if len(segs) == 4 and segs[-1].isdigit():
            segs[-1] = "%03d" % int(segs[-1])
        return "entity:QNL_" + "_".join(segs)
    if family == "DX":
        m = re.match(r"DX/([A-Z]+)/0*(\d+)$", t)
        return "entity:QNL_DX_%s%02d" % (m.group(1), int(m.group(2))) if m else None
    if family == "HEX":                       # PHX/B/0n -> HEX0n, by index
        m = re.match(r"PHX/B/0*(\d+)$", t)
        return "entity:QNL_HEX%02d" % int(m.group(1)) if m else None
    if family == "PUMP":                       # CHWP/B/0n -> CHW_P0n
        m = re.match(r"CHWP/B/0*(\d+)$", t)
        return "entity:QNL_CHW_P%02d" % int(m.group(1)) if m else None
    if family == "GEN":
        return "entity:QNL_ELEC_Gen"
    if family == "EF":
        if re.match(r"[TK]EF_", t):            # TEF_B01A, KEF_101 - direct
            return "entity:QNL_" + t
        m = re.match(r"EF/([A-Z0-9]+)/0*(\d+)$", t)
        if m:
            seg = "RP" if m.group(1) == "RP" else m.group(1)
            return "entity:QNL_EF_%s%02d" % (seg, int(m.group(2)))
        return None
    return None                                # CCU, CLIMATE, PU: no join


def fcu_units(tag):
    """An FCU Euroclima selection-sheet tag names one or more BMS FCUs, as a
    level plus singles, comma-lists and hyphen-ranges: 'B/07', 'B/03,04',
    '1F/04-54'. Expand to the ontology FCU entities it covers. 2F tags use a
    numbering that does not correspond to the ontology's 2F FCUs, so they expand
    to entities that do not exist and fall through to the unmatched report."""
    t = re.sub(r"\(.*?\)", "", tag).replace("FCU/", "").replace("FCU", "").strip()
    m = re.match(r"(1F|2F|B)\s*/?\s*(.*)$", t)
    if not m:
        return []
    lvl, rest = m.group(1), m.group(2).strip()
    nums = []
    for part in rest.split(","):
        part = part.strip()
        rng = re.match(r"(\d+)\s*-\s*(\d+)$", part)
        if rng:
            nums += list(range(int(rng.group(1)), int(rng.group(2)) + 1))
        elif part.isdigit():
            nums.append(int(part))
    return ["entity:QNL_FCU_%s_%03d" % (lvl, n) for n in nums]


def EXPAND(family, tag):
    """A datasheet tag -> the ontology entity id(s) it covers. One for most
    families; an FCU selection sheet covers several units."""
    if family == "FCU":
        return fcu_units(tag)
    eid = NORMALISE(family, tag)
    return [eid] if eid else []


SHEET_FAMILY = {
    "AHU": "AHU", "CAV Units": "CAV", "VAV Units": "VAV", "DX Units": "DX",
    "Heat Exchangers": "HEX", "Pumps": "PUMP", "Generators": "GEN",
    "Exhaust Fans": "EF",
    # deliberately unmapped - reported, not guessed:
    "FCU": "FCU", "Closed Control Units": "CCU",
    "Climate Control Units": "CLIMATE", "Pressurization Unit": "PU",
}

# Component (exact 'Component' cell) -> (id suffix, brick class, parent suffix|None).
# None parent means the equipment itself. Only families that model components.
COMPONENTS = {
    "AHU": {
        "Cooling coil": ("_CHW-Coil", "brick:Chilled_Water_Coil", None),
        "External auxiliary cooling coil": ("_Aux-Coil", "brick:Chilled_Water_Coil", None),
        "Electric coil": ("_Electric-Coil", "brick:Heating_Coil", None),
        "Supply fan": ("_SF", "brick:Supply_Fan", None),
        "Return fan": ("_RF", "brick:Return_Fan", None),
        "Supply fan drive motor": ("_SF_Motor", "brick:Motor", "_SF"),
        "Return fan drive motor": ("_RF_Motor", "brick:Motor", "_RF"),
    },
    "PUMP": {
        "Drive motor": ("_Motor", "brick:Motor", None),
    },
    "DX": {
        "Refrigeration": ("_OD", "brick:Condensing_Unit", None),  # outdoor unit
    },
}


def value_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# --------------------------------------------------------------------------- base
def load_base(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Ontology"] if "Ontology" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = [["" if c is None else str(c).strip() for c in r]
            for r in ws.iter_rows(values_only=True)]
    rows = [r + [""] * (27 - len(r)) for r in rows]     # pad short rows
    return rows[0], rows[1:]


def build(base_path, meta_path, out_dir):
    hdr, base = load_base(base_path)

    # entity -> class, and entity -> its first row (host for subject-literals)
    equip_class, host_row = {}, {}
    for r in base:
        s = r[0]
        if s.startswith("entity:QNL_") and s not in host_row:
            host_row[s] = r
            equip_class[s] = r[1]
    equip = set(host_row)

    new_rows = []
    subentities = OrderedDict = collections.OrderedDict()   # sub_id -> (parent, cls, parent_suffix, obj_literals{})
    qty_seen = set()
    lit_seen = set()
    report = []          # (sheet, tag, entity_or_blank, status, detail)
    meta = openpyxl.load_workbook(meta_path, data_only=True)

    def sub_entity(parent, suffix, cls, parent_suffix):
        sid = parent + suffix
        if sid not in subentities:
            subentities[sid] = {"parent": parent, "cls": cls,
                                "parent_suffix": parent_suffix, "lits": []}
        return sid

    for sheet, family in SHEET_FAMILY.items():
        ws = meta[sheet]
        rows = list(ws.iter_rows(values_only=True))
        hi = next(i for i, r in enumerate(rows)
                  if r and "Equipment Tag" in [str(c) for c in r])
        h = {str(c): i for i, c in enumerate(rows[hi])}
        ti, ci, pi = h["Equipment Tag"], h["Component"], h["Property"]
        vi, ui = h["Value (Dar Cairo)"], h["Unit (QUDT)"]

        matched, unmatched = {}, set()   # tag -> [entity ids] ; {tags}
        for r in rows[hi + 1:]:
            if not r or not r[ti]:
                continue
            tag = str(r[ti]).strip()
            comp = str(r[ci]).strip() if r[ci] else ""
            prop = str(r[pi]).strip() if r[pi] else ""
            pred, scope, _ = classify(comp, prop)
            if scope != "core" or not pred:
                continue
            eids = [e for e in EXPAND(family, tag) if e in equip]
            if not eids:
                unmatched.add(tag)
                continue
            matched[tag] = eids

            val = value_str(r[vi] if r[vi] is not None else r[h.get("Value (as printed)", vi)])
            unit = str(r[ui]).strip() if r[ui] else "unit:UNITLESS"

            # one selection sheet can name several identical units - write to each
            for eid in eids:
                spec = COMPONENTS.get(family, {}).get(comp)
                if spec:
                    suffix, cls, parent_suffix = spec
                    owner = sub_entity(eid, suffix, cls, parent_suffix)
                    owner_cls = cls
                else:
                    owner, owner_cls = eid, equip_class[eid]

                if is_literal(pred):
                    if spec:
                        if (pred, val) not in subentities[owner]["lits"]:
                            subentities[owner]["lits"].append((pred, val))
                    else:
                        key = (owner, pred, val)
                        if key not in lit_seen:
                            add_props(host_row[owner], "s", [(pred, val)])
                            lit_seen.add(key)
                else:
                    key = (owner, pred, val, unit)
                    if key not in qty_seen:
                        new_rows.append(row(owner, owner_cls, pred, "<blanknode>", "<blanknode>",
                                            [("o", "brick:value", val),
                                             ("o", "brick:hasUnit", unit)]))
                        qty_seen.add(key)

        for t in sorted(matched):
            report.append((sheet, t, ";".join(matched[t]), "matched",
                           "%d unit(s)" % len(matched[t])))
        for t in sorted(unmatched):
            report.append((sheet, t, "", "unmatched",
                           "2F numbering does not match the ontology's 2F FCUs"
                           if family == "FCU" else
                           "no ontology entity for this tag" if family in
                           ("EF", "CAV", "VAV", "DX", "HEX", "PUMP")
                           else "family tags do not correspond to ontology tags"))

    # emit component brick:hasPart rows - non-motor (fans/coils) before motors,
    # so a fan is declared before the motor that hangs on it.
    def part_row(sid):
        info = subentities[sid]
        parent = info["parent"] if not info["parent_suffix"] else info["parent"] + info["parent_suffix"]
        parent_cls = equip_class.get(parent) or subentities.get(parent, {}).get("cls") \
            or (subentities[info["parent"] + info["parent_suffix"]]["cls"]
                if info["parent_suffix"] else equip_class[info["parent"]])
        props = [("o", "rdfs:label_en", label(sid.replace("entity:QNL_", "")))]
        for pred, val in info["lits"]:
            props.append(("o", pred, val))
        return row(parent, parent_cls, "brick:hasPart", sid, info["cls"], props)

    part_rows = [sid for sid in subentities if not subentities[sid]["parent_suffix"]] + \
                [sid for sid in subentities if subentities[sid]["parent_suffix"]]
    haspart = [part_row(sid) for sid in part_rows]

    out = base + haspart + new_rows
    os.makedirs(out_dir, exist_ok=True)
    # xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ontology"
    ws.append(hdr)
    for r in out:
        ws.append(r)
    wb.save(os.path.join(out_dir, "QNL_Ontology.xlsx"))
    # csv
    with open(os.path.join(out_dir, "QNL_Ontology.csv"), "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(hdr)
        w.writerows(out)
    # join report
    with open(os.path.join(out_dir, "QNL_metadata_join_report.csv"), "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["sheet", "tag", "entity", "status", "detail"])
        w.writerows(report)

    # summary
    by_status = collections.Counter(x[3] for x in report)
    print("base rows:      ", len(base))
    print("component parts:", len(haspart))
    print("quantity rows:  ", len(new_rows))
    print("literal props:  ", len(lit_seen), "on equipment +",
          sum(len(v["lits"]) for v in subentities.values()), "on components")
    print("total out rows: ", len(out))
    print("join report:    ", dict(by_status))
    print("\nunmatched by sheet:")
    um = collections.Counter(x[0] for x in report if x[3] == "unmatched")
    for s, n in um.most_common():
        print("   %-24s %d" % (s, n))
    return out


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--base", default=os.path.join(HERE, "QNL_Ontology.xlsx"))
    ap.add_argument("--meta", default=os.path.join(REPO, "QNL_Needed_For_Ontology.xlsx"))
    ap.add_argument("--out", default=HERE)
    a = ap.parse_args()
    build(a.base, a.meta, a.out)
