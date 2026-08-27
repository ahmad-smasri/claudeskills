#!/usr/bin/env python3
"""Build the QNL (Qatar National Library) PARA/Brick ontology sheet.

Inputs  : QNL_Room_Names_for_Ontology.xlsx, QNL_Assets_Location_Relationships.xlsx
Outputs : QNL_Ontology.xlsx (27-column triple sheet), QNL_identifier_crosswalk.csv
"""
import csv, os, re, sys
from collections import OrderedDict
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "datapoints"))
import qnl_datapoints
import qnl_instrumentation
from step1_confirm_datapoints import load_selected as _load_selected, \
    load_historian as _load_historian

SRC = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sources") + "/"
ROOMS_XLSX = SRC + "QNL_Room_Names_for_Ontology.xlsx"
ASSETS_XLSX = SRC + "QNL_Assets_Location_Relationships.xlsx"
OUT = os.path.dirname(os.path.abspath(__file__)) + "/"

HEADER = ["subject", "subjectType", "predicate", "object", "objectType"] + \
    ["subject_prop_name", "subject_prop_val", "object_prop_name", "object_prop_val"] * 5 + \
    ["subject_prop_name", "subject_prop_val"]
assert len(HEADER) == 27


def label(text):
    """Label style: QF SSC. The source text, with underscores as spaces.

    The current SSC sheet labels rooms `1.001 CORRIDOR` and its own assets
    `SSC_CHW_CHWP01 Motor` - the raw schedule and register strings with `_`
    read as a word break, and every other mark left alone: the dot between
    level and room number survives, and so do dashes and slashes (`A / V ROOM`).
    That is the difference from the PARA label rule, which strips punctuation
    outright.
    """
    return " ".join(text.replace("_", " ").split())


def row(subject, stype, pred, obj, otype="", props=()):
    """props is a sequence of (side, name, value); side is 's' or 'o'."""
    cells = [subject, stype, pred, obj, otype] + [""] * 22
    slots = [(5, 6, "s"), (7, 8, "o"), (9, 10, "s"), (11, 12, "o"),
             (13, 14, "s"), (15, 16, "o"), (17, 18, "s"), (19, 20, "o"),
             (21, 22, "s"), (23, 24, "o"), (25, 26, "s")]
    used = set()
    for side, name, val in props:
        for idx, (n, v, sd) in enumerate(slots):
            if idx in used or sd != side:
                continue
            cells[n], cells[v] = name, val
            used.add(idx)
            break
        else:
            raise RuntimeError(f"no free {side} prop slot for {name}")
    return cells


# --------------------------------------------------------------------------- rooms
LEVEL_RE = re.compile(r"^(B|L1|L2|T1|P)[_-]?(.*)$")
LEVEL_TYPE = {"B": "rec:BasementLevel", "L1": "rec:Level",
              "L2": "rec:Level", "T1": "rec:Level", "P": "rec:Level"}
# Identifiers keep the source level codes - they are the segment the room tags
# carry. Labels spell the level out, because that is what the front end shows.
LEVEL_LABEL = {"B": "Basement", "L1": "Level 1", "L2": "Level 2",
               "T1": "Terrace 1", "P": "Roof Plant"}

# The room schedule was typed by hand and carries misspellings that would
# otherwise ride into both the identifier and the label a user reads. Each
# entry below was settled against the rest of the schedule, not guessed: the
# correct spelling appears on a sibling room, or the token is a run-together
# pair whose separator convention every other room follows. The map is applied
# to the underscore-separated tokens of the room *name* only - never to the
# level or the room number, which are the join key back to the drawings.
#
# Whole-word misspellings; the evidence for each is in QNL_handover-note.md.
TYPO_FIXES = {
    "CARRLES": "CARRELS",              # CARRELS on 15 other rooms
    "CTRCULATION": "CIRCULATION",
    "GREEM": "GREEN",
    "SPRIMKLERS": "SPRINKLERS",
    "ITTIGATION": "IRRIGATION",
    "WASING": "WASHING",
    "CONTOL": "CONTROL",               # CONTROL on 8 other rooms
    "LEVLEL": "LEVEL",
    "LOBY": "LOBBY",                   # L1_042_LOBBY
    "VENTILATON": "VENTILATION",
    "VENTLATON": "VENTILATION",
    "LIBRARIA": "LIBRARIAN",           # 14 LIBRARIAN rooms
    "PANKING": "PARKING",
    "WATING": "WAITING",
    "MULTPURPOSE": "MULTIPURPOSE",
    "ANGUAGE": "LANGUAGE",
    "BIBLIOGRANHER1": "BIBLIOGRAPHER1",   # L1_082_BIBLIOGRAPHER2 next door
    "PUBLIS": "PUBLIC",                # L1_080_PUBLIC_SERVICE
    "CSECURITY": "SECURITY",           # 8 SECURITY rooms, stray leading C
    "TRANSH": "TRASH",                 # only waste room in the building
    "PRESEARCHERS": "RESEARCHERS",     # stray leading P; B_151 RESEARCH
    # Run-together pairs. The separator, not the wording, is what is restored -
    # every sibling room writes these two tokens apart.
    "ROOMMEN": "ROOM_MEN",             # 9 x REST_ROOM_MEN
    "ABLUTIONMEN": "ABLUTION_MEN",
    "ABLUTIONWOMEN": "ABLUTION_WOMEN",
    "ADPUBLIC": "AD_PUBLIC",           # AD_COLLECTIONS, AD_OFFICE, AD_ADMIN
    "LIBDIRECTORS": "LIB_DIRECTORS",
    "INDIVISTUDY": "INDIVI_STUDY",     # sibling shape is GROUP_STUDY_ROOM
}


def fix_spelling(name):
    """Correct the schedule's typing errors in a room name, token by token.

    Whole tokens only, so a correction can never fire inside a longer word,
    and abbreviations the schedule uses deliberately - AD, SEC, RES, LIBR,
    PERS, ITT and the rest - are left exactly as the source wrote them.
    """
    return "_".join(TYPO_FIXES.get(t, t) for t in name.split("_"))

wb = openpyxl.load_workbook(ROOMS_XLSX, read_only=True, data_only=True)
room_src = []
for r in list(wb.worksheets[0].iter_rows(values_only=True))[1:]:
    if not r[0]:
        continue
    room_src.append((str(r[0]).strip(), str(r[1]).strip(), str(r[2]).strip()))

rooms = OrderedDict()      # source entity name -> dict
notes = []
for src_entity, number, name in room_src:
    m = LEVEL_RE.match(number)
    if m:
        level, num = m.group(1), m.group(2)
    else:
        level, num = "B", number
        notes.append(f"room number {number!r} carries no level prefix; placed on level B "
                     f"because its ST-nn siblings are all basement rooms")
    # The source room schedule is inconsistent about how the level is joined to
    # the room number: most rows write B_034, a minority write B036_REST,
    # B-ST-01, L1023_1. The user asked for one shape throughout, so both the
    # identifier and the label are rebuilt from the same (level, num, name)
    # triple and cannot drift apart:
    #
    #   identifier  entity:QNL_<level>_<num>_<name>     QNL_B_063_PLANT_ROOM_01
    #   label             <level>.<num> <name>          B.063 PLANT ROOM 01
    #
    # The dot between level and room number is the QF SSC label shape - SSC
    # writes 1.001 CORRIDOR for room 001 on level 1 - and label() turns the
    # remaining underscores into spaces. Nothing inside <num> or <name> is
    # touched; only the join between the segments is regularised.
    if not num:
        sys.exit("room %r has a level but no room number" % number)
    fixed = fix_spelling(name)
    if fixed != name:
        notes.append(f"room name {name!r} corrected to {fixed!r}")
    name = fixed
    ident = "entity:QNL_%s_%s_%s" % (level, num, name)
    rooms[src_entity] = {
        "id": ident, "level": level,
        "label": label("%s.%s_%s" % (level, num, name)),
        "number": number, "name": name,
    }

dupes = [i for i in rooms if list(x["id"] for x in rooms.values()).count(rooms[i]["id"]) > 1]
if dupes:
    sys.exit("duplicate room identifiers: %s" % dupes[:10])

# Rooms an equipment Feeds column names but the room schedule does not list. The
# user directed that these be created following the sibling naming, since the
# feeds relationship needs a real room to point at. Only one this round:
# TEF_B02B feeds "RESTROOM (WOMEN) B.033" and B.033 is absent, though its
# men's counterpart B_032_REST_ROOM_MEN is present - so B_033_REST_ROOM_WOMEN
# is coined to match. Listed in the handover note.
CREATED_ROOMS = [("B", "033", "REST_ROOM_WOMEN")]
for level, num, name in CREATED_ROOMS:
    ident = "entity:QNL_%s_%s_%s" % (level, num, name)
    key = ident  # no source entity - it is invented here
    if key not in rooms:
        rooms[key] = {"id": ident, "level": level,
                      "label": label("%s.%s_%s" % (level, num, name)),
                      "number": "%s_%s" % (level, num), "name": name,
                      "created": True}
        notes.append("room %s created: named by an equipment Feeds column but "
                     "absent from the room schedule" % ident)

# --------------------------------------------------------------------------- assets
CLASS = {"AHUB": "brick:Air_Handling_Unit",
         "VAV": "brick:Variable_Air_Volume_Box",
         "CAV": "brick:Constant_Air_Volume_Box",
         "FCU": "brick:Fan_Coil_Unit"}
TERMINAL = {"VAV", "CAV", "FCU"}
# QF SSC 0.5 names its loop entity:CHWS-MAIN-LOOP, types it
# para:Chilled_Water_Loop_Network and gives it two subject rows - rec:locatedIn
# the building, and an IFC reference. QNL follows that shape.
#
# The building code is the one departure. SSC's loop carries none, but it is
# rec:locatedIn entity:SSC, so a QNL loop under the same bare name would be one
# entity located in two buildings once the converter loads both sheets into one
# graph. Site-level systems - entity:HVAC, entity:QF - are genuinely shared and
# rightly bare; a per-building main loop is not. Flagged in the handover note.
LOOP = "entity:QNL_CHWS-MAIN-LOOP"

# The systems layer, which is what the front end's system tree is built from.
#
# Dar Cairo's shape: a top-level system is a subject carrying brick:isPartOf the
# SITE and a label; a sub-system is declared only as the object of its parent's
# brick:hasPart row, with its class in objectType and its label in an object
# prop; equipment points up with brick:isPartOf. No link is ever stated twice.
# QF SSC 0.5 agrees - entity:HVAC brick:isPartOf entity:QF.
#
# entity:CHW-System is not a node QNL invents - QF SSC 0.5 already declares it,
# bare and site-level under entity:HVAC, holding SSC's four chilled water booster
# pumps and five heat exchangers. QNL reusing it puts its loop in that same
# group rather than creating a one-child node, the same way both buildings share
# entity:QF and entity:HVAC. SSC leaves its own loop outside CHW-System, which
# looks like an oversight: a distribution loop is part of the chilled water
# system by any reading.
# brick:HVAC_System is a Brick 1.4 alias for
# brick:Heating_Ventilation_Air_Conditioning_System, and the ladder would take the
# preferred term. The house overrides that: both reference models write
# brick:HVAC_System and the front end keys off it, so consistency across the
# estate wins over the preferred spelling. The W-TYP-5 alias warning is accepted
# and recorded in the handover.
HVAC = "entity:HVAC"
CHW = "entity:CHW-System"
# The generator is electrical, not HVAC. QF SSC declares a shared site-level
# entity:Electrical_System (brick:System, isPartOf the site); QNL reuses it the
# same way it reuses entity:HVAC and entity:CHW-System, rather than minting a
# building-local electrical node for a single asset.
ELEC = "entity:Electrical_System"
ELEC_CLASS = "brick:System"

NEW_FAMILIES = {
    "DX":    {"cls": "para:DXUnit",                     "system": HVAC, "feeds": "self"},
    "CCU":   {"cls": "brick:CRAC",                      "system": HVAC, "feeds": "self"},
    "EF":    {"cls": "brick:Exhaust_Fan",               "system": HVAC, "feeds": "end"},
    "TEF":   {"cls": "brick:Exhaust_Fan",               "system": HVAC, "feeds": "end"},
    "KEF":   {"cls": "brick:Exhaust_Fan",               "system": HVAC, "feeds": "end"},
    "SEF":   {"cls": "para:Smoke_Extract_Fan",          "system": HVAC, "feeds": "end"},
    "HEX":   {"cls": "brick:Heat_Exchanger",            "system": CHW,  "feeds": "none"},
    "CHW":   {"cls": "brick:Chilled_Water_Booster_Pump","system": CHW,  "feeds": "loop"},
    "CHWPU": {"cls": "brick:Chilled_Water_Booster_Pump","system": CHW,  "feeds": "loop"},
    "ELEC":  {"cls": "para:Generator",                  "system": ELEC, "feeds": "none"},
}

NEW_EXT_CLASSES = [
    ("para:DXUnit", "brick:HVAC_Equipment", "DX Unit"),
    ("para:Generator", "brick:Electrical_Equipment", "Generator"),
    ("para:Smoke_Extract_Fan", "brick:Exhaust_Fan", "Smoke Extract Fan"),
]

CHW_CLASS = "brick:Chilled_Water_System"
HVAC_CLASS = "brick:HVAC_System"
LOOP_CLASS = "para:Chilled_Water_Loop_Network"


def equip_id(tag):
    """Prefix the register tag with the building code, QF SSC style.

    SSC subjects read entity:SSC_FCU0001 - building code, then the register tag.
    QNL follows: FCU_1F_056 becomes entity:QNL_FCU_1F_056. Tags are otherwise not
    touched, so they stay the BMS join key.

    The one exception is the AHUB family, at the user's direction. AHUB002 packs
    type and level into a single token and carries no separators, where VAV, CAV
    and FCU are all TYPE_LEVEL_COUNT; it is rewritten AHU_B_002 so all four
    families parse by one rule.
    """
    m = re.match(r"^AHUB(\d+)$", tag)
    if m:
        tag = "AHU_B_%s" % m.group(1)
    return "QNL_" + tag


wb = openpyxl.load_workbook(ASSETS_XLSX, read_only=True, data_only=True)
ahub_tags = {str(r[0]).strip() for r in list(wb["AHUB"].iter_rows(values_only=True))[1:]
             if r[0]}


def derive_fed_by(tag):
    """Recover the Fed By value the register's own formula computes.

    The current register does not store Fed By as text. Column C of the VAV and
    CAV sheets holds a formula that reads the `S<nn>` segment out of the
    equipment tag and looks up `AHUB<nnn>` - VAV_B_S11_024 resolves to AHUB011.
    The workbook was saved without a formula cache, so openpyxl's data_only read
    returns None for every one of those cells, which would silently drop 297
    rec:isFedBy rows. Deriving it here keeps the sheet identical either way.

    Checked against the previous register, which stored the same column as
    literals: the derivation reproduces all 297 values exactly, with no
    disagreements and no unresolved tags.
    """
    m = re.match(r"^[A-Z]+_[A-Z0-9]+_S(\d+)_", tag)
    if not m:
        return ""
    cand = "AHUB%03d" % int(m.group(1))
    return cand if cand in ahub_tags else ""


assets = []
unknown_rooms = set()
derived = 0
for ws in wb.worksheets:
    kind = ws.title.strip()
    if kind not in CLASS:
        # the register carries the reviewer's own working sheets - a Claude Log
        # and an AHU-VAV Check pivot. Only the four equipment sheets are data.
        if kind not in NEW_FAMILIES:
            notes.append(f"register sheet {kind!r} is not an equipment family; skipped")
        continue
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        tag, room_tag = str(r[0]).strip(), str(r[1]).strip()
        fed_by = str(r[2]).strip() if r[2] is not None else ""
        if not fed_by:
            fed_by = derive_fed_by(tag)
            if fed_by:
                derived += 1
        if room_tag not in rooms:
            # the room list carries one entry with a trailing space
            match = [k for k in rooms if k.strip() == room_tag.strip()]
            if not match:
                unknown_rooms.add(room_tag)
                continue
            room_tag = match[0]
        assets.append({"kind": kind, "tag": tag, "id": "entity:" + equip_id(tag),
                       "cls": CLASS[kind], "room": rooms[room_tag]["id"],
                       "fed_by": fed_by})
if unknown_rooms:
    sys.exit("assets reference rooms absent from the room list: %s" % sorted(unknown_rooms))
if derived:
    notes.append(f"{derived} Fed By values were empty in the register and were "
                 f"derived from the equipment tag; see derive_fed_by()")

# Resolve a free-text Feeds cell - "UPS BATERRY ROOM B.084", "Plant Zone P.004" -
# to a room entity. The trailing <level>.<number> token is the key; the room
# schedule's own number formats vary (B_084, B046_RES, P_004), so rooms are
# indexed by (level, integer) and the feeds token parsed the same way. Where a
# number maps to more than one room (B046 is both a restroom and a control
# room), a keyword shared between the Feeds text and the room name breaks the
# tie; anything still ambiguous or missing is reported, never guessed.
_ROOMBYNUM = {}
for _d in rooms.values():
    _m = LEVEL_RE.match(_d["number"])
    if not _m:
        continue
    _lvl = _m.group(1)
    _dig = re.search(r"\d+", _m.group(2) or "")
    if _dig:
        _ROOMBYNUM.setdefault((_lvl, int(_dig.group())), []).append(_d)

_FEEDS_REF = re.compile(r"\b(B|L1|L2|T1|P)\s*[.\-_]?\s*(\d+)\b", re.I)


def resolve_feeds(text, equip_tag):
    m = list(_FEEDS_REF.finditer(text))
    if not m:
        sys.exit("cannot find a room number in Feeds %r on %s" % (text, equip_tag))
    lvl, num = m[-1].group(1).upper(), int(m[-1].group(2))
    cands = _ROOMBYNUM.get((lvl, num), [])
    if not cands:
        sys.exit("Feeds %r on %s points at %s.%d, which no room matches - "
                 "define it or fix the cell" % (text, equip_tag, lvl, num))
    if len(cands) == 1:
        return cands[0]["id"]
    tw = {w for w in re.split(r"[^A-Za-z]+", text.upper()) if len(w) > 2}
    scored = sorted(cands, key=lambda d: -len(
        tw & {w for w in re.split(r"[^A-Za-z]+", d["name"].upper()) if len(w) > 2}))
    return scored[0]["id"]


# The new families: read each sheet, resolve locatedIn and (where applicable)
# the end room its Feeds column names. Tags keep the source name with the QNL_
# prefix, exactly as the historian writes them, so the datapoint join later is
# one-to-one.
new_assets = []
new_unknown = set()
for kind, spec in NEW_FAMILIES.items():
    if kind not in wb.sheetnames:
        notes.append("register has no %r sheet; skipped" % kind)
        continue
    for r in list(wb[kind].iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        tag = str(r[0]).strip()
        room_tag = str(r[1]).strip() if r[1] is not None else ""
        feeds_cell = str(r[2]).strip() if len(r) > 2 and r[2] is not None else ""
        if room_tag not in rooms:
            match = [k for k in rooms if k.strip() == room_tag.strip()]
            if not match:
                new_unknown.add(room_tag)
                continue
            room_tag = match[0]
        loc = rooms[room_tag]["id"]
        if spec["feeds"] == "self":
            feeds = loc
        elif spec["feeds"] == "loop":
            feeds = LOOP   # CHW pumps feed the chilled-water loop, per Dar Cairo
        elif spec["feeds"] == "end" and feeds_cell:
            feeds = resolve_feeds(feeds_cell, tag)
        else:
            feeds = ""     # plant equipment, or an area-only SEF with no Feeds
        new_assets.append({"kind": kind, "tag": tag, "id": "entity:QNL_" + tag,
                           "cls": spec["cls"], "system": spec["system"],
                           "room": loc, "feeds": feeds})
if new_unknown:
    sys.exit("new-family assets reference rooms absent from the room list: %s"
             % sorted(new_unknown))
_allids = [a["id"] for a in new_assets]
if len(set(_allids)) != len(_allids):
    sys.exit("duplicate new-family identifiers")

by_tag = {a["tag"]: a for a in assets}
for a in assets:
    if a["fed_by"].upper() == "CHILLED WATER LOOP":
        a["src"], a["src_cls"] = LOOP, LOOP_CLASS
    elif a["fed_by"] in by_tag:
        a["src"], a["src_cls"] = by_tag[a["fed_by"]]["id"], by_tag[a["fed_by"]]["cls"]
    else:
        sys.exit("unresolved Fed By value %r on %s" % (a["fed_by"], a["tag"]))

ids = [a["id"] for a in assets]
if len(set(ids)) != len(ids):
    sys.exit("duplicate equipment identifiers")

# --------------------------------------------------------------------------- new families
# The 21-Aug register added ten more equipment families beyond AHU/VAV/CAV/FCU.
# Each row below carries the class the ladder produced (Dar Cairo -> Brick 1.4 ->
# SSC -> para:, in that order) and how its rec:feeds is derived:
#   'self'  the unit conditions the room it sits in (DX, CCU) - feeds = locatedIn
#   'end'   an exhaust fan serving a room named in the register's Feeds column
#   'none'  plant equipment with no feeds (HEX, CHW pumps, generator), and the
#           three area-only SEFs the user flagged as representing zones, not feeds
#
# Classes, with where each came from:
#   brick:CRAC                     CCU  - Dar Cairo (118) and SSC (44)
#   brick:Exhaust_Fan              EF/TEF/KEF - Dar Cairo (169), SSC
#   brick:Heat_Exchanger           HEX  - Dar Cairo (10)
#   brick:Chilled_Water_Booster_Pump  CHW/CHWPU - Dar Cairo (77), SSC (36)
#   para:DXUnit                    DX   - SSC precedent (subclass of HVAC_Equipment)
#   para:Generator                 ELEC - Dar Cairo precedent (Electrical_Equipment)
#   para:Smoke_Extract_Fan         SEF  - newly minted; datapoints read "Smoke
#                                         Extract Fan"; no term in Dar Cairo,
#                                         Brick 1.4 or SSC, so a subclass of the
#                                         closest parent brick:Exhaust_Fan

# --------------------------------------------------------------------------- rows
out = []

# Extensions -----------------------------------------------------------------
out.append(row(LOOP_CLASS, "owl:Class", "rdfs:subClassOf", "brick:HVAC_Equipment", "",
               [("s", "rdfs:label_en", "Chilled Water Loop Network")]))

# Spatial --------------------------------------------------------------------
# The site is the organisation's code, not its spelled-out name, and the current
# QF SSC sheet writes exactly this row for its own building:
#   entity:SSC | rec:Building | rec:isPartOf | entity:QF | rec:Site
#              | rdfs:label_en | SSC Building | rdfs:label_en | Qatar Foundation
# QNL sits under the same site, so it reuses entity:QF rather than minting a
# second name for it - that is what lets the two buildings' sheets join.
out.append(row("entity:QNL", "rec:Building", "rec:isPartOf",
               "entity:QF", "rec:Site",
               [("s", "rdfs:label_en", "QNL Building"),
                ("o", "rdfs:label_en", "Qatar Foundation")]))

levels_used = sorted({d["level"] for d in rooms.values()})
for lvl in levels_used:
    out.append(row("entity:QNL_%s" % lvl, LEVEL_TYPE[lvl], "rec:isPartOf",
                   "entity:QNL", "rec:Building",
                   [("s", "rdfs:label_en", LEVEL_LABEL[lvl])]))

for d in rooms.values():
    out.append(row(d["id"], "rec:Room", "rec:isPartOf",
                   "entity:QNL_%s" % d["level"], LEVEL_TYPE[d["level"]],
                   [("s", "rdfs:label_en", d["label"])]))

# Chilled water loop ---------------------------------------------------------
out.append(row(HVAC, HVAC_CLASS, "brick:isPartOf", "entity:QF", "rec:Site",
               [("s", "rdfs:label_en", "HVAC System")]))

out.append(row(CHW, CHW_CLASS, "brick:isPartOf", HVAC, HVAC_CLASS,
               [("s", "rdfs:label_en", "Chilled Water System")]))
out.append(row(LOOP, LOOP_CLASS, "brick:isPartOf", CHW, CHW_CLASS))
out.append(row(LOOP, LOOP_CLASS, "rec:locatedIn", "entity:QNL", "rec:Building",
               [("s", "rdfs:label_en", label(LOOP.replace("entity:", "")))]))
out.append(row(LOOP, LOOP_CLASS, "ref:hasExternalReference",
               "<blanknode>", "ref:IFCReference",
               [("o", "para:IFC_ID", ""),
                ("o", "ref:ifcName", LOOP.replace("entity:", ""))]))

# Equipment ------------------------------------------------------------------
for kind in ("AHUB", "VAV", "CAV", "FCU"):
    for a in [x for x in assets if x["kind"] == kind]:
        bare = a["id"].replace("entity:", "")
        out.append(row(a["id"], a["cls"], "rec:locatedIn", a["room"], "rec:Room",
                       [("s", "rdfs:label_en", label(bare))]))
        out.append(row(a["id"], a["cls"], "brick:isPartOf", HVAC, HVAC_CLASS))
        out.append(row(a["id"], a["cls"], "rec:isFedBy", a["src"], a["src_cls"]))
        if kind in TERMINAL:
            out.append(row(a["id"], a["cls"], "rec:feeds", a["room"], "rec:Room"))
        # SSC shape: the IFC reference carries both properties. para:IFC_ID is the
        # slot for the real IFC GUID and is left empty until BIM supplies it;
        # ref:ifcName is the entity name, which is derivable.
        out.append(row(a["id"], a["cls"], "ref:hasExternalReference",
                       "<blanknode>", "ref:IFCReference",
                       [("o", "para:IFC_ID", ""), ("o", "ref:ifcName", bare)]))
        # No timeseries reference row here. A ref:TimeseriesReference belongs to a
        # POINT, never to the equipment: all 1,767 of them in QF SSC hang off a
        # brick:hasPoint object and none off a piece of equipment. QNL has no
        # points because no IO list was supplied, so it has no timeseries refs.

# New-family extension classes -----------------------------------------------
for cls, parent, lab in NEW_EXT_CLASSES:
    out.append(row(cls, "owl:Class", "rdfs:subClassOf", parent, "",
                   [("s", "rdfs:label_en", lab)]))

# The electrical system node the generator hangs under, declared the way SSC
# declares it - a site-level brick:System under entity:QF.
if any(a["system"] == ELEC for a in new_assets):
    out.append(row(ELEC, ELEC_CLASS, "brick:isPartOf", "entity:QF", "rec:Site",
                   [("s", "rdfs:label_en", "Electrical System")]))

# New-family equipment -------------------------------------------------------
SYS_CLASS = {HVAC: HVAC_CLASS, CHW: CHW_CLASS, ELEC: ELEC_CLASS}
for kind in NEW_FAMILIES:
    for a in [x for x in new_assets if x["kind"] == kind]:
        bare = a["id"].replace("entity:", "")
        out.append(row(a["id"], a["cls"], "rec:locatedIn", a["room"], "rec:Room",
                       [("s", "rdfs:label_en", label(bare))]))
        out.append(row(a["id"], a["cls"], "brick:isPartOf",
                       a["system"], SYS_CLASS[a["system"]]))
        if a["feeds"] == LOOP:
            out.append(row(a["id"], a["cls"], "rec:feeds", LOOP, LOOP_CLASS))
        elif a["feeds"]:
            out.append(row(a["id"], a["cls"], "rec:feeds", a["feeds"], "rec:Room"))
        out.append(row(a["id"], a["cls"], "ref:hasExternalReference",
                       "<blanknode>", "ref:IFCReference",
                       [("o", "para:IFC_ID", ""), ("o", "ref:ifcName", bare)]))

# Datapoints and parts for the new families -----------------------------------
# The ledger step is done internally, priority Dar Cairo -> Brick 1.4 -> SSC ->
# para:. See qnl_datapoints.py; the resolved ledger is written for review.
_dp_rows, _dp_ledger, _para_used = qnl_datapoints.build_datapoints(
    new_assets, _load_selected(), _load_historian(), row, label)

# declare every para: class the datapoints use, as an owl:Class subclass, so the
# sheet is self-contained. The three family classes (DXUnit/Generator/
# Smoke_Extract_Fan) were declared above; skip any repeats.
_declared = {c for c, _, _ in NEW_EXT_CLASSES}
for _cls in sorted(_para_used):
    if _cls in _declared:
        continue
    out.append(row(_cls, "owl:Class", "rdfs:subClassOf", _para_used[_cls], "",
                   [("s", "rdfs:label_en",
                     qnl_datapoints.PARA_LABEL.get(_cls,
                         label(_cls.split(":", 1)[-1])))]))
    _declared.add(_cls)
out.extend(_dp_rows)

with open(OUT + "QNL_newfamily_datapoint_ledger.csv", "w", newline="") as fh:
    w = csv.DictWriter(fh, fieldnames=["equip", "kind", "token", "class",
                                       "source", "unit", "desc"])
    w.writeheader()
    w.writerows(_dp_ledger)

# AHU/VAV/CAV/FCU points, taken verbatim from the reviewed sheet built for those
# four families earlier - the ledger step for them was already done and approved,
# so they are lifted in rather than regenerated. Only point rows are taken (the
# equipment rows already exist above), and only for equipment that exists in this
# build - the three orphan assets that carry telemetry but no register row are
# left out, since there is no subject here to hang their points on.
ORIG4_POINTS = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "assets", "QNL_original4_points.xlsx")
# Two orphan assets carry telemetry but have no register row. Yesterday modelled
# them (assumption QNL-021) with their points and no location; they are kept, not
# dropped - see the orphan handling and the assumption log.
ORIG4_ORPHANS = {"entity:QNL_VAV_B_S13_005", "entity:QNL_CAV_1F_S15_001"}
_cur_equip = {a["id"] for a in assets}          # the 449 AHU/VAV/CAV/FCU subjects
_owb = openpyxl.load_workbook(ORIG4_POINTS, read_only=True, data_only=True)
_ows = _owb["Ontology"]
_orows = [list(r) for r in _ows.iter_rows(values_only=True)][1:]
# pass 1: the hasPoint declarations whose subject is a current equipment or a
# part of one; collect the point-entity ids they introduce.
_pt_ids, _orig_rows, _skipped_equip = set(), [], set()
for r in _orows:
    if str(r[2]) == "brick:hasPoint":
        subj = str(r[0])
        equip = subj.split("_")[:0]  # placeholder
        # the owning equipment is the subject; keep only if it is a current unit
        if subj in _cur_equip or subj in ORIG4_ORPHANS:
            _orig_rows.append(r)
            _pt_ids.add(str(r[3]))
        else:
            _skipped_equip.add(subj)
# pass 2: every other row whose subject is one of those points (its timeseries
# reference, and any point attribute rows), plus the orphan units' own
# declaration rows (isPartOf HVAC, IFC) - yesterday modelled them with a class
# and an IFC reference but no rec:locatedIn/feeds/isFedBy.
for r in _orows:
    subj = str(r[0])
    if str(r[2]) != "brick:hasPoint" and (subj in _pt_ids or subj in ORIG4_ORPHANS):
        _orig_rows.append(r)
for r in _orig_rows:
    out.append([("" if c is None else c) for c in r] + [""] * (27 - len(r)))
notes.append("%d point rows for AHU/VAV/CAV/FCU lifted from the reviewed sheet; "
             "%d equipment skipped as not in this build (orphans)"
             % (len(_orig_rows), len(_skipped_equip)))

# declare the para: classes those points use that are not already declared.
ORIG4_PARA = {
    "para:Trip_Alarm": ("brick:Alarm", "Trip Alarm"),
    "para:Fail_Stop_Alarm": ("brick:Alarm", "Fail to Stop Alarm"),
    "para:Fail_Start_Alarm": ("brick:Alarm", "Fail to Start Alarm"),
    "para:Scheduled_Hrs_Duration": ("brick:Point", "Scheduled Hours Duration"),
    "para:UnScheduled_Hrs_Duration": ("brick:Point", "Unscheduled Hours Duration"),
    "para:Room_Air_Temperature": ("brick:Temperature_Sensor", "Room Air Temperature"),
}
for _cls, (_par, _lab) in ORIG4_PARA.items():
    if _cls not in _declared:
        out.append(row(_cls, "owl:Class", "rdfs:subClassOf", _par, "",
                       [("s", "rdfs:label_en", _lab)]))
        _declared.add(_cls)

# Orphan units named in the historian/selected list but absent from the register
# (assumption QNL-023). Modelled like yesterday's VAV/CAV orphans: a class, a
# label, an IFC reference and their datapoints, but NO rec:locatedIn / rec:feeds
# / rec:isFedBy - position and feeds are not asserted because the register, the
# only source for them, says nothing.
ORPHAN_UNITS = (
    [("CCU_%d" % n, "brick:CRAC", HVAC, HVAC_CLASS) for n in range(8081, 8087)]
    + [("CHWPU_P02", "brick:Chilled_Water_Booster_Pump", CHW, CHW_CLASS)]
)
_orphan_assets = [{"kind": "ORPHAN", "tag": t, "id": "entity:QNL_" + t, "cls": c}
                  for t, c, sysent, syscls in ORPHAN_UNITS]
_osys = {"entity:QNL_" + t: (sysent, syscls) for t, c, sysent, syscls in ORPHAN_UNITS}
for a in _orphan_assets:
    bare = a["id"].replace("entity:", "")
    sysent, syscls = _osys[a["id"]]
    out.append(row(a["id"], a["cls"], "brick:isPartOf", sysent, syscls,
                   [("s", "rdfs:label_en", label(bare))]))
    # A chilled-water pump feeds the loop by function, not by position, so that
    # much is asserted even for an orphan; a CCU feeds a specific room the
    # register would have named, so its feeds is left unasserted.
    if a["cls"] == "brick:Chilled_Water_Booster_Pump":
        out.append(row(a["id"], a["cls"], "rec:feeds", LOOP, LOOP_CLASS))
    out.append(row(a["id"], a["cls"], "ref:hasExternalReference",
                   "<blanknode>", "ref:IFCReference",
                   [("o", "para:IFC_ID", ""), ("o", "ref:ifcName", bare)]))
_odp_rows, _odp_ledger, _opara = qnl_datapoints.build_datapoints(
    _orphan_assets, _load_selected(), _load_historian(), row, label)
for _cls in sorted(_opara):
    if _cls not in _declared:
        out.append(row(_cls, "owl:Class", "rdfs:subClassOf", _opara[_cls], "",
                       [("s", "rdfs:label_en",
                         qnl_datapoints.PARA_LABEL.get(_cls,
                             label(_cls.split(":", 1)[-1])))]))
        _declared.add(_cls)
out.extend(_odp_rows)
notes.append("%d orphan units modelled (assumption QNL-023): %s"
             % (len(_orphan_assets),
                ", ".join(a["tag"] for a in _orphan_assets)))

# Orphan sensors: single room-condition points named in the historian/selected
# list under the CCU prefix but not tied to any register unit (assumption
# QNL-025). The user directed they be treated as orphans, not assumed onto the
# register CCUs that happen to share those rooms. Each tag is the sensor itself
# (a point), so it is typed as its measurement class, made brick:isPartOf the
# HVAC system, given its timeseries - and NO rec:locatedIn, exactly like the
# other orphans.
_ohist = _load_historian()
ORPHAN_SENSORS = sorted(
    t[:-3] for t in {s["tag"] for s in _load_selected()}
    if t.startswith("QNL_CCU_") and "Rm" in t and t.endswith(".PV"))
_osensor_rows = 0
for head in ORPHAN_SENSORS:
    tag = head + ".PV"
    if tag not in _ohist:
        continue
    cls = ("brick:Relative_Humidity_Sensor" if "Humd" in head
           else "brick:Temperature_Sensor" if "Temp" in head else "brick:Sensor")
    eid = "entity:" + head
    unit = qnl_datapoints.to_unit(_ohist[tag].get("unit", ""))
    out.append(row(eid, cls, "brick:isPartOf", HVAC, HVAC_CLASS,
                   [("s", "rdfs:label_en", label(head)),
                    ("s", "brick:hasUnit", unit)]))
    out.append(row(eid, cls, "ref:hasExternalReference", "<blanknode>",
                   "ref:TimeseriesReference",
                   [("o", "ref:hasTimeseriesId", tag),
                    ("o", "para:hasEntityId", tag)]))
    _osensor_rows += 2
notes.append("%d orphan CCU room sensors modelled as points isPartOf HVAC "
             "(assumption QNL-025)" % (len(ORPHAN_SENSORS)))

# Building/system-level instrumentation (assumption QNL-024): CHW plant, the
# electrical meters, orphan roof/control DX units, and loose room points -
# attached to the parents Dar Cairo/SSC use. See qnl_instrumentation.py.
_instr_n, _instr_para = qnl_instrumentation.build(
    out, row, label, qnl_datapoints.to_unit, _load_historian(),
    HVAC, HVAC_CLASS, CHW, CHW_CLASS, ELEC, ELEC_CLASS, LOOP, LOOP_CLASS,
    _declared)
_PARA_LABELS = dict(qnl_datapoints.PARA_LABEL)
_PARA_LABELS.update({c: v[1] for c, v in qnl_instrumentation.PARA_DECL.items()})
for _cls in sorted(_instr_para):
    if _cls not in _declared:
        out.append(row(_cls, "owl:Class", "rdfs:subClassOf",
                       _instr_para[_cls] or "brick:Point", "",
                       [("s", "rdfs:label_en",
                         _PARA_LABELS.get(_cls, label(_cls.split(":", 1)[-1])))]))
        _declared.add(_cls)
notes.append("%d instrumentation rows for the 68 building/system-level tags "
             "(assumption QNL-024, now modelled)" % _instr_n)

# --------------------------------------------------------------------------- write
wbo = openpyxl.Workbook()
ws = wbo.active
ws.title = "Ontology"
ws.append(HEADER)
for r in out:
    ws.append(r)
wbo.save(OUT + "QNL_Ontology.xlsx")

with open(OUT + "QNL_Ontology.csv", "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(HEADER)
    w.writerows(out)

with open(OUT + "QNL_identifier_crosswalk.csv", "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(["kind", "source_identifier", "ontology_identifier", "rdfs:label_en"])
    for src_entity, d in rooms.items():
        w.writerow(["room", src_entity, d["id"], d["label"]])
    for a in assets:
        w.writerow([a["kind"], a["tag"], a["id"],
                    label(a["id"].replace("entity:", ""))])
    for a in new_assets:
        w.writerow([a["kind"], a["tag"], a["id"],
                    label(a["id"].replace("entity:", ""))])

print("rows      :", len(out))
print("rooms     :", len(rooms), "levels:", levels_used)
print("assets    :", len(assets), {k: sum(1 for a in assets if a["kind"] == k) for k in CLASS})
print("new equip :", len(new_assets),
      {k: sum(1 for a in new_assets if a["kind"] == k) for k in NEW_FAMILIES})
print("datapoints:", len(_dp_rows), "rows;", len(_para_used), "para classes declared")
for n in sorted(set(notes)):
    print("note      :", n)
