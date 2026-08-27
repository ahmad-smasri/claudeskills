#!/usr/bin/env python3
"""Datapoint and part triples for QNL's new equipment families.

Given a list of new-family equipment (DX, CCU, EF, TEF, KEF, SEF, HEX, CHW,
CHWPU, generator), this builds the point and part triples for them, doing the
ledger step internally with the priority the user set:

    Dar Cairo  ->  Brick 1.4  ->  QF SSC  ->  para:

Each selected datapoint is a tag QNL_<equipment>[_<part>].<point>. Only tags
that exist in the historian are emitted (a point with no historian tag resolves
to an empty timeseries). A token before the dot is a PART only when it is one of
the curated component tokens for these families; otherwise it is part of the
point name and the point hangs directly on the equipment.

Every point carries: rdfs:label_en, brick:hasUnit (when the historian gives a
unit), and a ref:TimeseriesReference blank node with ref:hasTimeseriesId (the
historian tag) and para:hasEntityId - on the point, never the equipment.

The module exposes build_datapoints(); run it directly to dump the ledger.
"""
import collections
import csv
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.join(HERE, "..", "..")
sys.path.insert(0, os.path.join(HERE, "datapoints"))
sys.path.insert(0, os.path.join(REPO, "skills", "building-ontology", "scripts"))

DARCAIRO = os.path.join(REPO, "reference-models", "DarCairo_V93.csv")
BRICKVOCAB = os.path.join(REPO, "skills", "building-ontology",
                          "references", "data", "brick-vocab.txt")
SSC = os.path.join(REPO, "projects", "SSC",
                   "QF_SSC_Ontology_ver01_reviewed.xlsx")

# --------------------------------------------------------------------------- text
# BMS token -> English, longest first. Only what these families actually use.
ABBREV = [
    ("CalcEntryUnscheduledHrs", "unscheduled outage hours"),
    ("CalcEntryScheduledHrs", "scheduled maintenance hours"),
    ("StartStopCmdSts", "start stop command status"),
    ("StartStopCmd", "start stop command"),
    ("EngStartStopCmd", "engine start stop command"),
    ("EngStartedFbk", "engine started feedback"),
    ("EngAutoManSts", "engine auto manual status"),
    ("EngManOvrSts", "engine manual override status"),
    ("EngFailFbk", "engine fail feedback"),
    ("EngOilPrs", "engine oil pressure"),
    ("GenReadySts", "generator ready status"),
    ("BattChrgSts", "battery charge status"),
    ("MCBOnOffSts", "main circuit breaker on off status"),
    ("DayTnkHiLvlAlm", "day tank high level alarm"),
    ("DayTnkLoLvlAlm", "day tank low level alarm"),
    ("LineVoltageL1", "line voltage"), ("LineVoltageL2", "line voltage"),
    ("LineVoltageL3", "line voltage"),
    ("LineCurrentL1", "line current"), ("LineCurrentL2", "line current"),
    ("LineCurrentL3", "line current"),
    ("AutoManCmd", "auto manual command"),
    ("AutoManSts", "auto manual status"),
    ("OpenCmdSts", "open command status"),
    ("OpenCmd", "open command"), ("OpenSts", "open status"),
    ("CloseSts", "close status"),
    ("SpeedFbk", "speed feedback"),
    ("PosCmd", "position command"), ("PosFbk", "position feedback"),
    ("RmTempSPHys", "room temperature setpoint hysteresis"),
    ("RmTempSP", "room temperature setpoint"),
    ("RmTemp", "room temperature"),
    ("ChOverHrsSP", "changeover hours setpoint"),
    ("ZoneASpcTemp", "zone space temperature"),
    ("ZoneBSpcTemp", "zone space temperature"),
    ("ZoneCSpcTemp", "zone space temperature"),
    ("PriRtnTemp", "primary return temperature"),
    ("SecRtnTemp", "secondary return temperature"),
    ("PriSupTemp", "primary supply temperature"),
    ("DiffPrs", "differential pressure"),
    ("RunSts", "run status"), ("LocSts", "local status"),
    ("RemSts", "remote status"),
    ("TripAlm", "trip alarm"), ("CommAlm", "communication alarm"),
    ("FireAlm", "fire alarm"), ("RlyAlm", "relay alarm"),
    ("HiAlm", "high level alarm"), ("LoAlm", "low level alarm"),
    ("ECUSts", "engine control unit status"),
    ("EngFailFbk", "engine fail feedback"),
    ("FltRst", "fault reset"), ("Reset", "reset"),
    ("Frequency", "frequency"),
    ("RuntimeMtr", "runtime meter"), ("StartsCtr", "starts count"),
    ("TripCtr", "trip count"),
    ("OnOffSts", "on off status"), ("Sts", "status"),
    ("MWh", "electrical energy"), ("kWH", "electrical energy"),
    ("kWh", "electrical energy"), ("kW", "electric power"),
    ("PV", ""), ("SP", "setpoint"),
    # part tokens spelled out so the point phrase reads naturally
    ("IsoVlv", "isolation valve"), ("FTPmp", "fuel transfer pump"),
    ("ACB", "air circuit breaker"),
]

# Point kind from the token suffix (Brick names end in the same word).
KIND = [
    ("HiAlm", "Alarm"), ("LoAlm", "Alarm"), ("Alm", "Alarm"),
    ("SP", "Setpoint"), ("Setpoint", "Setpoint"),
    ("Cmd", "Command"), ("Ctrl", "Command"), ("Rst", "Command"),
    ("Reset", "Command"),
    ("Sts", "Status"), ("Fbk", "Sensor"), ("PV", "Sensor"),
    ("Ctr", "Sensor"), ("Mtr", "Sensor"),
    ("kWH", "Sensor"), ("kWh", "Sensor"), ("kW", "Sensor"), ("MWh", "Sensor"),
    ("Frequency", "Sensor"), ("Prs", "Sensor"), ("Temp", "Sensor"),
    ("Voltage", "Sensor"), ("Current", "Sensor"),
]

SYNONYM = {"rtn": "return", "sup": "supply", "spc": "space",
           "pri": "primary", "sec": "secondary", "tnk": "tank",
           "eng": "engine", "gen": "generator", "alm": "alarm",
           "sts": "status", "cmd": "command", "vlv": "valve"}

DEVICE = ("damper", "valve", "fan", "coil", "filter", "heater", "pump",
          "motor", "breaker", "exchanger", "compressor")

STOP = {"the", "a", "of", "and", "on", "off", "process", "value", "qnl",
        "zone", "level", "basement", "roof", "floor", "first"}
GENERIC = {"process value", "status", "command", "setpoint", "alarm", "value",
           "feedback", "control", "power", "energy", "", "setpoint status"}

# The only tokens-before-a-dot that are real components in these families. Each
# resolves through the same ladder, recorded in PART_LADDER for the handover.
PART_TOKENS = {"IsoVlv", "MV", "ACB14", "FTPmp1", "FTPmp2"}
PART_PHRASE = {"IsoVlv": "isolation valve", "MV": "modulating valve",
               "ACB14": "air circuit breaker",
               "FTPmp1": "fuel transfer pump", "FTPmp2": "fuel transfer pump"}


# Units the historian uses -> the canonical codes the estate already uses.
UNIT_MAP = {"°C": "unit:DEG_C", "%": "unit:PERCENT", "kW": "unit:KiloW",
            "Watt": "unit:W", "kWh": "unit:KiloW-HR", "kWH": "unit:KiloW-HR",
            "MWh": "unit:MegaW-HR", "Hrs": "unit:HR", "hrs": "unit:HR",
            "bar": "unit:BAR", "Hz": "unit:HZ", "A": "unit:A", "V": "unit:V",
            "kPa": "unit:KiloPA", "Pa": "unit:PA", "%rH": "unit:PERCENT_RH"}


def to_unit(u):
    if not u or u in ("None",):
        return "unit:UNITLESS"
    return UNIT_MAP.get(u, "unit:UNITLESS")


# The internal ledger, curated. Keyed by the point token as it appears
# (point, or part.point). Where the fuzzy ladder would mis-map or mint a clumsy
# name, this fixes it to the class the ladder *should* land on, priority
# Dar Cairo -> Brick 1.4 -> SSC -> para. Source is the winning step. Anything not
# listed falls through to the automatic resolver.
CANON = {
    # fans, pumps, generic actuation/telemetry
    "RunSts": ("brick:Run_Status", "Brick 1.4"),
    # orphan CCU points (Temp/Humidity/Alm room monitors)
    "Alm": ("brick:Alarm", "Dar Cairo"),
    "Humidity": ("brick:Relative_Humidity_Sensor", "Dar Cairo"),
    "Temp": ("brick:Temperature_Sensor", "Dar Cairo"),
    "TripAlm": ("para:Trip_Status", "Dar Cairo"),
    "LocSts": ("para:Local_Status", "para (minted)"),
    "RemSts": ("para:Remote_Status", "para (minted)"),
    "AutoManCmd": ("para:Auto_Manual_Command", "para (minted)"),
    "AutoManSts": ("brick:Manual_Auto_Status", "SSC"),
    "StartStopCmd": ("brick:Start_Stop_Command", "Brick 1.4"),
    "StartStopCmdSts": ("brick:Start_Stop_Status", "Brick 1.4"),
    "SpeedFbk": ("brick:Speed_Sensor", "Brick 1.4"),
    "kW": ("brick:Electric_Power_Sensor", "Dar Cairo"),
    "kWH": ("brick:Electrical_Energy_Usage_Sensor", "Dar Cairo"),
    "kWh": ("brick:Electrical_Energy_Usage_Sensor", "Dar Cairo"),
    "CommAlm": ("brick:Communication_Loss_Alarm", "SSC"),
    "FltRst": ("brick:Reset_Command", "Brick 1.4"),
    "Reset": ("brick:Reset_Command", "Brick 1.4"),
    "FireAlm": ("para:Fire_Alarm", "para (minted)"),
    "RlyAlm": ("para:Relay_Alarm", "para (minted)"),
    "CalcEntryScheduledHrs": ("para:Scheduled_Maintenance_Hours", "para (minted)"),
    "CalcEntryUnscheduledHrs": ("para:Unscheduled_Outage_Hours", "para (minted)"),
    "StartsCtr": ("para:Start_Count", "para (minted)"),
    "TripCtr": ("para:Trip_Count", "para (minted)"),
    "RuntimeMtr": ("para:Operation_Hours", "Dar Cairo"),
    "StartStopSts": ("brick:Start_Stop_Status", "Brick 1.4"),
    "Priority1": ("para:Command_Priority", "para (minted)"),
    "Priority2": ("para:Command_Priority", "para (minted)"),
    "Priority3": ("para:Command_Priority", "para (minted)"),
    # DX
    "RmTemp.PV": ("para:Room_Air_Temperature", "Dar Cairo"),
    "RmTempSP.SP": ("brick:Room_Air_Temperature_Setpoint", "Brick 1.4"),
    "RmTempSPHys.SP": ("para:Room_Air_Temperature_Setpoint_Hysteresis", "para (minted)"),
    "ChOverHrsSP.SP": ("para:Changeover_Hours_Setpoint", "para (minted)"),
    # EF zone temps
    "ZoneASpcTemp.PV": ("brick:Zone_Air_Temperature_Sensor", "Brick 1.4"),
    "ZoneBSpcTemp.PV": ("brick:Zone_Air_Temperature_Sensor", "Brick 1.4"),
    "ZoneCSpcTemp.PV": ("brick:Zone_Air_Temperature_Sensor", "Brick 1.4"),
    # HEX points and parts
    "DiffPrs.PV": ("brick:Differential_Pressure_Sensor", "Brick 1.4"),
    "PriRtnTemp.PV": ("para:Primary_Return_Temperature_Sensor", "para (minted)"),
    "SecRtnTemp.PV": ("para:Secondary_Return_Temperature_Sensor", "para (minted)"),
    "IsoVlv": ("brick:Isolation_Valve", "Dar Cairo"),
    "IsoVlv.AutoManCmd": ("para:Auto_Manual_Command", "para (minted)"),
    "IsoVlv.CloseSts": ("brick:Open_Close_Status", "Brick 1.4"),
    "IsoVlv.OpenCmd": ("brick:Open_Close_Command", "Brick 1.4"),
    "IsoVlv.OpenCmdSts": ("brick:Open_Close_Status", "Brick 1.4"),
    "IsoVlv.OpenSts": ("brick:Open_Close_Status", "Brick 1.4"),
    "MV": ("para:Modulating_Valve", "para (minted)"),
    "MV.AutoManCmd": ("para:Auto_Manual_Command", "para (minted)"),
    "MV.PosCmd": ("brick:Valve_Position_Command", "Dar Cairo"),
    "MV.PosFbk": ("brick:Valve_Position_Sensor", "Dar Cairo"),
    # generator
    "BattChrgSts": ("para:Battery_Charge_Status", "para (minted)"),
    "ECUSts": ("para:Engine_Control_Unit_Status", "para (minted)"),
    "EngAutoManSts": ("brick:Manual_Auto_Status", "SSC"),
    "EngFailFbk": ("para:Engine_Fail_Status", "para (minted)"),
    "EngManOvrSts": ("para:Engine_Manual_Override_Status", "para (minted)"),
    "EngOilPrs": ("para:Oil_Pressure_Sensor", "para (minted)"),
    "EngStartStopCmd": ("brick:Start_Stop_Command", "Brick 1.4"),
    "EngStartedFbk": ("para:Engine_Started_Status", "para (minted)"),
    "GenReadySts": ("para:Generator_Ready_Status", "para (minted)"),
    "Frequency": ("brick:Frequency_Sensor", "Brick 1.4"),
    "LineCurrentL1": ("brick:Current_Sensor", "Brick 1.4"),
    "LineCurrentL2": ("brick:Current_Sensor", "Brick 1.4"),
    "LineCurrentL3": ("brick:Current_Sensor", "Brick 1.4"),
    "LineVoltageL1": ("brick:Voltage_Sensor", "Brick 1.4"),
    "LineVoltageL2": ("brick:Voltage_Sensor", "Brick 1.4"),
    "LineVoltageL3": ("brick:Voltage_Sensor", "Brick 1.4"),
    "MCBOnOffSts": ("brick:On_Off_Status", "Dar Cairo"),
    "ACB14": ("brick:Circuit_Breaker", "Brick 1.4"),
    "ACB14.MWh": ("brick:Electrical_Energy_Usage_Sensor", "Dar Cairo"),
    "ACB14.OnOffSts": ("brick:On_Off_Status", "Dar Cairo"),
    "ACB14.Sts": ("brick:On_Off_Status", "Dar Cairo"),
    "ACB14.TripAlm": ("para:Trip_Status", "Dar Cairo"),
    "DayTnkHiLvlAlm.HiAlm": ("para:Day_Tank_High_Level_Alarm", "para (minted)"),
    "DayTnkLoLvlAlm.LoAlm": ("para:Day_Tank_Low_Level_Alarm", "para (minted)"),
    "FTPmp1": ("para:Fuel_Transfer_Pump", "para (minted)"),
    "FTPmp2": ("para:Fuel_Transfer_Pump", "para (minted)"),
    "FTPmp1.RunSts": ("brick:Run_Status", "Brick 1.4"),
    "FTPmp2.RunSts": ("brick:Run_Status", "Brick 1.4"),
    "FTPmp1.TripAlm": ("para:Trip_Status", "Dar Cairo"),
    "FTPmp2.TripAlm": ("para:Trip_Status", "Dar Cairo"),
}

# Parent for every para: class we emit, so the sheet declares each as an
# owl:Class subclass. The Dar-Cairo-defined ones (Trip_Status, Room_Air_Temperature,
# Operation_Hours) are declared too, to keep the sheet self-contained.
PARA_PARENT = {
    "para:Trip_Status": "brick:Status",
    "para:Room_Air_Temperature": "brick:Temperature_Sensor",
    "para:Operation_Hours": "brick:Point",
    "para:Local_Status": "brick:Status",
    "para:Remote_Status": "brick:Status",
    "para:Auto_Manual_Command": "brick:Command",
    "para:Fire_Alarm": "brick:Alarm",
    "para:Relay_Alarm": "brick:Alarm",
    "para:Scheduled_Maintenance_Hours": "brick:Point",
    "para:Unscheduled_Outage_Hours": "brick:Point",
    "para:Start_Count": "brick:Point",
    "para:Trip_Count": "brick:Point",
    "para:Command_Priority": "brick:Point",
    "para:Room_Air_Temperature_Setpoint_Hysteresis": "brick:Setpoint",
    "para:Changeover_Hours_Setpoint": "brick:Setpoint",
    "para:Primary_Return_Temperature_Sensor": "brick:Temperature_Sensor",
    "para:Secondary_Return_Temperature_Sensor": "brick:Temperature_Sensor",
    "para:Modulating_Valve": "brick:Valve",
    "para:Battery_Charge_Status": "brick:Status",
    "para:Engine_Control_Unit_Status": "brick:Status",
    "para:Engine_Fail_Status": "brick:Status",
    "para:Engine_Manual_Override_Status": "brick:Status",
    "para:Oil_Pressure_Sensor": "brick:Pressure_Sensor",
    "para:Engine_Started_Status": "brick:Status",
    "para:Generator_Ready_Status": "brick:Status",
    "para:Day_Tank_High_Level_Alarm": "brick:Alarm",
    "para:Day_Tank_Low_Level_Alarm": "brick:Alarm",
    "para:Fuel_Transfer_Pump": "brick:Pump",
}

PARA_LABEL = {
    "para:Trip_Status": "Trip Status", "para:Room_Air_Temperature": "Room Air Temperature",
    "para:Operation_Hours": "Operation Hours", "para:Local_Status": "Local Status",
    "para:Remote_Status": "Remote Status", "para:Auto_Manual_Command": "Auto Manual Command",
    "para:Fire_Alarm": "Fire Alarm", "para:Relay_Alarm": "Relay Alarm",
    "para:Scheduled_Maintenance_Hours": "Scheduled Maintenance Hours",
    "para:Unscheduled_Outage_Hours": "Unscheduled Outage Hours",
    "para:Start_Count": "Start Count", "para:Trip_Count": "Trip Count",
    "para:Command_Priority": "Command Priority",
    "para:Room_Air_Temperature_Setpoint_Hysteresis": "Room Air Temperature Setpoint Hysteresis",
    "para:Changeover_Hours_Setpoint": "Changeover Hours Setpoint",
    "para:Primary_Return_Temperature_Sensor": "Primary Return Temperature Sensor",
    "para:Secondary_Return_Temperature_Sensor": "Secondary Return Temperature Sensor",
    "para:Modulating_Valve": "Modulating Valve",
    "para:Battery_Charge_Status": "Battery Charge Status",
    "para:Engine_Control_Unit_Status": "Engine Control Unit Status",
    "para:Engine_Fail_Status": "Engine Fail Status",
    "para:Engine_Manual_Override_Status": "Engine Manual Override Status",
    "para:Oil_Pressure_Sensor": "Oil Pressure Sensor",
    "para:Engine_Started_Status": "Engine Started Status",
    "para:Generator_Ready_Status": "Generator Ready Status",
    "para:Day_Tank_High_Level_Alarm": "Day Tank High Level Alarm",
    "para:Day_Tank_Low_Level_Alarm": "Day Tank Low Level Alarm",
    "para:Fuel_Transfer_Pump": "Fuel Transfer Pump",
}

def words(text):
    out = []
    for w in re.split(r"[^a-z0-9]+", text.lower()):
        w = re.sub(r"(?<=[a-z])\d+$", "", w)
        w = SYNONYM.get(w, w)
        if w and w not in STOP:
            out.append(w)
    return out


def expand(token):
    out = token
    for abbr, word in ABBREV:
        out = re.sub(r"(?<![A-Za-z])" + abbr + r"(?![a-z])", " " + word + " ", out)
    out = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", out)
    out = re.sub(r"\d+", " ", out)
    return " ".join(out.replace("_", " ").replace(".", " ").lower().split())


def kind_of(point):
    for suffix, kind in KIND:
        if point == suffix or point.endswith(suffix):
            return kind
    if point.endswith("Hrs"):
        return "Sensor"
    return ""


def device_of(text):
    w = set(words(text))
    for d in DEVICE:
        if d in w:
            return d
    return ""


# --------------------------------------------------------------------------- indexes
def _label_on_row(r):
    for i in range(7, min(len(r), 27), 2):
        if i - 1 < len(r) and str(r[i - 1]).strip() == "rdfs:label_en":
            return str(r[i]).strip()
    return ""


def load_reference_classes():
    """Point and equipment classes from Dar Cairo and SSC, each with the English
    labels their own entities carry - the labels are what let an abbreviated BMS
    token match a spelled-out class."""
    pts = {"dar": collections.defaultdict(lambda: {"n": 0, "labels": collections.Counter()}),
           "ssc": collections.defaultdict(lambda: {"n": 0, "labels": collections.Counter()})}
    eqs = {"dar": collections.defaultdict(lambda: {"n": 0, "labels": collections.Counter()}),
           "ssc": collections.defaultdict(lambda: {"n": 0, "labels": collections.Counter()})}
    with open(DARCAIRO, encoding="utf-8-sig") as fh:
        for r in csv.reader(fh):
            if len(r) < 5 or not r[4]:
                continue
            lab = _label_on_row(r)
            if r[2] == "brick:hasPoint":
                pts["dar"][r[4]]["n"] += 1
                if lab:
                    pts["dar"][r[4]]["labels"][lab] += 1
            elif r[2] == "brick:hasPart":
                eqs["dar"][r[4]]["n"] += 1
                if lab:
                    eqs["dar"][r[4]]["labels"][lab] += 1
    wb = openpyxl.load_workbook(SSC, read_only=True, data_only=True)
    from validate_ontology import pick_ontology_sheet
    ws = pick_ontology_sheet(wb, SSC)
    hdr = [str(c or "") for c in next(ws.iter_rows(values_only=True))]
    pr, ot, ob = hdr.index("predicate"), hdr.index("objectType"), hdr.index("object")
    # label sits in the first object_prop pair when it is rdfs:label_en
    def ssc_label(r):
        for i in range(len(hdr)):
            if str(hdr[i]) == "object_prop_name" and i + 1 < len(r) \
                    and str(r[i] or "") == "rdfs:label_en":
                return str(r[i + 1] or "").strip()
        return ""
    for r in ws.iter_rows(min_row=2, values_only=True):
        if len(r) <= ot or not r[ot]:
            continue
        cls = str(r[ot]).strip()
        if str(r[pr] or "") == "brick:hasPoint":
            pts["ssc"][cls]["n"] += 1
            lab = ssc_label(r)
            if lab:
                pts["ssc"][cls]["labels"][lab] += 1
        elif str(r[pr] or "") == "brick:hasPart":
            eqs["ssc"][cls]["n"] += 1
    return pts, eqs


def load_brick():
    ok = {}
    for line in open(BRICKVOCAB):
        if line.startswith("#") or not line.strip():
            continue
        parts = line.rstrip("\n").split("\t")
        ok[parts[0]] = parts[1] if len(parts) > 1 else "OK"
    return ok


# --------------------------------------------------------------------------- scoring
def _score(phrase, cls, labels, kind="", device=""):
    pw = set(words(phrase))
    if not pw:
        return 0.0
    bare = cls.split(":", 1)[-1]
    cw = set(words(bare.replace("_", " ")))
    best = len(pw & cw) / max(len(pw | cw), 1)
    for lab, _ in labels.most_common(8):
        lw = set(words(lab))
        best = max(best, len(pw & lw) / max(len(pw | lw), 1))
    if kind and not bare.endswith(kind):
        best *= 0.30
    if device:
        cd = device_of(bare.replace("_", " "))
        if cd and cd != device:
            best *= 0.30
    return best


KIND_PARENT = {"Sensor": "brick:Sensor", "Setpoint": "brick:Setpoint",
               "Status": "brick:Status", "Command": "brick:Command",
               "Alarm": "brick:Alarm", "": "brick:Point"}


class Resolver:
    def __init__(self):
        self.pts, self.eqs = load_reference_classes()
        self.brick = load_brick()
        self.minted = {}

    def _search(self, index, phrase, kind, device, thresh):
        best, best_sc = "", 0.0
        for cls, info in index.items():
            sc = _score(phrase, cls, info["labels"], kind, device)
            if sc > best_sc:
                best, best_sc = cls, sc
        return (best, best_sc) if best_sc >= thresh else ("", best_sc)

    def _brick_exact(self, phrase, kind):
        stem = "_".join(w.capitalize() for w in words(phrase))
        for cand in ("brick:%s_%s" % (stem, kind), "brick:%s" % stem):
            if self.brick.get(cand) == "OK":
                return cand
        return ""

    def resolve(self, phrase, kind, device, equipment=False):
        """Walk Dar Cairo -> Brick -> SSC -> para and always return
        (class, source). Points search point indexes; parts search equipment."""
        dar = self.pts["dar"] if not equipment else self.eqs["dar"]
        ssc = self.pts["ssc"] if not equipment else self.eqs["ssc"]
        c, sc = self._search(dar, phrase, kind, device, 0.60)
        if c:
            return c, "Dar Cairo"
        bc = self._brick_exact(phrase, kind if not equipment else "")
        if bc:
            return bc, "Brick 1.4 exact"
        bfuzz, bsc = "", 0.0
        for term, st in self.brick.items():
            if st != "OK":
                continue
            s = _score(phrase, term, collections.Counter(), kind, device)
            if s > bsc:
                bfuzz, bsc = term, s
        if bsc >= 0.70:
            return bfuzz, "Brick 1.4"
        c, sc = self._search(ssc, phrase, kind, device, 0.60)
        if c:
            return c, "SSC"
        if bsc >= 0.55:
            return bfuzz, "Brick 1.4"
        # para: mint. Points subclass their kind's Brick parent; equipment the
        # closest device parent, or brick:Equipment.
        name = "_".join(w.capitalize() for w in words(phrase)) or "Unknown"
        if equipment:
            cls = "para:%s" % name
            parent = "brick:Equipment"
        else:
            cls = "para:%s_%s" % (name, kind) if kind else "para:%s" % name
            parent = KIND_PARENT.get(kind, "brick:Point")
        self.minted.setdefault(cls, parent)
        return cls, "para (minted)"


# --------------------------------------------------------------------------- build
def usable(desc):
    return bool(desc) and not desc.upper().startswith("QNL") \
        and desc.strip().lower() not in GENERIC


def build_datapoints(new_assets, selected, historian, row, label):
    """new_assets: list of {kind, tag(no QNL_ prefix... actually source tag),
    id (entity:QNL_...)}. Returns (rows, ledger, minted)."""
    R = Resolver()
    # index selected tags per equipment id
    by_equip = collections.defaultdict(list)
    idset = {a["id"] for a in new_assets}
    id_of = {a["id"].replace("entity:", ""): a for a in new_assets}
    for s in selected:
        tag = s["tag"]
        # equipment id = QNL_<...> up to the part/point; match against our ids
        # by longest known equipment prefix
        for a in new_assets:
            base = a["id"].replace("entity:", "")
            if tag == base or tag.startswith(base + ".") or tag.startswith(base + "_"):
                by_equip[a["id"]].append((a, tag, tag[len(base):]))
                break

    rows, ledger = [], []
    part_decl = set()          # (equip_id, part_id) already declared
    for a in new_assets:
        eid = a["id"]
        for _a, tag, rest in sorted(by_equip.get(eid, [])):
            if tag not in historian:
                continue
            rest = rest.lstrip("_")
            if rest.startswith("."):
                part, point = "", rest[1:]
            elif "." in rest:
                part, point = rest.split(".", 1)
            else:
                part, point = "", rest
            h = historian[tag]
            unit, desc = h.get("unit", ""), h.get("desc", "")
            is_part = part in PART_TOKENS
            token = (part + "." if part else "") + point

            kind = kind_of(point)
            device = device_of(PART_PHRASE.get(part, "")) if is_part else ""
            if token in CANON:
                pcls, psrc = CANON[token]
            else:
                if usable(desc):
                    phrase = desc
                elif is_part:
                    phrase = expand(point)
                else:
                    phrase = expand((part + " " + point) if part else point)
                pcls, psrc = R.resolve(phrase, kind, device)

            # the entity that owns the point: the part if there is one, else equip
            owner, owner_cls = eid, a["cls"]
            if is_part:
                part_id = "%s_%s" % (eid, part)
                if part in CANON:
                    part_cls, part_src = CANON[part]
                else:
                    part_phrase = PART_PHRASE[part]
                    part_cls, part_src = R.resolve(part_phrase, "",
                                                   device_of(part_phrase),
                                                   equipment=True)
                if (eid, part_id) not in part_decl:
                    rows.append(row(eid, a["cls"], "brick:hasPart",
                                    part_id, part_cls,
                                    [("o", "rdfs:label_en",
                                      label(part_id.replace("entity:", "")))]))
                    part_decl.add((eid, part_id))
                    ledger.append({"equip": eid, "kind": "PART", "token": part,
                                   "class": part_cls, "source": part_src,
                                   "unit": "", "desc": PART_PHRASE.get(part, part)})
                owner, owner_cls = part_id, part_cls

            # The point entity is derived from the datapoint's own tag, so it is
            # unique - several points on one unit can share a suffix (RmTempSP.SP,
            # RmTempSPHys.SP both end .SP) and must not collapse to one id.
            point_id = "entity:" + tag.replace(".", "_")
            eid_bare = eid.replace("entity:", "")
            local = tag[len(eid_bare):].lstrip("_")
            props = [("o", "rdfs:label_en", label(local.replace(".", "_")))]
            props.append(("o", "brick:hasUnit", to_unit(unit)))
            rows.append(row(owner, owner_cls, "brick:hasPoint",
                            point_id, pcls, props))
            # timeseries reference - on the point
            rows.append(row(point_id, pcls, "ref:hasExternalReference",
                            "<blanknode>", "ref:TimeseriesReference",
                            [("o", "ref:hasTimeseriesId", tag),
                             ("o", "para:hasEntityId", tag)]))
            ledger.append({"equip": eid, "kind": "POINT" if not part else "POINT/part",
                           "token": (part + "." if part else "") + point,
                           "class": pcls, "source": psrc,
                           "unit": unit, "desc": desc[:50]})
    para_used = {}
    for l in ledger:
        c = l["class"]
        if c.startswith("para:"):
            para_used[c] = PARA_PARENT.get(c, R.minted.get(c, "brick:Point"))
    for c, par in R.minted.items():
        para_used.setdefault(c, par)
    return rows, ledger, para_used


if __name__ == "__main__":
    from step1_confirm_datapoints import load_selected, load_historian

    NEW = {"DX", "EF", "SEF", "CCU", "TEF", "HEX", "KEF", "CHW", "CHWPU", "ELEC"}
    CLS = {"DX": "para:DXUnit", "CCU": "brick:CRAC", "EF": "brick:Exhaust_Fan",
           "TEF": "brick:Exhaust_Fan", "KEF": "brick:Exhaust_Fan",
           "SEF": "para:Smoke_Extract_Fan", "HEX": "brick:Heat_Exchanger",
           "CHW": "brick:Chilled_Water_Booster_Pump",
           "CHWPU": "brick:Chilled_Water_Booster_Pump", "ELEC": "para:Generator"}
    reg = os.path.join(HERE, "sources", "QNL_Assets_Location_Relationships.xlsx")
    wb = openpyxl.load_workbook(reg, read_only=True, data_only=True)
    new_assets = []
    for k in NEW:
        for r in list(wb[k].iter_rows(values_only=True))[1:]:
            if r[0]:
                new_assets.append({"kind": k, "id": "entity:QNL_" + str(r[0]).strip(),
                                   "cls": CLS[k]})

    def label(t):
        return " ".join(t.replace("_", " ").split())

    def row(subject, stype, pred, obj, otype="", props=()):
        return (subject, stype, pred, obj, otype, tuple(props))

    rows, ledger, minted = build_datapoints(load_selected(), None, None, row, label) \
        if False else (None, None, None)
    from step1_confirm_datapoints import load_historian as _lh
    rows, ledger, minted = build_datapoints(new_assets, load_selected(),
                                            _lh(), row, label)
    out = os.path.join(HERE, "datapoints", "ledger_newfamilies.csv")
    with open(out, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["equip", "kind", "token", "class",
                                           "source", "unit", "desc"])
        w.writeheader()
        w.writerows(ledger)
    print("datapoint rows:", len(rows), " ledger entries:", len(ledger))
    print("minted para classes:", len(minted))
    for c, p in sorted(minted.items()):
        print("   ", c, "subClassOf", p)
    import collections as _c
    print("\nby source:", dict(_c.Counter(l["source"] for l in ledger)))
    print("wrote", out)
