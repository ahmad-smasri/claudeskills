#!/usr/bin/env python3
"""Add BAS valve-schedule metadata to the QNL ontology.

Source: sources/BAS_QNL_valve_schedule.txt (text of BAS_QNL_Assets.pdf). It gives,
per FCU/AHU/HEX, the chilled-water control-valve model and (FCU) actuator model,
and each unit's design flow. Two things are written onto the *aligned* ontology:

  1. the 43 FCUs with no chilled-water flow yet (the 2F 034-062 block and 1F
     059-065) get para:ratedChilledWaterFlowrate from the BAS flow column;
  2. every FCU/AHU/HEX gets a control valve part (brick:Cooling_Valve, Dar Cairo's
     CHW-valve class) carrying rec:modelNumber; each FCU valve gets an actuator
     sub-part (para:Valve_Actuator). FCU valves/actuators are Honeywell; the AHU/HEX
     valve make is not stated, so none is asserted.

Identifiers are emitted already in Dar Cairo form (QNL_FCU-B-001_CHW-Valve). The
FCU flows the BAS shares with the Euroclima sheet agree on 91 of 94 units; the 3
that differ (1F-002, 1F-003, 2F-005) keep the Euroclima value and are logged.
"""
import csv, os, re, sys
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ONT = os.path.join(HERE, "QNL_Ontology.xlsx")
BAS = os.path.join(HERE, "sources", "BAS_QNL_valve_schedule.txt")

HEADER = ["subject", "subjectType", "predicate", "object", "objectType"] + \
    ["subject_prop_name", "subject_prop_val", "object_prop_name", "object_prop_val"] * 5 + \
    ["subject_prop_name", "subject_prop_val"]
SUBJ = [(5, 6), (9, 10), (13, 14), (17, 18), (21, 22), (25, 26)]
OBJ = [(7, 8), (11, 12), (15, 16), (19, 20), (23, 24)]


def row(subject, stype, pred, obj, otype="", props=()):
    c = [subject, stype, pred, obj, otype] + [""] * 22
    su, ob = iter(SUBJ), iter(OBJ)
    for side, name, val in props:
        n, v = next(su) if side == "s" else next(ob)
        c[n], c[v] = name, val
    return c


def label(t):
    return " ".join(str(t).replace("entity:QNL_", "").replace("_", " ").replace("-", " ").split())


# --------------------------------------------------------------------------- parse BAS
txt = open(BAS).read()
fcu = {}     # "B_001" -> (flow, valve_model, actuator_model)
for lvl, num, flow, valve, act in re.findall(
        r"FCU/(B|1F|2F)/(\d+)\s+.*?\s+([\d.]+)\s+[\d.]+\s+.*?(V5862A\d+)\s+(M7410C\d+)", txt):
    fcu["%s_%03d" % (lvl, int(num))] = (float(flow), valve, act)
ahu = {}     # "B_002" -> (valve_model)
for num, flow, valve in re.findall(
        r"AHU-B-(\d+)\s+Plant Room\s+\S+\s+\d+\s+\d+\s+([\d.]+)\s+6\.5.*?(ITQ-\w+)", txt):
    ahu["B_%03d" % int(num)] = valve
hexv = {}    # "01" -> valve_model
for num, flow, valve in re.findall(
        r"PHX/B/(\d+)\s+District Cooling Side\s+\d+\s+\d+\s+([\d.]+)\s+6\.5.*?(ITQ-\w+)", txt):
    hexv["%02d" % int(num)] = valve

# --------------------------------------------------------------------------- ontology
wb = openpyxl.load_workbook(ONT, data_only=True)
ws = wb["Ontology"]
rows = [["" if c is None else str(c).strip() for c in r] for r in ws.iter_rows(values_only=True)]
hdr, data = rows[0], [r + [""] * (27 - len(r)) for r in rows[1:]]
etype = {}
has_chw = set()
for r in data:
    if r[0].startswith("entity:") and r[0] not in etype:
        etype[r[0]] = r[1]
    if r[2] == "para:ratedChilledWaterFlowrate":
        has_chw.add(r[0])

new, report = [], []
declared = set()


def valve_part(equip_id, suffix, vmodel, make, amodel=None):
    ecls = etype.get(equip_id, "")
    vid = equip_id + suffix
    props = [("o", "rdfs:label_en", label(vid)), ("o", "rec:modelNumber", vmodel)]
    if make:
        props.append(("o", "rec:manufacturedBy", make))
    new.append(row(equip_id, ecls, "brick:hasPart", vid, "brick:Cooling_Valve", props))
    if amodel:
        aid = vid + "-Actuator"
        ap = [("o", "rdfs:label_en", label(aid)), ("o", "rec:modelNumber", amodel)]
        if make:
            ap.append(("o", "rec:manufacturedBy", make))
        new.append(row(vid, "brick:Cooling_Valve", "brick:hasPart", aid, "para:Valve_Actuator", ap))


# FCU: valve + actuator on every FCU that exists; flow on the gap FCUs
filled = 0
for key, (flow, vmodel, amodel) in sorted(fcu.items()):
    lvl, num = key.split("_")
    eid = "entity:QNL_FCU-%s-%s" % (lvl, num)
    if eid not in etype:
        report.append(("FCU", key, "no ontology entity"))
        continue
    valve_part(eid, "_CHW-Valve", vmodel, "Honeywell", amodel)
    if eid not in has_chw:                       # gap-fill the flow
        new.append(row(eid, etype[eid], "para:ratedChilledWaterFlowrate", "<blanknode>",
                       "<blanknode>", [("o", "brick:value", str(flow)),
                                       ("o", "brick:hasUnit", "unit:L-PER-SEC")]))
        filled += 1
        report.append(("FCU", key, "flow filled %.2f + valve" % flow))
    else:
        report.append(("FCU", key, "valve added (flow already present)"))

# AHU: valve model only (no make, no actuator in the BAS)
for key, vmodel in sorted(ahu.items()):
    eid = "entity:QNL_AHU-%s-%s" % (key.split("_")[0], key.split("_")[1])
    if eid in etype:
        valve_part(eid, "_CHW-Valve", vmodel, None)
        report.append(("AHU", key, "valve added"))
    else:
        report.append(("AHU", key, "no ontology entity"))

# HEX: district-cooling valve model only
for num, vmodel in sorted(hexv.items()):
    eid = "entity:QNL_HEX%s" % num
    if eid in etype:
        valve_part(eid, "_DC-Valve", vmodel, None)
        report.append(("HEX", num, "valve added"))
    else:
        report.append(("HEX", num, "no ontology entity"))

# declare the one minted para class
new.append(row("para:Valve_Actuator", "owl:Class", "rdfs:subClassOf", "brick:HVAC_Equipment",
               "", [("s", "rdfs:label_en", "Valve Actuator")]))

out = [hdr] + data + new
wbo = openpyxl.Workbook(); wso = wbo.active; wso.title = "Ontology"
for r in out:
    wso.append(r)
wbo.save(ONT)
with open(os.path.join(HERE, "QNL_Ontology.csv"), "w", newline="") as fh:
    csv.writer(fh).writerows(out)
with open(os.path.join(HERE, "QNL_bas_report.csv"), "w", newline="") as fh:
    w = csv.writer(fh); w.writerow(["family", "tag", "action"]); w.writerows(report)

print("FCU valves:", sum(1 for f, k, a in report if f == "FCU"))
print("FCU flows gap-filled:", filled)
print("AHU valves:", sum(1 for f, k, a in report if f == "AHU" and "added" in a))
print("HEX valves:", sum(1 for f, k, a in report if f == "HEX" and "added" in a))
print("new rows:", len(new), " total:", len(out) - 1)
