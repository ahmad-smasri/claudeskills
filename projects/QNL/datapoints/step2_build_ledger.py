#!/usr/bin/env python3
"""Step 2 of the datapoint exercise: the decision ledger.

2,230 selected datapoint rows in scope collapse to ~100 distinct
(family, part, point) signatures. Resolve each signature once and it applies to
every unit that carries it, so the class ladder is walked a hundred times rather
than two thousand, and the review is a hundred rows rather than a spreadsheet
nobody can check.

Two questions are kept apart, because they have different authorities:

  Does a part exist, and what hangs on it?
      From the tag. QNL_AHUB001_SupFan.kW says part SupFan, point kW;
      QNL_AHUB001.EnableDisableCmd says a point directly on the equipment.
      That is QNL's own BMS structure - better evidence than another
      building's modelling convention.

  What class is it?
      Dar Cairo first, then Brick 1.4, then para: with the user's permission.
      The existing ladder, unchanged.

A token before the dot is not automatically a part. It is a part only if it
resolves to an EQUIPMENT class; if it resolves to a POINT class it names the
measurement, and token and suffix together name a point on the parent. So
SupFan -> brick:Supply_Fan is a part, while MixAirTemp.PV ->
brick:Mixed_Air_Temperature_Sensor is a point on the AHU.

Writes ledger.csv (the reviewable artefact) and ledger-summary.md.
"""
import collections
import csv
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.join(HERE, "..", "..", "..")
sys.path.insert(0, HERE)
sys.path.insert(0, os.path.join(REPO, "skills", "building-ontology", "scripts"))

from step1_confirm_datapoints import (  # noqa: E402
    load_selected, load_historian, load_register, make_resolver,
    family_of, FAMILIES)

DARCAIRO = os.path.join(REPO, "reference-models", "DarCairo_V93.csv")
BRICKVOCAB = os.path.join(REPO, "skills", "building-ontology", "references",
                          "data", "brick-vocab.txt")
LEDGER = os.path.join(HERE, "ledger.csv")
SUMMARY = os.path.join(HERE, "ledger-summary.md")

# --------------------------------------------------------------------------- 1
# The BMS abbreviations, expanded. Only used where the historian description is
# unusable; the description is always the better key when it is real English.
# Longest first, so Rtn matches before Rt.
ABBREV = [
    ("CalcEntryUnscheduledHrs", "unscheduled outage hours"),
    ("CalcEntryScheduledHrs", "scheduled maintenance hours"),
    ("EnableDisableCmd", "enable disable command"),
    ("AutoManCmd", "auto manual command"),
    ("OverCurrentAlm", "over current alarm"),
    ("EarthFltAlm", "earth fault alarm"),
    ("OverloadAlm", "overload alarm"),
    ("HtrHiTempAlm", "heater high temperature alarm"),
    ("ElectHtrSts", "electric heater on off status"),
    ("CommAlm", "communication fail alarm"),
    ("PositionCtrl", "position command"),
    ("PositionFbk", "position feedback"),
    ("SpeedFbk", "speed feedback"),
    ("ValveFbk", "valve position feedback"),
    ("EffectiveSP", "effective temperature setpoint"),
    ("DuctAirFlow", "duct air flow"),
    ("DmprPos", "damper position"),
    ("StartsCtr", "start count"),
    ("TripCtr", "trip count"),
    ("TripAlm", "trip alarm"),
    ("RunSts", "run status"),
    ("LocSts", "local status"),
    ("FltRst", "fault reset"),
    ("FTST", "fail to start alarm"),
    ("FTSP", "fail to stop alarm"),
    ("FTO", "fail to open alarm"),
    ("FTC", "fail to close alarm"),
    ("PosFbk", "position feedback"),
    ("kWH", "electrical energy"),
    ("kW", "electric power"),
    ("RtnAirDuctPrs", "return air duct pressure"),
    ("SupAirDuctPrs", "supply air duct pressure"),
    ("FACHWRtnTemp", "fresh air chilled water return temperature"),
    ("FACHWSupTemp", "fresh air chilled water supply temperature"),
    ("CHWRtnTemp", "chilled water return temperature"),
    ("CHWSupTemp", "chilled water supply temperature"),
    ("IntrnlEADmpr", "internal exhaust air damper"),
    ("IntrnlFADmpr", "internal fresh air damper"),
    ("AvgSpcHumd", "average space humidity"),
    ("AvgSpcTemp", "average space temperature"),
    ("FrshAirTemp", "fresh air temperature"),
    ("MixAirTemp", "mixed air temperature"),
    ("RtnAirFlow", "return air flow"),
    ("RtnAirHumd", "return air humidity"),
    ("RtnAirTemp", "return air temperature"),
    ("RtnHumiditySP", "return air humidity setpoint"),
    ("RtnTempSP", "return air temperature setpoint"),
    ("SupAirFlow", "supply air flow"),
    ("SupAirHumd", "supply air humidity"),
    ("SupAirTemp", "supply air temperature"),
    ("DuctHumd", "duct humidity"),
    ("DuctTemp", "duct temperature"),
    ("SpcHumd", "space humidity"),
    ("SpcTemp", "space temperature"),
    ("RmTemp", "room temperature"),
    ("CoolVlv", "cooling valve"),
    ("SupFan", "supply fan"),
    ("RtnFan", "return fan"),
    ("Dmpr", "damper"),
    ("Vlv", "valve"),
    ("SP", "setpoint"),
    ("PV", ""),
]

# A token before the dot names a component, not a measurement, when it reads as
# a thing rather than a reading. Confirmed against QNL's own 65 AHU tokens: the
# split is 9 components and 56 measurements, with no borderline cases.
COMPONENT = re.compile(
    r"^(SupFan|RtnFan|CoolVlv|HeatVlv|IntrnlEADmpr|IntrnlFADmpr|FADmpr|EADmpr|"
    r"MixDmpr|RecDmpr|Filter|SupHtr)\d*$")

POINTISH = re.compile(
    r"(Sensor|Setpoint|Status|Command|Alarm|Count|Meter|Limit|Mode|Hours|"
    r"Target|Time|Array|Priority)$")


def expand(token):
    """Turn a BMS token into an English phrase using the abbreviation table."""
    out = token
    for abbr, word in ABBREV:
        out = re.sub(r"(?<![A-Za-z])" + abbr + r"(?![a-z])", " " + word + " ", out)
    out = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", out)
    out = re.sub(r"\d+", " ", out)
    return " ".join(out.replace("_", " ").replace(".", " ").lower().split())


STOP = {"the", "a", "of", "and", "on", "off", "process", "value", "qnl"}

# Descriptions that pass the "is it English" test and still say nothing. The
# historian writes "Process Value" against 300-odd distinct sensors; taken at
# face value it matches everything and identifies nothing.
GENERIC = {"process value", "status", "command", "setpoint", "alarm", "value",
           "feedback", "control", "power", "energy", ""}


def words(text):
    out = []
    for w in re.split(r"[^a-z0-9.]+", text.lower()):
        w = re.sub(r"(?<=[a-z])\d+$", "", w)   # Damper1 and Damper are one word
        w = SYNONYM.get(w, w)
        if w and w not in STOP:
            out.append(w)
    return out


# A point that hangs on a damper cannot be a valve point, however the words
# overlap. Dar Cairo uses brick:Valve_Position_Command 80 times against
# brick:Damper_Position_Command 29, so on the phrase "position command" alone
# the two tie and the valve wins on usage - which is how IntrnlEADmpr.PositionCtrl,
# a damper, came out as a valve command. The part names the device; the class
# has to agree with it.
DEVICE = ("damper", "valve", "fan", "coil", "filter", "heater", "pump", "motor")

# Where the BMS and Brick use different words for the same thing. Kept short and
# each one evidenced, because a synonym is a claim about meaning: "fresh air" and
# "outside air" are the same air in every HVAC text, and Brick standardised on
# "Outside"; Dar Cairo writes brick:Outside_Damper 18 times and has no fresh-air
# damper class at all. Applied after tokenising, so it cannot fire mid-word.
SYNONYM = {"fresh": "outside", "frsh": "outside", "rtn": "return",
           "sup": "supply", "spc": "space", "amb": "outside"}


def device_of(text):
    w = set(words(text))
    for d in DEVICE:
        if d in w:
            return d
    return ""


def usable_description(desc):
    """A historian description is evidence only if it names the measurement.

    24% of them just repeat the equipment name, and a few hundred more are
    "Process Value". Both look like English and neither identifies anything.
    """
    if not desc or desc.upper().startswith("QNL"):
        return False
    return desc.strip().lower() not in GENERIC


# The suffix of a BMS token says what KIND of point it is, and Brick encodes the
# same thing in the last word of the class name. Requiring the two to agree is
# what stops RtnHumiditySP - a setpoint - matching Return_Air_Humidity_Sensor,
# which it otherwise does at 0.60 on word overlap alone.
# Order matters: FTSP is "fail to stop", not a setpoint, so the four-letter
# fail codes are tested before the SP suffix that would otherwise swallow them.
KIND = [
    ("FTST", "Alarm"), ("FTSP", "Alarm"), ("FTO", "Alarm"), ("FTC", "Alarm"),
    ("Alm", "Alarm"),
    ("Setpoint", "Setpoint"), ("SP", "Setpoint"),
    ("Cmd", "Command"), ("Ctrl", "Command"), ("Rst", "Command"),
    ("Sts", "Status"),
    ("Fbk", "Sensor"), ("PV", "Sensor"), ("Ctr", "Sensor"),
    ("kWH", "Sensor"), ("kW", "Sensor"),
]


def kind_of(point):
    for suffix, kind in KIND:
        if point == suffix or point.endswith(suffix):
            return kind
    if point.endswith("Hrs"):
        return "Hours"
    return ""


# --------------------------------------------------------------------------- 2
def load_darcairo():
    """Dar Cairo's point and equipment classes, each with the English labels
    its own entities carry - the labels are what make matching possible, since
    Dar Cairo names points in dashed English and QNL in abbreviated camelCase."""
    points = collections.defaultdict(lambda: {"n": 0, "labels": collections.Counter()})
    equips = collections.defaultdict(lambda: {"n": 0, "labels": collections.Counter()})
    with open(DARCAIRO, encoding="utf-8-sig") as fh:
        for r in csv.reader(fh):
            if len(r) < 5:
                continue
            pred, obj_cls = r[2], r[4]
            label = ""
            for i in range(7, min(len(r), 27), 4):
                if i - 1 < len(r) and str(r[i - 1]).strip() == "rdfs:label_en":
                    label = str(r[i]).strip()
                    break
            if pred == "brick:hasPoint" and obj_cls:
                points[obj_cls]["n"] += 1
                if label:
                    points[obj_cls]["labels"][label] += 1
            elif pred == "brick:hasPart" and obj_cls:
                equips[obj_cls]["n"] += 1
                if label:
                    equips[obj_cls]["labels"][label] += 1
            if r[1]:
                (points if POINTISH.search(r[1]) else equips)[r[1]]["n"] += 0
    return points, equips


def load_brick():
    ok = {}
    with open(BRICKVOCAB) as fh:
        for line in fh:
            if line.startswith("#") or not line.strip():
                continue
            parts = line.rstrip("\n").split("\t")
            ok[parts[0]] = parts[1] if len(parts) > 1 else "OK"
    return ok


def score(phrase, cls, labels, kind="", device=""):
    """How well a class matches a phrase: token overlap against the class name,
    and against the English labels Dar Cairo's own entities carry.

    A class whose kind disagrees with the token's kind is knocked down hard
    rather than excluded, so it still shows up as a runner-up for review.
    """
    pw = set(words(phrase))
    if not pw:
        return 0.0, ""
    bare = cls.split(":", 1)[-1]
    cw = set(words(bare.replace("_", " ")))
    best = len(pw & cw) / max(len(pw | cw), 1)
    why = "class name"
    for lab, _ in labels.most_common(6):
        lw = set(words(lab))
        sc = len(pw & lw) / max(len(pw | lw), 1)
        if sc > best:
            best, why = sc, "label %r" % lab
    if kind and not bare.endswith(kind):
        best *= 0.30
        why += " (kind mismatch: wanted %s)" % kind
    if device:
        cd = device_of(bare.replace("_", " "))
        if cd and cd != device:
            best *= 0.30
            why += " (device mismatch: point is on a %s, class names a %s)" % (device, cd)
    return best, why


def brick_candidate(phrase, kind, brick):
    """Step 2 of the ladder: what does Brick 1.4 call this?

    Tried two ways. First the exact name the phrase implies, which is right
    where the BMS and Brick happen to agree - brick:Mixed_Air_Temperature_Sensor
    is a real term even though Dar Cairo has no mixed-air AHU to copy from.
    Then a search over the whole 2,800-term vocabulary with the same scorer,
    which is what surfaces a term the BMS words differently: "space temperature"
    finds brick:Zone_Air_Temperature_Sensor, and a reviewer can accept or reject
    it. Exact hits are returned as answers; searched hits as candidates.
    """
    if not phrase:
        return "", "", 0.0
    if kind:
        stem = "_".join(w.capitalize() for w in words(phrase))
        for cand in ("brick:%s_%s" % (stem, kind), "brick:%s" % stem):
            if cand in brick:
                return cand, brick[cand], 1.0
    best, best_sc = "", 0.0
    for term, st in brick.items():
        if st != "OK":
            continue
        sc, _ = score(phrase, term, collections.Counter(), kind)
        if sc > best_sc:
            best, best_sc = term, sc
    return best, brick.get(best, ""), best_sc


def main():
    selected = load_selected()
    hist = load_historian()
    reg = load_register()
    resolve = make_resolver(reg)
    dar_points, dar_equips = load_darcairo()
    brick = load_brick()

    # ------------------------------------------------ collapse to signatures
    sig = collections.OrderedDict()
    seen_tags = set()
    for s in selected:
        if family_of(s["tag"]) not in FAMILIES:
            continue
        if s["tag"] in seen_tags:      # the appended addendum repeats 15 tags
            continue
        seen_tags.add(s["tag"])
        eq, part, point = resolve(s["tag"])
        if eq is None:
            continue
        key = (reg[eq], part, point)
        d = sig.setdefault(key, {"units": set(), "tags": []})
        d["units"].add(eq)
        d["tags"].append(s["tag"])

    total_units = collections.Counter(reg.values())

    rows = []
    for (fam, part, point), d in sig.items():
        probe = d["tags"][0]
        h = hist.get(probe, {})
        desc = h.get("desc", "")
        is_part = bool(part) and bool(COMPONENT.match(part))
        kind = kind_of(point)

        # The phrase the class has to match. The historian description wins when
        # it names the measurement; otherwise the tag is expanded. When the point
        # hangs under a part, the part is already carrying that half of the
        # meaning, so only the point token is expanded.
        if usable_description(desc):
            point_phrase, phrase_src = desc, "historian description"
        else:
            point_phrase = expand(point if is_part
                                  else ((part + " " + point) if part else point))
            phrase_src = "expanded from the tag"

        # Each class is scored against the point alone AND against the part and
        # point together, and the better wins. The point alone is right for
        # SupFan.kW - "electric power" beats "supply fan electric power" against
        # brick:Electric_Power_Sensor, and Brick has no fan-specific power class.
        # The pair is what saves IntrnlEADmpr.PositionCtrl: on "position command"
        # alone Dar Cairo's brick:Valve_Position_Command wins on sheer usage, and
        # the point is on a damper. Whichever phrase matched is recorded, so the
        # reviewer can see which reading produced the answer.
        pair_phrase = (expand(part) + " " + point_phrase).strip() if part else point_phrase
        device = device_of(expand(part)) if is_part else ""
        cands = []
        for cls, info in dar_points.items():
            sc, why = score(point_phrase, cls, info["labels"], kind, device)
            sc2, why2 = score(pair_phrase, cls, info["labels"], kind, device)
            if sc2 > sc:
                sc, why = sc2, why2 + " [matched on part+point]"
            if sc > 0:
                cands.append((sc, info["n"], cls, why))
        cands.sort(reverse=True)
        best = cands[0] if cands else (0, 0, "", "")
        runner = cands[1] if len(cands) > 1 else (0, 0, "", "")

        # a part token gets looked up among Dar Cairo's equipment classes
        # The part walks the same ladder as the point, and for the same reason:
        # Dar Cairo has no brick:Return_Fan among its parts, only Supply_Fan, so
        # a Dar-Cairo-only lookup answers "supply fan" for a return fan. Brick
        # has the term; step 2 finds it.
        part_cls, part_sc, part_src = "", 0.0, ""
        if is_part:
            pphrase = expand(part)
            pdev = device_of(pphrase)
            pc = []
            for cls, info in dar_equips.items():
                sc, _ = score(pphrase, cls, info["labels"], "", pdev)
                if sc > 0:
                    pc.append((sc, info["n"], cls))
            pc.sort(reverse=True)
            if pc:
                part_sc, _, part_cls = pc[0]
                part_src = "Dar Cairo"
            if part_sc < 1.0:
                stem = "brick:" + "_".join(w.capitalize() for w in words(pphrase))
                if brick.get(stem) == "OK":
                    part_cls, part_sc, part_src = stem, 1.0, "Brick 1.4 exact"
                else:
                    for term, st in brick.items():
                        if st != "OK" or POINTISH.search(term):
                            continue
                        sc, _ = score(pphrase, term, collections.Counter(), "", pdev)
                        if sc > part_sc:
                            part_cls, part_sc, part_src = term, sc, "Brick 1.4 search"
        part_status = ("" if not is_part else
                       "ok" if part_sc >= 0.60 else "needs-you: weak part match")

        # step 2 of the ladder, walked only when step 1 did not answer
        bcand, bstatus, bsc = brick_candidate(pair_phrase if device else point_phrase,
                                              kind, brick)

        if best[0] >= 0.60:
            proposed, source, status = best[2], "Dar Cairo", "auto"
        elif bsc >= 1.0 and bstatus == "OK":
            proposed, source, status = bcand, "Brick 1.4 exact", "auto"
        elif best[0] >= 0.25:
            proposed, source = "", ""
            status = "needs-you: confirm the Dar Cairo candidate"
        elif bcand and bsc >= 0.40:
            proposed, source = "", ""
            status = "needs-you: confirm the Brick candidate"
        else:
            proposed, source = "", ""
            status = "needs-you: no match - propose a para: class"

        rows.append({
            "family": fam,
            "kind": "part+point" if is_part else "point",
            "part": part,
            "point": point,
            "units": len(d["units"]),
            "of": total_units[fam],
            "example_tag": probe,
            "historian_description": desc,
            "unit_of_measure": h.get("unit", ""),
            "signal": h.get("kind", ""),
            "match_phrase": point_phrase,
            "match_phrase_source": phrase_src,
            "point_kind": kind,
            "part_device": device,
            "part_class_darcairo": part_cls,
            "part_class_score": "%.2f" % part_sc if is_part else "",
            "part_class_from": part_src,
            "part_status": part_status,
            "darcairo_class": best[2],
            "darcairo_score": "%.2f" % best[0],
            "darcairo_uses": best[1],
            "darcairo_evidence": best[3],
            "runner_up": runner[2],
            "runner_up_score": "%.2f" % runner[0],
            "darcairo_class_in_brick": brick.get(best[2], "not in Brick 1.4") if best[2] else "",
            "brick_candidate": bcand,
            "brick_candidate_status": bstatus,
            "brick_candidate_score": "%.2f" % bsc,
            "proposed_class": proposed,
            "proposed_from": source,
            "status": status,
        })

    rows.sort(key=lambda r: (r["family"], -r["units"], r["part"], r["point"]))
    cols = list(rows[0].keys())
    with open(LEDGER, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)

    # An Excel view of the same rows, because the ledger exists to be reviewed:
    # needs-you rows filled yellow, the columns that carry the decision frozen
    # on the left. Same content as the CSV - this is a view, not a second source.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"
    ws.append(cols)
    for c in ws[1]:
        c.font = openpyxl.styles.Font(bold=True)
    fill = openpyxl.styles.PatternFill("solid", fgColor="FFFF00")
    for r in rows:
        ws.append([r[c] for c in cols])
        if r["status"].startswith("needs-you"):
            for c in ws[ws.max_row]:
                c.fill = fill
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions
    for i, col in enumerate(cols, start=1):
        width = max(len(col), *(len(str(r[col])) for r in rows)) + 2
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(width, 52)
    wb.save(os.path.join(HERE, "ledger.xlsx"))

    # ------------------------------------------------------------- summary
    out = []

    def say(x=""):
        print(x)
        out.append(x)

    auto = [r for r in rows if r["status"] == "auto"]
    confirm = [r for r in rows if "confirm" in r["status"]]
    nomatch = [r for r in rows if r["status"].startswith("needs-you") and r not in confirm]
    say("# Step 2 — the decision ledger")
    say()
    say("%d selected rows in scope collapse to **%d distinct signatures**."
        % (len(seen_tags), len(rows)))
    say()
    say("| Outcome | Signatures | Meaning |")
    say("|---|---|---|")
    say("| `auto` | %d | the ladder answered: Dar Cairo at 0.60+, or a real Brick 1.4 term |" % len(auto))
    say("| `needs-you: confirm` | %d | a plausible Dar Cairo candidate, below the bar |" % len(confirm))
    say("| `needs-you` | %d | no match, or the only Brick term is deprecated/an alias |" % len(nomatch))
    say()
    say("Of the %d resolved automatically, %d came from Dar Cairo and %d from Brick 1.4."
        % (len(auto), sum(1 for r in auto if r["proposed_from"] == "Dar Cairo"),
           sum(1 for r in auto if r["proposed_from"].startswith("Brick"))))
    say()
    say("Parts: %d signatures hang a point under a part, %d put the point "
        "directly on the equipment."
        % (sum(1 for r in rows if r["kind"] == "part+point"),
           sum(1 for r in rows if r["kind"] == "point")))
    say()
    # The same point name resolving two ways across families is the defect
    # check_consistency.py would raise later, so it is raised here instead -
    # while it is still one decision to make rather than hundreds of rows to
    # unpick. It happens because the evidence differs per family: FCU's RmTemp
    # description is unusable so the tag is expanded, VAV's is real English, and
    # the two phrases score differently against the same class.
    cross = collections.defaultdict(set)
    for r in rows:
        cross[(r["part"], r["point"])].add(
            (r["darcairo_class"] or r["brick_candidate"], r["status"] == "auto"))
    split = {k: v for k, v in cross.items() if len({c for c, _ in v}) > 1
             or len({a for _, a in v}) > 1}
    if split:
        say("## Same point, different answer across families — settle these first")
        say()
        say("| Part | Point | Families disagree on |")
        say("|---|---|---|")
        for (part, point), v in sorted(split.items()):
            say("| %s | `%s` | %s |" % (("`%s`" % part) if part else "—", point,
                "; ".join("%s (%s)" % (c or "no candidate",
                                       "auto" if a else "needs-you")
                          for c, a in sorted(v))))
        say()

    for fam in FAMILIES:
        fr = [r for r in rows if r["family"] == fam]
        if not fr:
            continue
        say("## %s — %d signatures" % (fam, len(fr)))
        say()
        say("| Part | Point | Units | Proposed class | Score | Dar Cairo uses | Status |")
        say("|---|---|---|---|---|---|---|")
        for r in fr[:40]:
            say("| %s | `%s` | %d/%d | %s | %s | %s | %s |" % (
                ("`%s`" % r["part"]) if r["part"] else "—", r["point"],
                r["units"], r["of"],
                r["proposed_class"] or "*%s?*" % (r["darcairo_class"] or r["brick_candidate"] or "none"),
                r["darcairo_score"], r["darcairo_uses"], r["status"]))
        if len(fr) > 40:
            say("| … %d more, see ledger.csv | | | | | | |" % (len(fr) - 40))
        say()

    with open(SUMMARY, "w") as fh:
        fh.write("\n".join(out) + "\n")
    print("\nwritten: %s\n         %s\n         %s"
          % (LEDGER, os.path.join(HERE, "ledger.xlsx"), SUMMARY))
    return 0


if __name__ == "__main__":
    sys.exit(main())
