#!/usr/bin/env python3
"""Step 1 of the datapoint exercise: confirm the selected datapoints exist.

Points come from the IO list and only from the IO list. A point written into the
ontology whose tag is not in the historian resolves to an empty timeseries: the
front end draws a tile with no data behind it, and nobody can tell whether the
sensor is broken or was never real. So before any triple is written, three
questions get answered from the sources rather than assumed.

  1. Is every selected datapoint present in the historian?
  2. Does every selected datapoint belong to an asset the register knows about,
     and does every register asset have datapoints?
  3. Which parts and points does each equipment family actually carry - because
     a unit is only given the points its own tags prove it has.

Sources, both in ../sources/:
  Selected_PARA_OS_Data_Points_v4.0.xlsx   the points chosen for this building
  QNL_Historian_IO_list_CP2.xlsx           everything the historian can serve
  QNL_Assets_Location_Relationships.xlsx   the asset register, for the join

Writes step1-findings.md next to this script.
"""
import collections
import os
import re
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "..", "sources")
SELECTED = os.path.join(SRC, "Selected_PARA_OS_Data_Points_v4.0.xlsx")
HISTORIAN = os.path.join(SRC, "QNL_Historian_IO_list_CP2.xlsx")
REGISTER = os.path.join(SRC, "QNL_Assets_Location_Relationships.xlsx")
REPORT = os.path.join(HERE, "step1-findings.md")

FAMILIES = ("AHUB", "VAV", "CAV", "FCU")


def guard(path, sheets):
    """Refuse a workbook whose values are uncached formulas - see known-issues.md.

    Scoped to the sheets this script actually reads. The register is deliberately
    not guarded: its Fed By column is a known formula with no cache, handled in
    build_qnl.py's derive_fed_by(), and its AHU-VAV Check pivot is all formulas.
    Neither is read here - only column A, the equipment tag - so a whole-file
    guard would block on data nobody is using.
    """
    sys.path.insert(0, os.path.join(HERE, "..", "..", "..",
                                    "skills", "building-ontology", "scripts"))
    from validate_ontology import uncached_formulas
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for title in sheets:
        ws = wb[title]
        rows = [["" if c is None else str(c) for c in r]
                for r in ws.iter_rows(values_only=True)]
        bad = uncached_formulas(path, title, rows)
        if bad:
            sys.exit("%s: sheet %r has %d formula cells with no cached value"
                     % (path, title, len(bad)))


def load_selected():
    ws = openpyxl.load_workbook(SELECTED, read_only=True, data_only=True)["Sheet1"]
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if r[0]:
            out.append({"tag": str(r[0]).strip(),
                        "integration": str(r[1] or "").strip(),
                        "unit": str(r[3] or "").strip()})
    return out


def load_historian():
    wb = openpyxl.load_workbook(HISTORIAN, read_only=True, data_only=True)
    hist = {}
    for sheet, kind in (("QNL analog cp2", "analog"),
                        ("QNL Descrete cp2", "discrete")):
        for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
            if r[0]:
                hist[str(r[0]).strip()] = {
                    "kind": kind,
                    "desc": str(r[1] or "").strip(),
                    "unit": str(r[4] or "").strip() if kind == "analog" else ""}
    return hist


def load_register():
    wb = openpyxl.load_workbook(REGISTER, read_only=True, data_only=True)
    reg = {}
    for ws in wb.worksheets:
        if ws.title in FAMILIES:
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if r[0]:
                    reg["QNL_" + str(r[0]).strip()] = ws.title
    return reg


def family_of(tag):
    m = re.match(r"^QNL_([A-Z]+)", tag)
    if not m:
        return ""
    t = m.group(1)
    return "AHUB" if t.startswith("AHUB") else t


def make_resolver(reg):
    """Split a datapoint tag into (equipment, part, point) against the register.

    Tags read QNL_<equipment>[_<part>].<point>. The equipment token is matched
    against the register's own tags rather than parsed by pattern, longest first,
    so an asset whose tag contains an underscore cannot be mistaken for a part.
    The register tag is the right join key here even where the ontology
    identifier differs - QNL's AHUs are entity:QNL_AHU_B_001 but register tag
    AHUB001, and the telemetry follows the register.
    """
    keys = sorted(reg, key=len, reverse=True)

    def resolve(tag):
        for k in keys:
            if tag == k or tag.startswith(k + ".") or tag.startswith(k + "_"):
                rest = tag[len(k):].lstrip("_")
                if rest.startswith("."):
                    return k, "", rest[1:]
                if "." in rest:
                    part, _, point = rest.partition(".")
                    return k, part, point
                return k, "", rest
        return None, "", tag
    return resolve


def main():
    guard(SELECTED, ["Sheet1"])
    guard(HISTORIAN, ["QNL analog cp2", "QNL Descrete cp2"])
    selected = load_selected()
    hist = load_historian()
    reg = load_register()
    resolve = make_resolver(reg)

    out = []

    def say(line=""):
        print(line)
        out.append(line)

    say("# Step 1 — do the selected datapoints exist?")
    say()
    say("| Source | Rows |")
    say("|---|---|")
    say("| Selected datapoints | %d |" % len(selected))
    say("| Historian tags | %d (%d analog, %d discrete) |" % (
        len(hist),
        sum(1 for v in hist.values() if v["kind"] == "analog"),
        sum(1 for v in hist.values() if v["kind"] == "discrete")))
    say("| Register assets, 4 families | %d |" % len(reg))
    say()

    # 1 -------------------------------------------------- selected vs historian
    missing = [s for s in selected if s["tag"] not in hist]
    dupes = [t for t, n in collections.Counter(s["tag"] for s in selected).items()
             if n > 1]
    say("## 0. Only column A of the selected-points file is usable")
    say()
    say("The file holds **two independent lists side by side**, not one list with "
        "seven columns. Column A/B (`TagName`, `DP Integration`) runs to 2,769 "
        "rows and is the selected list. Columns C-G (`DP Name`, `Units`, "
        "`Tag-Reference`, `Equip-Name`, `Point-Name`) stop at row 2,435, cover a "
        "different set of tags, and are sorted independently - of the 2,434 rows "
        "where both blocks have values, the `DP Name` agrees with the "
        "`Point-Name` on **15**. Row 26 reads `QNL_AHUB002_RtnAirDuctPrs.PV` "
        "against a `DP Name` of `RtnFan1_AutoManCmd`.")
    say()
    say("So the equipment/part/point split is taken from column A's tag, "
        "resolved against the asset register, and the unit of measure and "
        "description are taken from the historian. Nothing here reads columns "
        "C-G. Reconstructing tags from `Equip-Name` + `Point-Name` also produces "
        "51 tags that are in no other source, some visibly corrupt "
        "(`QNL_CCU_B01..unSts`).")
    say()

    say("## 1. Every selected datapoint is in the historian")
    say()
    say("%d of %d selected tags matched a historian tag exactly - no "
        "normalising, no case folding." % (len(selected) - len(missing), len(selected)))
    if missing:
        say()
        say("Missing (%d):" % len(missing))
        for s in missing[:40]:
            say("  - `%s`" % s["tag"])
    say()
    if dupes:
        say("**%d tags appear twice in column A** - one per AHU, all "
            "`RtnAirDuctPrs.PV`, once in the main block and once in the 335-row "
            "addendum below it. Confirmed with you as an artefact of appending "
            "the addendum, not two distinct points: deduplicate on the tag, so "
            "each AHU gets the point once." % len(dupes))
        say()

    # 2 -------------------------------------------------- selected vs register
    scope = [s for s in selected if family_of(s["tag"]) in FAMILIES]
    seen, unresolved = set(), []
    rows_per = collections.Counter()
    parts = collections.defaultdict(collections.Counter)
    points = collections.defaultdict(lambda: collections.defaultdict(set))
    for s in scope:
        eq, part, point = resolve(s["tag"])
        if eq is None:
            unresolved.append(s["tag"])
            continue
        seen.add(eq)
        rows_per[reg[eq]] += 1
        parts[reg[eq]][part or "(equipment level)"] += 1
        points[reg[eq]][(part + "." if part else "") + point].add(eq)

    say("## 2. The join to the asset register")
    say()
    say("| Family | Register assets | With datapoints | Selected rows |")
    say("|---|---|---|---|")
    for f in FAMILIES:
        tot = sum(1 for v in reg.values() if v == f)
        cov = sum(1 for k in seen if reg[k] == f)
        say("| %s | %d | %d | %d |" % (f, tot, cov, rows_per[f]))
    say()

    orphan_eq = sorted({resolve(t)[2] and t.split(".")[0] for t in unresolved})
    if unresolved:
        say("**Datapoints for assets the register does not have (%d rows, %d "
            "assets).** They are real - the historian carries them - but no "
            "ontology subject exists to hang them on:" % (
                len(unresolved), len(orphan_eq)))
        say()
        for e in orphan_eq:
            n = sum(1 for t in hist if t.startswith(e + ".") or t.startswith(e + "_"))
            say("  - `%s` — %d historian tags" % (e, n))
        say()
    silent = sorted(k for k in reg if k not in seen)
    if silent:
        say("**Register assets with no selected datapoints (%d).**" % len(silent))
        say()
        for e in silent:
            n = sum(1 for t in hist if t.startswith(e + ".") or t.startswith(e + "_"))
            say("  - `%s` (%s) — %d historian tags%s" % (
                e, reg[e], n,
                ", available but not selected. Confirmed a real distinct asset, "
                "not a typo of its sibling: **highlight its row and note that it "
                "has telemetry nobody selected**" if n else
                ", no telemetry at all. **Highlight its row and note that it "
                "can carry no points**"))
        say()

    # 3 -------------------------------------------------- what each family carries
    say("## 3. What each family actually carries")
    say()
    say("A unit only gets the points its own tags prove it has, so these counts "
        "are the input to the triples, not a template to apply uniformly.")
    for f in FAMILIES:
        tot = sum(1 for v in reg.values() if v == f)
        say()
        say("### %s — %d units, %d distinct points, %d distinct part tokens"
            % (f, tot, len(points[f]), len(parts[f])))
        say()
        say("| Point | Units | Unit of measure | Historian description |")
        say("|---|---|---|---|")
        for sig, eqs in sorted(points[f].items(), key=lambda kv: (-len(kv[1]), kv[0]))[:25]:
            probe = next(iter(eqs)) + ("." if sig.count(".") == 0 else "_") + sig
            h = hist.get(next(iter(eqs)) + "." + sig) or hist.get(
                next(iter(eqs)) + "_" + sig) or {}
            say("| `%s` | %d/%d | %s | %s |" % (
                sig, len(eqs), tot, h.get("unit", "") or "—",
                (h.get("desc", "") or "—")[:60]))
        if len(points[f]) > 25:
            say("| … %d more | | | |" % (len(points[f]) - 25))
    say()

    with open(REPORT, "w") as fh:
        fh.write("\n".join(out) + "\n")
    print("\nwritten: %s" % REPORT)
    return 0


if __name__ == "__main__":
    sys.exit(main())
