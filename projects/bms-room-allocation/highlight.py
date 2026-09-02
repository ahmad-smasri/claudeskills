"""Fill column J light blue on the rows still wanting a human eye.

Only the HQ floors already read - basement, ground, 1F and 2F. Rows the
reviewer has already coloured are left alone; their marking is the record of a
decision already made.
"""
import re, zipfile, os
import openpyxl
from needs_check import BLANK, NOTE
from bf_alloc import BF
from gf_alloc import GF
from ff_alloc import FF1
from ff2_alloc import FF2
from sf1_alloc import SF1
from sf2_alloc import SF2

SRC = '/home/user/claudeskills/projects/bms-room-allocation/Appendix_A_Asset_Register_SSC_HQ_BMS_rooms.xlsx'
OUT = '/tmp/claude-0/-home-user-claudeskills/7b732886-7f20-51be-97dc-21f5f8123adc/scratchpad/hl_out.xlsx'
BLUE = 'FFADD8E6'

read_here = {b.replace('-', '') for lst in (BF, GF, FF1, FF2, SF1, SF2) for b, *_ in lst}
read_here |= set(BLANK) | set(NOTE)

ws = openpyxl.load_workbook(SRC, data_only=True)['Controllable Asset Registry']
rows, why = {}, {}
for i in range(4, 765):
    t = ws.cell(i, 1).value
    if t:
        rows[str(t).strip().upper()] = i

def num(s):
    m = re.search(r'\b(\d{1,2}\.\d{3}[A-Z]?|[A-Z]\.\d{3})\b', str(s or '').upper())
    return m.group(1) if m else None

for tag, note in BLANK.items():
    why[tag] = note
for tag, note in NOTE.items():
    why[tag] = note
# written, but the room disagrees with column D
for tag in read_here:
    i = rows.get(tag)
    if not i or tag in why:
        continue
    j, d = ws.cell(i, 10).value, ws.cell(i, 4).value
    if not j:
        continue
    a, b = num(j), num(d)
    if a and b and a != b:
        why[tag] = 'BMS says %s, column D says %s' % (a, b)
    elif b and not a:
        why[tag] = 'the screen names the zone but gives no number; column D says %s' % b

targets = sorted({rows[t] for t in why if t in rows})

# ---- add the fill and a cell format that uses it
z = zipfile.ZipFile(SRC)
st = z.read('xl/styles.xml').decode('utf-8')
fills = re.search(r'<fills count="(\d+)">', st)
nfill = int(fills.group(1))
st = st.replace(fills.group(0), '<fills count="%d">' % (nfill + 1))
st = st.replace('</fills>', '<fill><patternFill patternType="solid">'
                            '<fgColor rgb="%s"/><bgColor indexed="64"/>'
                            '</patternFill></fill></fills>' % BLUE)
xfs = re.search(r'<cellXfs count="(\d+)">', st)
nxf = int(xfs.group(1))
st = st.replace(xfs.group(0), '<cellXfs count="%d">' % (nxf + 1))
st = st.replace('</cellXfs>', '<xf numFmtId="0" fontId="0" fillId="%d" borderId="0" '
                              'xfId="0" applyFill="1"/></cellXfs>' % nfill)

# ---- which existing styles already carry a fill, so those cells are left alone
base = zipfile.ZipFile(SRC).read('xl/styles.xml').decode('utf-8')
xf_list = re.findall(r'<xf [^>]*?/>|<xf [^>]*?>.*?</xf>', re.search(
    r'<cellXfs count="\d+">(.*?)</cellXfs>', base, re.S).group(1), re.S)
filled = {i for i, x in enumerate(xf_list)
          if (lambda m: m and m.group(1) not in ('0', '1'))(re.search(r'fillId="(\d+)"', x))}

sheet = z.read('xl/worksheets/sheet6.xml').decode('utf-8')
done = []

def fix(mo):
    head, body = mo.group(1), mo.group(2)
    rn = int(re.search(r'\br="(\d+)"', head).group(1))
    if rn not in targets:
        return mo.group(0)
    cell = re.search(r'<c r="J%d"([^>]*?)(/>|>.*?</c>)' % rn, body, re.S)
    if cell:
        cur = re.search(r's="(\d+)"', cell.group(1))
        if cur and int(cur.group(1)) in filled:
            return mo.group(0)                      # reviewer already coloured it
        attrs = re.sub(r'\s*s="\d+"', '', cell.group(1))
        new = '<c r="J%d"%s s="%d"%s' % (rn, attrs, nxf, cell.group(2))
        body = body[:cell.start()] + new + body[cell.end():]
    else:
        body += '<c r="J%d" s="%d"/>' % (rn, nxf)
        head = re.sub(r'spans="1:\d+"', 'spans="1:10"', head)
    done.append(rn)
    return '<row' + head + '>' + body + '</row>'

sheet2 = re.sub(r'<row([^>]*)>(.*?)</row>', fix, sheet, flags=re.S)

zin = zipfile.ZipFile(SRC)
zout = zipfile.ZipFile(OUT, 'w', zipfile.ZIP_DEFLATED)
for it in zin.infolist():
    data = zin.read(it.filename)
    if it.filename == 'xl/styles.xml':
        data = st.encode('utf-8')
    elif it.filename == 'xl/worksheets/sheet6.xml':
        data = sheet2.encode('utf-8')
    zout.writestr(it, data)
zout.close()
print('rows wanting a check: %d, highlighted %d (rest already coloured by the reviewer)'
      % (len(targets), len(done)))
for t in sorted(why, key=lambda k: rows.get(k, 0)):
    if t in rows:
        print('  %-5s %-9s %s' % (rows[t], t, why[t]))
