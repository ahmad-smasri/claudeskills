"""Build the delivered workbook in one pass.

Reading the sheet and writing it back used to be three scripts that each
started from the uploaded file, so whichever ran last silently threw the other
two away. This does the whole thing in order, on one copy of the XML:

  1  corrections to column J cells that already carry a value
  2  new column J values from the per-floor allocation lists
  3  expand the two collapsed outline groups, widen column J
  4  fill light blue the rows that still want a human eye

Everything is done on the sheet XML rather than through openpyxl, so the cell
comments, the web-extension task panes, the formulas on the other eight sheets
and - most of all - the reviewer's own fills survive untouched.
"""
import re, zipfile, os, sys
import openpyxl

from hq_alloc import HQ, SSC_EXTRA
from gf_alloc import GF
from bf_alloc import BF
from ff_alloc import FF1
from ff2_alloc import FF2
from sf1_alloc import SF1
from sf2_alloc import SF2
from f3_alloc import F3
from f4_alloc import F4
from f5_alloc import F5
from f6_alloc import F6
from f7_alloc import F7
from f89_alloc import F8, F9
from ssc_alloc import ALLOC as SSC_A, PLANT as SSC_P
from needs_check import BLANK, NOTE
from screens import image

SRC = '/root/.claude/uploads/7b732886-7f20-51be-97dc-21f5f8123adc/849d736f-Appendix_A_Asset_Register_SSC_HQ_BMS_rooms_1.xlsx'
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   'Appendix_A_Asset_Register_SSC_HQ_BMS_rooms.xlsx')
SHEET = 'xl/worksheets/sheet6.xml'
HQ_LO, HQ_HI = 4, 764
SSC_LO, SSC_HI = 1318, 1441
HEADER = 'ROOM PER BMS SCREEN'
K_HEADER = 'BMS SCREEN (IMAGE FILE)'
BLUE = 'FFADD8E6'

# --- corrections to values already written and reviewed ----------------------
# tag -> (value, why). '' clears the cell.
CORRECT = {
    'VAV0026': ('G.103 BMS ROOM',   'ends inside G.103; its name is printed to the right'),
    'VAV0028': ('G.003',            'real end (546,617), above the G.003/G.002 wall'),
    'VAV0029': ('G.002',            'real end (548,659), not the far side of the plan'),
    'VAV0030': ('',                 'ends in the unnamed corridor east of G.001-G.004'),
    'VAV0052': ('G.108 CONSULTANT SPACE', 'inside G.108, not at the Male Toilet text'),
    'VAV0053': ('',                 'also ends inside G.108 - left blank per review'),
    'FCU0010': ('',                 'runs past B.105 and stops inside B.014; D says B.124'),
    # the 3F bridge FCUs: their leaders end in the open bridge zone, which is
    # what column D already calls them. The toilet/pantry values delivered
    # earlier came from reading the nearest label instead of the endpoint.
    'FCU0064': ('3.63 CORRIDOR BRIDGE', 'endpoint is in the open bridge zone, not the toilet'),
    'FCU0065': ('3.63 CORRIDOR BRIDGE', 'endpoint is in the open bridge zone, not the toilet'),
    'FCU0066': ('3.63 CORRIDOR BRIDGE', 'endpoint is in the open bridge zone, not the pantry'),
}

ws = openpyxl.load_workbook(SRC, data_only=True)['Controllable Asset Registry']

def key(t):
    """one spelling for a tag: the register writes AHUB_0001, the screens
    AHU-B-0001, the VAV list 0050-VAV-0001"""
    return str(t).strip().upper().replace('-', '').replace('_', '').replace(' ', '')


def index(lo, hi):
    d = {}
    for i in range(lo, hi + 1):
        v = ws.cell(i, 1).value
        if v:
            d[key(v)] = i
    return d

hq_rows, ssc_rows = index(HQ_LO, HQ_HI), index(SSC_LO, SSC_HI)

FLOORS = list(GF) + list(BF) + list(FF1) + list(FF2) + list(SF1) + list(SF2) + list(F3) + list(F4) + list(F5) + list(F6) + list(F7) + list(F8) + list(F9)

targets = {}                       # row -> value to write where J is empty
shots = {}                         # row -> the screen the reading came from
for bms, room, screen, conf in HQ:
    t = key(bms)
    if t in hq_rows:
        targets[hq_rows[t]] = room
        shots[hq_rows[t]] = image(screen)
    else:
        print('!! no HQ register row for', bms)
for bms, room, screen, conf in FLOORS:
    t = key(bms)
    if t not in hq_rows:
        print('!! no HQ register row for', bms)
        continue
    # the screen is recorded even when it named no room - that is the picture
    # the reviewer has to open to settle the blank
    shots.setdefault(hq_rows[t], image(screen))
    if room:
        targets.setdefault(hq_rows[t], room)
for bms, room, screen, conf in SSC_EXTRA:
    t = key(bms)
    if t in ssc_rows:
        targets[ssc_rows[t]] = room
        shots[ssc_rows[t]] = image(screen)
    else:
        print('!! no SSC register row for', bms)
for bms, room, screen, conf in list(SSC_A) + list(SSC_P):
    t = key(bms)
    if t in ssc_rows:
        shots.setdefault(ssc_rows[t], image(screen))
targets[2] = HEADER
shots[2] = K_HEADER

corrections = {}                   # row -> value, overriding whatever is there
for tag, (val, _why) in CORRECT.items():
    tag = key(tag)
    if tag in hq_rows:
        corrections[hq_rows[tag]] = val
    else:
        print('!! no register row for correction', tag)

# --- rows that still want a human eye ---------------------------------------
read_here = {key(b) for b, *_ in FLOORS}
read_here |= set(BLANK) | set(NOTE)
why = dict(BLANK)
why.update(NOTE)
for bms, room, screen, conf in FLOORS:
    if conf in ('check', 'unlabelled', 'untraced', 'mis-traced') and room == '':
        why.setdefault(key(bms),
                       'the screen gives no room name at the endpoint')

def num(s):
    m = re.search(r'\b(\d{1,2}\.\d{2,3}[A-Z]?|[A-Z]\.\d{3})\b', str(s or '').upper())
    if not m:
        return None
    # the register writes some levels with a leading zero (04.004); the screens
    # never do, and 04.004 and 4.004 are the same room
    return re.sub(r'^0(?=\d)', '', m.group(1))

for bms, room, screen, conf in FLOORS:
    if conf == 'check':
        why.setdefault(key(bms), 'read from the screen but worth a look')
for tag in sorted(read_here):
    i = hq_rows.get(tag)
    if not i or tag in why:
        continue
    j = corrections.get(i, targets.get(i, ws.cell(i, 10).value))
    d = ws.cell(i, 4).value
    if not j:
        continue
    a, b = num(j), num(d)
    ju, du = str(j).upper(), str(d).upper()
    # a corridor, bridge or terrace is named, not numbered, on these screens;
    # when both sides call it the same kind of space there is nothing to check
    same_kind = any(w in ju and w in du for w in
                    ('CORRIDOR', 'BRIDGE', 'TERRACE', 'LOUNGE', 'ZONE'))
    if a and b and a != b:
        why[tag] = 'BMS says %s, column D says %s' % (a, b)
    elif b and not a and not same_kind:
        why[tag] = 'the screen names the zone but gives no number; column D says %s' % b
blue_rows = sorted({hq_rows[t] for t in why if t in hq_rows})

# --- rewrite the sheet -------------------------------------------------------
def esc(s):
    return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

z = zipfile.ZipFile(SRC)
xml = z.read(SHEET).decode('utf-8')
styles = z.read('xl/styles.xml').decode('utf-8')

m = re.search(r'<c r="D2"([^>]*)>', xml)
sm = re.search(r's="(\d+)"', m.group(1)) if m else None
hdr_style = ' s="%s"' % sm.group(1) if sm else ''

# a new fill plus the cell format that uses it, appended to styles.xml
fills = re.search(r'<fills count="(\d+)">', styles)
nfill = int(fills.group(1))
styles = styles.replace(fills.group(0), '<fills count="%d">' % (nfill + 1))
styles = styles.replace('</fills>', '<fill><patternFill patternType="solid">'
                                    '<fgColor rgb="%s"/><bgColor indexed="64"/>'
                                    '</patternFill></fill></fills>' % BLUE)
xfs = re.search(r'<cellXfs count="(\d+)">', styles)
nxf = int(xfs.group(1))
styles = styles.replace(xfs.group(0), '<cellXfs count="%d">' % (nxf + 1))
styles = styles.replace('</cellXfs>', '<xf numFmtId="0" fontId="0" fillId="%d" '
                                      'borderId="0" xfId="0" applyFill="1"/></cellXfs>' % nfill)

# which of the workbook's existing formats already carry a fill - a cell using
# one of those is the reviewer's marking and is never recoloured
base_xfs = re.findall(r'<xf [^>]*?/>|<xf [^>]*?>.*?</xf>',
                      re.search(r'<cellXfs count="\d+">(.*?)</cellXfs>',
                                z.read('xl/styles.xml').decode('utf-8'), re.S).group(1), re.S)
filled = {i for i, x in enumerate(base_xfs)
          if (lambda mm: mm and mm.group(1) not in ('0', '1'))(re.search(r'fillId="(\d+)"', x))}

wrote, fixed, blued, shot_rows = [], [], [], []

def row(mo):
    head, body = mo.group(1), mo.group(2)
    rn = int(re.search(r'\br="(\d+)"', head).group(1))

    if rn == 2:
        body = re.sub(r'<c r="J2"([^>]*?)(?: t="s")?(?:/>|>.*?</c>)',
                      lambda c: '<c r="J2"%s t="inlineStr"><is><t>%s</t></is></c>'
                                % (re.sub(r' t="[^"]*"', '', c.group(1)), esc(HEADER)),
                      body, count=1, flags=re.S)

    cell = re.search(r'<c r="J%d"([^>]*?)(/>|>.*?</c>)' % rn, body, re.S)
    style = ''
    if cell:
        s = re.search(r's="(\d+)"', cell.group(1))
        if s:
            style = ' s="%s"' % s.group(1)

    if rn in corrections:                       # replace what is there
        val = corrections[rn]
        new = ('<c r="J%d"%s t="inlineStr"><is><t>%s</t></is></c>' % (rn, style, esc(val))
               if val else '<c r="J%d"%s/>' % (rn, style))
        if cell:
            body = body[:cell.start()] + new + body[cell.end():]
        else:
            body += new
            head = re.sub(r'spans="1:\d+"', 'spans="1:10"', head)
        fixed.append(rn)
        cell = re.search(r'<c r="J%d"([^>]*?)(/>|>.*?</c>)' % rn, body, re.S)
    elif rn in targets and not cell:            # nothing there yet
        body += '<c r="J%d"%s t="inlineStr"><is><t>%s</t></is></c>' % (
            rn, hdr_style if rn == 2 else '', esc(targets[rn]))
        head = re.sub(r'spans="1:\d+"', 'spans="1:10"', head)
        wrote.append(rn)
        cell = re.search(r'<c r="J%d"([^>]*?)(/>|>.*?</c>)' % rn, body, re.S)

    if rn in shots and shots[rn]:
        kc = re.search(r'<c r="K%d"([^>]*?)(/>|>.*?</c>)' % rn, body, re.S)
        ks = ''
        if kc:
            sk = re.search(r's="(\d+)"', kc.group(1))
            if sk:
                ks = ' s="%s"' % sk.group(1)
        new = '<c r="K%d"%s t="inlineStr"><is><t>%s</t></is></c>' % (
            rn, ks or (hdr_style if rn == 2 else ''), esc(shots[rn]))
        if kc:
            body = body[:kc.start()] + new + body[kc.end():]
        else:
            body += new
        head = re.sub(r'spans="1:\d+"', 'spans="1:11"', head)
        shot_rows.append(rn)

    if rn in blue_rows:
        if cell:
            cur = re.search(r's="(\d+)"', cell.group(1))
            if not (cur and int(cur.group(1)) in filled):
                attrs = re.sub(r'\s*s="\d+"', '', cell.group(1))
                new = '<c r="J%d"%s s="%d"%s' % (rn, attrs, nxf, cell.group(2))
                body = body[:cell.start()] + new + body[cell.end():]
                blued.append(rn)
        else:
            body += '<c r="J%d" s="%d"/>' % (rn, nxf)
            head = re.sub(r'spans="1:\d+"', 'spans="1:10"', head)
            blued.append(rn)

    if HQ_LO <= rn <= HQ_HI or SSC_LO <= rn <= SSC_HI:
        head = head.replace(' hidden="1"', '')
    if rn in (HQ_LO - 1, SSC_LO - 1):
        head = head.replace(' collapsed="1"', '')
    return '<row' + head + '>' + body + '</row>'

xml2 = re.sub(r'<row([^>]*)>(.*?)</row>', row, xml, flags=re.S)
if re.search(r'<col min="10" max="10"[^>]*/>', xml2):
    xml2 = re.sub(r'<col min="10" max="10"[^>]*/>',
                  '<col min="10" max="10" width="40" customWidth="1"/>'
                  '<col min="11" max="11" width="26" customWidth="1"/>', xml2, count=1)
elif not re.search(r'<col min="11" max="11"', xml2):
    xml2 = re.sub(r'</cols>',
                  '<col min="11" max="11" width="26" customWidth="1"/></cols>', xml2, count=1)

zin = zipfile.ZipFile(SRC)
tmp = OUT + '.tmp'
zout = zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED)
for it in zin.infolist():
    data = zin.read(it.filename)
    if it.filename == SHEET:
        data = xml2.encode('utf-8')
    elif it.filename == 'xl/styles.xml':
        data = styles.encode('utf-8')
    zout.writestr(it, data)
zout.close()
os.replace(tmp, OUT)

print('new column J values: %d' % (len(wrote) - (1 if 2 in wrote else 0)))
print('corrected:           %d  %s' % (len(fixed), sorted(fixed)))
print('column K screen names: %d' % (len(shot_rows) - (1 if 2 in shot_rows else 0)))
print('flagged light blue:  %d of %d (the rest already carry the reviewer\'s fill)'
      % (len(blued), len(blue_rows)))
for t in sorted(why, key=lambda k: hq_rows.get(k, 0)):
    if t in hq_rows:
        print('  %-5s %-9s %s' % (hq_rows[t], t, why[t]))
