import re, os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from ssc_alloc import ALLOC
from write_j import reg_tag, SRC, SSC_LO, SSC_HI

OUT = '/home/user/claudeskills/projects/bms-room-allocation/SSC_BMS_room_findings.xlsx'

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['Controllable Asset Registry']
reg = {}
for i in range(SSC_LO, SSC_HI + 1):
    v = ws.cell(i, 1).value
    if v:
        reg[str(v).strip().upper()] = (i, ws.cell(i, 4).value, ws.cell(i, 8).value)

def num(s):
    m = re.search(r'(\d{1,2}|B)[.\s](\d{3}[A-Z]?)', str(s or '').upper())
    if not m:
        return None
    lvl = m.group(1)
    return (lvl.lstrip('0') or '0', m.group(2))

def words(s):
    return re.sub(r'[^A-Z0-9]+', ' ', str(s or '').upper()).split()

rowsout = []
for bms, room, screen, conf in ALLOC:
    t = reg_tag(bms)
    i, d, h = reg.get(t, (None, None, None))
    cur = h if (h not in (None, '')) else d
    nb, nc = num(room), num(cur)
    if i is None:
        verdict = 'not in the SSC controllable register'
    elif nb and nc and nb != nc:
        verdict = 'ROOM NUMBER DIFFERS'
    elif nb and nc and nb == nc:
        same = set(words(room)) & set(words(cur))
        verdict = 'confirms the register' if len(same) > 1 else 'same room number, different name'
    else:
        verdict = 'could not compare room numbers'
    rowsout.append([i, t, bms, room, d, h, cur, verdict, screen, conf])

out = openpyxl.Workbook()
s = out.active
s.title = 'SSC BMS allocation'
hdr = ['Register row', 'Register tag', 'BMS tag', 'Room per BMS (col J)',
       'Col D (drawings)', 'Col H (VAV/FCU list)', 'Compared against',
       'Finding', 'BMS screen', 'Read confidence']
s.append(hdr)
for c in s[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F4E79')
BAD = PatternFill('solid', fgColor='F8CBAD')
WARN = PatternFill('solid', fgColor='FFF2CC')
for r in sorted(rowsout, key=lambda x: (x[0] or 99999)):
    s.append(r)
    if r[7] == 'ROOM NUMBER DIFFERS':
        for c in s[s.max_row]:
            c.fill = BAD
    elif r[9] == 'check':
        for c in s[s.max_row]:
            c.fill = WARN
s.freeze_panes = 'A2'
s.auto_filter.ref = s.dimensions
for i, w in enumerate([12, 13, 14, 34, 32, 32, 32, 34, 14, 15], 1):
    s.column_dimensions[get_column_letter(i)].width = w

s2 = out.create_sheet('Still to do')
s2.append(['SSC screen', 'What is not yet allocated'])
for c in s2[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F4E79')
for row in [
    ('BF-part1', 'AHU-B-0001/0002/0003/0005, CCU-B-001A/001B/002A/002B/0003/0004/005A/005B/006A/006B/0007/0008, DX-B-0001/0002, FCU-0001/0002/0003, VAV-0001'),
    ('BF-part2', 'AHU-B-0004, DX-B-0003..0010, GEF-B-0001'),
    ('FF-part1', 'FCU-0004, VAV-0037, VAV-0038, VAV-0026, VAV-0022, VAV-0011'),
    ('FF-part2', 'FCU-0009, VAV-0020, FCU-0006, FCU-0008, TEF-1F-0102, VAV-0023'),
    ('Second Floor', 'VAV-0043, VAV-0044, VAV-0045, EF-RF-0001'),
    ('TF-part1', 'VAV-0096/0097/0098, VAV-0100/0102/0103, TEF-3F-0301, KEF-3F-0303, CCU-3F-0301, CCU-3F-0302'),
    ('TF-part2', 'VAV-0063..0070 top row, VAV-0073/0074/0076/0079, VAV-0084, VAV-0099, VAV-0101, VAV-0104/0105/0106, FCU-0011'),
    ('not on any floor plan', 'CHW Pump x8, HEX x5, Generator - these need the plant/system screens'),
]:
    s2.append(list(row))
s2.column_dimensions['A'].width = 22
s2.column_dimensions['B'].width = 120

os.makedirs(os.path.dirname(OUT), exist_ok=True)
out.save(OUT)
from collections import Counter
print(Counter(r[7] for r in rowsout))
print('wrote', OUT)
