#!/usr/bin/env python3
"""Derive Dar Cairo-shaped room subjects from the BMS room-allocation registers.

One source workbook per building in ``sources/``; one output workbook with one
sheet per building.  The rules the client set out, in the order they are applied:

  R1  a green-filled row is a row the BMS screen overruled - take the last of
      the three name columns (D drawings, H VAV/FCU list, J BMS screen).
  R5  otherwise pick between D and J on which of them carries a room number:
      H carries one            -> D
      H does not, J has num+name -> J
      H does not, J is empty     -> D
  R4  when the pick is D (or when D and the pick disagree), reconcile against
      the ontology name already in E: if the pick would lead to E, keep the
      pick's spelling; otherwise take the name from E.
  R2  same room number, different names -> the distinctive one wins.  E is what
      makes a name distinctive: where D matches E's name and the rule-5 pick
      does not, D wins (this is what keeps ROOM NO 42/43/44 apart from the
      SHELL SPACE they all sit in).
  R3  pad the room reference to the building's own width - SSC prints
      <level>.<3 digits>, so 1.027 and 3.041 become 01.027 and 03.041.

Green rows are the one place E is not consulted: green means the drawings and
therefore E were wrong, which is why the BMS screen was traced in the first
place.
"""
import argparse
import csv
import difflib
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

GREEN = "FF00B050"

# Initialisms that stay upper case in an identifier segment, the way Dar Cairo
# keeps UPS, RMU and MDB upper case while spelling Pump-Room in title case.
# TECH, MECH and ELEC are deliberately absent: on HQ they are shortened words
# sitting among other shortened words (STR. PL. DIR. TECH. STAFF), not
# initialisms, and upper-casing one of them alone reads as a mistake.
ACRONYMS = {
    "AV", "AVR", "BMS", "IDF", "MDF", "IT", "ST", "UPS", "LV", "HV", "MEP",
    "AHU", "FCU", "VAV", "WC", "TV", "CCTV", "ELV", "MV", "DB", "MDB", "EMDB",
    "RMU", "HVAC", "PABX", "SMATV", "MCC",
}

# Dar Cairo writes Boiler-And-Heat-Exchange, so an ampersand becomes a word.
WORD_FIXES = {"&": "And", "AND": "And"}


# --------------------------------------------------------------------------
# building profiles - room-reference shape, level segment, building code
# --------------------------------------------------------------------------
class Building:
    def __init__(self, code, level_style, room_width, pad, entity_style,
                 level_names=None, level_alias=None):
        self.code = code
        self.level_alias = level_alias or {}
        self.level_style = level_style      # "pad2" or "strip0"
        self.room_width = room_width
        self.pad = pad                      # "left" or "right"
        self.entity_style = entity_style    # "underscore" or "dotted"
        self.level_names = level_names or {}

    def norm_level(self, raw):
        if raw is None:
            return None
        raw = raw.strip()
        raw = self.level_alias.get(raw.upper(), raw)
        if not raw.isdigit():
            return raw
        if self.level_style == "pad2":
            return raw.zfill(2)
        return raw.lstrip("0") or "0"       # strip0: 04 -> 4

    def norm_room(self, num, suffix):
        """R3 - the room number is a fixed-width field.

        SSC prints it left-padded and never lost a digit.  HQ's numbers reached
        the sheet as decimals, so Excel ate the trailing zero: 1.020 came back
        as 1.02 and 3.360 as 3.36.  Those pad on the right.
        """
        if len(num) < self.room_width:
            num = (num.zfill(self.room_width) if self.pad == "left"
                   else num.ljust(self.room_width, "0"))
        return num + (suffix or "")

    def level_segment(self, raw):
        return self.level_names.get(raw, "Level-%s" % raw)


BUILDINGS = {
    "SSC": Building(
        code="SSC",
        level_style="pad2",
        room_width=3,
        pad="left",
        entity_style="underscore",          # entity:SSC_01_024_CORRIDOR
        level_alias={"B1": "B"},            # drawings and BMS both print B.013
        level_names={"B": "Level-B1", "B1": "Level-B1", "01": "Level-01",
                     "02": "Level-02", "03": "Level-03"},
    ),
    "HQ": Building(
        code="HQ",
        level_style="strip0",               # E writes 4.010, D writes 04.010
        room_width=3,
        pad="right",
        entity_style="dotted",              # entity:HQ_4.010_STR_PL_DIR
        level_alias={"B": "B1"},            # drawings print B.001, E prints B1.001
        level_names={"B1": "Level-B1", "G": "Level-G", "RF": "Level-RF"},
    ),
}

# A room reference is a level token, a dot, and a room number with an optional
# letter suffix: B.013, 01.024, 03.006A, 1.027.
REF_RE = re.compile(r"\b(B\d?|G|RF|\d{1,2})\.(\d{1,3})([A-Z])?\b")
# Some sources print the reference with no dot at all (ST-4, ROOM NO 42).
BARE_NUM_RE = re.compile(r"\b(\d{2,3})([A-Z])?\b")


def norm_ref(building, level, num, suffix):
    return building.norm_level(level), building.norm_room(num, suffix)


def split_ref(building, text):
    """Return (level, room, name) for a source cell, or (None, None, name)."""
    if text is None:
        return None, None, None
    s = str(text).strip()
    if not s:
        return None, None, None
    m = REF_RE.search(s)
    if not m:
        return None, None, s.strip()
    lvl, room = norm_ref(building, m.group(1), m.group(2), m.group(3))
    name = (s[: m.start()] + " " + s[m.end():]).strip()
    return lvl, room, name


def strip_ref(name, ref):
    """Drop a bare reference token the source ran into the name (ST-4 STAIR 4)."""
    if not name or not ref:
        return name
    out = re.sub(r"(?<![A-Za-z0-9])%s(?![A-Za-z0-9])" % re.escape(ref), " ", name)
    out = " ".join(out.split())
    return out or name


def parse_entity(building, text):
    """entity:SSC_01_024_CORRIDOR -> ('01', '024', 'CORRIDOR')."""
    if text is None:
        return None, None, None
    s = str(text).strip()
    if not s.startswith("entity:"):
        return None, None, None
    parts = s[len("entity:"):].split("_")
    if len(parts) < 2 or parts[0] != building.code:
        return None, None, None
    # Both shapes appear: the register writes entity:HQ_4.010_CA_COORD, the
    # delivered HQ draft writes entity:HQ_10_002A_OFFICE_SPACE.  Read either.
    if "." in parts[1]:
        lvl, room = parts[1].split(".", 1)
        name = "_".join(parts[2:])
    else:
        if len(parts) < 3:
            return None, None, None
        lvl, room, name = parts[1], parts[2], "_".join(parts[3:])
    m = re.match(r"^(\d+)([A-Z]?)$", room)
    if m:
        room = building.norm_room(m.group(1), m.group(2))
    return (building.norm_level(lvl), room,
            name.replace("_", " ").strip())


def key(name):
    """Comparison form - case and punctuation are not differences."""
    if not name:
        return ""
    s = str(name).upper().replace("&", " AND ")
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    return " ".join(s.split())


def squash(name):
    return re.sub(r"[^A-Z0-9]+", "", str(name or "").upper())


def same_name(a, b):
    """Two spellings of one room name.

    HQ's column E was built from the BMS tokens, so it writes
    STRPLDIRTECHNSTAFF where the drawings write STR. PL. DIR. TECH. STAFF.
    Comparing on punctuation alone calls those different names and hands the
    room the unreadable spelling.
    """
    if not a or not b:
        return False
    if key(a) == key(b):
        return True
    sa, sb = squash(a), squash(b)
    if not sa or not sb:
        return False
    return difflib.SequenceMatcher(None, sa, sb).ratio() >= 0.92


def split_prefix(name, candidates):
    """Break the department prefix off a name column E ran together.

    HQ's E writes EXECDIRLEGAL_ADVISOR where the drawings write LEGAL ADVISOR.
    The prefix is real and is what tells 11.017 from 11.018, so it is kept -
    but the boundary is visible wherever another column spells the tail out,
    and splitting there is the difference between Execdirlegal-Advisor and
    Execdir-Legal-Advisor.
    """
    sq = squash(name)
    best = None
    for c in candidates:
        sc = squash(c)
        if not sc or sc == sq or not sq.endswith(sc):
            continue
        if len(sq) - len(sc) < 3:       # a one- or two-letter prefix is noise
            continue
        if best is None or len(sc) > len(squash(best)):
            best = c
    if best is None:
        return name
    head = name[:len(name) - len(str(best).replace(" ", ""))]
    # walk back over the characters of the tail as they appear in `name`
    tail_len, i = len(squash(best)), len(name)
    seen = 0
    while i > 0 and seen < tail_len:
        i -= 1
        if name[i].isalnum():
            seen += 1
    head = name[:i].rstrip(" _-")
    return "%s %s" % (head, best) if head else name


def best_spelling(name, candidates):
    """R2 - of the spellings of one name, the clearest is the most spaced."""
    best, best_words = name, len(str(name).split())
    for c in candidates:
        if not c or not same_name(name, c):
            continue
        w = len(str(c).split())
        if w > best_words:
            best, best_words = c, w
    return best


def to_segment(name):
    """'VISITOR'S CUBICLE' -> 'Visitors-Cubicle'; keeps acronyms upper case."""
    if not name:
        return ""
    s = str(name).replace("&", " And ").replace("/", " ")
    s = s.replace("'", "").replace("’", "")
    s = re.sub(r"[^A-Za-z0-9\- ]+", " ", s)      # drops . , ( ) : etc.
    s = re.sub(r"\s*-\s*", "-", s)               # IDF - 1 -> IDF-1
    words = [w for w in re.split(r"\s+", s) if w]
    out = []
    for w in words:
        parts = []
        for p in w.split("-"):
            if not p:
                continue
            up = p.upper()
            if up in WORD_FIXES:
                parts.append(WORD_FIXES[up])
            elif up in ACRONYMS:
                parts.append(up)
            elif p.isdigit():
                parts.append(p)
            else:
                parts.append(p.capitalize())
        out.append("-".join(parts))
    return "-".join(out)


def cell_fill(c):
    try:
        if c.fill is not None and c.fill.fill_type == "solid":
            return str(c.fill.start_color.rgb)
    except Exception:
        pass
    return ""


# --------------------------------------------------------------------------
# the decision
# --------------------------------------------------------------------------
def decide(building, row):
    """Return (level, room, name, source, notes[])."""
    notes = []
    d_l, d_r, d_n = split_ref(building, row["D"])
    h_l, h_r, h_n = split_ref(building, row["H"])
    j_l, j_r, j_n = split_ref(building, row["J"])
    e_l, e_r, e_n = parse_entity(building, row["E"])

    if d_r is None and h_r is None and j_r is None and e_r is None:
        notes.append("no room reference in D, H, J or E - left unresolved")
        return None, None, None, "", notes

    if row["green"]:
        # R1 - last of the three names that is present.
        for tag, (l, r, n) in (("J", (j_l, j_r, j_n)),
                               ("H", (h_l, h_r, h_n)),
                               ("D", (d_l, d_r, d_n))):
            if n:
                notes.append("green row: took the last populated name (%s)" % tag)
                if e_n and key(n) != key(e_n):
                    notes.append("overrides column E (%s)" % e_n)
                if r is None and e_r:
                    # the screen names the room but not its number - only the
                    # name was in dispute, so E still supplies the reference.
                    l, r = e_l, e_r
                    n = strip_ref(n, e_r)
                    notes.append("no reference on the screen - taken from E "
                                 "(%s.%s)" % (e_l, e_r))
                return l, r, n, tag, notes
        notes.append("green row with no name in D, H or J")
        return None, None, None, "", notes

    # R5 - D or J, on whether H carries a room number.
    if h_r is not None:
        pick, tag = (d_l, d_r, d_n), "D"
        notes.append("H carries a room number, so column D stands")
    elif j_r is not None and j_n:
        pick, tag = (j_l, j_r, j_n), "J"
        notes.append("H has no room number and J has both, so column J stands")
    else:
        pick, tag = (d_l, d_r, d_n), "D"
        notes.append("no usable J, so column D stands")

    lvl, room, name = pick

    # R2 + R4 - E arbitrates the name.
    if e_n:
        if not name:
            notes.append("%s is a bare reference with no name - name taken "
                         "from E" % tag)
            name, tag = e_n, tag + "->E"
        elif not same_name(name, e_n):
            if d_n and same_name(d_n, e_n):
                notes.append("column D matches E (%s) and %s does not - D wins"
                             % (e_n, tag))
                lvl, room, name, tag = d_l, d_r, d_n, "D"
            else:
                notes.append("%s (%s) does not match E - name taken from E"
                             % (tag, name))
                name = e_n
                tag = tag + "->E"
        if room is None:
            notes.append("no room number in the picked column - taken from E")
        if e_r and room != e_r:
            if room is None:
                lvl, room = e_l, e_r
            else:
                notes.append("room number %s.%s disagrees with E (%s.%s)"
                             % (lvl, room, e_l, e_r))
        if lvl is None:
            lvl = e_l
    clearer = best_spelling(name, [d_n, h_n, j_n, e_n])
    if clearer != name:
        notes.append("spelled %r on the clearest source" % clearer)
        name = clearer
    split = split_prefix(name, [d_n, h_n, j_n])
    if split != name:
        notes.append("department prefix split off - %r" % split)
        name = split
    return lvl, room, name, tag, notes


# --------------------------------------------------------------------------
# R2 across rows - one room reference must not end up with two spellings
# --------------------------------------------------------------------------
def is_variant(a, b):
    """'ASSOC DIR OFFICE' and 'ASSOC DIRECTORS OFFICE' name the same room."""
    ta, tb = a.split(), b.split()
    if len(ta) != len(tb) or not ta:
        return False
    for x, y in zip(ta, tb):
        if x == y:
            continue
        lo, hi = (x, y) if len(x) < len(y) else (y, x)
        if len(lo) >= 3 and hi.startswith(lo):
            continue
        return False
    return True


def reconcile(rows, building):
    """Collapse abbreviation variants sharing one room reference.

    Only variants collapse.  01.029 really does hold ROOM NO 42, 43, 44 and a
    SHELL SPACE, and those stay four rooms; what must not survive is the same
    room spelled two ways.
    """
    groups = {}
    for r in rows:
        if r["name"] and r["ref"]:
            groups.setdefault((r["level"], r["ref"]), []).append(r)

    # What column E calls each reference, where it calls it one thing.  SSC's
    # 01.029 is deliberately four rooms in E, so E does not arbitrate there.
    e_name = {}
    for r in rows:
        el, er, en = parse_entity(building, r["E"])
        if er and en:
            e_name.setdefault((el, er), set()).add(en)
    e_name = {k: v.pop() for k, v in e_name.items() if len(v) == 1}

    collapsed, split = [], []
    for (lvl, ref), grp in sorted(groups.items()):
        names = {}
        for r in grp:
            names.setdefault(key(r["name"]), r["name"])
        if len(names) < 2:
            continue
        arb = e_name.get((lvl, ref))
        if arb:
            arb = best_spelling(arb, [r["name"] for r in grp] + [arb])
            for r in grp:
                if not same_name(r["name"], arb):
                    was = r["name"]
                    r["notes"].append(
                        "%s.%s is spelled %r on other rows; E names it %r, "
                        "which settles it" % (lvl, ref, was, arb))
                    collapsed.append((lvl, ref, was, arb))
                r["name"] = arb
            names = {key(arb): arb}
        ks = list(names)
        canon = dict()
        for i, a in enumerate(ks):
            for b in ks[i + 1:]:
                if is_variant(a, b):
                    win = a if len(a) >= len(b) else b
                    canon[a] = win
                    canon[b] = win
        if canon:
            for r in grp:
                k = key(r["name"])
                if k in canon and canon[k] != k:
                    was = r["name"]
                    r["name"] = names[canon[k]]
                    r["notes"].append(
                        "same room as another row spelled %r - taken to the "
                        "fuller spelling %r" % (was, r["name"]))
                    collapsed.append((lvl, ref, was, r["name"]))
        left = {key(r["name"]) for r in grp}
        if len(left) > 1:
            split.append((lvl, ref, sorted(names[k] for k in left)))
    return collapsed, split


def subject(building, lvl, room, name):
    if not name:
        return ""
    seg_name = to_segment(name)
    if not seg_name:
        return ""
    # entity:SSC_01-029_Shell-Space - the reference segment already carries the
    # level, so a separate Level-01 segment would only repeat it.
    parts = ["entity:%s" % building.code]
    if room:
        parts.append("%s-%s" % (lvl, room) if lvl else room)
    parts.append(seg_name)
    return "_".join(parts)


def label(lvl, room, name):
    """Reading form in the QF SSC house style - '1.024 CORRIDOR'."""
    n = key(name)
    if not (lvl and room):
        return n
    if not room[:1].isdigit():          # ST-4 and the like carry no level
        return "%s %s" % (room, n)
    shown = lvl if lvl.startswith("B") else lvl.lstrip("0") or "0"
    return "%s.%s %s" % (shown, room, n)


# --------------------------------------------------------------------------
def pick_sheet(wb):
    """The register sheet is the one headed ROOM NAME AS PER DRAWINGS."""
    for cand in wb:
        for r in cand.iter_rows(min_row=1, max_row=4, values_only=True):
            if r and any(isinstance(c, str) and "AS PER DRAWINGS" in c.upper()
                         for c in r):
                return cand
    return None


def load_one(path, building):
    wb = openpyxl.load_workbook(path)
    ws = pick_sheet(wb)
    if ws is None:
        raise SystemExit("no register sheet in %s" % path)
    rows = {}
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row):
        tag = r[0].value
        if not tag or not isinstance(tag, str):
            continue
        if not re.match(r"^[A-Z]+[A-Z_]*\d", tag.strip()):
            continue
        rows[r[0].row] = {
            "excel_row": r[0].row,
            "A": tag.strip(),
            "B": r[1].value,
            "C": r[2].value,
            "D": r[3].value,
            "E": r[4].value,
            "H": r[7].value,
            "J": r[9].value,
            "green": cell_fill(r[0]) == GREEN,
        }
    return rows


def merge(parts, building):
    """Combine the per-level-range splits of one building's register.

    HQ came as two workbooks holding the same 761 rows, each with the BMS
    screen read for its own levels only - B1 to 2 in one, 3 to Roof in the
    other.  No row carries a screen reading in both, so J is a plain union and
    the file that supplied it is that row's author.  Green is a union too:
    each file only marked its own range.

    Where the two disagree on the drawings, the ontology name or the VAV/FCU
    list, the file that owns the row's level wins and the disagreement is
    written into the row's notes rather than quietly dropped.
    """
    if len(parts) == 1:
        return sorted(parts[0][1].values(), key=lambda r: r["excel_row"])

    keys = set()
    for _, rows in parts:
        keys |= set(rows)
    out = []
    for k in sorted(keys):
        present = [(name, rows[k]) for name, rows in parts if k in rows]
        owner_name, base = None, None
        for name, r in present:
            if r["J"]:
                owner_name, base = name, r
                break
        if base is None:
            # No screen reading anywhere - fall back to the file whose level
            # range covers the row, read off whichever E is available.
            for name, r in present:
                lvl, _, _ = parse_entity(building, r["E"])
                if lvl is not None:
                    owner_name = owner_for_level(parts, lvl)
                    break
            for name, r in present:
                if name == owner_name:
                    base = r
                    break
            if base is None:
                owner_name, base = present[0]
        row = dict(base)
        row["source_file"] = owner_name
        row["green"] = any(r["green"] for _, r in present)
        for col in ("D", "E", "H"):
            vals = {name: r[col] for name, r in present}
            others = {n: v for n, v in vals.items()
                      if n != owner_name and v not in (None, "")}
            if row[col] in (None, "") and others:
                n, v = sorted(others.items())[0]
                row[col] = v
                row.setdefault("merge_notes", []).append(
                    "column %s was blank in %s, taken from %s" % (col, owner_name, n))
            else:
                for n, v in sorted(others.items()):
                    if str(v).strip() != str(row[col]).strip():
                        row.setdefault("merge_notes", []).append(
                            "%s disagrees between the two workbooks - %s says "
                            "%r, %s says %r; %s owns this level"
                            % (col, owner_name, row[col], n, v, owner_name))
        out.append(row)
    return out


def owner_for_level(parts, lvl):
    """Which split covers a level - by the level ranges the file names state."""
    low = lvl in ("B", "B1", "G", "1", "2", "01", "02")
    for name, _ in parts:
        stem = os.path.basename(name).upper()
        if low and ("B-2F" in stem or "B2F" in stem):
            return name
        if not low and ("3F" in stem):
            return name
    return parts[0][0]


def load(paths, building):
    parts = [(p, load_one(p, building)) for p in paths]
    return merge(parts, building)


DELIVERED = {"SSC": "QF_SSC_Ontology_ver02.xlsx",
             "HQ": "QF_HQ_Ontology_draft0.4.xlsx"}


def load_delivered(code):
    """Room reference -> name, from the previously delivered ontology."""
    name = DELIVERED.get(code)
    if not name:
        return None
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "..", "..", "reference-models", name)
    if not os.path.isfile(path):
        return None
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # Pick the ontology sheet by its header, never by name or position -
    # "SSC_Ontology_Ver0.6" contains "log" and the log tab is sheet one.
    ws = None
    for cand in wb:
        head = next(cand.iter_rows(min_row=1, max_row=1, values_only=True), ())
        if head and str(head[0]).strip().lower() == "subject":
            ws = cand
            break
    if ws is None:
        return None
    building = BUILDINGS[code]
    exact, refs = set(), {}
    for r in ws.iter_rows(values_only=True):
        if not r or len(r) < 5:
            continue
        for subj, typ in ((r[0], r[1]), (r[3], r[4])):
            if typ != "rec:Room":
                continue
            lvl, room, name = parse_entity(building, subj)
            if not room:
                continue
            exact.add((lvl, room, key(name)))
            refs.setdefault((lvl, room), name)
    return {"exact": exact, "refs": refs}


HEADERS = [
    "Tag No.", "Equipment Type", "Included/Not Included",
    "D - room per drawings", "E - existing ontology name",
    "H - room per VAV/FCU list", "J - room per BMS screen",
    "Green", "Level", "Room ref", "Room name",
    "Subject (Dar Cairo shape)", "rdfs:label_en", "Chosen from", "Notes",
]


def write(out_path, sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr_fill = PatternFill("solid", start_color="FF1F4E79")
    hdr_font = Font(color="FFFFFFFF", bold=True)
    warn = PatternFill("solid", start_color="FFFFF2CC")
    bad = PatternFill("solid", start_color="FFF8CBAD")
    for name, rows in sheets:
        ws = wb.create_sheet(name)
        ws.append(HEADERS)
        for c in ws[1]:
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(wrap_text=True, vertical="center")
        for r in rows:
            ws.append([
                r["A"], r["B"], r["C"], r["D"], r["E"], r["H"], r["J"],
                "green" if r["green"] else "",
                r["level"] or "", r["ref"] or "", r["name"] or "",
                r["subject"], r["label"], r["source"], "; ".join(r["notes"]),
            ])
            row = ws.max_row
            if not r["subject"]:
                for c in ws[row]:
                    c.fill = bad
            elif r["green"] or "disagrees" in "; ".join(r["notes"]):
                for c in ws[row]:
                    c.fill = warn
        widths = [12, 8, 12, 32, 40, 32, 30, 7, 9, 10, 28, 46, 34, 12, 70]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(HEADERS)), ws.max_row)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", action="append", default=[],
                    metavar="CODE=PATH", help="e.g. SSC=sources/SSC_rooms.xlsx")
    ap.add_argument("--out", required=True)
    ap.add_argument("--rooms-csv", help="distinct room list")
    args = ap.parse_args()

    sheets, all_rooms = [], []
    for spec in args.src:
        code, path = spec.split("=", 1)
        building = BUILDINGS[code]
        rows = load([p for p in path.split(",") if p], building)
        for r in rows:
            lvl, room, name, src, notes = decide(building, r)
            r["level"], r["ref"], r["name"] = lvl, room, name
            r["source"], r["notes"] = src, r.get("merge_notes", []) + notes

        collapsed, split = reconcile(rows, building)
        for lvl, ref, was, now in collapsed:
            print("  collapsed %s.%s  %r -> %r" % (lvl, ref, was, now))
        for lvl, ref, names in split:
            print("  %s.%s carries %d distinct names: %s"
                  % (lvl, ref, len(names), ", ".join(names)))

        known = load_delivered(code)
        for r in rows:
            r["subject"] = subject(building, r["level"], r["ref"], r["name"])
            r["label"] = label(r["level"], r["ref"], r["name"]) if r["name"] else ""
            if known and r["ref"]:
                k = (r["level"], r["ref"], key(r["name"]))
                if k not in known["exact"]:
                    if (r["level"], r["ref"]) in known["refs"]:
                        r["notes"].append(
                            "%s calls %s.%s %r"
                            % (DELIVERED[code], r["level"], r["ref"],
                               known["refs"][(r["level"], r["ref"])]))
                    else:
                        r["notes"].append(
                            "%s.%s is not in %s"
                            % (r["level"], r["ref"], DELIVERED[code]))
            if r["subject"]:
                all_rooms.append((code, r["subject"], r["label"],
                                  building.level_segment(r["level"] or ""),
                                  "%s.%s" % (r["level"], r["ref"])
                                  if r["ref"] else "", r["A"]))
        sheets.append((code, rows))
        blank = [r["A"] for r in rows if not r["subject"]]
        print("%s: %d asset rows, %d distinct rooms, %d unresolved%s"
              % (code, len(rows), len({r['subject'] for r in rows if r['subject']}),
                 len(blank), (" (%s)" % ", ".join(blank)) if blank else ""))

    write(args.out, sheets)
    print("wrote %s" % args.out)

    if args.rooms_csv:
        seen, out = set(), []
        for code, subj, lab, lvl, ref, tag in all_rooms:
            if subj in seen:
                continue
            seen.add(subj)
            out.append((code, subj, lab, lvl, ref))
        out.sort()
        with open(args.rooms_csv, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["building", "subject", "rdfs:label_en",
                        "level segment", "room ref"])
            w.writerows(out)
        print("wrote %s (%d rooms)" % (args.rooms_csv, len(out)))


if __name__ == "__main__":
    main()
