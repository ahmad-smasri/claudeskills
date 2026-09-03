"""Build the HQ and SSC equipment-and-room workbook.

Two sources, side by side, for every piece of equipment that is not an AHU, a
VAV or an FCU:

  * the room the BMS screens put it in - harvested from the pass in
    `../bms-room-allocation`, which covers SSC only;
  * the room the delivered ontology puts it in - `rec:locatedIn` on the
    equipment entity, with the room's own `rdfs:label_en` for a name.

A row is marked as needing revision when the two disagree, when the ontology
points at a room that has no name, or when equipment carrying points has no
location at all.
"""
import collections
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import ontology
import screens

REF = '/home/user/claudeskills/reference-models'
BOOKS = [('SSC', os.path.join(REF, 'QF_SSC_Ontology_ver02.xlsx')),
         ('HQ', os.path.join(REF, 'QF_HQ_Ontology_draft0.4.xlsx'))]
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   'HQ_SSC_equipment_rooms.xlsx')

GREEN = PatternFill('solid', fgColor='FF00B050')   # the house 'needs a look'
HEAD = PatternFill('solid', fgColor='FF1F4E79')
THIN = Side(style='thin', color='FFBFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ('Building', 10), ('Tag', 20), ('Equipment class', 30),
    ('Ontology entity', 30), ('Ontology room entity', 32),
    ('Room per ontology', 26), ('Room per BMS screen', 26),
    ('BMS screen', 16), ('Screen reading', 13),
    ('Needs revision', 15), ('Why', 62),
]

PLACEHOLDER = re.compile(r'^entity:<|^none$', re.I)


def tag_of(entity):
    """the tag as the screens write it, from the ontology entity name"""
    t = entity.split(':', 1)[-1]
    for pre in ('SSC_', 'HQ_'):
        if t.startswith(pre):
            t = t[len(pre):]
    return t


def room_key(text):
    """a room's number and its significant words, for comparing two spellings

    The screens write `B.022 SSP` where the ontology writes `B1.022 SSP` -
    the level segment is spelled differently and the digits are not.
    """
    if not text:
        return None, set()
    m = re.search(r'\b([A-Z]{0,2}\d*)\.(\d+[A-Z]?)\b', text.upper())
    num = m.group(2) if m else None
    words = {w for w in re.findall(r'[A-Z]+', text.upper()) if len(w) > 2}
    return num, words


def same_room(a, b):
    na, wa = room_key(a)
    nb, wb = room_key(b)
    if na and nb:
        return na == nb
    return bool(wa & wb)


def collect():
    reading = screens.readings()
    rows, unlocated = [], []
    for building, path in BOOKS:
        eq, label, cls = ontology.equipment(path)
        ws = ontology.sheet(path)
        raw = [r for r in ws.iter_rows(min_row=2, max_col=9, values_only=True) if r[0]]
        haspoint = {str(r[0]) for r in raw if r[2] == 'brick:hasPoint'}
        parts = {str(r[3]) for r in raw if r[2] == 'brick:hasPart'}
        located = {e['entity'] for e in eq}

        index = {}
        for e in eq:
            for k in screens.key(tag_of(e['entity'])):
                index.setdefault(k, e)
        matched = {}
        for tag, (bldg, room, screen, conf) in reading.items():
            if bldg != building:
                continue          # the same tag exists in both buildings
            hit = None
            for k in screens.key(tag):
                hit = hit or index.get(k)
            if hit:
                matched[hit['entity']] = (tag, room, screen, conf)

        for e in sorted(eq, key=lambda z: (z['class'], z['entity'])):
            tag, room, screen, conf = matched.get(e['entity'], (tag_of(e['entity']), '', '', ''))
            why = []
            if not e['room'] or e['room'] in ('None', 'entity:'):
                why.append('the rec:locatedIn row carries no room at all')
            elif PLACEHOLDER.match(e['room']):
                why.append('the ontology points at the placeholder %s rather'
                           ' than a room' % e['room'])
            elif not e['room_label']:
                why.append('the room entity %s carries no rdfs:label_en, so it'
                           ' has no readable name' % e['room'])
            if e['class'] == 'brick:CRAC' and 'PARKING' in e['room_label'].upper():
                why.append('a CRAC is a computer-room unit and the ontology'
                           ' puts it in a parking bay')
            agrees = room and same_room(room, e['room_label'])
            if room and not agrees:
                why.append('the BMS screen puts it in %s and the ontology in %s'
                           % (room, e['room_label'] or e['room']))
                if conf == 'check':
                    why.append('and the screen reading was itself marked as'
                               ' wanting a second look when it was made')
            rows.append([building, tag, e['class'], e['entity'], e['room'],
                         e['room_label'], room, screen, conf,
                         'YES' if why else '', '; '.join(why)])

        for ent in sorted(haspoint - located - parts):
            klass = cls.get(ent, '')
            if klass in ontology.SKIP or klass in ('rec:Building', 'rec:Site'):
                continue
            if re.search(r'_(VFD|Electrical-Meter)$', ent):
                continue          # hangs off a fan, located with its parent
            unlocated.append([building, tag_of(ent), klass, ent, '', '', '', '',
                              '', 'YES',
                              'the ontology gives this equipment no'
                              ' rec:locatedIn at all'])
    return rows + unlocated


def style_header(ws, row, cols):
    for i, (name, width) in enumerate(cols, 1):
        c = ws.cell(row, i, name)
        c.fill = HEAD
        c.font = Font(bold=True, color='FFFFFFFF')
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 30


def main():
    rows = collect()
    wb = openpyxl.Workbook()

    intro = wb.active
    intro.title = 'START HERE'
    intro.column_dimensions['A'].width = 116
    text = [
        ('HQ and SSC equipment and the rooms they are in', True),
        ('', False),
        ('One row per piece of equipment that is NOT an AHU, a VAV or an FCU -'
         ' those three were covered by the earlier BMS screen pass and are left'
         ' out here.', False),
        ('', False),
        ('Two sources, side by side:', True),
        ('  Room per BMS screen - the room the serving-area leader on the BMS'
         ' floor-plan screen points at. Harvested from the earlier pass in'
         ' projects/bms-room-allocation. It covers SSC only: the HQ pass'
         ' recorded VAVs and FCUs and did not read the plant icons, and the HQ'
         ' screen images are no longer in this session to re-read.', False),
        ('  Room per ontology - the rec:locatedIn on the equipment entity in'
         ' the delivered ontology, named by that room entity\'s own'
         ' rdfs:label_en.', False),
        ('', False),
        ('A row is filled GREEN where it needs revision, for one of three'
         ' reasons, spelled out in the Why column:', True),
        ('  1. the two sources name different rooms;', False),
        ('  2. the ontology points at a placeholder rather than a room -'
         ' entity:<Location> on HQ - or at a room entity that carries no'
         ' readable name, or at nothing at all;', False),
        ('  3. the equipment carries points but the ontology gives it no'
         ' rec:locatedIn at all;', False),
        ('  4. the class and the room do not go together - eight CRACs, which'
         ' are computer-room units, are located in VVIP parking bays.', False),
        ('', False),
        ('One pattern worth knowing about before you start. On SSC the four'
         ' DX units DXB0001 to DXB0004 all sit in B1.005 UPS in the ontology'
         ' while the screens spread them across three different rooms, and'
         ' DXB0007 and DXB0008 share B1.019 FM STORAGE the same way. A block'
         ' of units carrying one room where the screens disagree is the'
         ' signature of a fill-down, not of a reading.', False),
        ('', False),
        ('Where the Room per BMS screen column is empty there was no reading'
         ' to compare against, so the row is judged on the ontology alone.',
         False),
    ]
    for i, (line, bold) in enumerate(text, 1):
        c = intro.cell(i, 1, line)
        c.font = Font(bold=bold, size=13 if i == 1 else 11)
        c.alignment = Alignment(wrap_text=True, vertical='top')

    data = wb.create_sheet('Equipment and rooms')
    style_header(data, 1, COLUMNS)
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = data.cell(r, c, v)
            cell.border = BOX
            cell.alignment = Alignment(vertical='top', wrap_text=(c == 11))
        if row[9] == 'YES':
            for c in range(1, len(COLUMNS) + 1):
                data.cell(r, c).fill = GREEN
    data.freeze_panes = 'A2'
    data.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(COLUMNS)),
                                        len(rows) + 1)

    summary = wb.create_sheet('Summary')
    style_header(summary, 1, [('Building', 12), ('Equipment class', 32),
                              ('Rows', 8), ('Needing revision', 18)])
    counts = collections.Counter((r[0], r[2]) for r in rows)
    flagged = collections.Counter((r[0], r[2]) for r in rows if r[9] == 'YES')
    for i, (k, n) in enumerate(sorted(counts.items()), 2):
        summary.cell(i, 1, k[0]).border = BOX
        summary.cell(i, 2, k[1]).border = BOX
        summary.cell(i, 3, n).border = BOX
        c = summary.cell(i, 4, flagged.get(k, 0))
        c.border = BOX
        if flagged.get(k):
            for j in range(1, 5):
                summary.cell(i, j).fill = GREEN

    wb.save(OUT)
    print('%s: %d rows, %d need revision'
          % (os.path.basename(OUT), len(rows), sum(1 for r in rows if r[9] == 'YES')))


if __name__ == '__main__':
    main()
