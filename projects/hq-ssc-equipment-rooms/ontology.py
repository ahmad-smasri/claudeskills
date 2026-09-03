"""Pull equipment and the room it is located in out of a delivered ontology.

The sheet is one triple per row: subject, subjectType, predicate, object,
objectType. `rec:locatedIn` with an object of type `rec:Room` is the
equipment-to-room link; the room's readable name comes from its own
`rdfs:label_en` property row.

The ontology sheet is picked by its header, never by tab name or `.active` -
the HQ workbook's tab is spelled `HQ_Onotlogy_Draft_v0.4`.
"""
import openpyxl

HEADER = ('subject', 'subjecttype', 'predicate', 'object', 'objecttype')
# the three families the BMS screen pass already covered
SKIP = {'brick:Air_Handling_Unit',
        'brick:Variable_Air_Volume_Box',
        'brick:Fan_Coil_Unit'}


def sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        head = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if tuple(str(h or '').strip().lower() for h in head[:5]) == HEADER:
            return ws
    raise SystemExit('no ontology sheet in %s' % path)


def read(path):
    ws = sheet(path)
    rows = [r for r in ws.iter_rows(min_row=2, max_col=9, values_only=True) if r[0]]
    label, cls = {}, {}
    for s, st, p, o, ot, pn, pv, opn, opv in rows:
        if pn == 'rdfs:label_en' and pv:
            label.setdefault(str(s), str(pv))
        if opn == 'rdfs:label_en' and opv and o:
            label.setdefault(str(o), str(opv))
        if st:
            cls.setdefault(str(s), str(st))
        if ot and o:
            cls.setdefault(str(o), str(ot))
    located = []
    for s, st, p, o, ot, *_ in rows:
        if p != 'rec:locatedIn' or ot != 'rec:Room':
            continue
        located.append({'entity': str(s), 'class': str(st or cls.get(str(s), '')),
                        'room': str(o), 'room_label': label.get(str(o), ''),
                        'label': label.get(str(s), '')})
    seen, out = set(), []
    for e in located:                     # an entity can carry the row twice
        k = (e['entity'], e['room'])
        if k in seen:
            continue
        seen.add(k)
        out.append(e)
    return out, label, cls


def equipment(path):
    """the located equipment that is not an AHU, a VAV or an FCU"""
    located, label, cls = read(path)
    return [e for e in located if e['class'] not in SKIP], label, cls
