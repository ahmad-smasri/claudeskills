#!/usr/bin/env python3
"""Generate Assumption_Log.xlsx from the per-building CSVs beside this script.

    python3 build.py                  # rebuild the workbook
    python3 build.py --add RDC2       # start a sheet for a new building
    python3 build.py --check          # rebuild to a temp file and diff, write nothing

**The CSVs are the source. The workbook is output.** Edit `QNL.csv` and rebuild;
never edit the workbook and expect it to survive. This replaces the older
format_assumption_log.py, which formatted the workbook in place - that made the
workbook itself the source, and a workbook is a binary blob git cannot merge, so
two people editing it meant one of them silently lost their entries.

One CSV per building, named for the sheet it becomes. All nine columns are text
in the CSV; `Rows affected` is written back as a number where the whole value is
an integer and as text otherwise, because seven entries legitimately say things
like `n/a` and `564 (574 errors -> 10)`.

A CSV with a header and no rows becomes a sheet with the italic
`(no entries yet)` placeholder, which is how RDC currently stands.
"""
import argparse
import csv
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
BOOK = HERE.parent / "Assumption_Log.xlsx"

# Sheet order in the workbook. A CSV not named here is appended after these, so
# adding a building is a matter of dropping in the file - but the four that
# exist keep their established order.
ORDER = ["SSC", "HQ", "QNL", "RDC"]

COLUMNS = [
    ("ID", 8.21875), ("Date", 10.33203125), ("Category", 12.21875),
    ("Layer", 9.6640625), ("Entity / Scope", 32.77734375),
    ("What the source says", 46.0), ("What we did", 45.0),
    ("Why / basis", 52.0), ("Rows affected", 12.88671875),
]

HEADER_FILL = "FF1F4E78"
HEADER_HEIGHT = 30.0
BODY_HEIGHT = 58.05
GRID = "FFBFBFBF"
PLACEHOLDER = "(no entries yet)"

# Category -> fill. A category with no entry here simply gets no fill, so
# inventing one in a CSV is not a crash.
CATEGORY_FILL = {
    "Identifier": "DDEBF7", "Location": "FCE4D6", "Units": "E2EFDA",
    "Spelling": "FFF2CC", "Class": "E4DFEC", "Scope": "F2F2F2",
    "Source defect": "FBE5D6", "Structure": "DEEAF6",
}

HEADERS = [name for name, _ in COLUMNS]


def read(path):
    """One CSV -> its rows, with the header checked against COLUMNS."""
    with path.open(newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        sys.exit("%s is empty - it needs at least the header row" % path.name)
    if rows[0] != HEADERS:
        sys.exit("%s has the wrong header.\n  expected %s\n  found    %s"
                 % (path.name, HEADERS, rows[0]))
    out = []
    for n, row in enumerate(rows[1:], 2):
        if not any(cell.strip() for cell in row):
            continue                      # a blank spacer line is not an entry
        if len(row) != len(HEADERS):
            sys.exit("%s line %d has %d fields, expected %d"
                     % (path.name, n, len(row), len(HEADERS)))
        out.append(row)
    return out


def value(text, column):
    """Text from the CSV as the cell should hold it.

    Only `Rows affected` is ever a number, and only when the whole field is an
    integer - `8 dropped` and `n/a` stay text.
    """
    text = text.strip()
    if not text:
        return None
    if column == len(HEADERS) - 1 and text.lstrip("-").isdigit():
        return int(text)
    return text


def sheet(wb, name, rows):
    ws = wb.create_sheet(name)
    edge = Side(style="thin", color=GRID)
    box = Border(left=edge, right=edge, top=edge, bottom=edge)

    for c, (label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c, label)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(bold=True, color="FFFFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = HEADER_HEIGHT

    body = Alignment(vertical="top", wrap_text=True)
    for r, row in enumerate(rows, 2):
        for c, text in enumerate(row):
            cell = ws.cell(r, c + 1, value(text, c))
            cell.alignment = body
            cell.border = box
        fill = CATEGORY_FILL.get(row[2].strip())
        if fill:
            ws.cell(r, 3).fill = PatternFill("solid", fgColor=fill)
        ws.row_dimensions[r].height = BODY_HEIGHT

    if not rows:
        cell = ws.cell(2, 1, PLACEHOLDER)
        cell.font = Font(italic=True, color="FF808080")
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(2, c).alignment = body
            ws.cell(2, c).border = box
        ws.row_dimensions[2].height = BODY_HEIGHT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(COLUMNS)),
                                      max(len(rows) + 1, 2))
    return ws


def build(target):
    names = [p.stem for p in sorted(HERE.glob("*.csv"))]
    ordered = [n for n in ORDER if n in names] + [n for n in names if n not in ORDER]
    if not ordered:
        sys.exit("no CSVs beside %s - nothing to build" % Path(__file__).name)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    total = 0
    for name in ordered:
        rows = read(HERE / ("%s.csv" % name))
        sheet(wb, name, rows)
        total += len(rows)
        print("  %-5s %s" % (name, "no entries" if not rows
                             else "%d entries" % len(rows)))
    wb.save(target)
    return total


def compare(fresh, existing):
    """Cell values that differ between a freshly built workbook and one on disk.

    Byte comparison is useless here - two builds of the same CSVs differ on the
    embedded timestamp alone - so this compares what the sheets actually say.
    """
    a = openpyxl.load_workbook(fresh)
    b = openpyxl.load_workbook(existing)
    out = []
    for name in sorted(set(a.sheetnames) | set(b.sheetnames)):
        if name not in a.sheetnames or name not in b.sheetnames:
            out.append("sheet %s is in only one of the two" % name)
            continue
        sa, sb = a[name], b[name]
        for r in range(1, max(sa.max_row, sb.max_row) + 1):
            for c in range(1, len(COLUMNS) + 1):
                x, y = sa.cell(r, c).value, sb.cell(r, c).value
                if ("" if x is None else str(x)) != ("" if y is None else str(y)):
                    out.append("%s!%s%d: workbook has %r, CSVs give %r"
                               % (name, get_column_letter(c), r, y, x))
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--add", metavar="NAME",
                    help="write an empty CSV for a new building")
    ap.add_argument("--check", action="store_true",
                    help="report whether the workbook is up to date; write nothing")
    args = ap.parse_args()

    if args.add:
        path = HERE / ("%s.csv" % args.add)
        if path.exists():
            sys.exit("%s already exists" % path.name)
        with path.open("w", newline="") as f:
            csv.writer(f, lineterminator="\n").writerow(HEADERS)
        print("wrote %s - add entries to it, then rebuild" % path.name)

    if args.check:
        if not BOOK.exists():
            sys.exit("%s does not exist - run without --check to build it"
                     % BOOK.name)
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
            build(Path(tmp.name))
            stale = compare(Path(tmp.name), BOOK)
        if stale:
            print("\n%s is STALE - %d cells differ from the CSVs:" % (BOOK.name,
                                                                      len(stale)))
            for line in stale[:12]:
                print("  " + line)
            if len(stale) > 12:
                print("  ... and %d more" % (len(stale) - 12))
            sys.exit(1)
        print("\n%s is up to date with the CSVs" % BOOK.name)
        return

    total = build(BOOK)
    print("\nwrote %s - %d entries" % (BOOK.name, total))


if __name__ == "__main__":
    main()
