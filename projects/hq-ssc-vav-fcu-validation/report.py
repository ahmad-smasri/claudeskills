import openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from compare import *

OUT = '/home/user/claudeskills/out/VAV_FCU_validation_findings.xlsx'
os.makedirs(os.path.dirname(OUT), exist_ok=True)

def overlap(r):
    a = set(norm_name(r.get('dname') or '').split())
    b = set(norm_name(r.get('sname') or '').split())
    if not a or not b:
        return None
    return len(a & b) / min(len(a), len(b))

def bucket(r):
    if r['status'] == 'DIFF_NUMBER':
        return 'Room number changed'
    if r['status'] == 'DIFF_BOTH':
        return 'Room number AND name changed'
    if r['status'] == 'DIFF_NAME':
        o = overlap(r)
        if o and o >= 0.5:
            return 'Same room - wording differs'
        if o:
            return 'Same room number - name partly different'
        return 'Same room number - name completely different'
    return {'MATCH': 'Matches the new list',
            'PARTIAL_MATCH': 'Matches on the part the list gives',
            'SOURCE_BLANK': 'New list gives no room - nothing to validate',
            'REGISTRY_BLANK': 'Register had no room name - now filled',
            'NO_SOURCE_ROW': 'Unit not in the new VAV/FCU list'}[r['status']]

HDR = Font(bold=True, color='FFFFFF')
HFILL = PatternFill('solid', fgColor='1F4E79')
WARN = PatternFill('solid', fgColor='FFF2CC')
BAD = PatternFill('solid', fgColor='F8CBAD')

wb = openpyxl.Workbook()

# ---------- summary ----------
ws = wb.active
ws.title = 'Summary'
lines = [
    ('VAV / FCU location validation - Controllable Asset Registry', ''),
    ('', ''),
    ('Sources', ''),
    ('HQ & SSC VAV LIST (HQ_&_SSC_VAV_LIST_2058.xls), column D = VAV CONTROLLER TAG', 708),
    ('HQ FCU LIST (HQ_&_SSC_FCU_LIST_1397.xls), column D = FCU CONTROLLER TAG', 133),
    ('SSC FCU LIST (HQ_&_SSC_FCU_LIST_1397.xls), column D = FCU CONTROLLER TAG', 11),
    ('', ''),
    ('Target', ''),
    ('Controllable Asset Registry, HQ + SSC sections, VAV and FCU rows', len([r for r in results])),
    ('  matched to a row in the new lists', len([r for r in results if r['status'] != 'NO_SOURCE_ROW'])),
    ('  not present in the new lists', len([r for r in results if r['status'] == 'NO_SOURCE_ROW'])),
    ('', ''),
    ('Outcome', ''),
]
b = Counter(bucket(r) for r in results)
order = ['Room number changed', 'Room number AND name changed',
         'Same room number - name completely different',
         'Same room number - name partly different',
         'Same room - wording differs', 'Matches the new list',
         'Matches on the part the list gives',
         'New list gives no room - nothing to validate',
         'Unit not in the new VAV/FCU list']
for k in order:
    if b.get(k):
        lines.append(('  ' + k, b[k]))
lines += [
    ('', ''),
    ('Column H written on', len([r for r in results if r['status'] in ('DIFF_NAME', 'DIFF_NUMBER', 'DIFF_BOTH')])),
    ('Column D left alone on', len([r for r in results if r['status'] not in ('DIFF_NAME', 'DIFF_NUMBER', 'DIFF_BOTH')])),
]
for i, (a, c) in enumerate(lines, 1):
    ws.cell(i, 1, a)
    if c != '':
        ws.cell(i, 2, c)
ws['A1'].font = Font(bold=True, size=13)
for rr in (3, 8, 13):
    ws.cell(rr, 1).font = Font(bold=True)
ws.column_dimensions['A'].width = 78
ws.column_dimensions['B'].width = 12

# ---------- full comparison ----------
ws = wb.create_sheet('All VAV & FCU rows')
cols = ['Register row', 'Building', 'Tag No. (col A)', 'Type',
        'Room name in col D (current)', 'Written to col H (new)', 'Finding',
        'New list - sheet', 'New list - row', 'New list - ref',
        'New list - room no.', 'New list - room name', 'Matched by']
ws.append(cols)
for c in ws[1]:
    c.font = HDR; c.fill = HFILL
for r in sorted(results, key=lambda x: x['row']):
    s = r['src']
    h = new_value(r) if r['status'] in ('DIFF_NAME', 'DIFF_NUMBER', 'DIFF_BOTH') else ''
    ws.append([r['row'], r['sec'], r['tag'], r['type'],
               r['D'], h, bucket(r),
               s['sheet'] if s else '', s['row'] if s else '', s['ref'] if s else '',
               s['room_no'] if s else '', ' '.join(s['room_name'].split()) if s else '',
               r['via'] if s else ''])
    if r['status'] in ('DIFF_NUMBER', 'DIFF_BOTH', 'NO_SOURCE_ROW'):
        for c in ws[ws.max_row]:
            c.fill = BAD
    elif r['status'] in ('DIFF_NAME',):
        for c in ws[ws.max_row]:
            c.fill = WARN
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
for i, w in enumerate([12, 9, 12, 7, 42, 42, 40, 20, 13, 12, 14, 34, 11], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---------- room number changes ----------
ws = wb.create_sheet('Room number changed')
ws.append(['Register row', 'Building', 'Tag No.', 'Type', 'Col D (current)',
           'Col H (new)', 'New list room no.', 'New list room name',
           'New list sheet', 'New list row'])
for c in ws[1]:
    c.font = HDR; c.fill = HFILL
for r in sorted([x for x in results if x['status'] in ('DIFF_NUMBER', 'DIFF_BOTH')], key=lambda x: x['row']):
    s = r['src']
    ws.append([r['row'], r['sec'], r['tag'], r['type'], r['D'], new_value(r),
               s['room_no'], ' '.join(s['room_name'].split()), s['sheet'], s['row']])
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
for i, w in enumerate([12, 9, 12, 7, 42, 42, 16, 34, 20, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---------- exceptions ----------
ws = wb.create_sheet('Exceptions')
ws.append(['Kind', 'Register row', 'Tag / ref', 'Detail'])
for c in ws[1]:
    c.font = HDR; c.fill = HFILL
for r in sorted([x for x in results if x['status'] == 'NO_SOURCE_ROW'], key=lambda x: x['row']):
    ws.append(['In the register, not in the new VAV/FCU list', r['row'], r['tag'],
               'col D currently: %s' % (r['D'] or '(blank)')])
regkeys = set()
for r in results:
    regkeys.add((r['sec'], norm_tag(r['tag'])))
    regkeys.add((r['sec'], norm_ref(r['tag'])))
for s in src:
    if (s['building'], norm_tag(s['tag'])) not in regkeys and (s['building'], norm_ref(s['ref'])) not in regkeys:
        ws.append(['In the new list, not in the register', '', s['tag'],
                   '%s row %s - %s, %s %s' % (s['sheet'], s['row'], s['ref'], s['room_no'], s['room_name'])])
for r in sorted([x for x in results if x['status'] == 'SOURCE_BLANK'], key=lambda x: x['row']):
    ws.append(['New list leaves room no. and name blank', r['row'], r['tag'],
               '%s row %s (%s) - col D kept as: %s' % (r['src']['sheet'], r['src']['row'], r['src']['ref'], r['D'])])
for r in sorted([x for x in results if x['status'] == 'PARTIAL_MATCH'], key=lambda x: x['row']):
    ws.append(['New list gives only part of the room', r['row'], r['tag'],
               'list has room no. "%s" / name "%s"; col D: %s' % (r['snum'], r['sname'], r['D'])])
for r in sorted([x for x in results if x['via'] == 'ref' and x['src']], key=lambda x: x['row']):
    ws.append(['Register row keyed by IFC ref, duplicates a controller-tag row', r['row'], r['tag'],
               'same unit as %s (register rows 738-750)' % r['src']['tag']])
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
for i, w in enumerate([52, 13, 14, 100], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print('wrote', OUT)
for s in wb.worksheets:
    print(' ', s.title, s.max_row - 1, 'rows')
