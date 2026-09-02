"""Compare the room the BMS screen gives against the register's column D.

Prints one line per QNL row that carries a column J reading, and writes the
same as a CSV. Room numbers are compared on the number alone - the register
writes `Break Out Area B.001A`, the screen writes only the name - so a row is
only called a difference when the *name* cannot be reconciled.
"""
import csv
import re
import sys

import openpyxl

import alloc
import write_j

REG = write_j.register_tags()


def key(screen, tag):
    """the register tag a screen tag maps to, so the note is found under it"""
    return write_j.reg_tag(tag, screen, REG)


ROWS = [(key(s, t), n)
        for s, rows in alloc.SCREENS.items() for t, _r, n in rows]
# a note starting with `!` says the screen disagrees with column D outright;
# any other note says the reading carries a caveat the word test cannot see
FLAGGED = {t: n.lstrip('!') for t, n in ROWS if n.startswith('!')}
NOTED = {t: n for t, n in ROWS if n and not n.startswith('!')}

BOOK = ('/home/user/claudeskills/projects/qnl-bms-room-allocation/'
        'Appendix_A_Asset_Register_QNL_BMS_rooms.xlsx')
OUT = ('/home/user/claudeskills/projects/qnl-bms-room-allocation/'
       'QNL_BMS_screen_findings.csv')
QNL_LO, QNL_HI = 766, 1316


STOP = {'the', 'and', 'of', 'to', 'area', 'rm', 'room', 'office', 'b'}


def words(s):
    """significant words, cut to four letters.

    The screen and the register spell the same room differently often enough
    that whole-word matching invents differences: `Analogue Resources` against
    `Analog Resource`, and `Receiving Area` against the screen's own
    `Receving Area`. Four letters is enough to keep `Storage` apart from
    `Stores` while letting those two pairs match.
    """
    ws = set(re.findall(r'[a-z]+', (s or '').lower())) - STOP
    return {w[:4] for w in ws if len(w) > 2}


def verdict(d, j, tag=None):
    if not j:
        return ''
    if tag in FLAGGED:
        return 'DIFF'
    if 'open plan' in j:
        # one open space carrying two labels: the screen cannot separate them
        return 'OPEN' if words(d) & words(j) else 'DIFF'
    wd, wj = words(d), words(j)
    if not wd:
        return 'D-BLANK'
    if tag in NOTED and 'column D' in NOTED[tag]:
        # the note says in so many words that column D names something else.
        # The word test cannot see it - `Unlabelled open area, west half of
        # Zone 1` and `OPEN READING AREA L1.001` share "open" and "area" and
        # were being called an agreement.
        return 'CHECK'
    if wd & wj:
        return 'SAME'
    # the reading came with a caveat, so the mismatch is not evidence of one
    return 'CHECK' if tag in NOTED else 'DIFF'


def findings():
    """[row, tag, D, J, K, verdict, note] for every QNL row carrying a reading"""
    wb = openpyxl.load_workbook(BOOK, data_only=True)
    ws = wb['Controllable Asset Registry']
    rows = []
    for i in range(QNL_LO, QNL_HI + 1):
        j = ws.cell(i, 10).value
        if not j:
            continue
        d = ws.cell(i, 4).value
        tag = str(ws.cell(i, 1).value or '').strip().upper()
        rows.append([i, ws.cell(i, 1).value, d, j, ws.cell(i, 11).value,
                     verdict(d, j, tag),
                     FLAGGED.get(tag) or NOTED.get(tag, '')])
    return rows


def main():
    rows = findings()
    with open(OUT, 'w', newline='') as fh:
        w = csv.writer(fh)
        w.writerow(['row', 'tag', 'D room as per drawings', 'J room per BMS screen',
                    'K screen', 'verdict', 'note'])
        w.writerows(rows)
    for r in rows:
        print(r[5].ljust(8), str(r[1]).ljust(16), '|', str(r[2])[:44].ljust(44),
              '|', str(r[3])[:52])
    n = {}
    for r in rows:
        n[r[5]] = n.get(r[5], 0) + 1
    print('\n', n, '->', OUT)


if __name__ == '__main__':
    main()
