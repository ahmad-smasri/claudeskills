import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT='projects/Assumption_Log.xlsx'
COLS=[("ID",8),("Date",12),("Category",20),("Layer",14),("Entity / Scope",34),
      ("What the source says",46),("What we did",46),("Why / basis",52),
      ("Rows affected",14),("Status",22),("Raised with client",18)]

HDR_FILL=PatternFill('solid', fgColor='1F4E78')
HDR_FONT=Font(bold=True, color='FFFFFF', size=11)
THIN=Side(style='thin', color='BFBFBF')
BORDER=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
CAT_FILL={
 'Identifier':'DDEBF7','Location':'FCE4D6','Units':'E2EFDA','Spelling':'FFF2CC',
 'Class':'E4DFEC','Scope':'F2F2F2','Source defect':'FBE5D6','Structure':'DEEAF6',
}

QNL=[
 # ---- Identifiers ----
 ("Identifier","Equipment","All 449 assets",
  "Register tags carry no building code (AHUB001, VAV_B_S11_024)",
  "Prefixed every asset with QNL_ (entity:QNL_VAV_B_S11_024)",
  "QF SSC house style writes entity:SSC_FCU0001. Prefix added, tag otherwise untouched so it stays the BMS join key.",
  "449","Accepted - client directed","Yes"),
 ("Identifier","Equipment","15 AHU units",
  "AHUB001 packs type+level in one token; other families are TYPE_LEVEL_COUNT",
  "Rewrote to entity:QNL_AHU_B_001",
  "Client direction, so all four families parse by one rule. Crosswalk carries the mapping; BMS still knows them as AHUB001.",
  "15","Accepted - client directed","Yes"),
 ("Identifier","Rooms","51 of 336 rooms",
  "Level and room number joined inconsistently (B036_REST, B-ST-01, L1023_1)",
  "Rebuilt to one shape entity:QNL_<level>_<num>_<name>",
  "285 rooms already used separate segments; the 51 were brought to the majority shape. Only the join changed - no name text, no room number.",
  "51","Accepted - client directed","Yes"),
 ("Identifier","Rooms","entity:QNL_L2_018_GROUP_STUDY_ROOM_8",
  "Room name carries a trailing space",
  "Stripped the trailing space",
  "The validator rejects it outright (E-WS-1); a trailing space cannot survive as an identifier.",
  "1","Accepted","Yes"),
 ("Identifier","Spatial","7 invented identifiers",
  "No source supplies entity:QF, entity:QNL, the 4 levels, or the CHW loop",
  "Minted them; levels match the level segment in the room tags",
  "Required for the spatial chain to reach rec:Building. entity:QF reuses the SSC site entity so the two buildings join.",
  "7","Accepted","Yes"),
 ("Identifier","Systems","entity:QNL_CHWS-MAIN-LOOP",
  "SSC writes entity:CHWS-MAIN-LOOP with no building code, yet rec:locatedIn entity:SSC",
  "Added the QNL_ prefix",
  "A bare name would load as one loop located in two buildings once both sheets enter the same graph. Departure from SSC is deliberate.",
  "3","Open - confirm if district loop","Yes"),
 # ---- Spelling ----
 ("Spelling","Rooms","24 room names",
  "Hand-typed schedule carries misspellings (CARRLES, VENTILATON, GREEM)",
  "Corrected whole tokens only, on the name segment",
  "Each settled against a sibling room that spells it correctly - never guessed. Names reach users through rdfs:label_en.",
  "24","Accepted - client directed","Yes"),
 ("Spelling","Rooms","11 room names",
  "Run-together words (REST_ROOMMEN, ABLUTIONMEN, INDIVISTUDY)",
  "Restored the separator only; wording unchanged",
  "Every sibling room writes the two tokens apart. No abbreviation was expanded.",
  "11","Accepted - client directed","Yes"),
 ("Spelling","Rooms","INDIVI_STUDY_ROOM (6 rooms), B046_ITT",
  "INDIVI probably means INDIVIDUAL; ITT sits in the room-number segment",
  "Left as-is",
  "Expanding an abbreviation is a rewrite, not a correction; ITT may be a drawing code and a join key. Client's call.",
  "7","Open - awaiting client","Yes"),
 ("Spelling","Equipment","All asset tags",
  "Register tags may contain errors",
  "Left completely untouched",
  "Asset tags are the BMS join key. Only room names were corrected.",
  "0","Accepted","Yes"),
 # ---- Units ----
 ("Units","Points","20 brick:Electric_Power_Sensor points",
  "IO list gives '%' on 20 of its 24 .kW tags, against a description reading 'Power'",
  "Overrode to unit:KiloW",
  "A power sensor cannot read a percentage. Dar Cairo uses unit:KiloW on all 559 of its Electric_Power_Sensor points.",
  "20","Accepted - source defect","Yes - IO list needs correcting"),
 ("Units","Points","22 AHU air-flow points",
  "IO list gives 'm/s' (a velocity) on Supply_/Return_Air_Flow_Sensor",
  "Overrode to unit:L-PER-SEC",
  "Dar Cairo uses L-PER-SEC on all 51 air-flow sensors, SSC on all 118; unit:M-PER-SEC occurs 0 times in either model, and no point in the IO list is named 'velocity'.",
  "22","Accepted - source defect","Yes - IO list needs correcting"),
 ("Units","Points","30 analog rows in the Selected sheet",
  "Humidity and temperature units transposed (AvgSpcHumd as degC, AvgSpcTemp as %rH)",
  "Took units from the IO list instead",
  "The IO list has both right. The Selected sheet is not used as the unit authority anywhere.",
  "30","Accepted - source defect","Yes - Selected sheet needs correcting"),
 ("Units","Points","64 brick:Relative_Humidity_Sensor points",
  "Dar Cairo uses unit:PERCENT on this class",
  "Used unit:PERCENT_RH",
  "Dar Cairo contradicts itself - its Return_ and Supply_Air_Humidity_Sensor both use PERCENT_RH. SSC and the IO list agree with PERCENT_RH.",
  "64","Accepted - deliberate departure","Yes"),
 ("Units","Points","22 brick:Speed_Sensor points",
  "SSC uses unit:RPM on this class",
  "Used unit:PERCENT",
  "SSC's are Motor_Speed_Fbk (shaft RPM); QNL's are fan VFD SpeedFbk in % of max. Dar Cairo's Speed_Command is PERCENT on all 88.",
  "22","Accepted - deliberate departure","Yes"),
 ("Units","Points","295 brick:Air_Flow_Sensor points",
  "Dar Cairo uses unit:UNITLESS on this class",
  "Used unit:L-PER-SEC",
  "A flow is not dimensionless - Dar Cairo's value is a missing unit, not a convention. Its own specific flow classes all use L-PER-SEC.",
  "295","Accepted - deliberate departure","Yes"),
 ("Units","Points","14 brick:Damper_Position_Command points",
  "Dar Cairo is mixed: UNITLESS x28, PERCENT x1",
  "Used unit:PERCENT",
  "Dar Cairo's unitless ones read as binary open/close. Its Speed_Command is PERCENT on all 88 and Damper_Position_Sensor PERCENT on all 74.",
  "14","Accepted - deliberate departure","Yes"),
 # ---- Class ----
 ("Class","Points","4 CHW temperature point signatures",
  "brick:Chilled_Water_Supply/Return_Temperature_Sensor are deprecated in Brick 1.4",
  "Kept the deprecated Dar Cairo class; recorded the replacement in the ledger note",
  "Client directed - the class is the established join key. Each raises an expected W-TYP-4 naming the entering/leaving replacement.",
  "64","Accepted - client directed","Yes"),
 ("Class","Points","22 fan trip-alarm points",
  "SSC types its own _TripAlm points as the bare brick:Alarm",
  "Coined para:Trip_Alarm (subClassOf brick:Alarm), defined in row 3",
  "brick:Alarm is reserved for a literal general/summary alarm; SSC's bare typing makes trips indistinguishable from every other alarm.",
  "22","Open - PARA team review","Yes"),
 ("Class","Points","VAV/CAV/FCU RmTemp, alarms, hours",
  "No single reference model covers every point signature",
  "Ran the ladder: Dar Cairo -> Brick 1.4 -> SSC -> new para:",
  "101 signatures resolved; 71 confirmed by an exact SSC match. Reused SSC's para:Fail_Start_Alarm, Fail_Stop_Alarm, Scheduled_Hrs_Duration, UnScheduled_Hrs_Duration rather than re-coining.",
  "2224","Accepted","No"),
 # ---- Location / rule 1 ----
 ("Location","Equipment","entity:QNL_CAV_1F_S15_001, entity:QNL_VAV_B_S13_005",
  "Named in the Selected points sheet but absent from the asset register - no room, no Fed By",
  "Modelled both units and their 6 points WITHOUT rec:locatedIn, rec:feeds or rec:isFedBy",
  "Rule 1: model the asset and its datapoints, assert nothing about position. Inventing a location would read as surveyed fact. Raises 2 E-FEED-1 and 2 W-GR-2, accepted here.",
  "10","Open - register missing 2 units","Yes"),
 ("Scope","Points","entity:QNL_VAV_1F_S15_039S",
  "In the asset register and the historian, but the Selected sheet lists no points for it",
  "Added the 3 points its 245 siblings carry, taken from the historian",
  "Rule 1: points exist in the historian, so they are modelled. The family's own selected signature decides which, so the unit matches its siblings. Its other 5 historian points are not selected on any sibling either.",
  "6","Open - confirm selection","Yes"),
 ("Source defect","Points","VAV_1F_S15_039 / 039S",
  "Earlier handover flagged the trailing S as a possible typo",
  "Overturned: modelled as two separate units",
  "The IO list gives each a full, separate point set - they are two real units, not one mistyped.",
  "0","Resolved","Yes"),
 ("Source defect","Points","15 tags in the Selected sheet",
  "RtnAirDuctPrs.PV listed twice per AHU",
  "Emitted once each",
  "A tag names one physical point. Emitting twice would produce duplicate rows (W-DUP-1) and two points sharing one timeseries id.",
  "15","Accepted - source defect","Yes - Selected sheet needs correcting"),
 ("Source defect","Points","Selected sheet column C (DP Name)",
  "Column C is misaligned against columns A/F/G from row 17 onward, and degrades into IO descriptions",
  "Used columns A/F/G only for membership; ignored column C",
  "A/F/G are internally consistent (A = Equip-Name + Point-Name). Column C is a deduplicated catalogue of point types, not a per-row value.",
  "0","Open - source needs correcting","Yes"),
 # ---- Structure / scope ----
 ("Structure","Points","All 2,224 points",
  "No historian ids were available at first modelling",
  "ref:hasTimeseriesId now filled from the IO list SourceTag on every point",
  "SSC hangs the timeseries reference on the point, never the equipment. All 2,224 carry a real id; none is a placeholder.",
  "2224","Accepted","No"),
 ("Structure","Equipment","449 assets + the CHW loop",
  "No BIM GUIDs supplied",
  "para:IFC_ID written empty on every ref:IFCReference row",
  "Deliberate placeholder, as directed. Accounts for 452 of the 454 E-PAIR-1 findings; paste the GUIDs and they clear.",
  "452","Open - awaiting BIM","Yes"),
 ("Scope","Equipment","CCU, EF, DX, HEX, KEF, SEF, TEF, GEN",
  "The Selected sheet lists 533 points for these families",
  "Not modelled",
  "Client scoped the ontology to AHU, VAV, CAV and FCU. No asset register exists for the other families.",
  "0","Accepted - client directed","Yes"),
 ("Scope","Points","Calculated points (CalcAvailability, CalcReliability, RuntimeMtr, StartsCtr, TripCtr)",
  "918 calculated point instances exist in the historian",
  "Not modelled - on hold",
  "Calc points are driven by prebuilt FDD rules and must map to real platform widgets; inventing classes would produce unusable points. Awaiting the FDD documents.",
  "0","On hold - awaiting FDD docs","Yes"),
 ("Scope","Equipment","Virtual meters",
  "SSC carries 45 per-equipment meters (electrical + cooling thermal)",
  "Not yet modelled",
  "SSC pattern chosen. QNL supports only AHU fan electrical meters: FCUs have no .kW, and no per-AHU CHW flow exists, so cooling thermal meters cannot be computed.",
  "0","Open - next task","Yes"),
 ("Scope","Spatial","Zones",
  "No zone data supplied",
  "Rooms sit directly under their level; no rec:Zone layer",
  "Satisfies the spatial-connectivity rule. Dar Cairo normally interposes a per-floor parent zone - worth adding if zoning drawings arrive.",
  "0","Open - awaiting drawings","Yes"),
 ("Scope","Equipment","Nameplate properties",
  "No manufacturer datasheets supplied",
  "No rated power, flow, capacity, model or manufacturer written",
  "House rule: nameplate properties come from datasheets only. Nothing is guessed or filled with a typical value.",
  "0","Open - awaiting datasheets","Yes"),
 ("Structure","Equipment","53 VAVs",
  "Asset level token differs from the level of the room it is tagged against",
  "Both rec:locatedIn and rec:feeds point at the same room, as written",
  "Client confirmed these are open-roof spaces running from the served level up to the level the box sits on, so the unit is genuinely in and feeding the same volume.",
  "106","Accepted - client confirmed","Yes"),
]

wb=openpyxl.Workbook()
for i,(name,rowsrc) in enumerate([("SSC",[]),("HQ",[]),("QNL",QNL),("RDC",[])]):
    ws = wb.active if i==0 else wb.create_sheet()
    ws.title=name
    ws.append([c for c,_ in COLS])
    for j,(c,w) in enumerate(COLS,1):
        cell=ws.cell(1,j); cell.fill=HDR_FILL; cell.font=HDR_FONT
        cell.alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width=w
    for n,rec in enumerate(rowsrc,1):
        cat,layer,ent,src,did,why,rows_,status,raised = rec
        ws.append([f"{name}-{n:03d}","2026-08-24",cat,layer,ent,src,did,why,rows_,status,raised])
    ws.freeze_panes='A2'
    ws.row_dimensions[1].height=30
    for r in range(2, ws.max_row+1):
        for c in range(1, len(COLS)+1):
            cell=ws.cell(r,c)
            cell.alignment=Alignment(vertical='top', wrap_text=True)
            cell.border=BORDER
        cat=ws.cell(r,3).value
        if cat in CAT_FILL: ws.cell(r,3).fill=PatternFill('solid', fgColor=CAT_FILL[cat])
        st=str(ws.cell(r,10).value or '')
        if st.startswith('Open') or st.startswith('On hold'):
            ws.cell(r,10).fill=PatternFill('solid', fgColor='FFF2CC')
        elif st.startswith('Accepted') or st.startswith('Resolved'):
            ws.cell(r,10).fill=PatternFill('solid', fgColor='E2EFDA')
        ws.row_dimensions[r].height=58
    if not rowsrc:
        ws.cell(2,1,"(no entries yet)").font=Font(italic=True, color='808080')
wb.save(OUT)
print('written', OUT)
import collections
print('QNL entries:', len(QNL))
print('by category:', dict(collections.Counter(r[0] for r in QNL)))
print('by status  :', dict(collections.Counter(r[7].split(' - ')[0] for r in QNL)))
