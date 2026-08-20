#!/usr/bin/env python3
"""The highlighter must mark findings, attach them, add no sheet, and never
modify its input."""
import csv
import pathlib
import subprocess
import sys
import tempfile

HERE = pathlib.Path(__file__).resolve().parent
SKILL = HERE.parent

try:
    import openpyxl
except ImportError:
    print("skipped (no openpyxl)")
    raise SystemExit(0)

with tempfile.TemporaryDirectory() as tmp:
    tmp = pathlib.Path(tmp)
    rows = list(csv.reader((HERE / "broken-sample.csv").open()))
    wb = openpyxl.Workbook()
    for r in rows:
        wb.active.append(r)
    src, out = tmp / "in.xlsx", tmp / "out.xlsx"
    wb.save(src)
    before = src.read_bytes()

    r = subprocess.run([sys.executable, str(SKILL / "scripts" / "highlight_findings.py"),
                        str(src), "--out", str(out), "--severity", "ERROR"],
                       capture_output=True, text=True)
    assert r.returncode == 0, r.stdout + r.stderr
    assert src.read_bytes() == before, "the highlighter modified its input"

    got = openpyxl.load_workbook(out)
    ws = got.active
    lit = [x for x in range(1, ws.max_row + 1)
           if ws.cell(x, 1).fill.start_color.rgb in ("FFFFFF00", "00FFFF00")]
    assert lit, "nothing was highlighted"
    assert any(ws.cell(x, 1).comment for x in lit), "no finding comments attached"
    assert got.sheetnames == wb.sheetnames, "the highlighter added a sheet"
print(f"({len(lit)} rows highlighted)")
