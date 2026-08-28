#!/usr/bin/env python3
"""Align QNL entity identifiers to Dar Cairo's naming convention.

Dar Cairo names entities as segments joined by '_', with words inside a segment
joined by '-', points written in dashed English, no camelCase, no dots:

    entity:HRAHU-R3_CHW-Coil                      equip _ component
    entity:FCU-9_F5_SF-Motor                      equip _ floor _ component
    entity:UPS-02_Trip-Status                     equip _ dashed-English point
    entity:Zone-C_Occupancy-Virtual-Sensor_Occupancy-Flag

QNL had segments right but wrote equipment tags with '_' between word-parts
(AHU_B_001) and points as camelCase BMS tokens (AvgSpcHumd_PV). This rewrites:

  * equipment tag        AHU_B_001         -> AHU-B-001        (one dashed segment)
  * component/part       _SF_Motor,_IsoVlv -> _SF-Motor,_Iso-Vlv
  * datapoint            _AvgSpcHumd_PV     -> _Average-Space-Humidity
                         (from the point's own rdfs:label_en - already English)

Only the subject and object identifier columns change. The BMS join keys stay
put: ref:hasTimeseriesId / para:hasEntityId keep the raw historian tag, and the
identifier crosswalk records old -> new. Rooms, levels, systems, the site and the
building keep their identifiers (they are not datapoints and several, like
entity:Electrical_System, already match Dar Cairo).

Prefer naming identifiers this way at build time (name each point
<owner>_<Dashed-English> from its label or class as you emit it). This script is
the RETROFIT for a sheet already built with raw/BMS identifiers: run it ONCE. It
is not idempotent - after a pass, points nest under their renamed owners, so a
second run reshuffles the instrumentation-point names.

Input  : an ontology .xlsx      Output: same, plus QNL_naming_crosswalk.csv
"""
import argparse
import collections
import csv
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))

# Types whose entities are equipment or components - their tags get dashed.
EQUIP_KEYWORDS = ("Air_Handling_Unit", "Variable_Air_Volume_Box",
                  "Constant_Air_Volume_Box", "Fan_Coil_Unit", "DXUnit", "CRAC",
                  "Computer_Room", "Exhaust_Fan", "Smoke_Extract_Fan",
                  "Heat_Exchanger", "Booster_Pump", "Pump", "Generator", "Fan",
                  "Coil", "Motor", "Condensing_Unit", "Valve", "Circuit_Breaker",
                  "Meter", "Tank", "Compressor", "Damper", "Filter", "VFD")
# Types to leave alone (spatial, systems, references).
KEEP_KEYWORDS = ("Room", "Level", "Building", "Site", "System", "Loop",
                 "Zone", "Storey", "Space")


def objprop(hdr, r, name):
    for i, c in enumerate(hdr):
        if c in ("object_prop_name", "subject_prop_name") and i < len(r) \
                and r[i] == name and i + 1 < len(r):
            return r[i + 1]
    return ""


def seg_format(text):
    """Any phrase, label or token -> one Dar-Cairo segment: letters/digits only,
    words and camelCase runs joined by '-', no camelCase left.
    'Average Space Humidity'->'Average-Space-Humidity' ; 'RunSts'->'Run-Sts' ;
    'Auto_Manual_Command'->'Auto-Manual-Command'."""
    t = re.sub(r"[^A-Za-z0-9]+", "-", str(text)).strip("-")
    t = re.sub(r"(?<=[a-z])(?=[A-Z])", "-", t)         # camel -> dash (not 1F -> 1-F)
    t = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", "-", t)    # ABCdef -> AB-Cdef
    return re.sub(r"-+", "-", t)


# a label is clean English when it has a lowercase letter and no camelCase run
def is_english(lbl):
    return bool(lbl) and bool(re.search(r"[a-z]", lbl)) \
        and not re.search(r"[a-z][A-Z]", lbl) and not re.search(r"[A-Za-z][0-9]", lbl)


dash_tag = seg_format          # equipment/part tags use the same formatter


def build(path, out_dir):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Ontology"] if "Ontology" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = [["" if c is None else str(c).strip() for c in r]
            for r in ws.iter_rows(values_only=True)]
    hdr, data = rows[0], rows[1:]
    data = [r + [""] * (27 - len(r)) for r in data]

    # ---- classify entities --------------------------------------------------
    etype = {}                       # id -> its subjectType (first seen)
    point_owner, point_label, point_class = {}, {}, {}
    part_parent = {}
    ts_points = set()                # ids carrying a timeseries reference
    for r in data:
        s, st, p, o, ot = r[0], r[1], r[2], r[3], r[4]
        if s.startswith("entity:") and s not in etype:
            etype[s] = st
        if o.startswith("entity:") and o not in etype:
            etype[o] = ot
        if p == "brick:hasPoint":
            point_owner[o] = s
            point_class[o] = ot
            lbl = objprop(hdr, r, "rdfs:label_en")
            if lbl:
                point_label[o] = lbl
        elif p == "brick:hasPart":
            part_parent[o] = s
        if p == "ref:hasExternalReference" and ot == "ref:TimeseriesReference":
            ts_points.add(s)

    all_ids = set(etype)

    def is_equipish(eid):
        t = etype.get(eid, "")
        if any(k in t for k in KEEP_KEYWORDS):
            return False
        return any(k in t for k in EQUIP_KEYWORDS)

    # ---- compute new ids (equipment -> parts -> points) ---------------------
    new = {}
    used_seg = collections.defaultdict(set)      # owner_new -> {segments taken}

    def rename_equip(eid):
        if eid in new:
            return new[eid]
        body = eid[len("entity:"):]
        if not body.startswith("QNL_"):
            new[eid] = eid                      # QF, HVAC, Electrical_System, ...
            return new[eid]
        tag = body[4:]                          # after QNL_
        new[eid] = "entity:QNL_" + dash_tag(tag)
        return new[eid]

    def rename_part(eid):
        if eid in new:
            return new[eid]
        parent = part_parent[eid]
        pnew = resolve(parent)
        suffix = eid[len(parent) + 1:] if eid.startswith(parent + "_") else \
            eid[len("entity:"):]
        # a part of a part extends the parent's segment with '-' (Dar Cairo's
        # FCU-9_F5_SF-Motor); a part of the equipment starts a new '_' segment.
        sep = "-" if parent in part_parent else "_"
        new[eid] = pnew + sep + dash_tag(suffix)
        return new[eid]

    def rename_point(eid):
        if eid in new:
            return new[eid]
        owner = point_owner[eid]
        onew = resolve(owner)
        suffix = eid[len(owner) + 1:] if eid.startswith(owner + "_") \
            else eid[len("entity:"):]
        owner_body = owner[len("entity:"):]
        lbl = re.sub(r"^\s*QNL[\s_-]+", "", point_label.get(eid, ""))
        cls = point_class.get(eid, "")
        if is_english(lbl):
            base = lbl                                   # AHU/VAV/CAV/FCU: real English
        elif eid.startswith(owner + "_"):
            # nested under its owner (DX/EF/HEX/ELEC): the class name is English
            base = cls.split(":", 1)[-1].replace("_", " ") if ":" in cls else suffix
        else:
            # instrumentation hung off a loop/system: use its own local token
            base = re.sub(r"^QNL[_-]", "", eid[len("entity:"):])
        seg = seg_format(base)
        if seg in used_seg[onew]:                        # keep every point distinct
            seg = seg + "-" + seg_format(suffix)
        used_seg[onew].add(seg)
        new[eid] = onew + "_" + seg
        return new[eid]

    def resolve(eid):
        if eid in new:
            return new[eid]
        if eid in point_owner:
            return rename_point(eid)
        if eid in part_parent:
            return rename_part(eid)
        return rename_equip(eid)

    for eid in all_ids:
        # standalone points (timeseries, no hasPoint parent): dash camelCase in place
        if eid in ts_points and eid not in point_owner:
            body = eid[len("entity:"):]
            segs = body.split("_")
            new[eid] = "entity:" + "_".join(dash_tag(s) for s in segs)
            continue
        if eid in point_owner:
            rename_point(eid)
        elif eid in part_parent:
            rename_part(eid)
        elif is_equipish(eid):
            rename_equip(eid)
        else:
            new[eid] = eid                      # rooms, levels, systems: keep

    # ---- collision guard: keep old id for any that would clash --------------
    forward = {}
    target = collections.Counter(v for v in new.values())
    collisions = [k for k, v in new.items() if target[v] > 1 and k != v]
    for k in collisions:
        new[k] = k                              # never merge two entities
    changed = {k: v for k, v in new.items() if v != k}

    # ---- apply --------------------------------------------------------------
    def remap(v):
        return new.get(v, v)

    ifc_bodies = {k[len("entity:"):]: v[len("entity:"):] for k, v in changed.items()}
    out = [hdr]
    for r in data:
        r = list(r)
        r[0] = remap(r[0])
        r[3] = remap(r[3])
        # ref:ifcName repeats the entity body - keep it in step (it is derivable)
        for i, c in enumerate(hdr):
            if c in ("object_prop_name", "subject_prop_name") and r[i] == "ref:ifcName" \
                    and i + 1 < len(r) and r[i + 1] in ifc_bodies:
                r[i + 1] = ifc_bodies[r[i + 1]]
        out.append(r)

    os.makedirs(out_dir, exist_ok=True)
    wbo = openpyxl.Workbook()
    wso = wbo.active
    wso.title = "Ontology"
    for r in out:
        wso.append(r)
    wbo.save(os.path.join(out_dir, "QNL_Ontology.xlsx"))
    with open(os.path.join(out_dir, "QNL_Ontology.csv"), "w", newline="") as fh:
        csv.writer(fh).writerows(out)
    with open(os.path.join(out_dir, "QNL_naming_crosswalk.csv"), "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["old_identifier", "new_identifier", "kind"])
        for k in sorted(changed):
            kind = ("point" if k in point_owner or k in ts_points else
                    "part" if k in part_parent else "equipment")
            w.writerow([k, changed[k], kind])

    print("entities total:      ", len(all_ids))
    print("renamed:             ", len(changed))
    print("  points:            ", sum(1 for k in changed if k in point_owner or k in ts_points))
    print("  parts:             ", sum(1 for k in changed if k in part_parent))
    print("  equipment:         ", sum(1 for k in changed
                                        if k not in point_owner and k not in part_parent
                                        and k not in ts_points))
    print("kept (rooms/levels/systems):", len(all_ids) - len(changed))
    print("collisions avoided:  ", len(collisions))
    return changed


if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="Align an ontology's identifiers to Dar Cairo's convention.")
    ap.add_argument("--in", dest="inp", required=True,
                    help="ontology .xlsx (a sheet named 'Ontology', or the first sheet)")
    ap.add_argument("--out", default=None,
                    help="output directory (default: alongside the input)")
    a = ap.parse_args()
    build(a.inp, a.out or os.path.dirname(os.path.abspath(a.inp)))
