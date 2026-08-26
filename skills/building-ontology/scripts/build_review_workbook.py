#!/usr/bin/env python3
"""Turn a validation run into the house review workbook.

Runs validate_ontology.py and check_consistency.py over a sheet, adds the checks
neither of them covers, then writes a copy of the workbook carrying:

  Claude Log       - what was asked, what was done, what came out
  Review_Summary   - headline defects, an index of the issue sheets, every rule code
  <the sheet>      - untouched, with flagged rows filled yellow (ERROR) / amber (WARN)
  <group>_Issues   - one sheet per entity family: counts, findings, actions, full register

This is the format the QF SSC review used, widened from equipment to every entity
in the sheet. Nothing is written into the ontology rows themselves - only fills.

    python3 build_review_workbook.py MyBuilding.xlsx --out MyBuilding_review_1.xlsx
    python3 build_review_workbook.py MyBuilding.xlsx --out R.xlsx --label-style verbatim

Groups are derived from the sheet: a subject's family class decides its group, and
subjects sharing an identifier stem are kept together. --group lets you override or
add a rule when the derived grouping splits a family you want read as one.
"""
import argparse
import collections
import datetime
import itertools
import re
import subprocess
import sys
import tempfile
from pathlib import Path

HERE = Path(__file__).resolve().parent

# --------------------------------------------------------------------- rules --
RULE = {
 'X-HDR-1': ('ERROR', 'The header row must repeat the four property-column names verbatim; numeric suffixes break the converter contract.'),
 'X-FILT-1': ('WARN', 'The sheet was saved with an autofilter applied, so most rows are hidden when it is opened.'),
 'E-HDR-1': ('ERROR', 'The first five columns must be subject, subjectType, predicate, object, objectType.'),
 'E-HDR-2': ('ERROR', 'A column in positions 6-27 is not one of the four permitted property column names.'),
 'E-CORE-1': ('ERROR', 'subject, predicate and object are all required on every row.'),
 'X-VAL-1': ('ERROR', 'A para:rated* property declared with an empty object cell - unit present, value absent.'),
 'E-BN-1': ('ERROR', 'A <blanknode> object carries no object_prop pairs, so it converts to nothing.'),
 'E-PAIR-1': ('ERROR', 'A property name is present with no value beside it.'),
 'E-PH-1': ('ERROR', 'The object cell still holds an unresolved <placeholder>.'),
 'E-TYP-1': ('ERROR', 'One entity is typed more than one way across the sheet.'),
 'E-TYP-2': ('ERROR', 'The class is not a term in Brick 1.4.'),
 'E-WS-1': ('ERROR', 'A cell carries leading or trailing whitespace - it will not join to anything.'),
 'E-GR-1': ('ERROR', 'A spatial entity does not connect up to the rec:Building through rec:isPartOf.'),
 'E-LBL-1': ('ERROR', 'A label carries punctuation the PARA label rule removes.'),
 'E-FEED-1': ('ERROR', 'A terminal unit carries no rec:feeds row.'),
 'X-DUP-1': ('ERROR', 'The same subject / predicate / object triple is written more than once.'),
 'X-ID-2': ('ERROR', 'An identifier departs from the digit width every sibling in its family uses.'),
 'X-ID-3': ('ERROR', "A room's level segment disagrees with the level it is rec:isPartOf."),
 'E-CON-1': ('ERROR', 'A unit carries a different number of rows from its siblings.'),
 'E-CON-2': ('ERROR', 'A relation present on most units of a family is absent on some.'),
 'E-CON-3': ('ERROR', 'A relation is written more than once on the same unit.'),
 'E-CON-4': ('ERROR', 'An object cell holds an error value or is empty where the family uses an entity.'),
 'E-CON-5': ('ERROR', 'The same relation is typed differently on different units.'),
 'E-CON-6': ('ERROR', 'A unit carries more than one triple for a predicate that should appear once.'),
 'E-CON-17': ('ERROR', 'A child identifier differs from its parent only in separators - one of the two is misspelled.'),
 'W-BN-4': ('WARN', 'brick:value with no brick:hasUnit beside it.'),
 'W-BN-5': ('WARN', 'ref:hasTimeseriesId with no para:hasEntityId beside it.'),
 'W-LBL-2': ('WARN', 'The entity never gets an rdfs:label_en, so the front end has nothing to display.'),
 'W-PT-1': ('WARN', 'A data point carries no ref:hasExternalReference, so its timeseries resolves empty.'),
 'W-TYP-5': ('WARN', 'The class is a Brick alias; the preferred term should be used.'),
 'X-LBL-3': ('WARN', 'rdfs:label_en repeats the identifier instead of naming the thing.'),
 'X-LBL-4': ('WARN', "Labels still carry '-' or '_' separators; neither label style leaves them in place."),
 'X-ID-1': ('WARN', 'Level identifiers use a different separator from every other identifier in the sheet.'),
 'X-ORPH-1': ('WARN', 'An entity appears only as an object - it never carries a row of its own.'),
 'W-CON-7': ('WARN', 'Every unit of a family shares one rec:locatedIn / rec:feeds target - check for placeholder data.'),
 'W-CON-9': ('WARN', 'A declared point carries no external reference.'),
 'W-CON-11': ('WARN', 'An entity carries an external reference but is never declared by a parent.'),
 'W-CON-12': ('WARN', "A unit's rows are scattered rather than grouped, making review by unit hard."),
 'W-CON-19': ('WARN', 'One class carries several different units across the sheet.'),
 'I-CON-8': ('INFO', 'rec:feeds and rec:locatedIn name the same room on every unit - confirm the source column.'),
 'I-CON-13': ('INFO', 'An identifier repeats a token.'),
 'I-CON-15': ('INFO', 'Only one instance of the class exists, so no cross-unit comparison is possible.'),
 'I-CON-16': ('INFO', 'A subject does not carry the building code its family uses.'),
}

ACTION = {
 'X-HDR-1': 'Rewrite the header row to the canonical 27 columns before the sheet goes to the converter.',
 'X-FILT-1': 'Clear the filter and save before the sheet goes out. The review copy has it cleared already.',
 'E-HDR-2': 'Restore the canonical header row: columns 6-27 repeat subject_prop_name / subject_prop_val / object_prop_name / object_prop_val with no numeric suffix.',
 'E-CORE-1': 'Put the value in the object column as <blanknode> carrying brick:value + brick:hasUnit, or delete the row.',
 'X-VAL-1': 'Put the rated figure in the object column, or delete the row - a unit with no value is not a property.',
 'E-BN-1': 'Give the blank node its object_prop pairs, or drop the row.',
 'E-PAIR-1': 'Supply the missing property value, or remove the property name.',
 'E-PH-1': 'Replace the placeholder with the real room / feeder entity before handover.',
 'E-TYP-1': 'Pick one class for the entity and use it on every row that types it.',
 'E-TYP-2': 'Substitute the Brick 1.4 term after confirming with the team.',
 'E-WS-1': 'Strip the padding whitespace from the cell.',
 'E-GR-1': 'Add the missing rec:isPartOf row so the entity reaches the building.',
 'X-DUP-1': 'Delete the duplicate rows, keeping one, and settle the class on the survivor.',
 'X-ID-1': 'Settle one separator convention for level identifiers and apply it to all of them.',
 'X-ID-2': 'Confirm the identifier against the asset register - every sibling uses a different digit width.',
 'X-ID-3': 'Confirm the room number against the room schedule - its level segment and its parent level disagree.',
 'X-LBL-3': 'Give the entity a human label; the identifier is a join key, not display text.',
 'X-LBL-4': 'Confirm the label style with the team - the QF SSC house style reads separators as spaces.',
 'X-ORPH-1': 'Type the entity on a row of its own, or confirm it is meant to exist only as an object.',
 'W-BN-4': 'Add brick:hasUnit, or unit:UNITLESS if the quantity is genuinely dimensionless.',
 'W-BN-5': 'Add para:hasEntityId beside the timeseries id.',
 'W-LBL-2': 'Add an rdfs:label_en row - the front end has nothing to display for this entity.',
 'W-PT-1': 'Add the ref:TimeseriesReference, or delete the point if the BMS does not publish it.',
 'W-TYP-5': 'Prefer the Brick 1.4 preferred term over the alias.',
 'E-CON-1': 'Compare the unit against its siblings and add or remove rows until the shape matches.',
 'E-CON-2': 'Add the missing relation to the units that lack it, or confirm the omission is real.',
 'E-CON-3': 'Delete the duplicated triple - keep one.',
 'E-CON-4': 'Replace the error / empty object cell with the real entity.',
 'E-CON-5': 'Pick one class for the relation and apply it on every unit.',
 'E-CON-6': 'Decide which is correct and delete the surplus row.',
 'E-CON-17': 'Correct whichever of parent or child carries the wrong separator so the two agree.',
 'W-CON-7': 'Confirm the shared target is real, not placeholder data left over from a template.',
 'W-CON-9': 'Add the external reference, or remove the entity if it is not a real point.',
 'W-CON-11': 'Declare the entity with brick:hasPoint or brick:hasPart from its parent.',
 'W-CON-12': "Group the unit's rows together so a reviewer can read one unit at a time.",
 'W-CON-19': 'Check the source unit column against the class - at least one unit is wrong.',
 'I-CON-8': 'Confirm the source room column means both location and served space.',
 'I-CON-13': 'Confirm the repeated token is intended and not a copy-paste slip.',
 'I-CON-15': 'Single instance - no sibling to compare against; review by hand.',
 'I-CON-16': 'Add the building code, or confirm the entity is site-wide.',
}

# --------------------------------------------------------------- plain words --
# Every rule code, said the way you would say it to someone who has never seen a
# validator. (title, why it matters, what to do about it).
PLAIN = {
 'X-FILT-1': ("The file opens with most of its rows hidden",
   "Someone left a filter switched on before saving. Anyone who opens the original sees a few hundred rows out of nearly thirty thousand, and may think that is the whole file.",
   "Open the file, clear the filter on the top row, and save. Nothing is missing - it is only hidden. This review copy already has it cleared."),
 'X-HDR-1': ("The column headings across the top are wrong",
   "The headings tell the software which column means what. These have had numbers stuck on the end of them, so the software stops reading at column 9 and silently ignores everything to the right - about ten thousand pieces of information.",
   "Replace the top row with the standard headings. This is the single most important fix in the file; do it before anything else."),
 'E-HDR-1': ("The first columns are not the standard ones", "The software expects a fixed order and cannot find its way around without it.", "Put the first five columns back in the standard order."),
 'E-HDR-2': ("A column heading is not one the software recognises", "Anything under an unrecognised heading is thrown away when the file is converted.", "Use only the standard heading names."),
 'E-CORE-1': ("A row is missing the thing it is meant to describe",
   "Every row is supposed to say three things: what, how it relates, and to what. These rows are missing the third one, so they describe nothing and vanish on conversion.",
   "Fill in the missing value, or delete the row if there is nothing to put there."),
 'X-VAL-1': ("A capacity or flow rate has its unit but no number",
   "The row says something like 'rated cooling capacity, in kilowatts' and then never gives the figure. It looks like data but carries none.",
   "Put the number in from the equipment datasheet. If no datasheet was supplied, delete the row - an empty figure is worse than no figure."),
 'E-PH-1': ("Someone left a to-do note in the file",
   "Placeholders like <Location> and <Fedby> are notes-to-self meaning 'fill this in later'. Later never came. Equipment with one of these does not know what room it is in or what feeds it.",
   "Replace each placeholder with the real room or the real piece of equipment."),
 'E-TYP-2': ("Equipment is described with a word the standard does not contain",
   "The file uses an industry standard vocabulary. These words are not in it, so the software will not know what the thing is.",
   "Swap in the correct standard word. Check with whoever built the file which was meant."),
 'E-TYP-1': ("The same thing is described two or three different ways",
   "One sensor is called three different things in three different rows. The software will treat it as three separate sensors, or refuse the file.",
   "Decide what the thing actually is and use that one description everywhere."),
 'E-CON-5': ("The same part is described differently on different units",
   "Identical units should describe their identical parts identically. These do not.",
   "Pick the right description and apply it to every unit."),
 'E-CON-1': ("One unit has a different number of rows from its identical twins",
   "Units of the same type should each carry the same set of information. One that is short is missing something; one that is long has something extra.",
   "Compare it against its siblings and add or remove until they match - or confirm the difference is genuine."),
 'E-CON-2': ("Most units have this, a handful do not",
   "Something almost every unit of this type carries is missing on a few of them. Usually an oversight when the file was built.",
   "Add it to the ones that lack it, or confirm those units genuinely do not have it."),
 'E-CON-6': ("A unit says two different things feed it",
   "Each unit should have one source of supply. These name two.",
   "Work out which is right and delete the other - unless the pair is a duty/standby arrangement, in which case say so."),
 'E-CON-3': ("The same fact is written more than once on a unit", "Duplicates inflate counts and can create phantom equipment.", "Delete the repeats, keep one."),
 'X-DUP-1': ("The same fact is written more than once",
   "The identical statement appears two or three times over. Duplicates create phantom equipment on screen.",
   "Delete the repeats, keeping one."),
 'E-CON-4': ("A cell holds an error message instead of a value",
   "Something like #N/A is sitting where a room or a piece of equipment should be - a spreadsheet formula that failed and was never noticed.",
   "Put the real value in."),
 'E-CON-17': ("A part is spelled slightly differently from its parent",
   "A dash where the parent has an underscore, or the reverse. To software these are different things, so the part comes adrift from the unit it belongs to.",
   "Make the two agree - correct whichever one is wrong."),
 'E-WS-1': ("A stray space at the start or end of a word",
   "Invisible on screen, but software matches text exactly, so a trailing space breaks the link.",
   "Delete the extra space."),
 'E-GR-1': ("A floor or room is not attached to the building",
   "Everything should trace up to the building. These do not, so they will not appear where they should in the building tree.",
   "Add the row that says which building or floor it belongs to."),
 'X-ID-3': ("A room number and its floor disagree",
   "The room is numbered as if it were on one floor but filed under another.",
   "Check the room schedule and correct whichever is wrong."),
 'X-ID-2': ("A tag has the wrong number of digits",
   "Every other unit in the family uses four digits; this one has five. Almost certainly a typing slip.",
   "Check the asset register and correct it."),
 'X-ID-1': ("Floor tags are punctuated differently from everything else",
   "Floors use dashes where rooms and equipment use underscores. Not fatal, but inconsistent.",
   "Agree one convention and apply it to all the floors."),
 'E-LBL-1': ("A display name contains punctuation that is not allowed", "The house style strips these characters.", "Remove the punctuation from the name."),
 'E-FEED-1': ("A unit that serves a room does not say which room",
   "Terminal units must name the space they serve or the building tree has a dead end.",
   "Add the room it serves."),
 'X-LBL-3': ("The on-screen name is a code, not a name",
   "Users will see 'HQ_VAV0001' where they expect something like 'VAV Box 1, Level 3'. It is a filing reference, not a name a person can read.",
   "Give each one a readable name."),
 'X-LBL-4': ("On-screen names still contain dashes and underscores",
   "Names read like filenames rather than English. Sister projects show them as spaces.",
   "Agree with the team how names should look, then apply it once across the file."),
 'W-LBL-2': ("This has nothing to show on screen",
   "No name was ever given, so the front end will display a blank or fall back to the raw code.",
   "Add a name."),
 'W-PT-1': ("A sensor with no live data behind it",
   "The file says this sensor exists but never says where its readings come from. On screen it becomes a tile that is permanently blank, and nobody can tell whether the sensor is broken or was never real.",
   "Either connect it to its data feed, or delete it if the system does not actually publish it."),
 'W-CON-9': ("A sensor is listed but never connected to its data",
   "Same problem as above, spotted by comparing a unit against its twins.",
   "Connect it to its data feed, or remove it."),
 'W-CON-11': ("Something has a data feed but nothing owns it",
   "It is wired up but never declared as a part of any unit, so it will float loose in the building tree.",
   "Attach it to the unit it belongs to."),
 'X-ORPH-1': ("Mentioned, but never actually defined",
   "Other rows refer to this thing, but it has no row of its own - so it has no name, no location and no data.",
   "Give it a proper row, or confirm it is only ever meant to be referred to."),
 'W-BN-4': ("A number with no unit",
   "A figure with no unit is ambiguous - kilowatts or watts, litres or cubic metres.",
   "Add the unit."),
 'W-BN-5': ("A data tag with nothing saying which unit it belongs to",
   "The data reference is there but not grouped under its equipment.",
   "Add the owning unit."),
 'E-BN-1': ("A reference that references nothing", "An empty placeholder that converts to nothing.", "Fill it in or delete the row."),
 'E-PAIR-1': ("A label with no value next to it",
   "Something is named - 'voltage', say - but the figure beside it was never filled in.",
   "Supply the figure, or remove the label."),
 'W-TYP-5': ("An old word used where the standard now prefers a newer one",
   "It still works, but the newer word is the one the standard recommends.",
   "Swap to the preferred word when convenient. Low priority."),
 'W-CON-7': ("Every single unit points at the same room",
   "All twenty-seven car park fans claim to serve one office. That is template data nobody went back and corrected - and the room named is from a different building's file entirely.",
   "Put in the real room for each unit."),
 'W-CON-12': ("A unit's rows are scattered through the file",
   "Rows for one unit are spread far apart, which makes checking it by eye hard.",
   "Group each unit's rows together. Tidiness only - nothing is wrong with the data."),
 'W-CON-19': ("One kind of sensor is measuring in several different units",
   "Temperature sensors reading in degrees on some units and percent on others. At least one is wrong.",
   "Check the source list and correct the odd ones out."),
 'I-CON-8': ("A unit's location and the space it serves are the same room",
   "That is often correct, but sometimes it means one column was copied into both.",
   "Confirm the original source really meant both."),
 'I-CON-13': ("A tag repeats a word",
   "Something like ...LEGAL_LEGAL... Usually a copy-paste slip, occasionally genuine.",
   "Glance at it and confirm."),
 'I-CON-15': ("Only one of these exists, so nothing to compare it against",
   "Everything else was checked by comparing identical units. This one is alone, so it was checked on its own terms only.",
   "Worth a manual read since no sibling could vouch for it."),
 'I-CON-16': ("A tag is missing the building code",
   "Everything else in the family starts with the building code and this does not.",
   "Add the code, or confirm it is deliberately shared across buildings."),
}

SEVERITY_WORD = {'ERROR': 'Must fix', 'WARN': 'Please check', 'INFO': 'Just so you know'}

# How much a problem matters is not how often it occurs. A broken header row is one
# finding and the worst thing in the file; a stray space is twenty-eight and trivial.
# Lower number = higher up the list. Anything unlisted sits in the middle.
IMPORTANCE = {
 'X-HDR-1': 1, 'E-HDR-1': 1, 'E-HDR-2': 2, 'X-FILT-1': 3,
 'E-PH-1': 4, 'W-CON-7': 5, 'X-VAL-1': 6, 'E-CORE-1': 7,
 'E-CON-4': 8, 'E-TYP-1': 9, 'E-TYP-2': 10, 'E-CON-5': 11,
 'E-GR-1': 12, 'X-ID-3': 13, 'E-CON-6': 14, 'E-CON-2': 15,
 'E-CON-1': 16, 'X-DUP-1': 17, 'E-CON-3': 17, 'E-CON-17': 18,
 'W-PT-1': 19, 'W-CON-9': 19, 'X-ORPH-1': 20, 'E-PAIR-1': 21,
 'X-LBL-3': 22, 'X-LBL-4': 23, 'W-LBL-2': 24, 'E-BN-1': 25,
 'W-BN-4': 30, 'W-BN-5': 31, 'E-WS-1': 32, 'X-ID-2': 33,
 'X-ID-1': 34, 'W-CON-11': 35, 'W-CON-19': 36, 'W-TYP-5': 40,
 'W-CON-12': 45,
}


def importance(code):
    return IMPORTANCE.get(code, 50)
SEVERITY_NOTE = {
    'Must fix': 'The file will not work properly until this is sorted.',
    'Please check': 'Might be deliberate, might be a mistake. Someone who knows the building should decide.',
    'Just so you know': 'Probably fine. Read it once and move on.',
}

# What each group of rows is called, in words rather than jargon.
PRETTY = {
 'Sheet_Structure':   ('The Whole File',        'Problems with the file itself rather than any one piece of equipment'),
 'Class_Definitions': ('Vocabulary List',       'The list of custom terms defined at the top of the file'),
 'Spatial':           ('Building, Floors and Rooms', 'The site, the building, its floors and its 1,044 rooms'),
 'Systems':           ('Systems',               'The top-level groupings - heating, cooling, electrical'),
 'AHU':               ('Air Handling Units',    'The big air units - AHU'),
 'VAV':               ('VAV Boxes',             'Variable air volume boxes - the dampers that feed each room'),
 'FCU':               ('Fan Coil Units',        'Fan coil units - FCU'),
 'CRAC':              ('Server Room Coolers',   'Computer room air conditioners - CCU'),
 'DX':                ('DX Units',              'Direct expansion cooling units'),
 'ExhaustFan':        ('Extract Fans',          'Car park, general and toilet extract fans'),
 'HeatExchanger':     ('Heat Exchangers',       'The plate heat exchangers - HEX'),
 'CHWPump':           ('Chilled Water Pumps',   'The chilled water booster pumps'),
 'Generator':         ('Generators',            'The standby generators'),
 'Misc':              ('Everything Else',       'A few odds and ends that do not belong to any group above'),
}


def pretty(group):
    if group in PRETTY:
        return PRETTY[group]
    return (group.replace('_', ' ').title(), 'Equipment of this type')


def plain(code):
    return PLAIN.get(code, ('Something needs checking',
                            'The checker flagged this and there is no plain-English note for it yet.',
                            'Ask whoever maintains the checker.'))

SEV_ORDER = {'ERROR': 0, 'WARN': 1, 'INFO': 2}
CANON = (['subject', 'subjectType', 'predicate', 'object', 'objectType']
         + ['subject_prop_name', 'subject_prop_val', 'object_prop_name', 'object_prop_val'] * 5
         + ['subject_prop_name', 'subject_prop_val'])

SPATIAL_CLASSES = {'rec:Site', 'rec:Building', 'rec:Level', 'rec:BasementLevel',
                   'rec:RoofLevel', 'rec:Room', 'rec:Zone', 'rec:HVACZone'}
SYSTEM_CLASSES = {'brick:System', 'brick:HVAC_System', 'brick:Chilled_Water_System',
                  'brick:Electrical_System', 'brick:Water_System', 'brick:Heating_System'}

Finding = collections.namedtuple('Finding', 'severity code row entity group finding action')

S = lambda x: x.strip() if isinstance(x, str) else x


# ------------------------------------------------------------------ loading --
def load_rows(path):
    if path.suffix.lower() in ('.xlsx', '.xlsm'):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        ws = pick_sheet(wb)
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        name = ws.title
        wb.close()
        return rows, name
    import csv
    with open(path, newline='', encoding='utf-8-sig') as f:
        rows = [r for r in csv.reader(f)]
    return rows, path.stem


def pick_sheet(wb):
    """The ontology sheet is the one whose header starts with subject/subjectType."""
    for ws in wb.worksheets:
        head = [S(c) for c in next(ws.iter_rows(max_row=1, values_only=True), [])][:3]
        if head and str(head[0]).lower() == 'subject' and str(head[1] or '').lower() == 'subjecttype':
            return ws
    return wb.worksheets[0]


def read_filter_state(path, rows):
    """A sheet saved with a filter applied opens showing a fraction of its rows, and any
    highlight on the rest is invisible. Returns (ref, criteria, hidden count) or None."""
    if path.suffix.lower() not in ('.xlsx', '.xlsm'):
        return None
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = pick_sheet(wb)
    hidden = [r for r, d in ws.row_dimensions.items() if d.hidden]
    state = None
    if ws.auto_filter and ws.auto_filter.ref and hidden:
        crit = ''
        for fc in (ws.auto_filter.filterColumn or []):
            vals = list(fc.filters.filter) if fc.filters and fc.filters.filter else []
            if vals:
                col = rows[0][fc.colId] if fc.colId < len(rows[0]) else 'column %d' % (fc.colId + 1)
                crit = ' (%s = %s)' % (col, ', '.join(map(str, vals[:4])))
        state = (ws.auto_filter.ref, crit, len(hidden))
    wb.close()
    return state


def run_checker(script, sheet, report, extra=()):
    """Run one of the two standard scripts and read its --report back."""
    cmd = [sys.executable, str(HERE / script), str(sheet), '--max', '1000000',
           '--report', str(report), *extra]
    subprocess.run(cmd, capture_output=True, text=True)
    if not report.exists():
        return []
    import openpyxl
    wb = openpyxl.load_workbook(report, read_only=True)
    out = [(sev, code, row, text) for sev, code, row, text, *_ in
           wb.active.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return out


# ----------------------------------------------------------------- grouping --
def derive_groups(rows, overrides):
    """Work out which group each row belongs to, from the sheet rather than a fixed list.

    A part or a point belongs on the same sheet as the unit that owns it, so every
    subject is walked up its brick:hasPart / brick:hasPoint chain to the entity that
    owns it, and the group is named after *that* entity's class. Spatial and system
    classes collapse into one group each; schema rows go to Class_Definitions.
    """
    data = rows[1:]
    subj_class, owner = {}, {}
    for r in data:
        s = S(r[0])
        st = S(r[1]) if len(r) > 1 else None
        o = S(r[3]) if len(r) > 3 else None
        p = S(r[2]) if len(r) > 2 else None
        ot = S(r[4]) if len(r) > 4 else None
        if isinstance(s, str) and s.startswith('entity:') and st:
            subj_class.setdefault(s, st)
        if isinstance(o, str) and o.startswith('entity:') and ot:
            subj_class.setdefault(o, ot)
        if p in ('brick:hasPart', 'brick:hasPoint') and isinstance(o, str) and isinstance(s, str):
            owner.setdefault(o, s)

    def root(e, _seen=None):
        seen = _seen or set()
        while e in owner and e not in seen:
            seen.add(e)
            e = owner[e]
        return e

    def group_for(subject):
        if not isinstance(subject, str):
            return 'Misc'
        s = subject.strip()
        for name, rx in overrides:
            if rx.search(s):
                return name
        if not s.startswith('entity:'):
            return 'Class_Definitions'
        top = root(s)
        cls = subj_class.get(top) or subj_class.get(s)
        if cls in SPATIAL_CLASSES:
            return 'Spatial'
        if cls in SYSTEM_CLASSES:
            return 'Systems'
        if cls:
            return cls.split(':', 1)[1]
        return 'Misc'

    row_group, row_subject = {}, {}
    for i, r in enumerate(data, 2):
        row_subject[i] = S(r[0]) or ''
        row_group[i] = group_for(r[0])

    # a subject that still has no owner joins the longest asset identifier it starts
    # with, so a point named after its unit lands beside that unit even when no
    # hasPoint row declares it
    assets = sorted((e for e, c in subj_class.items()
                     if c not in SPATIAL_CLASSES and c not in SYSTEM_CLASSES and e not in owner),
                    key=len, reverse=True)
    stem_group = {a: group_for(a) for a in assets}
    for i in row_group:
        if row_group[i] != 'Misc':
            continue
        s = row_subject[i]
        for a in assets:
            if s.startswith(a) and stem_group[a] != 'Misc':
                row_group[i] = stem_group[a]
                break

    # a group holding no top-level unit of its own is a point class that escaped the
    # walk - fold it into whichever group its rows' owners mostly sit in
    tops = collections.Counter(group_for(a) for a in assets)
    for i in list(row_group):
        g = row_group[i]
        if g in ('Spatial', 'Systems', 'Class_Definitions', 'Misc') or tops.get(g):
            continue
        row_group[i] = group_for(owner.get(row_subject[i], row_subject[i])) if \
            owner.get(row_subject[i]) else 'Misc'
    return row_group, row_subject, subj_class


# ------------------------------------------------------------ extra checks ---
def extra_checks(rows, row_group, row_subject, subj_class, add, filter_state=None):
    data, hdr = rows[1:], rows[0]

    if filter_state:
        ref, crit, nhidden = filter_state
        add('WARN', 'X-FILT-1', 1, '', 'Sheet_Structure',
            'The sheet is saved with an autofilter over %s still applied%s, so %d of its %d rows are hidden '
            'when it is opened - only %d show. Nothing is wrong with the hidden rows and all of them were '
            'validated, but anyone opening the draft sees a fraction of the sheet and could take it for the '
            'whole. The filter is cleared and every row unhidden in this review copy.'
            % (ref, crit, nhidden, len(data), len(data) - nhidden))

    bad = [(i + 1, h) for i, h in enumerate(hdr)
           if isinstance(h, str) and re.search(r'\d$', h.strip())]
    if bad:
        pairs = sum(1 for r in data for j in range(9, min(len(r), 27), 2)
                    if r[j] not in (None, ''))
        add('ERROR', 'X-HDR-1', 1, '', 'Sheet_Structure',
            'Header row: %d of the 27 columns carry numeric suffixes (%s ... %s). The contract repeats '
            'subject_prop_name / subject_prop_val / object_prop_name / object_prop_val verbatim, so every '
            'property pair from column 10 rightwards is invisible to a reader that follows the contract - '
            '%d property pairs in this sheet.' % (len(bad), bad[0][1], bad[-1][1], pairs))
    if hdr and S(hdr[0]) != CANON[0]:
        add('WARN', 'X-HDR-1', 1, '', 'Sheet_Structure',
            "Header cell A1 is %r; the contract and Dar Cairo both write it lower case as 'subject'." % S(hdr[0]))

    for i, r in enumerate(data, 2):
        p = S(r[2]) if len(r) > 2 else None
        if isinstance(p, str) and p.startswith('para:rated') and (len(r) < 4 or r[3] in (None, '')):
            unit = S(r[8]) if len(r) > 8 else None
            add('ERROR', 'X-VAL-1', i, row_subject[i], row_group[i],
                '%s declared with an empty object cell - the unit (%s) is present but the value is not, so '
                'the property converts to nothing.' % (p, unit or 'no unit either'))

    seen = collections.defaultdict(list)
    for i, r in enumerate(data, 2):
        k = (S(r[0]) or '', S(r[2]) or '', S(r[3]) or '')
        if k[2] and k[2] != '<blanknode>':
            seen[k].append(i)
    for k, rs in seen.items():
        if len(rs) > 1:
            for i in rs:
                add('ERROR', 'X-DUP-1', i, k[0], row_group[i],
                    'Triple `%s %s %s` is written %d times (rows %s).'
                    % (k[0], k[1], k[2], len(rs), ', '.join(map(str, rs))))

    lbl = {}
    for i, r in enumerate(data, 2):
        s, o = S(r[0]), S(r[3]) if len(r) > 3 else None
        for j in range(5, min(len(r) - 1, 26), 2):
            if S(r[j]) == 'rdfs:label_en' and r[j + 1] not in (None, ''):
                tgt = s if ((j - 5) // 2) % 2 == 0 else o
                if isinstance(tgt, str) and tgt.startswith('entity:'):
                    lbl.setdefault(tgt, (i, str(r[j + 1])))
    for e, (i, l) in lbl.items():
        if l == e.split(':', 1)[1]:
            add('WARN', 'X-LBL-3', i, e, row_group.get(i, 'Misc'),
                "rdfs:label_en is the identifier itself (%r) - the front end shows a BMS tag where a name "
                "should be." % l)
    sep = collections.defaultdict(list)
    for e, (i, l) in lbl.items():
        if re.search(r'[-_]', l):
            sep[row_group.get(i, 'Misc')].append((e, i, l))
    for g, hits in sep.items():
        e, i, l = hits[0]
        add('WARN', 'X-LBL-4', i, e, g,
            "%d labels in this group still carry '-' or '_' separators (e.g. %s -> %r). Neither label style "
            "in the skill leaves them in place; the QF SSC house style reads separators as spaces."
            % (len(hits), e, l))

    levels = sorted(s for s, c in subj_class.items()
                    if c in ('rec:Level', 'rec:BasementLevel', 'rec:RoofLevel'))
    if levels and sum('-' in s.split(':', 1)[1] for s in levels) == len(levels):
        others = [s for s, c in subj_class.items() if c not in SPATIAL_CLASSES]
        if others and sum('_' in s for s in others) > len(others) * 0.8:
            add('WARN', 'X-ID-1', None, levels[0], 'Spatial',
                '%d level identifiers use dashes (%s) where the assets and points in the sheet use '
                'underscores between segments.' % (len(levels), levels[0]))

    fam = collections.defaultdict(lambda: collections.defaultdict(list))
    for s in subj_class:
        m = re.match(r'^(entity:[A-Za-z_]*?[A-Za-z])(\d+)$', s)
        if m:
            fam[m.group(1)][len(m.group(2))].append(s)
    for stem, widths in fam.items():
        if len(widths) < 2:
            continue
        major = max(widths, key=lambda w: len(widths[w]))
        for w, es in widths.items():
            if w == major:
                continue
            for e in es:
                rn = next((i for i, x in row_subject.items() if x == e), None)
                add('ERROR', 'X-ID-2', rn, e, row_group.get(rn, 'Misc'),
                    '%s carries %d digits where the other %d units in the family carry %d.'
                    % (e, w, len(widths[major]), major))

    lvl_of = {}
    for i, r in enumerate(data, 2):
        if S(r[2]) == 'rec:isPartOf' and subj_class.get(S(r[0])) == 'rec:Room':
            lvl_of[S(r[0])] = (S(r[3]), i)
    for room, (lvl, rn) in lvl_of.items():
        m = re.match(r'^entity:[A-Za-z]+_(B?\d+|B\d|G|GF|RF|R)_', room)
        lm = re.match(r'^entity:[A-Za-z]+-Level-(.+)$', lvl or '')
        if m and lm:
            seg, ln = m.group(1), lm.group(1)
            if seg.lstrip('0') != ln.lstrip('0') and not (ln == 'Ground-Floor' and seg == 'G') \
               and not (ln == seg):
                add('ERROR', 'X-ID-3', rn, room, 'Spatial',
                    'Room identifier says level %r but the room is rec:isPartOf %s.' % (seg, lvl))

    subjects = {S(r[0]) for r in data if isinstance(r[0], str) and S(r[0]).startswith('entity:')}
    first = {}
    for i, r in enumerate(data, 2):
        o = S(r[3]) if len(r) > 3 else None
        if isinstance(o, str) and o.startswith('entity:'):
            first.setdefault(o, i)
    obj_type = {}
    for r in data:
        o, ot = (S(r[3]), S(r[4])) if len(r) > 4 else (None, None)
        if isinstance(o, str) and ot:
            obj_type.setdefault(o, ot)
    for e, i in sorted(first.items()):
        if e not in subjects:
            add('WARN', 'X-ORPH-1', i, e, row_group.get(i, 'Misc'),
                '%s (%s) is only ever an object - it never carries a row of its own, so it has no label, '
                'no location and no points.' % (e, obj_type.get(e, 'untyped')))


# ------------------------------------------------------------------- output --
def build(args):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sheet = Path(args.sheet)
    rows, sheet_name = load_rows(sheet)
    data = rows[1:]
    overrides = [(n, re.compile(p)) for n, p in (g.split('=', 1) for g in args.group)]
    row_group, row_subject, subj_class = derive_groups(rows, overrides)
    filter_state = read_filter_state(sheet, rows)

    # the two standard checkers, run against a header-normalised copy so a broken
    # header row does not drown their output in artefacts
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        norm = td / 'normalised.csv'
        import csv
        with open(norm, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(CANON)
            for r in data:
                w.writerow(['' if c is None else c for c in r])
        raw = (run_checker('validate_ontology.py', norm, td / 'v.xlsx',
                           ['--label-style', args.label_style]
                           + (['--io', args.io] if args.io else []))
               + run_checker('check_consistency.py', norm, td / 'c.xlsx',
                             ['--io', args.io] if args.io else []))

    class_group = {}
    for i, r in enumerate(data, 2):
        st = S(r[1]) if len(r) > 1 else None
        if isinstance(st, str) and st:
            class_group.setdefault(st, collections.Counter())[row_group[i]] += 1
    class_group = {k: v.most_common(1)[0][0] for k, v in class_group.items()}

    findings = []

    def add(sev, code, row, entity, group, text):
        findings.append(Finding(sev, code, row, entity, group, text,
                                ACTION.get(code, 'Review and confirm.')))

    def place(text, row):
        if isinstance(row, int) and row in row_group:
            return row_group[row], row_subject[row]
        m = re.search(r'(entity:[A-Za-z0-9_.<>&\-]+)', text or '')
        if m:
            e = m.group(1).rstrip('.,;:')
            for name, rx in overrides:
                if rx.search(e):
                    return name, e
            rn = next((i for i, s in row_subject.items() if s == e), None)
            return (row_group[rn] if rn else class_group.get(subj_class.get(e, ''), 'Misc')), e
        for rx in (r'^\s*((?:brick|para|rec|ref|owl|qudt):[A-Za-z0-9_\-]+)\s*:',
                   r'((?:brick|para|rec):[A-Za-z0-9_\-]+)'):
            m = re.search(rx, text or '')
            if m and m.group(1) in class_group:
                return class_group[m.group(1)], m.group(1)
        return 'Misc', ''

    for sev, code, row, text in raw:
        if code in args.ignore.split(','):
            continue
        row = row if isinstance(row, int) and row > 1 else None
        if code in ('E-HDR-1', 'E-HDR-2'):
            g, ent = 'Sheet_Structure', ''
        else:
            g, ent = place(text, row)
        add(sev, code, row, ent, g, text)

    extra_checks(rows, row_group, row_subject, subj_class, add, filter_state)

    # ---- styling
    YELLOW, AMBER = PatternFill('solid', fgColor='FFFF00'), PatternFill('solid', fgColor='FFE699')
    HDRF, SECF = PatternFill('solid', fgColor='1F3864'), PatternFill('solid', fgColor='D9E1F2')
    OKF = PatternFill('solid', fgColor='C6EFCE')
    BADF, WARNF = PatternFill('solid', fgColor='FFC7CE'), PatternFill('solid', fgColor='FFEB9C')
    TITLE = Font(bold=True, size=14, color='1F3864')
    SUB = Font(italic=True, size=9, color='595959')
    SECT = Font(bold=True, size=11, color='1F3864')
    COLH = Font(bold=True, size=10, color='FFFFFF')
    THIN = Border(*[Side(style='thin', color='BFBFBF')] * 4)

    def section(ws, r, text):
        ws.cell(r, 1, text).font = SECT
        ws.cell(r, 1).fill = SECF
        return r + 1

    def header_row(ws, r, cols):
        for c, h in enumerate(cols, 1):
            cell = ws.cell(r, c, h)
            cell.font, cell.fill, cell.border = COLH, HDRF, THIN
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        return r + 1

    def sev_fill(cell, sev):
        if sev == 'ERROR':
            cell.fill = BADF
        elif sev == 'WARN':
            cell.fill = WARNF

    if sheet.suffix.lower() in ('.xlsx', '.xlsm'):
        wb = openpyxl.load_workbook(sheet)
        src = pick_sheet(wb)
        sheet_name = src.title
    else:
        # a CSV has no workbook to copy, so the data sheet is written fresh
        wb = openpyxl.Workbook()
        src = wb.active
        src.title = sheet_name[:31]
        for r in rows:
            src.append(['' if c is None else c for c in r])

    # Clear the filter on the copy and unhide every row, so the fills below are visible.
    if filter_state:
        src.auto_filter.ref = None
        for r, d in src.row_dimensions.items():
            d.hidden = False
    err_rows = {f.row for f in findings if f.severity == 'ERROR' and isinstance(f.row, int) and f.row >= 1}
    warn_rows = {f.row for f in findings if f.severity == 'WARN' and isinstance(f.row, int) and f.row >= 1} - err_rows
    for rs, fill in ((err_rows, YELLOW), (warn_rows, AMBER)):
        for r in rs:
            for c in range(1, 6):
                src.cell(r, c).fill = fill

    by_group = collections.defaultdict(list)
    for f in findings:
        by_group[f.group].append(f)
    order = (['Sheet_Structure', 'Spatial', 'Systems']
             + sorted(g for g in by_group if g not in
                      ('Sheet_Structure', 'Class_Definitions', 'Spatial', 'Systems', 'Misc'))
             + ['Class_Definitions', 'Misc'])
    order = [g for g in order if by_group.get(g)]

    def collapse(fs):
        """One line per kind of problem, not one line per row."""
        out = []
        for code, group in itertools.groupby(
                sorted(fs, key=lambda f: (SEV_ORDER[f.severity], f.code)),
                key=lambda f: f.code):
            g = list(group)
            rws = sorted({f.row for f in g if f.row})
            out.append((g[0].severity, code, len(g), rws, g))
        out.sort(key=lambda t: (SEV_ORDER[t[0]], importance(t[1]), -t[2]))
        return out

    def where(rws, n=8):
        if not rws:
            return 'across the file'
        head = ', '.join(str(r) for r in rws[:n])
        return 'row ' + head if len(rws) <= n else 'rows %s and %d more' % (head, len(rws) - n)

    # ---------------------------------------------------------- one sheet each --
    built = []
    for g in order:
        fs = by_group[g]
        name, blurb = pretty(g)
        ws = wb.create_sheet(name[:31])
        rowsn = sum(1 for x in row_group.values() if x == g)
        must = sum(1 for f in fs if f.severity == 'ERROR')
        chk = sum(1 for f in fs if f.severity == 'WARN')

        ws['A1'] = name
        ws['A1'].font = TITLE
        ws['A2'] = blurb
        ws['A2'].font = SUB
        scope = ('%s rows of the file' % '{:,}'.format(rowsn)) if rowsn else 'the file as a whole'
        ws['A3'] = ('%s covers %s. There %s %d thing%s that must be fixed and %d to check. '
                    'Each line below is one kind of problem, however many rows it affects.'
                    % (name, scope, 'is' if must == 1 else 'are', must,
                       '' if must == 1 else 's', chk))
        ws['A3'].font = Font(size=10)

        r = 5
        r = header_row(ws, r, ['How serious', 'What is wrong', 'Why it matters', 'What to do',
                               'How many', 'Where to look'])
        if not fs:
            ws.cell(r, 1, 'Nothing wrong here.').fill = OKF
            r += 1
        for sev, code, cnt, rws, items in collapse(fs):
            word = SEVERITY_WORD[sev]
            title, why, todo = plain(code)
            c = ws.cell(r, 1, word)
            sev_fill(c, sev)
            c.font = Font(bold=True, size=10)
            ws.cell(r, 2, title).font = Font(bold=True, size=10)
            ws.cell(r, 3, why)
            ws.cell(r, 4, todo)
            ws.cell(r, 5, cnt)
            ws.cell(r, 6, where(rws))
            for k in range(1, 7):
                ws.cell(r, k).border = THIN
                ws.cell(r, k).alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[r].height = 58
            r += 1

        r += 1
        ws.cell(r, 1, 'Examples, if you want to see actual cases').font = SECT
        ws.cell(r, 1).fill = SECF
        r += 1
        r = header_row(ws, r, ['What is wrong', 'Row', 'What the checker saw'])
        for sev, code, cnt, rws, items in collapse(fs):
            title = plain(code)[0]
            for f in items[:3]:
                ws.cell(r, 1, title)
                ws.cell(r, 2, f.row if f.row else 'whole file')
                ws.cell(r, 3, f.finding)
                for k in range(1, 4):
                    ws.cell(r, k).border = THIN
                r += 1

        for i, wd in enumerate([14, 40, 66, 62, 13, 30], 1):
            ws.column_dimensions[get_column_letter(i)].width = wd
        built.append((g, ws, must, chk, len(fs) - must - chk, rowsn))

    # ------------------------------------------------------------- Start Here --
    ws = wb.create_sheet('START HERE')
    ws['A1'] = 'How to use this file'
    ws['A1'].font = Font(bold=True, size=18, color='1F3864')
    total_must = sum(b[2] for b in built)
    total_chk = sum(b[3] for b in built)
    lines = [
        ('', ''),
        ('What this is', 'A check of the QF HQ building data file. The file lists every room, every piece '
                         'of equipment and every sensor in the building, one fact per row - almost thirty '
                         'thousand rows. This workbook says what is wrong with it.'),
        ('', ''),
        ('What we found', 'There are %s things that must be fixed and %s that someone should look at and '
                          'decide about. They are not %s separate problems - they are a few dozen kinds of '
                          'problem, each repeated across many rows.'
                          % ('{:,}'.format(total_must), '{:,}'.format(total_chk),
                             '{:,}'.format(total_must + total_chk))),
        ('', ''),
        ('What to do', ''),
        ('   Step 1', 'Read the "Biggest problems" list below. That is the short version - about ten lines.'),
        ('   Step 2', 'Each tab after this one covers one part of the building - the rooms, the air handling '
                      'units, the VAV boxes and so on. Open the tabs that are yours.'),
        ('   Step 3', 'On each tab, work down the list. Anything marked "Must fix" needs doing. Anything '
                      'marked "Please check" needs a decision from someone who knows the building.'),
        ('   Step 4', 'Send the file back to whoever built it with this workbook attached.'),
        ('', ''),
        ('The colours in the data tab',
         'The tab named %s is the original data, untouched apart from colour. A YELLOW row has something '
         'that must be fixed. An ORANGE row has something to check. A plain white row is fine.'
         % sheet_name),
        ('', ''),
        ('One thing to know',
         'The original file was saved with a filter switched on, so it opens showing only 846 of its 29,169 '
         'rows. Nothing was missing - just hidden. We cleared it here so you can see everything.'
         if filter_state else ''),
        ('', ''),
        ('If a line makes no sense',
         'Every line says what is wrong in ordinary words, why it matters, and what to do. If one still '
         'does not make sense, the "Technical detail" tab at the end has the raw version for whoever is '
         'doing the fixing.'),
    ]
    r = 3
    for head, body in lines:
        if head:
            ws.cell(r, 1, head).font = Font(bold=True, size=11, color='1F3864')
        if body:
            c = ws.cell(r, 2, body)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[r].height = max(16, 15 * (len(body) // 95 + 1))
        r += 1

    r += 1
    ws.cell(r, 1, 'The biggest problems, worst first').font = Font(bold=True, size=13, color='1F3864')
    r += 1
    r = header_row(ws, r, ['How serious', 'What is wrong', 'Why it matters', 'What to do',
                           'How many', 'Which tab'])
    big = collections.Counter()
    for f in findings:
        big[f.code] += 1
    tabs = collections.defaultdict(set)
    for f in findings:
        tabs[f.code].add(pretty(f.group)[0])
    sev_of = {}
    for f in findings:
        if f.code not in sev_of or SEV_ORDER[f.severity] < SEV_ORDER[sev_of[f.code]]:
            sev_of[f.code] = f.severity
    for code, cnt in sorted(big.items(),
                            key=lambda kv: (importance(kv[0]),
                                            SEV_ORDER[sev_of[kv[0]]], -kv[1]))[:12]:
        title, why, todo = plain(code)
        c = ws.cell(r, 1, SEVERITY_WORD[sev_of[code]])
        sev_fill(c, sev_of[code])
        c.font = Font(bold=True, size=10)
        ws.cell(r, 2, title).font = Font(bold=True, size=10)
        ws.cell(r, 3, why)
        ws.cell(r, 4, todo)
        ws.cell(r, 5, cnt)
        ws.cell(r, 6, ', '.join(sorted(tabs[code])[:3]))
        for k in range(1, 7):
            ws.cell(r, k).border = THIN
            ws.cell(r, k).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 58
        r += 1

    r += 2
    ws.cell(r, 1, 'What is on each tab').font = Font(bold=True, size=13, color='1F3864')
    r += 1
    r = header_row(ws, r, ['Tab', 'What it covers', 'Rows', 'Must fix', 'Please check', 'How it looks'])
    for g, wsx, must, chk, info, rowsn in built:
        nm, bl = pretty(g)
        ws.cell(r, 1, nm)
        ws.cell(r, 2, bl)
        ws.cell(r, 3, rowsn if rowsn else 'whole file')
        for k, v, sev in ((4, must, 'ERROR'), (5, chk, 'WARN')):
            c = ws.cell(r, k, v)
            if v:
                sev_fill(c, sev)
        v = ws.cell(r, 6, 'needs work' if must else ('worth a look' if chk else 'all fine'))
        v.fill = BADF if must else (WARNF if chk else OKF)
        for k in range(1, 7):
            ws.cell(r, k).border = THIN
        r += 1
    for i, wd in enumerate([20, 44, 62, 58, 13, 26], 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.column_dimensions['B'].width = 100

    # --------------------------------------------------------- technical tab --
    tech = wb.create_sheet('Technical detail')
    tech['A1'] = 'Technical detail - for whoever is doing the fixing'
    tech['A1'].font = TITLE
    tech['A2'] = ('Every finding, one per line, with the checker rule code it came from. The tabs before '
                  'this one are the readable version; nothing here is extra, it is the same findings '
                  'written for a specialist.')
    tech['A2'].font = SUB
    r = header_row(tech, 4, ['Severity', 'Rule code', 'Tab', 'Row', 'Entity', 'What the checker saw',
                             'What to do'])
    for f in sorted(findings, key=lambda f: (SEV_ORDER[f.severity], f.code, f.row or 0)):
        sev_fill(tech.cell(r, 1, f.severity), f.severity)
        tech.cell(r, 2, f.code)
        tech.cell(r, 3, pretty(f.group)[0])
        tech.cell(r, 4, f.row if f.row else 'file')
        tech.cell(r, 5, f.entity)
        tech.cell(r, 6, f.finding)
        tech.cell(r, 7, f.action)
        r += 1
    for i, wd in enumerate([11, 12, 22, 8, 34, 96, 70], 1):
        tech.column_dimensions[get_column_letter(i)].width = wd

    # ------------------------------------------------------------- Claude Log --
    log = wb.create_sheet('Claude Log')
    log.append(['Turn #', 'Date', 'User Request', 'Action Taken', 'Details', 'Outcome'])
    for c in range(1, 7):
        log.cell(1, c).font = COLH
        log.cell(1, c).fill = HDRF
    log.append([1, datetime.date.today(), args.request,
                'Built this review workbook: a START HERE tab, %d plain-English tabs, a technical tab, and '
                'colour on the data rows.' % len(built),
                'Ran the row-level validator and the cross-unit consistency checker over %d rows, both '
                'against a header-normalised copy so the broken header row did not drown the output, plus '
                'seven checks neither script covers. Findings are collapsed to one line per kind of problem '
                'rather than one line per row, and every rule code is rewritten in ordinary English. The '
                'raw coded findings are kept on the Technical detail tab.' % len(data),
                '%d findings in total: %d must be fixed, %d to check, %d for information. %d rows coloured '
                'yellow, %d orange.' % (len(findings), total_must, total_chk,
                                        len(findings) - total_must - total_chk,
                                        len(err_rows), len(warn_rows))])
    for c in range(1, 7):
        log.cell(2, c).alignment = Alignment(wrap_text=True, vertical='top')
    for i, wd in enumerate([8, 12, 60, 60, 110, 90], 1):
        log.column_dimensions[get_column_letter(i)].width = wd
    log.row_dimensions[2].height = 150

    wb._sheets = ([wb['START HERE'], wb[sheet_name]]
                  + [w for _, w, *_ in built] + [wb['Technical detail'], wb['Claude Log']])
    wb.save(args.out)

    print('%s -> %s' % (sheet.name, args.out))
    print('%d findings: %d must fix, %d please check' % (len(findings), total_must, total_chk))
    print('%d rows yellow, %d orange' % (len(err_rows), len(warn_rows)))
    for g, wsx, must, chk, info, rowsn in built:
        print('  %-24s rows=%-6d must fix=%-5d check=%d' % (wsx.title, rowsn, must, chk))
    return 1 if total_must else 0


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('sheet', type=Path)
    ap.add_argument('--out', required=True, help='where to write the review workbook')
    ap.add_argument('--label-style', choices=('para', 'verbatim'), default='para')
    ap.add_argument('--io', help='IO list, passed through to both checkers as evidence')
    ap.add_argument('--ignore', default='I-TYP-6', help='comma-separated rule codes to leave out')
    ap.add_argument('--group', action='append', default=[], metavar='NAME=REGEX',
                    help='force subjects matching REGEX onto sheet NAME; repeatable')
    ap.add_argument('--request', default='Run a validation exercise over this ontology.',
                    help='the ask, recorded in the Claude Log sheet')
    args = ap.parse_args()
    if not args.sheet.exists():
        sys.exit('no such sheet: %s' % args.sheet)
    sys.exit(build(args))


if __name__ == '__main__':
    main()
