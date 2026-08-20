#!/usr/bin/env python3
"""Validate a PARA Brick ontology sheet (.csv or .xlsx) before handover.

    python3 validate_ontology.py MyBuilding.xlsx
    python3 validate_ontology.py MyBuilding.csv --strict --max 5
    python3 validate_ontology.py MyBuilding.xlsx --label-style verbatim
    python3 validate_ontology.py MyBuilding.xlsx --preflight

Exit status: 0 clean (warnings allowed), 1 errors found, 2 with --strict if
anything at all was reported, 3 if the file could not be read.

Every finding carries a stable code so a rule can be discussed, suppressed
(--ignore W-TYP-2) or looked up in references/known-issues.md.
"""
from __future__ import annotations

import argparse
import collections
import csv
import re
import sys
from pathlib import Path

DATA = Path(__file__).resolve().parent.parent / "references" / "data"

CORE = ["subject", "subjecttype", "predicate", "object", "objecttype"]
PAIR_COLS = {"subject_prop_name", "subject_prop_val", "object_prop_name", "object_prop_val"}

ALLOWED_PREFIXES = {
    "entity", "brick", "rec", "ref", "unit", "qudt", "para",
    "rdf", "rdfs", "owl", "xsd", "skos", "bacnet",
}

# Equipment that terminates in a conditioned space: it must declare what it feeds.
TERMINAL_EQUIPMENT = {
    "brick:Variable_Air_Volume_Box", "brick:Fan_Coil_Unit", "brick:CRAC",
    "brick:Exhaust_Fan", "brick:Radiator", "brick:Chilled_Beam",
    "brick:Terminal_Unit", "para:Pressure_Independent_Module", "para:DXUnit",
}
FEED_TARGETS = {"rec:Room", "rec:HVACZone", "rec:Zone", "rec:Space"}

SPATIAL_PARENTS = {"rec:isPartOf", "rec:locatedIn", "brick:isPartOf"}
SPATIAL_CLASSES = {
    "rec:Room", "rec:Level", "rec:BasementLevel", "rec:RoofLevel",
    "rec:GroundLevel", "rec:Zone", "rec:HVACZone", "rec:Building", "rec:Site",
}

PLACEHOLDER = re.compile(r"<(?!blanknode>)[^>]*>")


def label_issues(text: str) -> list[str]:
    """Return the characters in a label that the PARA label rule disallows.

    Labels may carry spaces and a decimal point between two digits (PM2.5);
    every other punctuation mark is stripped.
    """
    bad = []
    for i, ch in enumerate(text):
        if ch.isalnum() or ch == " ":
            continue
        if ch == "." and 0 < i < len(text) - 1 and text[i - 1].isdigit() and text[i + 1].isdigit():
            continue
        bad.append(ch)
    return sorted(set(bad))


def clean_label(text: str) -> str:
    """Rewrite a raw name into a compliant rdfs:label_en value."""
    out = []
    for i, ch in enumerate(text):
        if ch.isalnum():
            out.append(ch)
        elif ch == "." and 0 < i < len(text) - 1 and text[i - 1].isdigit() and text[i + 1].isdigit():
            out.append(ch)
        else:
            out.append(" ")
    return " ".join("".join(out).split())


def load_vocab() -> dict[str, tuple[str, str]]:
    """Load the Brick/REC term list: term -> (status, note)."""
    path = DATA / "brick-vocab.txt"
    if not path.exists():
        return {}
    out = {}
    for line in path.read_text().splitlines():
        if not line.strip() or line.startswith("#"):
            continue
        parts = line.split("\t")
        out[parts[0]] = (parts[1] if len(parts) > 1 else "OK",
                         parts[2] if len(parts) > 2 else "")
    return out


def load_list(name: str) -> set[str]:
    path = DATA / name
    if not path.exists():
        return set()
    if path.suffix == ".csv":
        with open(path, newline="") as fh:
            return {r[0].strip() for r in list(csv.reader(fh))[1:] if r and r[0].strip()}
    return {
        line.strip() for line in path.read_text().splitlines()
        if line.strip() and not line.startswith("#")
    }


def pick_ontology_sheet(wb, path):
    """Choose the sheet holding the triples.

    A reference model can carry review and comparison sheets alongside the
    ontology, and the workbook's *active* sheet is simply whichever tab was
    open when it was last saved - `QF_SSC_Ontology_draft0.5_review.xlsx` saves
    with `VAV_Comparison` selected. Reading `.active` therefore reads a random
    sheet. Pick by the header contract instead: the first sheet whose first row
    starts with the five core columns.
    """
    for ws in wb.worksheets:
        head = [str(c).strip().lower() if c is not None else ""
                for c in next(ws.iter_rows(max_row=1, values_only=True), ())][:5]
        if head == ["subject", "subjecttype", "predicate", "object", "objecttype"]:
            return ws
    return wb.active

def read_sheet(path: Path):
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:
            sys.exit("openpyxl is needed to read .xlsx - pip install openpyxl")
        ws = pick_ontology_sheet(
            openpyxl.load_workbook(path, data_only=True), path)
        rows = [
            ["" if c is None else str(c) for c in r]
            for r in ws.iter_rows(values_only=True)
        ]
    else:
        with open(path, encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.reader(fh))
    if not rows:
        sys.exit(f"{path} is empty")
    header = [h.strip() for h in rows[0]]
    body = [(i + 2, r) for i, r in enumerate(rows[1:]) if any(c.strip() for c in r)]
    return header, body


class Report:
    def __init__(self, max_per_code: int):
        self.findings: list[tuple[str, str, int, str]] = []
        self.all: list[tuple[str, str, int, str]] = []
        self.max = max_per_code
        self.counts: collections.Counter = collections.Counter()

    def add(self, severity: str, code: str, row: int, message: str):
        self.counts[code] += 1
        self.all.append((severity, code, row, message))
        if self.counts[code] <= self.max:
            self.findings.append((severity, code, row, message))

    def write(self, path: Path, source: Path, ignore: set[str]) -> int:
        """Write every finding to a file of its own.

        Never into the ontology workbook. The deliverable is one sheet of
        triples and nothing else - a converter that meets a second sheet has to
        be told which one to read, and a reviewer diffing two versions has to
        skip it. Findings go somewhere separate or they go to stdout.
        """
        rows = [f for f in self.all if f[1] not in ignore]
        if not rows:
            return 0
        rows.sort(key=lambda f: (f[1], f[2]))
        header = ["severity", "code", "row", "finding", "source"]
        body = [[sev, code, row or "", msg, source.name] for sev, code, row, msg in rows]

        if path.suffix.lower() in (".xlsx", ".xlsm"):
            try:
                import openpyxl
            except ImportError:
                sys.exit("openpyxl is needed to write .xlsx - pip install openpyxl")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Findings"
            ws.append(header)
            for r in body:
                ws.append(r)
            wb.save(path)
        else:
            with open(path, "w", newline="", encoding="utf-8") as fh:
                w = csv.writer(fh)
                w.writerow(header)
                w.writerows(body)
        return len(rows)

    def emit(self, ignore: set[str]) -> tuple[int, int]:
        for severity, code, row, message in sorted(
                (f for f in self.findings if f[1] not in ignore),
                key=lambda f: (f[1], f[2])):
            where = f"row {row}" if row else "file"
            print(f"{severity:5} {code:10} {where:>10}  {message}")
        errors = warns = 0
        for code, n in sorted(self.counts.items()):
            if code in ignore:
                continue
            if code.startswith("E-"):
                errors += n
            elif code.startswith("W-"):
                warns += n
            if n > self.max:
                print(f"      {code:10} {'':>10}  ... and {n - self.max} more")
        return errors, warns


def validate(path: Path, report: Report, label_style: str = "para", io=None):
    header, body = read_sheet(path)
    low = [h.lower() for h in header]

    if low[:5] != CORE:
        report.add("ERROR", "E-HDR-1", 0,
                   f"first five headers must be {CORE}, found {low[:5]}")
    for i in range(5, len(header)):
        if header[i] and header[i] not in PAIR_COLS:
            report.add("ERROR", "E-HDR-2", 0,
                       f"column {i + 1} header {header[i]!r} is not a prop column")

    vocab = load_vocab()
    precedent = load_list("brick-rec-vocab.txt")
    para_known = load_list("para-classes.csv") | load_list("para-properties.csv")
    units_known = load_list("units.csv") | {t for t in vocab if t.startswith("unit:")}

    types: dict[str, set[str]] = collections.defaultdict(set)
    labelled: set[str] = set()
    mentioned: set[str] = set()
    subjects: set[str] = set()
    referenced: set[str] = set()
    para_defined: set[str] = set()
    parent_of: dict[str, str] = {}
    feeds: dict[str, list[str]] = collections.defaultdict(list)
    located: set[str] = set()
    points: set[str] = set()
    has_ext_ref: set[str] = set()
    seen_rows: dict[tuple, int] = {}

    for rownum, raw in body:
        r = [c if c is not None else "" for c in raw] + [""] * (len(header) - len(raw))
        subj, stype, pred, obj, otype = (c.strip() for c in r[:5])

        for i, cell in enumerate(r[:len(header)]):
            if cell != cell.strip():
                report.add("ERROR", "E-WS-1", rownum,
                           f"{header[i] or f'col{i + 1}'} has padding whitespace: {cell!r}")
            if PLACEHOLDER.search(cell.strip()):
                report.add("ERROR", "E-PH-1", rownum,
                           f"{header[i] or f'col{i + 1}'} still holds a placeholder: {cell.strip()!r}")

        if not subj or not pred or not obj:
            report.add("ERROR", "E-CORE-1", rownum,
                       "subject, predicate and object are all required")
            continue

        key = tuple(c.strip() for c in r[:len(header)])
        if key in seen_rows:
            report.add("WARN", "W-DUP-1", rownum, f"identical to row {seen_rows[key]}")
        else:
            seen_rows[key] = rownum

        # --- prefixes and identifier hygiene --------------------------------
        for col, val in (("subject", subj), ("subjectType", stype),
                         ("predicate", pred), ("object", obj), ("objectType", otype)):
            if not val or val == "<blanknode>":
                continue
            if ":" not in val:
                if col in ("subjectType", "predicate", "objectType"):
                    report.add("ERROR", "E-PFX-2", rownum, f"{col} {val!r} has no prefix")
                continue
            prefix = val.split(":", 1)[0]
            if prefix not in ALLOWED_PREFIXES:
                report.add("ERROR", "E-PFX-1", rownum, f"{col} uses unknown prefix {prefix!r}")
            if " " in val:
                report.add("ERROR", "E-SPACE-1", rownum, f"{col} {val!r} contains a space")
            # These apply to namespaced cells only, never to a *_prop_val literal:
            # a label may legitimately contain anything a label may contain.
            bad = [c for c in val if ord(c) < 32 or c in "\u00a0\u200b\u2060\ufeff"]
            if bad:
                report.add("ERROR", "E-CELL-1", rownum,
                           f"{col} {val!r} contains {[hex(ord(c)) for c in bad]} - a "
                           f"control character or invisible space")
            if val.count(":") > 1:
                report.add("ERROR", "E-CELL-2", rownum,
                           f"{col} {val!r} has {val.count(':')} colons; a term is "
                           f"prefix:localName")
            if prefix != prefix.lower():
                report.add("ERROR", "E-CELL-3", rownum,
                           f"{col} prefix {prefix!r} must be lower case")

        # --- vocabulary ------------------------------------------------------
        for col, val in (("subjectType", stype), ("predicate", pred), ("objectType", otype)):
            if not val or val == "<blanknode>":
                continue
            if val.startswith(("brick:", "rec:", "ref:")) and vocab:
                status, note = vocab.get(val, ("MISSING", ""))
                if status == "MISSING":
                    report.add("ERROR", "E-TYP-2", rownum,
                               f"{col} {val} is not a term in Brick 1.4 - "
                               "check the spelling on ontology.brickschema.org")
                elif status == "ALIAS":
                    report.add("WARN", "W-TYP-5", rownum, f"{col} {val} is an alias - {note}")
                elif status == "DEPRECATED":
                    report.add("WARN", "W-TYP-4", rownum,
                               f"{col} {val} is deprecated - {note or 'find a REC replacement'}")
                elif precedent and val not in precedent:
                    report.add("INFO", "I-TYP-6", rownum,
                               f"{col} {val} is valid Brick but has no precedent in Dar Cairo")
            if val.startswith("para:"):
                mentioned.add(val)

        if pred == "rdfs:subClassOf" and subj.startswith("para:"):
            para_defined.add(subj)
            parent_of[subj] = obj
            if stype in ("owl:DatatypeProperty", "owl:ObjectProperty", "brick:EntityProperty"):
                report.add("WARN", "W-EXT-5", rownum,
                           f"{subj} is a property, not a class - relate it with "
                           "rdfs:subPropertyOf rather than rdfs:subClassOf")
            elif stype != "owl:Class":
                report.add("ERROR", "E-EXT-1", rownum,
                           f"{subj} is a class definition, subjectType must be owl:Class "
                           f"(found {stype!r})")
            elif not obj.startswith(("brick:", "rec:", "para:")):
                report.add("ERROR", "E-EXT-2", rownum,
                           f"{subj} needs a brick:/rec:/para: parent class, found {obj!r}")
            if obj == subj:
                report.add("ERROR", "E-EXT-3", rownum, f"{subj} is its own parent")
        if pred == "rdfs:subPropertyOf" and subj.startswith("para:"):
            para_defined.add(subj)
            if obj == subj:
                report.add("WARN", "W-EXT-4", rownum,
                           f"{subj} declares itself as its own super-property")

        # --- type consistency -------------------------------------------------
        if subj.startswith("entity:") and stype:
            types[subj].add(stype)
            mentioned.add(subj)
        subjects.add(subj)
        if obj.startswith("entity:"):
            referenced.add(obj)
            mentioned.add(obj)
            if otype and otype != "<blanknode>":
                types[obj].add(otype)

        # --- relationship bookkeeping ------------------------------------------
        if pred in ("rec:feeds", "brick:feeds") and subj.startswith("entity:"):
            feeds[subj].append(otype or "?")
            if stype in TERMINAL_EQUIPMENT and otype and otype not in FEED_TARGETS:
                report.add("WARN", "W-FEED-2", rownum,
                           f"{stype} feeds {otype} - a terminal unit should feed "
                           f"one of {sorted(FEED_TARGETS)}")
        if pred in ("rec:isFedBy", "brick:isFedBy") and obj.startswith("entity:"):
            feeds[obj].append(stype or "?")
        if pred in ("rec:locatedIn", "brick:hasLocation"):
            located.add(subj)
        if pred in SPATIAL_PARENTS and stype in SPATIAL_CLASSES:
            parent_of.setdefault(subj, obj)
        if pred == "brick:hasPoint":
            points.add(obj)
        if pred == "ref:hasExternalReference":
            has_ext_ref.add(subj)
            if otype not in ("ref:TimeseriesReference", "ref:IFCReference", "ref:BACnetReference"):
                report.add("WARN", "W-BN-3", rownum,
                           f"external reference should be typed in objectType, found {otype!r}")

        # --- property pairs ------------------------------------------------------
        object_props = 0
        row_props: set[str] = set()
        i = 5
        while i + 1 < len(header) + 1 and i < len(header):
            name_col = header[i].lower()
            val_col = header[i + 1].lower() if i + 1 < len(header) else ""
            if not name_col.endswith("_name") or not val_col.endswith("_val"):
                i += 1
                continue
            pname, pval = r[i].strip(), r[i + 1].strip()
            if pname and not pval:
                report.add("ERROR", "E-PAIR-1", rownum, f"{pname} has no value")
            if pval and not pname:
                report.add("ERROR", "E-PAIR-2", rownum, f"value {pval!r} has no property name")
            if pname:
                row_props.add(pname)
                side = "subject" if name_col.startswith("subject") else "object"
                target = subj if side == "subject" else obj
                if side == "object":
                    object_props += 1
                if pname in ("rdfs:label_en", "skos:prefLabel"):
                    labelled.add(target)
                    bad = label_issues(pval) if label_style == "para" else []
                    if bad:
                        report.add("ERROR", "E-LBL-1", rownum,
                                   f"label {pval!r} contains {bad} - use {clean_label(pval)!r}")
                if pname == "brick:hasUnit":
                    if not pval.startswith(("unit:", "para:")):
                        report.add("ERROR", "E-UNIT-1", rownum,
                                   f"unit {pval!r} must be a unit: or para: term")
                    elif units_known and pval not in units_known:
                        report.add("WARN", "W-UNIT-2", rownum,
                                   f"unit {pval} is in no known unit list - confirm at "
                                   "ontology.brickschema.org/qudt/Unit.html")
                if pname.startswith("para:"):
                    mentioned.add(pname)
            i += 2

        if obj == "<blanknode>" and object_props == 0:
            report.add("ERROR", "E-BN-1", rownum,
                       "<blanknode> object carries no object_prop pairs")
        # A half-filled blank node is worse than an empty one: it looks finished.
        if "brick:value" in row_props and "brick:hasUnit" not in row_props:
            report.add("WARN", "W-BN-4", rownum,
                       "brick:value with no brick:hasUnit - use unit:UNITLESS if the "
                       "quantity is genuinely dimensionless")
        if "ref:hasTimeseriesId" in row_props and "para:hasEntityId" not in row_props:
            report.add("WARN", "W-BN-5", rownum,
                       "ref:hasTimeseriesId with no para:hasEntityId - the telemetry "
                       "key has nothing to say which entity it is grouped under")
        if pred == "brick:aggregate":
            for want in ("brick:aggregationFunction", "brick:aggregationInterval"):
                if want not in row_props:
                    report.add("WARN", "W-AGG-1", rownum,
                               f"brick:aggregate with no {want}")
        if obj == "<blanknode>" and otype not in ("<blanknode>", "ref:TimeseriesReference",
                                                  "ref:IFCReference", "ref:BACnetReference"):
            report.add("WARN", "W-BN-2", rownum,
                       f"blank-node objectType should be <blanknode> or a ref: class, found {otype!r}")

    # ---- whole-file checks ---------------------------------------------------
    for entity, ts in sorted(types.items()):
        if len(ts) > 1:
            report.add("ERROR", "E-TYP-1", 0,
                       f"{entity} is typed {len(ts)} different ways: {sorted(ts)}")

    # A dangling reference is an entity that is mentioned and never given a
    # class - a typo, a renamed entity, or a placeholder resolved on one row and
    # not the others. Being object-only is not itself a defect: the house style
    # declares sub-systems, parts and the site entirely on the row that
    # references them, with the class in objectType and the label in an object
    # prop. Requiring a subject row would flag every one of those.
    for entity in sorted(referenced - subjects - set(types)):
        report.add("WARN", "W-REF-1", 0,
                   f"{entity} is referenced but never given a class, on this row or "
                   f"any other - dangling reference")

    for term in sorted(mentioned):
        if term.startswith("para:") and term not in para_defined and term not in para_known:
            report.add("ERROR", "E-TYP-3", 0,
                       f"{term} is used but never defined here and is not in the para registry")

    for entity in sorted(types):
        if entity not in labelled:
            report.add("WARN", "W-LBL-2", 0, f"{entity} never gets an rdfs:label_en")

    for entity, ts in sorted(types.items()):
        stype = sorted(ts)[0] if ts else ""
        if stype in TERMINAL_EQUIPMENT and entity not in feeds:
            report.add("ERROR", "E-FEED-1", 0,
                       f"{entity} ({stype}) never declares what it feeds - rec:feeds "
                       "must point at the room or HVAC zone it serves")
        if stype in TERMINAL_EQUIPMENT and entity not in located:
            report.add("WARN", "W-GR-2", 0, f"{entity} has no rec:locatedIn")

    # A point with no external reference is a defect when the BMS publishes a key
    # for it and a fact when it does not. With an IO list to hand, ask rather
    # than flag - that is the pass a reviewer would otherwise do by hand.
    for point in sorted(points):
        if point in has_ext_ref:
            continue
        verdict = None
        if io is not None:
            local = point.split(":", 1)[-1]
            parent, _, suffix = local.rpartition("_")
            verdict = io.timeseries_id(parent, suffix)
        if verdict == "":
            report.add("INFO", "I-PT-3", 0,
                       f"{point} has no ref:hasExternalReference, and the IO list has "
                       f"no timeseries id for it either - confirmed, not a defect")
        elif verdict:
            report.add("ERROR", "E-PT-4", 0,
                       f"{point} has no ref:hasExternalReference but the IO list gives "
                       f"it {verdict!r} - the reference is missing from the sheet")
        else:
            report.add("WARN", "W-PT-1", 0,
                       f"{point} is a data point with no ref:hasExternalReference")

    # spatial connectivity: walk each spatial entity up to a rec:Building / rec:Site
    roots = {e for e, ts in types.items() if ts & {"rec:Building", "rec:Site"}}
    for entity, ts in sorted(types.items()):
        if not (ts & SPATIAL_CLASSES) or entity in roots:
            continue
        seen, cur = set(), entity
        while cur in parent_of and cur not in seen:
            seen.add(cur)
            cur = parent_of[cur]
        if cur not in roots:
            report.add("ERROR", "E-GR-1", 0,
                       f"{entity} does not connect up to a rec:Building "
                       f"(chain ends at {cur or 'nothing'})")

    print(f"\n{len(body)} rows, {len(types)} typed entities, "
          f"{len(para_defined)} para: definitions\n")


def preflight(path: Path):
    """Discover and show: print what this sheet actually contains, before judging it.

    Every rule below runs against what was found here, not against a list of one
    building's facts carried in from another. Read this first and confirm
    anything that looks wrong - a class you do not recognise, a unit nobody
    uses, a prefix that should not be there - because a sheet can validate clean
    and still model the wrong building.
    """
    header, body = read_sheet(path)
    low = [h.strip().lower() for h in header]
    idx = [low.index(c) for c in CORE]

    prefixes = collections.Counter()
    classes = collections.defaultdict(collections.Counter)
    predicates = collections.Counter()
    props = collections.Counter()
    units = collections.Counter()
    entities = set()

    for _, r in body:
        cells = [r[i].strip() if i < len(r) else "" for i in idx]
        subj, stype, pred, obj, otype = cells
        for v in (subj, stype, pred, obj, otype):
            if v and ":" in v and v != "<blanknode>":
                prefixes[v.split(":", 1)[0]] += 1
        if subj.startswith("entity:"):
            entities.add(subj)
        if pred:
            predicates[pred] += 1
        for v in (stype, otype):
            if v and v != "<blanknode>":
                kind = ("spatial" if v in SPATIAL_CLASSES else
                        "extension" if v.startswith("para:") else
                        "schema" if v.startswith(("owl:", "rdfs:")) else
                        "reference" if v.startswith("ref:") else "equipment or point")
                classes[kind][v] += 1
        for i in range(5, len(header) - 1, 2):
            if i + 1 < len(r) and r[i].strip():
                props[r[i].strip()] += 1
                if r[i].strip() == "brick:hasUnit":
                    units[r[i + 1].strip()] += 1

    print(f"\n{path.name}: {len(body)} rows, {len(entities)} entities\n")
    print("prefixes  ", ", ".join(f"{k}({v})" for k, v in prefixes.most_common()))
    print("predicates", ", ".join(f"{k}({v})" for k, v in predicates.most_common(12)))
    print("properties", ", ".join(f"{k}({v})" for k, v in props.most_common(12)))
    print("units     ", ", ".join(f"{k}({v})" for k, v in units.most_common(12)) or "none")
    for kind in ("spatial", "equipment or point", "extension", "reference", "schema"):
        got = classes.get(kind)
        if got:
            print(f"\n{kind} classes ({len(got)})")
            for cls, n in got.most_common(15):
                print(f"  {n:>6}  {cls}")
            if len(got) > 15:
                print(f"         ... and {len(got) - 15} more")
    print("\nConfirm anything above that does not look like this building before "
          "reading the findings.\n")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("sheet", type=Path)
    ap.add_argument("--max", type=int, default=15, help="max findings shown per rule code")
    ap.add_argument("--strict", action="store_true", help="fail on warnings too")
    ap.add_argument("--io", type=Path, metavar="PATH",
                    help="the IO list. Supplied, it is used as evidence: findings it "
                         "can adjudicate are resolved against it instead of flagged.")
    ap.add_argument("--preflight", action="store_true",
                    help="print what the sheet contains - prefixes, classes, predicates, "
                         "properties, units - and stop. Discover and show: run this "
                         "first and confirm the picture before trusting any finding.")
    ap.add_argument("--ignore", default="", help="comma-separated rule codes to suppress")
    ap.add_argument("--report", type=Path, metavar="PATH",
                    help="also write every finding to PATH (.xlsx or .csv), a file of "
                         "its own. Nothing is ever written into the ontology sheet.")
    ap.add_argument("--label-style", choices=("para", "verbatim"), default="para",
                    help="para (default): enforce the PARA label rule - letters, digits "
                         "and spaces only. verbatim: labels carry the source text as "
                         "written, the QF SSC house style; E-LBL-1 is not applied. Pick "
                         "the one the user chose at intake and say which in the handover.")
    args = ap.parse_args()

    if not args.sheet.exists():
        print(f"no such file: {args.sheet}", file=sys.stderr)
        return 3

    if args.preflight:
        preflight(args.sheet)
        return 0

    io = None
    if args.io:
        if not args.io.exists():
            print(f"no such file: {args.io}", file=sys.stderr)
            return 3
        import io_list
        io = io_list.load(args.io)
        print(f"IO list: {io.describe()}")

    report = Report(args.max)
    validate(args.sheet, report, args.label_style, io)
    ignore = {c.strip() for c in args.ignore.split(",") if c.strip()}
    errors, warns = report.emit(ignore)

    infos = sum(n for c, n in report.counts.items()
                if c not in ignore and c.startswith("I-"))
    print(f"\n{errors} errors, {warns} warnings, {infos} advisories")
    if args.report:
        n = report.write(args.report, args.sheet, ignore)
        print(f"{n} findings written to {args.report}" if n
              else "nothing to report, no file written")
    if errors:
        return 1
    if args.strict and warns:
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
