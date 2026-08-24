#!/usr/bin/env python3
"""Regenerate references/data/* from the reference models.

Run this after a new reference model lands in reference-models/ so the
validator's vocabulary stays in step with what the team actually ships.

    python3 scripts/build_vocab.py

Writes:
    references/data/brick-rec-vocab.txt   verified brick:/rec:/ref:/qudt: terms
    references/data/para-classes.csv      para: classes + their declared parent
    references/data/para-properties.csv   para: entity properties
    references/data/units.csv             units seen with brick:hasUnit
"""
import collections
import csv
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
MODELS = ROOT.parent.parent / "reference-models"
OUT = ROOT / "references" / "data"

# Only the primary reference model seeds the "verified vocabulary" list. The other
# models are drafts: their typos must not become accepted terms.
PRIMARY = {"DarCairo_V93.csv"}

# Units observed in Dar Cairo that are fill-down artifacts, not real QUDT units.
# MicroGM-PER-M3 is micrograms per cubic metre; M4..M27 do not exist.
BOGUS_UNIT = re.compile(r"^unit:MicroGM-PER-M(?!3$)\d+$")


def pick_ontology_sheet(wb, path):
    """Choose the sheet holding the triples.

    A reference model can carry review and comparison sheets alongside the
    ontology, and the workbook's *active* sheet is simply whichever tab was
    open when it was last saved - earlier SSC drafts saved on a review tab such
    as `VAV_Comparison`. Reading `.active` therefore risks reading the wrong
    sheet. Pick by the header contract instead: the first sheet whose first row
    starts with the five core columns.
    """
    for ws in wb.worksheets:
        head = [str(c).strip().lower() if c is not None else ""
                for c in next(ws.iter_rows(max_row=1, values_only=True), ())][:5]
        if head == ["subject", "subjecttype", "predicate", "object", "objecttype"]:
            return ws
    return wb.active


def read_rows(path):
    """Yield (header, rows) from a .csv or .xlsx reference model."""
    if path.suffix.lower() == ".csv":
        with open(path, encoding="utf-8-sig", newline="") as fh:
            rows = [r for r in csv.reader(fh) if any(c.strip() for c in r)]
    else:
        import openpyxl

        ws = pick_ontology_sheet(
            openpyxl.load_workbook(path, data_only=True), path)
        rows = [
            ["" if c is None else str(c) for c in r]
            for r in ws.iter_rows(values_only=True)
        ]
        rows = [r for r in rows if any(c.strip() for c in r)]
    return rows[0], rows[1:]


def main():
    models = sorted(
        p for p in MODELS.iterdir() if p.suffix.lower() in (".csv", ".xlsx")
    )
    if not models:
        sys.exit(f"no reference models found in {MODELS}")

    vocab = set()
    para_classes = {}
    para_props = {}
    units = collections.Counter()

    for path in models:
        header, rows = read_rows(path)
        if [h.strip().lower() for h in header[:3]] != ["subject", "subjecttype", "predicate"]:
            print(f"skipping {path.name}: not a triple sheet", file=sys.stderr)
            continue
        for r in rows:
            r = r + [""] * (len(header) - len(r))
            subj, stype, pred, obj, otype = (c.strip() for c in r[:5])

            if path.name in PRIMARY:
                for term in (stype, pred, otype):
                    if term.split(":")[0] in (
                            "brick", "rec", "ref", "qudt", "rdfs", "rdf", "owl", "skos"):
                        vocab.add(term)

            if pred == "rdfs:subClassOf" and subj.startswith("para:"):
                label = ""
                for i in range(5, len(r) - 1):
                    if header[i].strip() == "subject_prop_name" and r[i].strip() in (
                        "rdfs:label_en",
                        "skos:prefLabel",
                    ):
                        label = r[i + 1].strip()
                        break
                para_classes.setdefault(subj, (obj, label, path.name))
            if pred == "rdfs:subPropertyOf" and subj.startswith("para:"):
                para_props.setdefault(subj, (stype, obj, path.name))
            # any para: property used as a predicate is an entity property
            if pred.startswith("para:") and pred[5].islower():
                para_props.setdefault(pred, ("brick:EntityProperty", "", path.name))

            for i in range(5, len(r) - 1):
                if header[i].strip().endswith("_name") and r[i].strip() == "brick:hasUnit":
                    unit = r[i + 1].strip()
                    if unit and not BOGUS_UNIT.match(unit):
                        units[unit] += 1

    OUT.mkdir(parents=True, exist_ok=True)

    (OUT / "brick-rec-vocab.txt").write_text(
        "# brick:/rec:/ref: terms verified in the primary reference model (Dar Cairo).\n"
        "# A term outside this list is not wrong, it just has no precedent here, so\n"
        "# confirm it on ontology.brickschema.org before using it.\n"
        "# Regenerate with scripts/build_vocab.py\n"
        + "\n".join(sorted(vocab))
        + "\n"
    )

    with open(OUT / "para-classes.csv", "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["class", "subClassOf", "label", "source"])
        for k in sorted(para_classes):
            w.writerow([k, *para_classes[k]])

    with open(OUT / "para-properties.csv", "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["property", "type", "subPropertyOf", "source"])
        for k in sorted(para_props):
            w.writerow([k, *para_props[k]])

    with open(OUT / "units.csv", "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["unit", "uses"])
        for k, v in sorted(units.items(), key=lambda kv: (-kv[1], kv[0])):
            w.writerow([k, v])

    print(
        f"{len(vocab)} vocab terms, {len(para_classes)} para classes, "
        f"{len(para_props)} para properties, {len(units)} units"
    )


if __name__ == "__main__":
    main()
