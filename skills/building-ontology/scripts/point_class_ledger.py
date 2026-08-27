#!/usr/bin/env python3
"""Resolve a BMS point or part token to a Brick / para class, with provenance.

This is the datapoint half of the class-resolution ladder made mechanical. Where
`lookup_reference.py` answers "does this exact term exist", this takes an
*abbreviated BMS token* - `RmTempSP`, `SupFan.kW`, `IsoVlv.OpenSts` - and walks
the four-source ladder to a class, in the priority the workflow uses:

    Dar Cairo  ->  Brick 1.4  ->  QF SSC  ->  para: (minted)

and returns the class together with the step that settled it, so the resolution
is a reviewable ledger rather than a guess. See `references/datapoints.md` for
the workflow this belongs to.

Two ways to use it.

CLI - hand it a list of tokens and read the ledger:

    # one token
    point_class_ledger.py --token RmTempSP --kind Setpoint

    # a file of tokens, one per line (optionally  token<TAB>description<TAB>kind),
    # printed as a ledger CSV
    point_class_ledger.py --tokens points.txt > ledger.csv

Library - a project's build script drives it directly, layering its own curated
overrides on top of the automatic pass:

    from point_class_ledger import Resolver, kind_of, device_of
    R = Resolver()                       # loads Dar Cairo, Brick, (SSC if present)
    cls, source = R.resolve("room temperature setpoint", kind="Setpoint",
                            device="")

The automatic result is a FIRST PASS. Inspect every `auto` decision by hand and
curate the handful it mis-maps into a project `CANON` map keyed by the token,
each entry tagged with the ladder step it represents - three of the first
twenty-four were wrong on one building before the guards below went in.

Three guards stop the fuzzy match from confidently returning the wrong class,
each earning its place by catching a real mistake:

  * Kind      - a token's suffix fixes what KIND of point it is, and the Brick
                class name ends in the same word; a setpoint cannot match a
                `..._Sensor` however the words overlap.
  * Device    - a point on a damper cannot be a valve point; the part's device
                has to agree with the class's device.
  * Generic   - `"Process Value"` identifies nothing; when a historian
                description is generic or self-referential, expand the tag
                instead of trusting the description.

Dar Cairo (a CSV) is the only source needed for import and for the CLI on
tokens; Brick is read from the generated vocab; SSC is read only if the sample
workbook is present and openpyxl is installed, and its absence just drops step 3.
"""
from __future__ import annotations

import argparse
import collections
import csv
import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
MODELS = ROOT.parent.parent / "reference-models"
DATA = ROOT / "references" / "data"
DARCAIRO = MODELS / "DarCairo_V93.csv"
BRICKVOCAB = DATA / "brick-vocab.txt"
# The sample sheet is optional - step 3 of the ladder is skipped when it or
# openpyxl is missing. A project with its own reviewed SSC sheet passes ssc_path.
SSC_DEFAULT = MODELS / "QF_SSC_Ontology_draft0.5_review.xlsx"


# --------------------------------------------------------------------------- text
# A general BMS abbreviation table: token fragment -> English, longest first.
# It is deliberately broad, not building-specific; a project extends it by
# passing extra_abbrev= to expand(), but most tokens resolve without additions.
ABBREV = [
    ("StartStopCmdSts", "start stop command status"),
    ("StartStopCmd", "start stop command"),
    ("OnOffSts", "on off status"),
    ("OpenCmdSts", "open command status"),
    ("OpenCmd", "open command"), ("OpenSts", "open status"),
    ("CloseSts", "close status"), ("CloseCmd", "close command"),
    ("AutoManCmd", "auto manual command"), ("AutoManSts", "auto manual status"),
    ("PosCmd", "position command"), ("PosFbk", "position feedback"),
    ("SpeedFbk", "speed feedback"), ("SpeedCmd", "speed command"),
    ("DiffPrs", "differential pressure"), ("DiffPress", "differential pressure"),
    ("RmTempSP", "room temperature setpoint"), ("RmTemp", "room temperature"),
    ("SupTemp", "supply temperature"), ("RtnTemp", "return temperature"),
    ("SpcTemp", "space temperature"), ("SupAir", "supply air"),
    ("RtnAir", "return air"), ("MixAir", "mixed air"), ("OutAir", "outside air"),
    ("Humidity", "humidity"), ("Hum", "humidity"),
    ("RunSts", "run status"), ("Sts", "status"),
    ("TripAlm", "trip alarm"), ("CommAlm", "communication alarm"),
    ("FireAlm", "fire alarm"), ("HiAlm", "high alarm"), ("LoAlm", "low alarm"),
    ("Alm", "alarm"),
    ("FltRst", "fault reset"), ("Rst", "reset"), ("Reset", "reset"),
    ("Frequency", "frequency"), ("Freq", "frequency"),
    ("Voltage", "voltage"), ("Current", "current"),
    ("kWH", "electrical energy"), ("kWh", "electrical energy"),
    ("MWh", "electrical energy"), ("kW", "electric power"),
    ("Prs", "pressure"), ("Press", "pressure"), ("Temp", "temperature"),
    ("Flow", "flow"), ("Lvl", "level"),
    ("IsoVlv", "isolation valve"), ("Vlv", "valve"), ("Dmpr", "damper"),
    ("Fan", "fan"), ("Pmp", "pump"), ("Cmp", "compressor"),
    ("Fbk", "feedback"), ("Cmd", "command"), ("PV", ""), ("SP", "setpoint"),
]

# Point kind from the token suffix - the Brick class name ends in the same word.
# Order matters: FTSP (fail-to-stop) must be read before SP (setpoint).
KIND = [
    ("HiAlm", "Alarm"), ("LoAlm", "Alarm"), ("Alm", "Alarm"),
    ("FTSP", "Status"),
    ("SP", "Setpoint"), ("Setpoint", "Setpoint"),
    ("Cmd", "Command"), ("Ctrl", "Command"), ("Rst", "Command"),
    ("Reset", "Command"),
    ("Sts", "Status"), ("Fbk", "Sensor"), ("PV", "Sensor"),
    ("Ctr", "Sensor"), ("Mtr", "Sensor"),
    ("kWH", "Sensor"), ("kWh", "Sensor"), ("kW", "Sensor"), ("MWh", "Sensor"),
    ("Frequency", "Sensor"), ("Prs", "Sensor"), ("Temp", "Sensor"),
    ("Voltage", "Sensor"), ("Current", "Sensor"), ("Humidity", "Sensor"),
]

KIND_PARENT = {"Sensor": "brick:Sensor", "Setpoint": "brick:Setpoint",
               "Status": "brick:Status", "Command": "brick:Command",
               "Alarm": "brick:Alarm", "": "brick:Point"}

SYNONYM = {"rtn": "return", "sup": "supply", "spc": "space", "rm": "room",
           "pri": "primary", "sec": "secondary", "tnk": "tank",
           "eng": "engine", "gen": "generator", "alm": "alarm",
           "sts": "status", "cmd": "command", "vlv": "valve",
           "dmpr": "damper", "pmp": "pump", "prs": "pressure",
           "temp": "temperature", "hum": "humidity"}

DEVICE = ("damper", "valve", "fan", "coil", "filter", "heater", "pump",
          "motor", "breaker", "exchanger", "compressor")

STOP = {"the", "a", "of", "and", "on", "off", "process", "value",
        "zone", "level", "basement", "roof", "floor", "first"}

# Descriptions that pass an is-it-English test but identify nothing.
GENERIC = {"process value", "status", "command", "setpoint", "alarm", "value",
           "feedback", "control", "power", "energy", "", "setpoint status"}


def words(text, synonym=None):
    syn = synonym or SYNONYM
    out = []
    for w in re.split(r"[^a-z0-9]+", str(text).lower()):
        w = re.sub(r"(?<=[a-z])\d+$", "", w)
        w = syn.get(w, w)
        if w and w not in STOP:
            out.append(w)
    return out


def expand(token, extra_abbrev=None):
    """Abbreviated camelCase token -> a spelled-out English phrase."""
    out = str(token)
    for abbr, word in (extra_abbrev or []) + ABBREV:
        out = re.sub(r"(?<![A-Za-z])" + re.escape(abbr) + r"(?![a-z])",
                     " " + word + " ", out)
    out = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", out)
    out = re.sub(r"\d+", " ", out)
    return " ".join(out.replace("_", " ").replace(".", " ").lower().split())


def kind_of(point, kind_table=None):
    for suffix, kind in (kind_table or KIND):
        if point == suffix or point.endswith(suffix):
            return kind
    if point.endswith("Hrs"):
        return "Sensor"
    return ""


def device_of(text):
    w = set(words(text))
    for d in DEVICE:
        if d in w:
            return d
    return ""


def is_generic(desc):
    """A historian description too generic or self-referential to name a point.
    Pass the equipment tag prefix (e.g. "QNL") so "QNL AHU..." descriptions -
    which merely repeat the equipment name - are rejected too."""
    if not desc:
        return True
    return desc.strip().lower() in GENERIC


# --------------------------------------------------------------------------- indexes
def _label_on_row(r):
    """rdfs:label_en from a wide CSV row's object_prop pairs (cols 7.. in pairs)."""
    for i in range(7, min(len(r), 27), 2):
        if i - 1 < len(r) and str(r[i - 1]).strip() == "rdfs:label_en":
            return str(r[i]).strip()
    return ""


def load_darcairo_classes(path=DARCAIRO):
    """Point and part classes Dar Cairo actually uses, each with the English
    labels its own entities carry - the labels are what let an abbreviated BMS
    token match a spelled-out class. Returns (points, parts)."""
    pts = collections.defaultdict(lambda: {"n": 0, "labels": collections.Counter()})
    parts = collections.defaultdict(lambda: {"n": 0, "labels": collections.Counter()})
    with open(path, encoding="utf-8-sig") as fh:
        for r in csv.reader(fh):
            if len(r) < 5 or not r[4]:
                continue
            lab = _label_on_row(r)
            if r[2] == "brick:hasPoint":
                pts[r[4]]["n"] += 1
                if lab:
                    pts[r[4]]["labels"][lab] += 1
            elif r[2] == "brick:hasPart":
                parts[r[4]]["n"] += 1
                if lab:
                    parts[r[4]]["labels"][lab] += 1
    return pts, parts


def load_ssc_classes(path):
    """Point and part classes the SSC sample uses. Needs openpyxl and the
    workbook; returns two empty indexes when either is missing so the ladder
    simply drops step 3."""
    empty = (collections.defaultdict(lambda: {"n": 0, "labels": collections.Counter()}),
             collections.defaultdict(lambda: {"n": 0, "labels": collections.Counter()}))
    if not path or not Path(path).exists():
        return empty
    try:
        import openpyxl
        from lookup_reference import pick_ontology_sheet
    except Exception:
        return empty
    pts, parts = empty
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = pick_ontology_sheet(wb, path)
    hdr = [str(c or "") for c in next(ws.iter_rows(values_only=True))]
    try:
        pr, ot = hdr.index("predicate"), hdr.index("objectType")
    except ValueError:
        return empty

    def ssc_label(r):
        for i in range(len(hdr)):
            if str(hdr[i]) == "object_prop_name" and i + 1 < len(r) \
                    and str(r[i] or "") == "rdfs:label_en":
                return str(r[i + 1] or "").strip()
        return ""

    for r in ws.iter_rows(min_row=2, values_only=True):
        if len(r) <= ot or not r[ot]:
            continue
        cls = str(r[ot]).strip()
        if str(r[pr] or "") == "brick:hasPoint":
            pts[cls]["n"] += 1
            lab = ssc_label(r)
            if lab:
                pts[cls]["labels"][lab] += 1
        elif str(r[pr] or "") == "brick:hasPart":
            parts[cls]["n"] += 1
    return pts, parts


def load_brick(path=BRICKVOCAB):
    ok = {}
    with open(path) as fh:
        for line in fh:
            if line.startswith("#") or not line.strip():
                continue
            parts = line.rstrip("\n").split("\t")
            ok[parts[0]] = parts[1] if len(parts) > 1 else "OK"
    return ok


# --------------------------------------------------------------------------- scoring
def _score(phrase, cls, labels, kind="", device=""):
    pw = set(words(phrase))
    if not pw:
        return 0.0
    bare = cls.split(":", 1)[-1]
    cw = set(words(bare.replace("_", " ")))
    best = len(pw & cw) / max(len(pw | cw), 1)
    for lab, _ in labels.most_common(8):
        lw = set(words(lab))
        best = max(best, len(pw & lw) / max(len(pw | lw), 1))
    if kind and not bare.endswith(kind):      # kind guard
        best *= 0.30
    if device:                                # device guard
        cd = device_of(bare.replace("_", " "))
        if cd and cd != device:
            best *= 0.30
    return best


class Resolver:
    """Walks Dar Cairo -> Brick 1.4 -> SSC -> para and always returns
    (class, source). Points search the point indexes; parts (equipment=True)
    search the part indexes. Minted para classes accumulate in .minted, keyed
    class -> parent, for the sheet's Extensions layer and the handover."""

    def __init__(self, darcairo=DARCAIRO, ssc_path=SSC_DEFAULT, brickvocab=BRICKVOCAB):
        self.pts_dar, self.parts_dar = load_darcairo_classes(darcairo)
        self.pts_ssc, self.parts_ssc = load_ssc_classes(ssc_path)
        self.brick = load_brick(brickvocab)
        self.minted = {}

    def _search(self, index, phrase, kind, device, thresh):
        best, best_sc = "", 0.0
        for cls, info in index.items():
            sc = _score(phrase, cls, info["labels"], kind, device)
            if sc > best_sc:
                best, best_sc = cls, sc
        return best if best_sc >= thresh else ""

    def _brick_exact(self, phrase, kind):
        stem = "_".join(w.capitalize() for w in words(phrase))
        for cand in ("brick:%s_%s" % (stem, kind), "brick:%s" % stem):
            if self.brick.get(cand) == "OK":
                return cand
        return ""

    def _brick_fuzzy(self, phrase, kind, device):
        best, best_sc = "", 0.0
        for term, st in self.brick.items():
            if st != "OK":
                continue
            s = _score(phrase, term, collections.Counter(), kind, device)
            if s > best_sc:
                best, best_sc = term, s
        return best, best_sc

    def resolve(self, phrase, kind="", device="", equipment=False):
        dar = self.parts_dar if equipment else self.pts_dar
        ssc = self.parts_ssc if equipment else self.pts_ssc
        c = self._search(dar, phrase, kind, device, 0.60)
        if c:
            return c, "Dar Cairo"
        bc = self._brick_exact(phrase, "" if equipment else kind)
        if bc:
            return bc, "Brick 1.4 exact"
        bfuzz, bsc = self._brick_fuzzy(phrase, kind, device)
        if bsc >= 0.70:
            return bfuzz, "Brick 1.4"
        c = self._search(ssc, phrase, kind, device, 0.60)
        if c:
            return c, "SSC"
        if bsc >= 0.55:
            return bfuzz, "Brick 1.4"
        # para: mint. A point subclasses its kind's Brick parent; equipment the
        # closest device parent, else brick:Equipment.
        name = "_".join(w.capitalize() for w in words(phrase)) or "Unknown"
        if equipment:
            cls, parent = "para:%s" % name, "brick:Equipment"
        else:
            cls = "para:%s_%s" % (name, kind) if kind else "para:%s" % name
            parent = KIND_PARENT.get(kind, "brick:Point")
        self.minted.setdefault(cls, parent)
        return cls, "para (minted)"


# --------------------------------------------------------------------------- CLI
def _resolve_token(R, token, desc="", kind=""):
    """Resolve one token the way a build script would: split part.point, pick the
    phrase (usable description, else the expanded token), resolve through the
    ladder. Returns a ledger dict."""
    part, point = (token.split(".", 1) + [""])[:2] if "." in token else ("", token)
    point = point or part
    k = kind or kind_of(point)
    if desc and not is_generic(desc):
        phrase = desc
    else:
        phrase = expand(point if part else token)
    cls, src = R.resolve(phrase, k, "")
    return {"token": token, "kind": k, "class": cls, "source": src,
            "phrase": phrase, "desc": desc}


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--token", help="a single BMS point token, e.g. RmTempSP")
    ap.add_argument("--kind", default="",
                    help="override the point kind (Sensor/Setpoint/Status/Command/Alarm)")
    ap.add_argument("--desc", default="", help="the historian description, if any")
    ap.add_argument("--tokens", help="a file of tokens, one per line, "
                    "optionally  token<TAB>description<TAB>kind")
    ap.add_argument("--ssc", default=str(SSC_DEFAULT),
                    help="path to an SSC sample sheet for ladder step 3 "
                    "(default: the reference-models sample; pass '' to skip)")
    args = ap.parse_args(argv)

    R = Resolver(ssc_path=(args.ssc or None))
    rows = []
    if args.token:
        rows.append(_resolve_token(R, args.token, args.desc, args.kind))
    if args.tokens:
        for line in Path(args.tokens).read_text().splitlines():
            line = line.rstrip("\n")
            if not line.strip() or line.lstrip().startswith("#"):
                continue
            cols = line.split("\t")
            tok = cols[0].strip()
            desc = cols[1].strip() if len(cols) > 1 else ""
            kind = cols[2].strip() if len(cols) > 2 else ""
            rows.append(_resolve_token(R, tok, desc, kind))
    if not rows:
        ap.error("give --token or --tokens")

    w = csv.DictWriter(sys.stdout,
                       fieldnames=["token", "kind", "class", "source", "phrase", "desc"])
    w.writeheader()
    w.writerows(rows)
    if R.minted:
        sys.stderr.write("\nminted para classes (%d) - review before handover:\n"
                         % len(R.minted))
        for c, p in sorted(R.minted.items()):
            sys.stderr.write("    %s  subClassOf  %s\n" % (c, p))
    return 0


if __name__ == "__main__":
    sys.exit(main())
