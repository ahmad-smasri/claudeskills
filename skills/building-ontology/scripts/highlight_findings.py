#!/usr/bin/env python3
"""Fill every row that still carries a finding yellow, for a manual pass.

    python3 highlight_findings.py In.xlsx --out Reviewed.xlsx --label-style verbatim
    python3 highlight_findings.py In.xlsx --out Reviewed.xlsx --io IO_List.xlsx

Some findings need a human. This writes a copy of the workbook with those rows
filled `#FFFF00` and the finding text attached as a cell comment on the subject,
so the reviewer works in the sheet rather than alongside a separate report.

Each flagged row also gets the finding written into two columns past the data -
`validator_code` and `validator_finding` - so the reviewer can read, sort and
filter on it instead of hovering over every cell. They sit beyond the 27-column
contract, so the converter still reads the sheet, but **delete them and the fills
before handover**: the deliverable holds triples and nothing else.

**It writes a copy and never touches the input**, and it adds no sheet: the
review marks live on the ontology rows themselves, which is the one place they
belong.

File-level findings - a type clash, a terminal unit with no feeds - name an
entity rather than a row. Those are resolved to every row where that entity is
the subject, so nothing is reported at a row number the reviewer cannot find.
"""
from __future__ import annotations

import argparse
import collections
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from validate_ontology import Report, read_sheet, pick_ontology_sheet, validate  # noqa: E402
import check_consistency                                                        # noqa: E402

YELLOW = "FFFF00"
ENTITY = re.compile(r"entity:[A-Za-z0-9_\-.]+")
# The consistency checker names units without the prefix - "absent on SSC_AHUB0001,
# SSC_AHUB0002 and 3 more" - so fall back to bare names that the sheet knows.
NAKED = re.compile(r"\b[A-Za-z][A-Za-z0-9]*(?:[_\-][A-Za-z0-9]+){1,}\b")


def collect(sheet: Path, label_style: str, io) -> Report:
    report = Report(10 ** 6)
    validate(sheet, report, label_style, io)
    check_consistency.run(sheet, report, None, io)
    return report


def first_row_of(sheet: Path) -> dict[str, int]:
    """Each entity's defining row - the first row where it is the subject.

    A file-level finding is placed there and nowhere else. Marking every row an
    entity owns would paint 300 rows yellow for 15 findings and bury the ones
    that point at a single cell.

    Indexed by full name and by local name, because the consistency checker
    names units without the `entity:` prefix.
    """
    header, body = read_sheet(sheet)
    low = [h.strip().lower() for h in header]
    i = low.index("subject")
    out: dict[str, int] = {}
    for rownum, r in body:
        subj = r[i].strip() if i < len(r) else ""
        if subj:
            out.setdefault(subj, rownum)
            out.setdefault(subj.split(":", 1)[-1], rownum)
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("sheet", type=Path)
    ap.add_argument("--out", type=Path, required=True, help="where to write the copy")
    ap.add_argument("--label-style", choices=("para", "verbatim"), default="para")
    ap.add_argument("--io", type=Path, help="IO list, used as evidence")
    ap.add_argument("--ignore", default="",
                    help="comma-separated rule codes to leave unhighlighted")
    ap.add_argument("--severity", default="ERROR,WARN",
                    help="which severities to highlight (default ERROR,WARN)")
    ap.add_argument("--no-annotate", action="store_true",
                    help="fill the rows but do not add the validator_code and "
                         "validator_finding columns")
    args = ap.parse_args()

    if not args.sheet.exists():
        print(f"no such file: {args.sheet}", file=sys.stderr)
        return 3
    try:
        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("openpyxl is needed - pip install openpyxl")

    io = None
    if args.io:
        import io_list
        io = io_list.load(args.io)
        print(f"IO list: {io.describe()}")

    report = collect(args.sheet, args.label_style, io)
    ignore = {c.strip() for c in args.ignore.split(",") if c.strip()}
    want = {s.strip().upper() for s in args.severity.split(",") if s.strip()}
    by_entity = first_row_of(args.sheet)

    marks: dict[int, list[str]] = collections.defaultdict(list)
    unplaced = 0
    for severity, code, row, message in report.all:
        if code in ignore or severity.upper() not in want:
            continue
        if row:
            marks[row].append(f"{code}: {message}")
            continue
        # file-level: place it on the defining row of every entity it names
        names = ENTITY.findall(message) or NAKED.findall(message)
        targets = {by_entity[n] for n in names if n in by_entity}
        if not targets:
            unplaced += 1
            continue
        for r in sorted(targets):
            marks[r].append(f"{code}: {message}")

    wb = openpyxl.load_workbook(args.sheet)
    ws = pick_ontology_sheet(wb, args.sheet)
    fill = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
    width = min(ws.max_column, 27)

    # Two columns past the data, so the 27-column contract still holds.
    col_code, col_text = width + 1, width + 2
    if not args.no_annotate:
        ws.cell(1, col_code, "validator_code").font = Font(bold=True)
        ws.cell(1, col_text, "validator_finding").font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_code)].width = 12
        ws.column_dimensions[get_column_letter(col_text)].width = 110

    for row, notes in sorted(marks.items()):
        if row > ws.max_row:
            continue
        for col in range(1, width + 1):
            ws.cell(row, col).fill = fill
        seen = sorted(set(notes))
        text = "\n".join(seen[:12])
        ws.cell(row, 1).comment = Comment(text[:3000], "ontology validator",
                                          height=220, width=520)
        if args.no_annotate:
            continue
        codes = sorted({n.split(":", 1)[0] for n in seen})
        # Strip the repeated "CODE: " prefix - the code has its own column - and
        # the family prefix the consistency checker adds, which is already
        # obvious from the row.
        bodies = [re.sub(r"^[EWI]-[A-Z]+-\d+:\s*(?:[a-z]+:[A-Za-z_]+:\s*)?", "", n)
                  for n in seen]
        for col, val in ((col_code, ", ".join(codes)),
                         (col_text, "  |  ".join(bodies)[:2000])):
            c = ws.cell(row, col, val)
            c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=(col == col_text))
    wb.save(args.out)

    codes = collections.Counter(
        n.split(":", 1)[0] for notes in marks.values() for n in set(notes))
    print(f"\n{len(marks)} rows highlighted in {ws.title}, written to {args.out}")
    for code, n in sorted(codes.items()):
        print(f"  {code:<10} on {n} row(s)")
    if unplaced:
        print(f"  {unplaced} finding(s) name no entity in this sheet and were not placed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
